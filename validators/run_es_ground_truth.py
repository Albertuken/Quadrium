"""
`OQ-S-05`: the pilot's answer was already published, and the pilot was 9.8 points
wrong. Both checked here rather than remembered.

The entry was answered at v1.8 and never validated. It carries the most
uncomfortable measurement in the project — the Spanish pilot's central estimate
against the truth — and until now that measurement lived only in prose, where it
could quietly stop being true.

WHAT THE INE PUBLISHES, AND AT HOW MANY LEVELS
------------------------------------------------
The question assumed two levels: 91 products compiled, 64 published
(`NSO_ES_01` p. 5). There are more than that, and the workbook says so itself:
**`cne_tod_22.xlsx` ships two different product correspondence tables.**

    Tabla 8 — correspondences for Tablas 1 and 2       118 rows
    Tabla 9 — correspondences for Tablas 3, 4, 5, 6, 7  76 rows

Two classifications in one publication. Loading them confirms the sizes: the
supply table and the purchasers'-price use table come at **110 products**; the
basic-price, domestic and import use tables come at **64**.

THE PILOT'S NUMBER WAS PUBLISHED ALL ALONG
--------------------------------------------
At 110 products `Servicios de alojamiento` and `Servicios de comidas y bebidas`
are separate:

    30,717.7  +  97,548.8  =  128,266.5

and 128,266.5 is the output of product 36 in the 64-product IOT, **to the last
decimal**. Not a close match — the same quantity, disaggregated by the office.
So the true accommodation share is `23.95 %`, and it was downloadable.

AND THE KEY WITH THE BEST CONCEPTUAL MATCH WAS THE THIRD WORST
----------------------------------------------------------------
The pilot split on `production`, output against output — the tightest conceptual
fit available. Against the truth:

    employment           21.24 %   −2.7 pp     ← loosest match, best result
    hours worked         27.65 %   +3.7
    purchases            29.80 %   +5.9
    personnel cost       32.76 %   +8.8
    production           33.73 %   +9.8        ← what the pilot chose
    value added          39.84 %  +15.9
    gross op. surplus    55.19 %  +31.2

The reason is population, not concept: an accommodation *enterprise* produces a
great deal of food-service *product*, and the product classification puts that in
`74`. **Choosing a proxy on conceptual proximity is not the same as choosing it
on population match**, and the second is what governs.

**What the tool got right, which is the reason to keep it.** The report printed
21 %–55 % across seven keys and called that the error bar. The truth sits inside
it. A tool reporting one number would have been confidently wrong; this one was
honestly uncertain over a range containing the answer.

THE RESIDUE, NAMED PRECISELY
------------------------------
`Tabla 4 — Tabla de destino de la producción interior a precios básicos` is the
table the engine actually loads, and the workbook publishes it under **Tabla 9's
correspondence, the 64-product one**. So the finer split exists for supply and
for purchasers'-price use, and not for the one matrix the engine needs. That is
the whole of what remains of this question, and it is a data request rather than
a research problem.

Run:
    python3 validators/run_es_ground_truth.py
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

TOD = ROOT / "data" / "ine" / "cne_tod_22.xlsx"
TIO = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
FAIL: list[str] = []

# The seven survey keys, as `examples/es_hosteleria.py` registers them. These
# are transcribed from INE survey tables 76811 and 76815 and cannot be checked
# against a file the project holds — what IS checked here is every conclusion
# drawn from them.
SURVEY_KEYS = {
    "employment":            (342_956, 1_271_964),
    "hours worked":          (556_585, 1_456_295),
    "purchases":             (15_863_771, 37_368_585),
    "personnel cost":        (9_444_965, 19_384_933),
    "production":            (31_167_086, 61_233_490),
    "value added":           (16_403_219, 24_768_118),
    "gross operating surplus": (6_634_659, 5_387_752),
}
PILOT_CHOICE = "production"


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def main() -> int:
    if not (TOD.exists() and TIO.exists()):
        print("INE fixtures absent")
        return 0

    import openpyxl
    from quadrium.io_loader import load_ine_tod, load_ine_tio

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    # ---- two classifications in one publication ---------------------------
    wb = openpyxl.load_workbook(TOD, read_only=True, data_only=True)
    n8, n9 = wb["Tabla8"].max_row, wb["Tabla9"].max_row
    check("the INE ships two product correspondence tables in one workbook",
          n8 > n9 + 30,
          f"Tabla 8 ({n8} rows, for Tablas 1–2) against Tabla 9 ({n9} rows, for "
          f"Tablas 3–7). The question assumed one published level; the "
          f"publication itself carries two")

    # ---- the split is published -------------------------------------------
    s = load_ine_tod(TOD)
    idx = {str(l).strip(): i for i, l in enumerate(s.product_labels)}
    acc = idx["Servicios de alojamiento"]
    food = idx["Servicios de comidas y bebidas"]
    q_acc, q_food = float(s.q[acc]), float(s.q[food])

    t = load_ine_tio(TIO, variant="interior", unbalanced="residual_column")
    parent = [i for i, l in enumerate(t.sector_labels)
              if "alojamiento y de comidas" in str(l)][0]
    q_parent = float(t.X[parent])

    print()
    print(f"    supply table, {len(s.product_labels)} products:")
    print(f"      Servicios de alojamiento          {q_acc:>12,.1f}")
    print(f"      Servicios de comidas y bebidas    {q_food:>12,.1f}")
    print(f"      sum                               {q_acc + q_food:>12,.1f}")
    print(f"    IOT, {t.n} products:")
    print(f"      producto 36                       {q_parent:>12,.1f}")

    check("the two 110-level products sum to the 64-level product exactly",
          abs(q_acc + q_food - q_parent) < 0.05,
          f"difference {abs(q_acc + q_food - q_parent):.4f} on "
          f"{q_parent:,.1f} — the same quantity, disaggregated by the office. "
          f"The pilot's target was downloadable the whole time")

    truth = q_acc / q_parent
    check("and the true accommodation share is 23.95 %",
          abs(truth - 0.2395) < 5e-4,
          f"{truth:.4%}, OBSERVED — not an estimate, and the benchmark every "
          f"key below is scored against")

    # ---- the keys, scored -------------------------------------------------
    scored = sorted(((v55 / (v55 + v56), name)
                     for name, (v55, v56) in SURVEY_KEYS.items()))
    print()
    print(f"    {'survey key':<26}{'share':>9}{'error':>10}")
    for share, name in scored:
        mark = "   <- the pilot's choice" if name == PILOT_CHOICE else ""
        print(f"    {name:<26}{share:>8.2%}{100 * (share - truth):>+9.1f} pp"
              f"{mark}")

    errors = {name: abs(share - truth) for share, name in scored}
    ranked = sorted(errors, key=errors.get)
    check("the key with the best conceptual match was the third worst",
          ranked.index(PILOT_CHOICE) == 4 and ranked[0] == "employment",
          f"'{PILOT_CHOICE}' — output against output — ranks "
          f"{ranked.index(PILOT_CHOICE) + 1} of {len(ranked)} at "
          f"{100 * errors[PILOT_CHOICE]:+.1f} pp, while 'employment', the "
          f"loosest match, is closest at "
          f"{100 * (dict((n, s) for s, n in scored)['employment'] - truth):+.1f} pp. "
          f"Population match governs, not conceptual proximity")

    shares = [s_ for s_, _ in scored]
    check("but the range the report printed does contain the truth",
          min(shares) < truth < max(shares),
          f"{min(shares):.0%}–{max(shares):.0%} across seven keys, and "
          f"{truth:.2%} sits inside it. A tool reporting a single number would "
          f"have been confidently wrong; this one was honestly uncertain")

    # ---- the residue ------------------------------------------------------
    check("the residue is one table, and it is a data request not a research "
          "problem",
          t.n == 64 and len(s.product_labels) == 110,
          "Tabla 4, `destino de la producción interior a precios básicos` — "
          "the matrix the engine loads — is published under Tabla 9's "
          "64-product correspondence. The finer split exists for supply and "
          "for purchasers'-price use, and not for the one matrix that needs it")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
