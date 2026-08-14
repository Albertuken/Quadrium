"""
Run the identity validators against the project's real table:
    <project root>/UK_IOAT_2023_domestic_ixi.xlsx

CAVEAT ON THAT FILE, established by reading it:
  * it is not a CSV. It is an XLSX workbook (PK zip header, Excel 2007+);
  * its Menu sheet reads "United Kingdom Input-Output Analytical Tables, 2023
    (consistent with UK National Accounts Blue Book 2025 & UK Balance of
    Payments Pink Book 2025)". The reference year is 2023, not 2022;
  * the IOT sheet is "Input-Output table (domestic use, basic prices,
    industry by industry), GBP million" -- so it is an IOT_IXI, DOMESTIC,
    at BASIC PRICES. It is not a supply-use pair.

Consequences for what can be tested here:
  * ID-11 (IOT balance) and ID-12 (Leontief) apply directly.
  * ID-02 and ID-03 apply to the value-added block.
  * ID-01, ID-07, ID-08, ID-10 need a SUPPLY table and margin/tax columns.
    This file has none: margins have already been reallocated by the time an
    analytical IOT at basic prices exists. Those identities are reported as
    NOT APPLICABLE rather than silently skipped.

Usage:
    python3 run_uk_iot.py [path-to-workbook]
"""

from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import quadrium.identities as ID  # noqa: E402

DEFAULT = (Path(__file__).resolve().parents[1] / "UK_IOAT_2023_domestic_ixi.xlsx")


def load_iot(path: Path):
    """Read the IOT sheet. openpyxl refuses a .csv extension, so copy to a
    temporary .xlsx first rather than renaming the user's file."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "wb.xlsx"
        shutil.copy(path, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        # The Menu banner is not in column A; scan the first few columns.
        menu = [next((c for c in r if c), None) for r in wb["Menu"].iter_rows(
            min_row=1, max_row=6, max_col=8, values_only=True)]
        R = [list(r) for r in wb["IOT"].iter_rows(values_only=True)]

    def num(v):
        return float(v) if isinstance(v, (int, float)) else np.nan

    # Layout, established by inspection (0-based indices):
    #   row 3  : SIC codes across;      row 4  : industry names
    #   rows 6..109  : 104 industry rows of the domestic Z matrix
    #   cols 2..105  : 104 industry columns
    #   col 106 : '_T' total intermediate use at basic prices
    #   cols 107..116 : final demand;  col 117 : 'TU' total use
    #   row 110 : total intermediate use at basic prices
    #   row 111 : use of imported products, cif
    #   row 112 : taxes less subsidies on products
    #   row 113 : total intermediate use at purchaser's prices
    #   row 114 : D1 compensation of employees
    #   row 115 : gross operating surplus and mixed income
    #   row 116 : D29-D39 taxes less subsidies on production
    #   row 117 : GVA
    #   row 118 : P1 total output at basic prices
    rows = range(6, 110)
    cols = range(2, 106)

    d = {
        "title": R[1][0],
        "menu": [str(m) for m in menu if m],
        "codes": [R[i][0] for i in rows],
        "names": [R[i][1] for i in rows],
        "Z": np.array([[num(R[i][j]) for j in cols] for i in rows]),
        "TU": np.array([num(R[i][117]) for i in rows]),
        "imports": np.array([num(R[111][j]) for j in cols]),
        "taxes_products": np.array([num(R[112][j]) for j in cols]),
        "int_purch": np.array([num(R[113][j]) for j in cols]),
        "compensation": np.array([num(R[114][j]) for j in cols]),
        "gos_mi": np.array([num(R[115][j]) for j in cols]),
        "other_taxes": np.array([num(R[116][j]) for j in cols]),
        "gva": np.array([num(R[117][j]) for j in cols]),
        "x": np.array([num(R[118][j]) for j in cols]),
    }
    fd_codes = [R[3][j] for j in range(107, 117)]
    d["fd_codes"] = fd_codes
    d["FD"] = {code: np.array([num(R[i][107 + k]) for i in rows])
               for k, code in enumerate(fd_codes)}
    return d


def main(path: Path):
    d = load_iot(path)
    Z, x = d["Z"], d["x"]

    print("=" * 78)
    print("UK Input-Output table -- identity validation")
    print("=" * 78)
    print(f"file      : {path.name}")
    for m in d["menu"]:
        print(f"menu      : {m}")
    print(f"sheet     : {d['title']}")
    print(f"dimensions: Z is {Z.shape[0]} x {Z.shape[1]} industries")
    print()

    # 'P3 S1' is the total of S13+S14+S15; using all four would double count.
    F = d["FD"]
    assert np.abs(F["P3 S1"] - (F["P3 S13"] + F["P3 S14"] + F["P3 S15"])).max() < 1e-6, \
        "P3 S1 is not the sum of the household/government/NPISH columns"
    fd_used = ["P3 S1", "P51G", "P52", "P53", "P61EU", "P61RW", "P62"]
    y = sum(F[k] for k in fd_used)

    results = []

    # ID-11 -- the defining property of an IOT.
    results.append(ID.id11_iot_balance(Z, y, d["imports"] + d["taxes_products"] + d["gva"], x))

    # ID-02 -- column identity. In this DOMESTIC table the column also carries
    # imported intermediate use and taxes less subsidies on products, both of
    # which sit outside Z.
    results.append(ID.id02_industry_identity(
        Z.sum(0) + d["imports"] + d["taxes_products"], d["gva"], x))

    # ID-03 -- value-added decomposition.
    results.append(ID.id03_value_added_decomposition(
        d["gva"], [d["compensation"], d["gos_mi"], d["other_taxes"]]))

    # ID-14 -- total intermediate demand vs total intermediate consumption.
    results.append(ID.id14_intermediate_totals(Z, Z.sum(0)))

    # ID-12 -- Leontief. Computed from the DOMESTIC table, as required by
    # CORE_006 par. 9.58, p. 289.
    r12 = ID.id12_leontief(Z, x, y)
    results.append(r12)

    for r in results:
        print(r)
        print()

    # Identities that need objects this file does not contain.
    print("-" * 78)
    print("NOT APPLICABLE to this file (no supply table, no margin or tax columns):")
    for iid, nm in [("ID-01", "product balance"),
                    ("ID-07", "q and g from the supply matrix V"),
                    ("ID-08", "margins sum to zero"),
                    ("ID-09", "margins by product vs by industry"),
                    ("ID-10", "CIF/FOB adjustment"),
                    ("ID-13", "value added preserved under transformation")]:
        print(f"   {iid}  {nm}")
    print("   This is an analytical IOT at basic prices: trade and transport")
    print("   margins have already been reallocated, so the zero-sum margin")
    print("   structure of CORE_003 par. 15.56, p. 488 is no longer visible.")
    print()

    # ------------------------------------------------------------------
    # Negative census. The point of this section: negatives are expected.
    # ------------------------------------------------------------------
    print("-" * 78)
    print("NEGATIVE-VALUE CENSUS  (see A_core_accounting_spec.md A.8.1)")
    print("-" * 78)
    census = ID.negative_census(
        Z=Z,
        taxes_less_subsidies_on_products=d["taxes_products"],
        other_taxes_less_subsidies_on_production=d["other_taxes"],
        gross_operating_surplus_and_mixed_income=d["gos_mi"],
        imports_of_intermediate_products=d["imports"],
        **{f"FD_{k}": v for k, v in F.items() if k != "P3 S1"},
    )
    for k, v in census.items():
        if v["n_negative"]:
            print(f"  {k:52s} n={v['n_negative']:5d}  min={v['min']:>14,.2f}  "
                  f"sum={v['sum_negative']:>14,.2f}")
    print()

    i, j = np.unravel_index(np.argmin(Z), Z.shape)
    print(f"  largest negative in Z: {Z[i, j]:,.2f}")
    print(f"     row {d['codes'][i]}  {str(d['names'][i]).strip()}")
    print(f"     col {d['codes'][j]}  {str(d['names'][j]).strip()}")
    print()

    A = r12.info["A"]
    print(f"  negative coefficients in A : {int((A < 0).sum())}")
    print(f"  max column sum of A        : {A.sum(0).max():.6f}  "
          f"(<1 required for a viable economy -- project test, not cited)")
    print()

    print("-" * 78)
    n_fail = sum(1 for r in results if not r.passed)
    print(f"{len(results) - n_fail}/{len(results)} identities passed.")
    print("Convergence and identity satisfaction are NECESSARY BUT NOT")
    print("SUFFICIENT for statistical validity (CORE_006 par. 9.51, p. 288).")
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    if not p.exists():
        sys.exit(f"not found: {p}")
    sys.exit(main(p))
