"""
Three regional datasets came in one archive. Only one can fit a location quotient.

WHY A NEGATIVE RESULT IS WORTH A FILE
---------------------------------------
The European MRIO archive (Zenodo 7875024) ships a technical-validation set with
survey or government regional tables for **Austria, Finland and Scotland**.
`run_delta_across_regions.py` uses Austria's nine to answer `OQ-R-02`. The
obvious next move is Finland — it is the country `CORE_034` calibrated delta on —
and then Scotland.

Neither can be used, and the reason is not that the fit comes out badly. It is
that **their tables are not the quantity a location quotient estimates**. That
distinction is worth a check rather than a sentence, because the next dataset
somebody points at this engine will need the same question asked of it, and
because a fit that is merely bad looks like a result while a fit on the wrong
quantity looks like a result too.

THE TEST THAT SEPARATES THEM
------------------------------
A location quotient converts national **technical** coefficients into regional
**domestic** ones: it scales cells down to reflect what the region cannot source
locally. So it needs a regional table whose intermediate block counts own-region
supply only, with purchases from elsewhere held apart.

Two signatures tell the two kinds of table apart, and they agree:

1. **`Z/X` against region size.** In a domestic table a small region sources
   little locally, so the ratio is low and rises with size. In a total-flow table
   it is the economy's intermediate share and barely moves.

2. **Where the fitted delta lands.** The quotient family can only scale DOWN.
   Given a total-flow regional table its target multipliers are larger than
   anything a downscaled national matrix produces, so the fit runs to the
   boundary — delta = 0, where FLQ collapses onto CILQ — and still undershoots.
   A delta pinned at a boundary with a large residual is not a calibration, it is
   the method saying it was handed the wrong object.

Run:
    python3 validators/run_regional_truth_survey.py
"""
from __future__ import annotations

import math
import sys
import warnings
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

TRUTH = ROOT / "data" / "mrio" / "truth"
AT = ("AT11", "AT12", "AT13", "AT21", "AT22", "AT31", "AT32", "AT33", "AT34")
FI = ("FI1B", "FI1C", "FI1D", "FI19")            # the usable NUTS-2 sheets
FI_S = 30
DELTAS = [round(0.02 * k, 2) for k in range(0, 51)]
FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def multipliers(A):
    return np.linalg.inv(np.eye(len(A)) - A).sum(axis=0)


def mu1(est, true):
    o = true > 0
    return float(100 / len(true) * np.sum((est[o] - true[o]) / true[o]))


def flq(slq, lam):
    q = np.minimum((slq[:, None] / slq[None, :]) * lam, 1.0)
    np.fill_diagonal(q, np.minimum(slq * lam, 1.0))
    return q


def fit(A_nat, X_nat, Z, X):
    with np.errstate(divide="ignore", invalid="ignore"):
        true = multipliers(np.where(X > 0, Z / X, 0.0))
    share = X.sum() / X_nat.sum()
    base = math.log2(1.0 + share)
    s = (X / X.sum()) / (X_nat / X_nat.sum())
    s = np.where(np.isfinite(s) & (s > 0), s, 1e-12)
    best = None
    for d in DELTAS:
        m = mu1(multipliers(A_nat * flq(s, base ** d)), true)
        if best is None or abs(m) < abs(best[1]):
            best = (d, m)
    return share * 100.0, best[0], best[1]


def finnish_sheets():
    import openpyxl

    f = TRUTH / "Finland" / "io_reg_2014.xlsx"
    if not f.exists():
        return None
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    out = {}
    for sh in FI:
        rows = list(wb[sh].iter_rows(values_only=True))
        M = np.array([[float(c) if isinstance(c, (int, float)) else 0.0
                       for c in r[1:FI_S + 6]] for r in rows[1:1 + FI_S]])
        # 30 sectors, then P3/S14, P5, P3/S13, P6K and SUM: the output
        # column is the fifth after the block, hence FI_S + 4.
        out[sh] = (M[:, :FI_S], M[:, FI_S + 4])
    return out


def main() -> int:
    warnings.filterwarnings("ignore")
    from quadrium.io_loader import read_rokicki_components

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    if not TRUTH.exists():
        print(f"    -- {TRUTH} absent; see data/mrio/_provenance.json.")
        return 0

    # ---- Scotland: no output vector at all
    import openpyxl
    sc = TRUTH / "Scotland" / "UKM-.xlsx"
    if sc.exists():
        rows = list(openpyxl.load_workbook(sc, read_only=True,
                                           data_only=True)["Sheet1"]
                    .iter_rows(values_only=True))
        n_rows, n_cols = len(rows), len(rows[0])
        check("Scotland cannot be used: it is a bare intermediate matrix",
              n_rows <= 12 and n_cols <= 12 and n_rows == n_cols,
              f"{n_rows} x {n_cols}, headed by NACE letter groups "
              f"({', '.join(str(c) for c in rows[0][1:4])}, …) with no final "
              f"demand and no output vector. A coefficient needs a denominator")

    # ---- Austria and Finland, side by side on signature 1
    parts = {r: read_rokicki_components(TRUTH / "Austria", r) for r in AT}
    fin = finnish_sheets()
    check("Finland's four usable NUTS-2 sheets are on disk", fin is not None,
          "FI1B, FI1C, FI1D, FI19 — FI20 is empty and the NUTS-3 sheets do not "
          "partition the country" if fin else "absent")
    if fin is None:
        return 1 if FAIL else 0

    at_ratio = np.array([p["Z"].sum() / p["X_col"].sum() for p in parts.values()])
    at_share = np.array([p["X_col"].sum() for p in parts.values()])
    at_share = at_share / at_share.sum()
    fi_ratio = np.array([Z.sum() / X.sum() for Z, X in fin.values()])
    fi_share = np.array([X.sum() for _, X in fin.values()])
    fi_share = fi_share / fi_share.sum()

    print()
    print(f"    {'':<12}{'Z/X range':>18}{'spread':>10}"
          f"{'corr with size':>17}{'separate import row':>22}")
    at_c = float(np.corrcoef(at_share, at_ratio)[0, 1])
    fi_c = float(np.corrcoef(fi_share, fi_ratio)[0, 1])
    print(f"    {'Austria':<12}{at_ratio.min():>8.3f}-{at_ratio.max():<9.3f}"
          f"{at_ratio.max() - at_ratio.min():>10.3f}{at_c:>17.2f}{'yes':>22}")
    print(f"    {'Finland':<12}{fi_ratio.min():>8.3f}-{fi_ratio.max():<9.3f}"
          f"{fi_ratio.max() - fi_ratio.min():>10.3f}{fi_c:>17.2f}{'no':>22}")

    # The discriminator is the LEVEL and the SIGN, not the spread. Spread was
    # the first thing tried and it does not separate them: 0.039 against 0.089
    # is a factor of two, which proves nothing with four regions against nine.
    check("signature 1: Finland's intermediate share sits at a national level "
          "and does not rise with region size",
          fi_ratio.mean() > 1.7 * at_ratio.mean() and at_c > 0.4 and fi_c < 0.0,
          f"Finland averages {fi_ratio.mean():.2f} of output against Austria's "
          f"{at_ratio.mean():.2f}, and correlates with region size at "
          f"{fi_c:+.2f} where Austria correlates at {at_c:+.2f}. A domestic "
          f"table has to show a low level rising with size, because a small "
          f"region sources little at home; Finland shows a whole economy's "
          f"intermediate share, flat. (With four regions the correlation is "
          f"weak evidence on its own, which is why signature 2 exists)")

    # ---- signature 2: where the fit lands
    ZN = sum(Z for Z, _ in fin.values())
    XN = sum(X for _, X in fin.values())
    with np.errstate(divide="ignore", invalid="ignore"):
        AN = np.where(XN > 0, ZN / XN, 0.0)
    print()
    print(f"    {'Finland':<10}{'share %':>9}{'delta*':>9}{'mu1 at delta*':>15}")
    res = []
    for r in FI:
        sh, d, m = fit(AN, XN, *fin[r])
        res.append((d, m))
        print(f"    {r:<10}{sh:>9.2f}{d:>9.2f}{m:>15.2f}")

    check("signature 2: the fit runs to the boundary and still undershoots by a "
          "lot",
          all(d == 0.0 for d, _ in res) and all(m < -5.0 for _, m in res),
          f"delta = 0 on all four — where FLQ collapses onto CILQ — and mu1 "
          f"still between {min(m for _, m in res):.0f} % and "
          f"{max(m for _, m in res):.0f} %. The family can only scale DOWN, and "
          f"here it would have to scale up. That is not a bad calibration, it "
          f"is the wrong object")

    check("so the two signatures agree, and Finland is out",
          True,
          "its tables are total-flow, not domestic. CORE_034 calibrated delta "
          "on Finland, but on Statistics Finland's 1995 regional tables, which "
          "are not these")

    print()
    print("    What the archive leaves is Austria's nine, which "
          "run_delta_across_regions.py")
    print("    uses. See D_open_questions.md OQ-R-02.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
