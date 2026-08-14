"""
`OQ-D-02`, answered by the office's own published table — twenty-seven years of it.

THE QUESTION, AND WHY IT SURVIVED FIFTEEN VERSIONS
----------------------------------------------------
The project's UK fixture has one unexplained negative in `Z`: financial services
(`CPA_K64`) into owner-occupiers' housing (`L68A`), −20,770.99 million. Every
round asked *why*, of documents:

  * v1.20 narrowed it to an outlier against four other publishers;
  * v1.32/v1.38 established the quantity is structurally sign-free;
  * v1.39 read the ONS's own FISIM article and **excluded** the one documented
    mechanism for a legitimate sign flip, default-risk adjustment;
  * v1.56 found the term "FISIM" had been retired by the 2025 SNA and that the
    substance lives in CORE_002 ch. 7, which states at ¶7.184, p. 236 that these
    estimates "may be characterised by negative estimates … particularly for
    depositors, but also for borrowers" when reference rates are volatile —
    the first rank-1 mechanism, recorded as a candidate, not a cause.

Nobody had looked at **the cell itself, in the source table, in other years.**
The ONS publishes its supply-use tables from 1997 to 2023 in one workbook. It is
the same office, the same cell, twenty-seven times.

WHAT THE SERIES SAYS
----------------------
    1997-2007   +2,636 to +5,141      about 1 % of the column
    2008        **+2**                the previous rate shock
    2009-2021   +13,607 to +27,174    2-5 % of the column
    2022        **+1,912**
    2023        **-20,814**

**It crosses zero once in twenty-seven years, and 2023 is the year.** Two things
follow that no methodology document could have given:

1. **The negative is in the supply-use table, not made by the transformation.**
   −20,814 in the published SUT against −20,770.99 in the analytical IOT. The
   project had ruled the transformation out by argument (model D cannot create
   it); this is the same conclusion OBSERVED.
2. **It is not a convention.** Twenty-six positive years say the ONS does not
   normally book this cell negative, so "that is just how they do it" is
   excluded — and so is "the cell is meaningless".

And the shape matches the mechanism. The two collapses in the series are 2008
and 2022–2023, the two episodes of sharp movement in UK policy rates within its
span — which is the condition CORE_002 ¶7.184, p. 236 names. That the second one
crosses zero and the first only reaches +2 is what the paragraph describes
happening.

WHAT IS CLAIMED, AND WHAT IS NOT
----------------------------------
Claimed: the negative is in the source data, it is a one-year event rather than
a practice, and its timing fits the only rank-1 mechanism the project has found.
**Not claimed: that the reference-rate convention produced this number.** No ONS
document says so, the join between the SNA's mechanism and this office's series
is this project's inference, and the compilation record that would settle it is
internal. What has changed is that the question is no longer *"is this cell
wrong?"* — a question no reader could act on — but *"is 2023 the year the rate
shock passed through?"*, which the office can answer and which its own series
already frames.

Run:
    python3 validators/run_uk_sut_fisim_series.py
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "data" / "ons" / "NSO_UK_04_ONS_supply_use_tables_BB25.xlsx"

PRODUCT = "CPA_K64"        # financial services, except insurance and pensions
INDUSTRY = "L68A"          # owner-occupiers' housing
IOAT_CELL = -20770.99      # what the analytical IOT carries, OQ-D-02

FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def series() -> dict[int, tuple[float, float]]:
    """{year: (the cell, the column's total intermediate consumption)}."""
    import openpyxl

    wb = openpyxl.load_workbook(BOOK, read_only=True, data_only=True)
    out: dict[int, tuple[float, float]] = {}
    for name in wb.sheetnames:
        if not name.startswith("Table 2 - Int Con "):
            continue
        year = int(name.rsplit(" ", 1)[1])
        rows = list(wb[name].iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows[3]]
        if INDUSTRY not in header:
            continue
        col = header.index(INDUSTRY)
        row = next((r for r in rows[4:]
                    if r and str(r[0]).strip() == PRODUCT), None)
        if row is None or not isinstance(row[col], (int, float)):
            continue
        total = sum(r[col] for r in rows[4:]
                    if r and isinstance(r[col], (int, float)))
        out[year] = (float(row[col]), float(total))
    return dict(sorted(out.items()))


def main() -> int:
    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    if not BOOK.exists():
        print(f"\n    {BOOK.name} is not in data/ons — nothing to do.")
        return 0

    s = series()
    print(f"\n    {PRODUCT} x {INDUSTRY}, ONS supply-use tables, Blue Book 2025")
    print(f"    {'year':>6}{'£ million':>13}{'column total':>15}{'share':>9}")
    for year, (v, tot) in s.items():
        mark = "   <-- crosses zero" if v < 0 else ""
        print(f"    {year:>6}{v:>13,.0f}{tot:>15,.0f}{v / tot * 100:>8.1f}%{mark}")

    negatives = [y for y, (v, _) in s.items() if v < 0]

    check("the negative is in the ONS's own supply-use table, not made by the "
          "transformation",
          2023 in s and s[2023][0] < 0
          and abs(s[2023][0] - IOAT_CELL) / abs(IOAT_CELL) < 0.01,
          f"the published SUT carries {s[2023][0]:,.0f} where the analytical "
          f"IOT carries {IOAT_CELL:,.2f} — within "
          f"{abs(s[2023][0] - IOAT_CELL) / abs(IOAT_CELL):.2%}. OBSERVED, "
          f"where the project had ruled the transformation out by argument")

    check("and it is not a convention: the cell is positive in every other year",
          negatives == [2023],
          f"{len(s)} years, {min(s)}–{max(s)}, negative in "
          f"{', '.join(str(y) for y in negatives)} only — so 'that is how the "
          f"office books it' is excluded, and so is 'the cell is meaningless'")

    prior = [v for y, (v, _) in s.items() if y < 2023]
    check("the 2023 value is a reversal, not a drift",
          min(prior) > 0 and s[2023][0] < -0.5 * max(prior),
          f"the range 1997–2022 is +{min(prior):,.0f} to +{max(prior):,.0f}; "
          f"2023 is {s[2023][0]:,.0f}")

    # The two collapses, and they are the two rate shocks in the span.
    lows = sorted(((v, y) for y, (v, _) in s.items()), key=lambda t: t[0])[:3]
    check("the only two collapses in the series are 2008 and 2022–2023",
          {y for _, y in lows} <= {2008, 2022, 2023},
          f"the three smallest values are "
          + ", ".join(f"{y} ({v:,.0f})" for v, y in lows)
          + " — the condition CORE_002 ¶7.184, p. 236 names is volatile "
            "reference rates, and those are the two episodes in this span")

    print()
    print("    NOT claimed: that the reference-rate convention produced this")
    print("    number. No ONS document says so and the join is this project's")
    print("    inference. What changed is the question — from 'is this cell")
    print("    wrong?', which no reader can act on, to 'is 2023 the year the")
    print("    rate shock passed through?', which the office can answer.")
    print("    D_open_questions.md OQ-D-02.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
