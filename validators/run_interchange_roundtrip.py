"""
The interchange format — the only way in for a table from anywhere else.

WHY THIS EXISTS
----------------
`io_loader.load_io_table` reads "the project's own interchange format": the
route documented for an analyst whose table is not the ONS workbook and not the
INE workbook, which is to say every analyst outside the United Kingdom and
Spain. It is named in the configuration template, in `config.TABLE_KINDS`, and
now in `docs/GUIDE.md` §3.

**Nothing exercised it.** No test, no validator, no example, no fixture — the
project's four worked pilots all build their tables in Python or read a
publisher's own workbook. Writing the user guide on 2026-08-25 ran it for the
first time, by following the guide's own instructions literally, and it
rejected a well-formed file:

    no value-added rows found between the sector block and the Output row

The cause was in how the sector block was located. The rows between the header
and `Output` hold the sectors and then the value-added components, with nothing
marking the join, and the loader counted *every* non-empty row label as a
sector. So four sectors and two value-added rows came out as six sectors, which
left no rows for value added, which produced a complaint about the very rows it
had just miscounted. Any file in the documented format failed, in every case.

The boundary is now found the way §4.1 says all blocks are found — by the
labels. The sector rows are the longest run whose labels match the header's
leading codes, one for one and in order; what follows down the page is value
added, and what follows across the header is final demand.

Two smaller defects surfaced in the same run and are guarded here too:

  * `label_<code>` rows in the metadata sheet were looked up with the code's
    own case against a dictionary whose keys had been lowercased, so every
    sector label was silently dropped and the report named sectors by their
    codes;
  * a scenario rejected as infeasible was left out of `results`, and both the
    exit code and `project.json`'s `all_passed` read only `results` — so a run
    that produced one table out of two reported success.

WHAT IS CHECKED
----------------
A synthetic four-sector table is written to the documented format, loaded back,
and every block compared against what was written. Then it is put through a
real split, end to end, as a user would: configuration workbook, `load_config`,
`IOProject`. Then five malformed files confirm the loader refuses each one for
the right reason rather than for an accidental one.

The fixture is deliberately small enough to check by hand: output 100, 120, 110
and 90, integers throughout, balancing exactly.

Run:
    python3 validators/run_interchange_roundtrip.py
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

FAIL: list[str] = []

# The fixture, in full, so that every number in this file is visible here.
CODES = ["A", "B", "C", "D"]
LABELS = ["Agriculture", "Manufacturing", "Services", "Public administration"]
Z = np.array([[10, 20, 15, 5],
              [30, 10, 25, 10],
              [5, 15, 10, 20],
              [10, 5, 20, 10]], float)
Y = np.array([[30, 20], [25, 20], [40, 20], [25, 20]], float)
Y_LABELS = ["P3_HH", "P6_EXP"]
VA = np.array([[25, 40, 25, 25],
               [20, 30, 15, 20]], float)
VA_LABELS = ["Compensation of employees", "Gross operating surplus"]
X = np.array([100, 120, 110, 90], float)


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def write_table(path: Path, *, rows=None, header=None, meta_drop=None,
                va_rows=True, stray=False) -> Path:
    """Write the fixture, optionally deformed, to `path`."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "table"
    ws.append([""] + (header if header is not None else CODES) + Y_LABELS)
    body = rows if rows is not None else [
        [CODES[i]] + list(Z[i]) + list(Y[i]) for i in range(4)]
    for r in body:
        ws.append(r)
    if stray:
        # A sector code below the block: the run that locates the block breaks
        # at the first non-match, so this must be caught explicitly.
        ws.append([VA_LABELS[0]] + list(VA[0]))
        ws.append(["C"] + list(VA[1]))
    elif va_rows:
        for lab, row in zip(VA_LABELS, VA):
            ws.append([lab] + list(row))
    ws.append(["Output"] + list(X))

    ms = wb.create_sheet("metadata")
    meta = [("country", "Ruritania"), ("year", 2022),
            ("unit", "EUR million, current prices, basic prices"),
            ("classification", "NACE Rev.2 (4 industries, synthetic)"),
            ("source", "Synthetic fixture — run_interchange_roundtrip.py")]
    meta += [(f"label_{c}", lab) for c, lab in zip(CODES, LABELS)]
    for k, v in meta:
        if meta_drop and k == meta_drop:
            continue
        ms.append([k, v])
    wb.save(path)
    return path


def write_config(path: Path, table: Path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    p = wb.active
    p.title = "project"
    for k, v in [("project_id", "interchange_roundtrip"),
                 ("table_path", table.name),
                 ("table_kind", "interchange"),
                 ("title", "Interchange round trip"),
                 ("notes", "Synthetic. A check, not an economy.")]:
        p.append([k, v])
    s = wb.create_sheet("splits")
    s.append(["sector_code", "new_code", "new_label", "key_id"])
    s.append(["B", "B1", "Heavy manufacturing", "kB"])
    s.append(["B", "B2", "Light manufacturing", "kB"])
    k = wb.create_sheet("keys")
    k.append(["key_id", "new_sector_code", "value", "source", "source_year",
              "strength"])
    for code, val in (("B1", 70), ("B2", 30)):
        k.append(["kB", code, val, "Synthetic employment proxy", 2022, "medium"])
    for code, val in (("B1", 65), ("B2", 35)):
        k.append(["kB2", code, val, "Synthetic turnover proxy — registered as "
                  "a corroborating key, not used", 2022, "medium"])
    sc = wb.create_sheet("scenarios")
    sc.append(["scenario_id", "label", "description", "internal_block_alpha"])
    sc.append(["S1_plain", "Size only", "Parent's purchasing pattern.", None])
    sc.append(["S2_hard", "Impossible on purpose",
               "Intensities the sector's own budget cannot absorb.", None])
    pr = wb.create_sheet("profiles")
    pr.append(["scenario_id", "subsector_code", "supplier_code", "intensity"])
    for row in (["S2_hard", "B1", "C", 1.6], ["S2_hard", "B1", "A", 0.5],
                ["S2_hard", "B2", "A", 1.7]):
        pr.append(row)
    wb.save(path)
    return path


def refuses(path: Path) -> str:
    """Load `path` and return the refusal message, or '' if it loaded."""
    from quadrium.io_loader import LoaderError, load_io_table
    try:
        load_io_table(path)
    except LoaderError as exc:
        return str(exc)
    return ""


def main() -> int:
    from quadrium.config import load_config
    from quadrium.io_loader import load_io_table
    from quadrium.project import IOProject

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        good = write_table(tmp / "ruritania.xlsx")

        # 1 -- it loads at all. This is the check that was failing.
        t = load_io_table(good)
        check("a file in the documented interchange format loads",
              t.n == 4 and len(t.VA_labels) == 2,
              f"{t.n} sectors and {len(t.VA_labels)} value-added rows read from "
              f"a sheet holding {4 + 2} labelled rows above `Output` — the "
              f"boundary is found by matching labels, not by counting them")

        # 2 -- and every block came back as written.
        blocks = {"Z": (t.Z, Z), "Y": (t.Y, Y), "VA": (t.VA, VA),
                  "X": (t.X, X)}
        worst = max(float(np.abs(got - want).max()) for got, want in
                    blocks.values())
        check("every block is recovered exactly", worst == 0.0,
              "Z, Y, VA and Output all reproduce the written values to 0.0")

        check("the column headers split into sectors and final demand at the "
              "same place the rows do",
              t.Y_labels == Y_LABELS and t.sector_codes == CODES,
              f"final demand {t.Y_labels}, sectors {t.sector_codes}")

        # 3 -- the labels, which were being dropped in silence.
        check("`label_<code>` metadata rows reach the table",
              t.sector_labels == LABELS,
              f"{t.sector_labels[1]!r} for B, not {'B'!r} — the lookup is "
              f"case-folded to match the metadata dictionary")

        # 4 -- and it balances, which the loader asserts on the way out.
        row_dev = float(np.abs(t.Z.sum(1) + t.Y.sum(1) - t.X).max())
        col_dev = float(np.abs(t.Z.sum(0) + t.VA.sum(0) - t.X).max())
        check("the loaded table satisfies both accounting identities",
              max(row_dev, col_dev) == 0.0,
              f"row max dev {row_dev:g}, column max dev {col_dev:g}")

        # 5 -- end to end, the way the guide tells a user to do it.
        cfg_path = write_config(tmp / "config.xlsx", good)
        cfg = load_config(cfg_path)
        project = IOProject(
            project_id=cfg["project_id"], table=cfg["table"],
            splits=cfg["splits"], scenarios=cfg["scenarios"], keys=cfg["keys"],
            ledger=cfg["ledger"], title=cfg["title"],
            source_file=cfg["source_file"], root=tmp / "outputs",
            preamble=cfg["notes"] or "")
        project.run().write()

        ran = [r.scenario_id for r in project.results]
        check("a split runs end to end from an interchange table",
              ran == ["S1_plain"] and project.results[0].report.passed,
              f"{ran[0]} balanced and passed every validation check; "
              f"S2_hard was rejected before balancing, on purpose")

        res = project.results[0]
        check("the disaggregated table reaggregates to the original",
              res.report.reaggregation_error_pct < 1e-9,
              f"max reaggregation error {res.report.reaggregation_error_pct:.2e} % — "
              f"the 4-sector table is recovered from the 5-sector one")

        # 6 -- a rejected scenario must not be reported as a passed one.
        import json
        meta = json.loads((project.dir / "project.json").read_text())
        rejected = [r["scenario_id"] for r in
                    meta["outcome"]["scenarios_rejected"]]
        check("a rejected scenario makes `all_passed` false",
              rejected == ["S2_hard"] and meta["outcome"]["all_passed"] is False,
              "one table of two produced, and the machine-readable outcome "
              "says so — it read only the scenarios that ran until 2026-08-25")
        check("and the report explains the rejection rather than omitting it",
              "Scenarios that were rejected" in
              (project.dir / "report.md").read_text(),
              "with the subsector named and the parent's own budget printed "
              "beside it")

        # 7 -- the refusals. Each must name its own cause.
        print()
        bad_no_va = write_table(tmp / "no_va.xlsx", va_rows=False)
        m = refuses(bad_no_va)
        check("a table with no value-added rows is refused",
              "value-added" in m and "Output" in m, m[:96])

        bad_stray = write_table(tmp / "stray.xlsx", stray=True)
        m = refuses(bad_stray)
        check("a sector code below the sector block is refused",
              "below the sector block" in m, m[:96])

        bad_header = write_table(tmp / "mismatch.xlsx",
                                 header=["W", "X", "Y", "Z"])
        m = refuses(bad_header)
        check("row labels that do not match the header are refused",
              "no sector rows found" in m, m[:96])

        bad_meta = write_table(tmp / "no_unit.xlsx", meta_drop="unit")
        m = refuses(bad_meta)
        check("metadata missing the price basis is refused",
              "unit" in m and "missing" in m, m[:96],)

        bad_unbal = write_table(
            tmp / "unbalanced.xlsx",
            rows=[[CODES[i]] + list(Z[i]) + list(Y[i] + (7 if i == 0 else 0))
                  for i in range(4)])
        m = refuses(bad_unbal)
        check("a table that does not balance is refused on load",
              bool(m), m[:96] or "IT LOADED — a 7-unit hole went unnoticed")

    print()
    print("    The documented route in for a table from anywhere other than")
    print("    the UK or Spain now has a fixture, and the guide's §3 is")
    print("    checked against the loader rather than against its docstring.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
