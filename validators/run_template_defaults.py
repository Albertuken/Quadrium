"""
The workbook's own instructions, checked against what the code actually does.

WHAT THIS IS FOR
-----------------
`quadrium --template` writes the first thing a user reads. Every sheet carries
its instructions at the bottom in grey, and those instructions are prose in
`config.py` -- a list of strings, next to but not derived from the defaults they
describe. Nothing made them agree.

WHAT IT FOUND ON THE DAY IT WAS WRITTEN
-----------------------------------------
The `scenarios` sheet told the user:

    Leave internal_block_alpha blank for the default 0.5
    (MVP_0.1 6.3 -- a project convention, not a source).

`Scenario.internal_block_alpha` has been **1.0** since v1.12, when the 0.5
was withdrawn: it rested on the intuition that a subsector buys from itself less
than proportionality implies, and measurement on 1,403 sibling pairs in three
published tables showed the diagonal of a real block is about 1.5x the outer
product. The intuition had the sign backwards. The guide says so at length. The
module docstring of `config.py`, sixty lines above the offending string, already
said 1.0 and cited CORE_031 eq. (14).

So the template kept telling users the withdrawn number, with the withdrawn
justification, for every release since. A user who left the cell blank got 1.0
while the workbook in their hands said they would get 0.5 -- and 0.5 vs 1.0 is
not a formatting difference, it is how much of a parent sector's trade with
itself stays inside each subsector.

**Worse, the same string was in the assumption ledger**, which is not an
instruction but a RESULT: every report produced through the workbook route
printed `source: MVP_0.1 6.3 -- project convention, no methodological source`
for the internal block, when the default it actually used is sourced and its
reproduction is checked by `check_wolsky_internal_block.py`. The engine was
understating its own provenance in the artefact the user keeps.

WHY IT WAS INVISIBLE
---------------------
Every existing check on the template is structural: `run_config_workbook.py`
checks the sheets exist and round-trip, the unit tests check the seeded values
parse. Nobody compared the ENGLISH against the FLOAT, because the English is a
comment and comments are not executed. This project has now found the same shape
several times -- a document, a docstring or a count that describes something the
code stopped doing -- and each time the fix was to make the description
derivable or checkable rather than to correct it once.

WHAT THIS CHECKS
-----------------
Every number the template's instructions state about a default is parsed out of
the generated workbook and compared with the default itself, read from
`Scenario`. It reads the WORKBOOK, not the source strings, so it fails the
same way a user would see it.
"""
import re
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402

from quadrium.config import write_template  # noqa: E402
from quadrium.models import Scenario  # noqa: E402

FAIL = []


def check(label, ok, detail=""):
    print(f"  [{'ok' if ok else 'FAIL'}] {label}")
    if detail:
        print(f"         {detail}")
    if not ok:
        FAIL.append(label)


def comments(ws):
    """Every instruction line of a sheet, joined -- the grey block at the foot."""
    out = []
    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if isinstance(first, str) and first.lstrip().startswith("#"):
            out.append(first.lstrip("# ").rstrip())
    return " ".join(out)


def main():
    print(__doc__.strip().split("\n")[0])
    print()

    cfg = Scenario(scenario_id="probe", label="probe")
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "t.xlsx"
        write_template(path)
        wb = openpyxl.load_workbook(path)
        text = {s: comments(wb[s]) for s in wb.sheetnames}

    # 1 -- the one that was wrong.
    scen = text.get("scenarios", "")
    stated = re.findall(r"default ([0-9]*\.?[0-9]+)", scen)
    check("the scenarios sheet states a default for internal_block_alpha",
          bool(stated), scen[:120] or "no 'default N' phrase found at all")
    if stated:
        check("and it is the default the engine actually uses",
              float(stated[0]) == cfg.internal_block_alpha,
              f"sheet says {stated[0]}, Scenario says "
              f"{cfg.internal_block_alpha}")

    # 2 -- the withdrawn number must not be offered as current anywhere.
    #      0.5 may be MENTIONED as history; it may not be called the default.
    for name, body in text.items():
        bad = re.search(r"default 0\.5", body)
        check(f"the {name} sheet does not offer the withdrawn 0.5 as current",
              bad is None,
              "found: " + body[max(0, bad.start() - 40):bad.end() + 40]
              if bad else "")

    # 3 -- the sourced default cites its source where it is stated.
    check("the stated default carries the source that justifies it",
          "CORE_031" in scen,
          "a number with no source beside it is the convention this engine "
          "exists to refuse")

    # 4 -- delta is the other REQUIRED-with-no-default, and the sheet must
    #      not invent one. The guide's argument is that a default would be a
    #      guess wearing a number.
    reg = text.get("regionalise", "")
    check("the regionalise sheet still refuses to give delta a default",
          "no default" in reg.lower() or "REQUIRED" in reg,
          reg[:150])

    print()
    print("    The instructions a user reads are the only part of this engine")
    print("    that nothing executed. Now something does.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
