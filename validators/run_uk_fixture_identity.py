"""
Dating the fixture from the fixture, so the question left for the owner is the
one that is actually theirs.

`OQ-D-01` recorded that the fixture then shipping as **`UK IO 2022.csv`** was an
XLSX, not a CSV; that its Menu sheet said 2023, not the 2022 in its name; and
that at v1.4 the data owner, asked directly, recalled 2018 — a third year. The
entry escalated to P1 on the right grounds: a pilot run on the wrong vintage
would be arithmetically perfect and economically about the wrong year. Its action
was "show the owner the workbook's own Menu sheet rather than ask again".

**The owner confirmed 2023 at v1.25 and the file was renamed to
`UK_IOAT_2023_domestic_ixi.xlsx`.** This file is what dated it first, and it
stays as the standing check: the evidence below is what the name now rests on, so
if the fixture is ever swapped the name stops being true and this goes red.

**A Menu sheet is a cover page, and a cover page is the weakest evidence in the
file.** It is exactly what would survive from a previous release if someone
rebuilt a workbook. So this checks the claim where it cannot be a leftover.

THE YEAR IS ON THE DATA SHEETS, NOT JUST THE COVER
---------------------------------------------------
Scanning every sheet for a four-digit year finds **thirteen occurrences of 2023
across all ten sheets** — every sheet in the workbook, including the two that
carry the actual matrices:

    [IOT]  "2023 Input-Output Analytical Tables"
    [A]    "2023 Input-Output Analytical Tables"
    [Menu] "United Kingdom Input-Output Analytical Tables, 2023"
    [Menu] "consistent with the 2025 edition ... Blue Book 2025"

**Not one occurrence of 2018. Not one of 2022.** The only other years anywhere in
the file are 2007 and 2008, and both are classification vintages — "SIC 2007
industry & CPA 2008 product code". The old filename's 2022 appeared nowhere
inside the file it named.

AND THE MAGNITUDES AGREE, WHICH LABELS CANNOT FAKE
----------------------------------------------------
Summing the domestic final-demand columns the loader reads:

    P3 S13  general government final consumption      567.4
    P3 S14  households                              1,186.3
    P51G    gross fixed capital formation             390.6
    P62     exports of services                       459.9    (GBP billion)

These are `OBSERVED` — computed from the file. Whether they are 2018 or 2023
magnitudes is a comparison against published UK aggregates, and **this project
holds no UK macro series**, so that half is `ESTIMATED` from general
national-accounts knowledge and is labelled as such: government consumption near
570 and services exports near 460 are 2023-scale, roughly 40 per cent above 2018.
The conclusion does not rest on the precision of that recall — a 40 per cent gap
is far outside any plausible error in it, and the label evidence is decisive on
its own. The magnitudes are corroboration from a direction that cannot be a
leftover cover sheet.

WHAT THIS DOES AND DOES NOT SETTLE
-----------------------------------
**Settled: the file in hand is the 2023 table.** Beyond reasonable doubt, from
two independent directions. The v1.4 worry that this might be a mislabelled 2018
table is closed — no 2018 table has government consumption of 567.

**Answered by the owner, not by this file.** Whether 2023 was the *intended*
year was never settleable from any workbook: the three v1.4 explanations — a
different ONS release, the survey year rather than the reference year, or the
wrong file entirely — are all about a recollection. Put as one clean question,
the owner confirmed 2023 on 2026-08-11. `OQ-D-01` is closed and the file
renamed.

Run:
    python3 validators/run_uk_fixture_identity.py
"""

from __future__ import annotations

import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

FIXTURE = ROOT / "UK_IOAT_2023_domestic_ixi.xlsx"
FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def _years() -> dict[str, list[tuple[str, str]]]:
    """Every four-digit year mentioned anywhere in the workbook's headers."""
    import openpyxl

    tmp = Path(tempfile.mkdtemp()) / "uk.xlsx"
    shutil.copy(FIXTURE, tmp)                     # leave the original alone
    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    pat = re.compile(r"\b(20[0-2]\d)\b")
    found: dict[str, list[tuple[str, str]]] = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(max_row=60, max_col=20, values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    for y in pat.findall(cell):
                        found.setdefault(y, []).append((sheet, cell.strip()))
    os.unlink(tmp)
    return found


def main() -> int:
    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    if not FIXTURE.exists():
        print("fixture absent")
        return 0

    # ---- the file is not what its name says -------------------------------
    header = FIXTURE.open("rb").read(2)
    check("the name now tells the truth about the format",
          header == b"PK" and FIXTURE.suffix == ".xlsx",
          "PK header — a zip-container workbook, and the extension finally "
          "says so. This was the first of OQ-D-01's three mislabellings")

    # ---- the year, where it cannot be a leftover cover page ----------------
    years = _years()
    on_data_sheets = {s for y, hits in years.items() if y == "2023"
                      for s, _ in hits} & {"IOT", "A"}
    print()
    for y in sorted(years):
        sheets = sorted({s for s, _ in years[y]})
        print(f"    {y}: {len(years[y]):>2} mention(s)  {', '.join(sheets)}")

    check("the workbook says 2023, and says it on the sheets holding the "
          "matrices",
          len(years.get("2023", [])) >= 10 and on_data_sheets == {"IOT", "A"},
          f"{len(years['2023'])} occurrences across "
          f"{len(set(s for s, _ in years['2023']))} sheets, including IOT and "
          f"A — not only the Menu cover page, which is the one place a stale "
          f"label could survive a rebuild")

    check("and neither 2018 nor 2022 appears anywhere in the file",
          "2018" not in years and "2022" not in years,
          "2022 was the old filename's year and appeared nowhere inside the "
          "file it named; 2018 was the year the owner recalled at v1.4 and "
          "appears nowhere either. The name now agrees with the contents")

    check("the only other years present are classification vintages",
          set(years) - {"2023", "2025"} <= {"2007", "2008"},
          "SIC 2007 and CPA 2008 — the classifications, not the reference "
          "period")

    # ---- magnitudes, which a stale label cannot produce --------------------
    from quadrium.io_loader import load_uk_analytical_iot

    t = load_uk_analytical_iot(FIXTURE)
    col = {lab: float(t.Y[:, i].sum()) / 1e3
           for i, lab in enumerate(t.Y_labels)}
    print()
    print("    domestic final demand, GBP billion, computed from the file:")
    for lab in ("P3 S13", "P3 S14", "P51G", "P62"):
        print(f"      {lab:<8} {col[lab]:>9,.1f}")

    check("the magnitudes are 2023-scale, not 2018-scale",
          col["P3 S13"] > 500 and col["P62"] > 400,
          f"government final consumption {col['P3 S13']:,.1f} and services "
          f"exports {col['P62']:,.1f}. The comparison year is ESTIMATED — the "
          f"project holds no UK macro series — but the gap to 2018 levels is "
          f"~40 %, far outside any error in that recall")

    print()
    print("    SETTLED: the file in hand is the 2023 table, from two")
    print("    independent directions. NOT SETTLED, and not settleable from")
    print("    any file: whether 2023 is the year the owner wants. That is a")
    print("    question about an intention, not about this workbook.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
