"""
Where the INE's interior table stops balancing, and it is one cell.

`OQ-D-04` recorded that the Spanish input-output table's domestic half falls
4,921.6 million EUR short of domestic output for agricultural products, while
its import half exceeds recorded imports by exactly the same amount. The total
table balances to zero, so the two errors cancel. The cause was `NOT SPECIFIED`:
the methodological note names a mechanism — the domestic/import split is
inherited from the use table and then transformed — but never mentions a
residual.

The supply and use tables settle it, because they are the step BEFORE the
transformation. This file runs the comparison.

WHAT IT ESTABLISHES, ALL OF IT ARITHMETIC
-----------------------------------------
1. The **supply-use tables balance exactly**, both halves, all 64 products.
   Domestic uses equal domestic output; imported uses equal imports; max
   deviation 0.0 in each.
2. The **input-output table does not**, for agriculture alone.
3. The difference is **not in intermediate demand**, which is identical in both:
   36,884.0 domestic and 9,196.7 imported.
4. The whole of it sits in **one final-demand cell pair** — final consumption of
   non-profit institutions serving households, agricultural products:

        supply-use, domestic     7.3        input-output, domestic  -4,914.3
        supply-use, imported     0.0        input-output, imported  +4,921.6

   7.3 - 4,921.6 = -4,914.3, and 0.0 + 4,921.6 = +4,921.6. The total table keeps
   the supply-use value of 7.3, which is why it balances.

WHAT IT SUGGESTS, WHICH IS JUDGEMENT AND IS LABELLED AS SUCH
------------------------------------------------------------
NPISH consumption of agricultural products of 4,921.6 million EUR is not
plausible. In the supply-use tables NPISH consumption of ALL products is 15,954.0
domestically and **25.0** imported. The input-output table puts 4,921.6 of
imported NPISH consumption into agriculture alone — nearly two hundred times the
whole imported NPISH column of the source it was built from — and its domestic
counterpart is the only negative in that column.

A negative NPISH final consumption is not a recognised construct, unlike the
negatives in inventory change and net taxes that appear legitimately throughout
the same table.

On that evidence this looks like a **publication error in the TIO 2022**, not a
convention. This file does not assert that; it establishes the arithmetic and
lets the reader weigh the rest. The INE is the only party who can settle it.

Run:
    python3 validators/run_ine_tod_vs_tio.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
TIO = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
TOD = ROOT / "data" / "ine" / "cne_tod_22.xlsx"

N = 64
FIRST_ROW = 9
COL_INTERMEDIATE, COL_NPISH, COL_USES = 67, 70, 79
FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def main() -> int:
    if not (TIO.exists() and TOD.exists()):
        print(f"fixture absent: {TIO.name} and/or {TOD.name}")
        return 0
    import openpyxl

    def sheets(path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return {n: list(wb[n].iter_rows(values_only=True)) for n in wb.sheetnames}

    tio, tod = sheets(TIO), sheets(TOD)
    col = lambda g, c: np.array([_num(g[FIRST_ROW + i][c]) for i in range(N)])

    g1 = tio["Tabla1"]
    X = np.array([_num(g1[86][c]) for c in range(2, 2 + N)])   # production
    M = np.array([_num(g1[87][c]) for c in range(2, 2 + N)])   # imports

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 72)

    # 1 -- the supply-use tables balance, both halves.
    for name, sheet, target, what in (
            ("supply-use, domestic", tod["Tabla4"], X, "domestic output"),
            ("supply-use, imported", tod["Tabla5"], M, "imports")):
        dev = float(np.abs(col(sheet, COL_USES) - target).max())
        check(f"{name} uses equal {what}, all 64 products", dev < 0.05,
              f"max deviation {dev:.4f}")

    # 2 -- the input-output table does not, and only for one product.
    d_dom = col(tio["Tabla2"], COL_USES) - X
    d_imp = col(tio["Tabla3"], COL_USES) - M
    off = np.flatnonzero(np.abs(d_dom) > 1.0)
    check("the input-output table fails for exactly one product",
          off.size == 1 and off[0] == 0,
          f"product(s) {off.tolist()}, worst {d_dom[off].tolist() if off.size else '—'}")
    check("and the two halves are equal and opposite",
          abs(d_dom[0] + d_imp[0]) < 0.05,
          f"domestic {d_dom[0]:,.1f}, imported {d_imp[0]:,.1f}")

    # 3 -- intermediate demand is untouched.
    for name, a, b in (("domestic", tod["Tabla4"], tio["Tabla2"]),
                       ("imported", tod["Tabla5"], tio["Tabla3"])):
        ia = _num(a[FIRST_ROW][COL_INTERMEDIATE])
        ib = _num(b[FIRST_ROW][COL_INTERMEDIATE])
        check(f"agriculture's {name} INTERMEDIATE demand is identical in both",
              abs(ia - ib) < 0.05, f"{ia:,.1f} in both")

    # 4 -- the whole discrepancy is one final-demand cell pair.
    pairs = {
        "domestic": (_num(tod["Tabla4"][FIRST_ROW][COL_NPISH]),
                     _num(tio["Tabla2"][FIRST_ROW][COL_NPISH])),
        "imported": (_num(tod["Tabla5"][FIRST_ROW][COL_NPISH]),
                     _num(tio["Tabla3"][FIRST_ROW][COL_NPISH])),
    }
    print()
    for k, (sut, iot) in pairs.items():
        print(f"    NPISH consumption of agricultural products, {k:<9} "
              f"supply-use {sut:>10,.1f}   input-output {iot:>10,.1f}")
    moved = pairs["imported"][1] - pairs["imported"][0]
    check("the entire discrepancy is that one cell pair",
          abs(moved - abs(d_dom[0])) < 0.05,
          f"{moved:,.1f} moved from the domestic table to the import table")
    check("the TOTAL table keeps the supply-use value, which is why it balances",
          abs(pairs["domestic"][1] + pairs["imported"][1]
              - pairs["domestic"][0]) < 0.05,
          f"{pairs['domestic'][1] + pairs['imported'][1]:,.1f} against "
          f"{pairs['domestic'][0]:,.1f}")

    # 5 -- the implausibility, stated as measurement rather than as opinion.
    npish_dom_sut = col(tod["Tabla4"], COL_NPISH).sum()
    npish_imp_sut = col(tod["Tabla5"], COL_NPISH).sum()
    npish_imp_iot = col(tio["Tabla3"], COL_NPISH).sum()
    print(f"\n    NPISH consumption of ALL products, supply-use: "
          f"{npish_dom_sut:,.1f} domestic, {npish_imp_sut:,.1f} imported")
    print(f"    the input-output table's imported NPISH column: {npish_imp_iot:,.1f}, "
          f"of which agriculture {pairs['imported'][1]:,.1f}")
    check("the imported NPISH cell dwarfs the entire imported NPISH column it came from",
          pairs["imported"][1] > 50 * npish_imp_sut,
          f"{pairs['imported'][1]:,.1f} against {npish_imp_sut:,.1f} — "
          f"{pairs['imported'][1] / npish_imp_sut:.0f}x")
    dom_col = col(tio["Tabla2"], COL_NPISH)
    check("and its domestic counterpart is the only negative in that column",
          int((dom_col < -0.05).sum()) == 1 and dom_col[0] < 0,
          f"{int((dom_col < -0.05).sum())} negative(s), agriculture "
          f"{dom_col[0]:,.1f} against a column total of {dom_col.sum():,.1f}")

    print("\n" + "=" * 72)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
