"""
Two doors into the same method, and neither is more permissive than the other.

WHY THE SECOND DOOR EXISTS
----------------------------
`docs/GUIDE.md` opens by promising **"No Python: you fill in a spreadsheet and
run one command"**, and every part of the engine kept that promise except
regionalisation, which existed only as command-line flags. The one feature added
most recently was the one that broke the product's own first sentence.

So a `regionalise` sheet was added to the configuration workbook at v1.86. The
`project` sheet already says which table to use and how to load it, so the new
sheet only says what to do with it — which means `table_kind: eurostat` works
through the workbook and cannot be expressed on the command line at all.

WHAT A SECOND DOOR RISKS
--------------------------
Two entry points to one method is two chances for them to disagree, and the
dangerous direction is the quiet one: a route that accepts what the other
refuses. So both call `config.read_activity` and both call `regionalise()`, and
this file checks that the refusals arrive through both — not that they exist.

It also checks the one thing that broke while this was being written: an
untouched template must still be a **split** configuration. The first version
read the template's own `key | value` header row as a setting named "key", so a
workbook nobody had edited looked like a regionalisation with nothing in it.

Run:
    python3 validators/run_regionalise_config.py
"""
from __future__ import annotations

import csv
import io
import shutil
import sys
import tempfile
import warnings
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

ES = ROOT / "data" / "ine" / "cne_tio_21.xlsx"
FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def run(argv):
    from quadrium.cli import main

    out, err = io.StringIO(), io.StringIO()
    with redirect_stdout(out), redirect_stderr(err):
        code = main(argv)
    return code, out.getvalue(), err.getvalue()


def workbook(path, *, reg_rows, table_path, splits=True):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "project"
    for k, v in (("project_id", "wbtest"), ("table_path", str(table_path)),
                 ("table_kind", "ine_interior")):
        ws.append([k, v])
    ws = wb.create_sheet("regionalise")
    ws.append(["key", "value"])          # the template's own header
    for row in reg_rows:
        ws.append(list(row))
    if splits:
        wb.create_sheet("splits")
    wb.save(path)
    return path


def main_() -> int:
    warnings.filterwarnings("ignore")
    from quadrium.config import ConfigError, load_config
    from quadrium.io_loader import load_ine_tio

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    check("a national table to regionalise", ES.exists(), ES.name)
    if not ES.exists():
        return 1

    tmp = Path(tempfile.mkdtemp(prefix="quadrium_regcfg_"))
    try:
        table = load_ine_tio(ES, variant="interior")
        act = tmp / "activity.csv"
        rng = np.random.default_rng(11)
        share = np.clip(0.20 * np.exp(rng.normal(0, 0.5, table.n)), 0.02, 0.85)
        with act.open("w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["sector_code", "regional"])
            for c, x, s in zip(table.sector_codes, table.X, share):
                w.writerow([c, f"{x * s:.6f}"])

        good = workbook(tmp / "reg.xlsx",
                        reg_rows=[("method", "FLQ"), ("delta", 0.25),
                                  ("activity_path", str(act))],
                        table_path=ES.resolve())

        # ---- the two doors, same answer
        c1, o1, _ = run([str(good), "--outputs", str(tmp / "a")])
        c2, o2, _ = run(["--regionalise", str(act), "--national", str(ES),
                         "--national-kind", "ine", "--method", "FLQ",
                         "--delta", "0.25", "--outputs", str(tmp / "b"),
                         "--name", "wbtest"])
        check("both doors run, and the workbook needs no --national of its own",
              c1 == 0 and c2 == 0,
              f"workbook exit {c1}, flags exit {c2}. The `project` sheet already "
              f"names the table, so the same sheet that feeds a split feeds a "
              f"regionalisation — including table_kind: eurostat, which the "
              f"flags cannot express")

        a = np.loadtxt(tmp / "a/wbtest/coefficients.csv", delimiter=",",
                       comments="#")
        b = np.loadtxt(tmp / "b/wbtest/coefficients.csv", delimiter=",",
                       comments="#")
        check("and they produce the same coefficients, to the last bit",
              a.shape == b.shape and float(np.abs(a - b).max()) == 0.0,
              f"{a.shape[0]}x{a.shape[1]}, identical. A result that depended on "
              f"which door it came through would be a defect nobody would "
              f"think to look for")

        for f in ("report.md", "assumption_ledger.json", "regional_table.xlsx",
                  "implicit_imports.csv"):
            same = (tmp / "a/wbtest" / f).exists() and (tmp / "b/wbtest" / f).exists()
            if not same:
                check(f"both write {f}", False)
                break
        else:
            check("and both write every output, including the cost",
                  "28.3" in (tmp / "a/wbtest/report.md").read_text(),
                  "coefficients, implicit imports, the table, the report and "
                  "the ledger — and the measured cost is in the workbook route "
                  "too, which is the one a non-programmer will use")

        # ---- the refusals, through both
        print()
        cases = []
        wb_nodelta = workbook(tmp / "nodelta.xlsx",
                              reg_rows=[("method", "FLQ"),
                                        ("activity_path", str(act))],
                              table_path=ES.resolve())
        c, _, e = run([str(wb_nodelta), "--outputs", str(tmp)])
        _, _, e2 = run(["--regionalise", str(act), "--national", str(ES),
                        "--national-kind", "ine", "--outputs", str(tmp)])
        cases.append(("the FLQ without a delta",
                      "no defensible default" in e and
                      "no defensible default" in e2))

        short = tmp / "short.csv"
        short.write_text("sector_code,regional\n1,10\n")
        wb_short = workbook(tmp / "short.xlsx",
                            reg_rows=[("method", "SLQ"),
                                      ("activity_path", str(short))],
                            table_path=ES.resolve())
        c, _, e = run([str(wb_short), "--outputs", str(tmp)])
        _, _, e2 = run(["--regionalise", str(short), "--national", str(ES),
                        "--national-kind", "ine", "--method", "SLQ",
                        "--outputs", str(tmp)])
        cases.append(("an activity file missing sectors",
                      "do not describe the same sectors" in e
                      and "do not describe the same sectors" in e2))

        fired = [n for n, ok in cases if ok]
        check("the same mistake is refused through both doors, in the same words",
              len(fired) == len(cases),
              f"{len(fired)} of {len(cases)}: " + "; ".join(fired)
              + ". Both call config.read_activity and both call regionalise(), "
                "so a route that quietly accepted what the other refused would "
                "have to be written on purpose")

        # ---- and the workbook's own refusals
        print()
        own = []
        wb_noact = workbook(tmp / "noact.xlsx", reg_rows=[("method", "SLQ")],
                            table_path=ES.resolve())
        try:
            load_config(wb_noact)
        except ConfigError as exc:
            own.append(("no activity_path", "activity_path" in str(exc)))

        wb_bad = workbook(tmp / "badmethod.xlsx",
                          reg_rows=[("method", "XLQ"),
                                    ("activity_path", str(act))],
                          table_path=ES.resolve())
        try:
            load_config(wb_bad)
        except ConfigError as exc:
            own.append(("an unknown method", "must be SLQ" in str(exc)))

        wb_none = workbook(tmp / "none.xlsx", reg_rows=[],
                           table_path=ES.resolve(), splits=False)
        try:
            load_config(wb_none)
        except ConfigError as exc:
            own.append(("a workbook that describes no job",
                        "describes no job" in str(exc)))

        check("and the sheet has refusals of its own",
              len(own) == 3 and all(ok for _, ok in own),
              "; ".join(n for n, _ in own))

        # ---- the template must still be a split
        tpl = tmp / "template.xlsx"
        run(["--template", str(tpl)])
        import openpyxl
        wb = openpyxl.load_workbook(tpl)
        wb["project"]["B2"] = str(
            (ROOT / "UK_IOAT_2023_domestic_ixi.xlsx").resolve())
        wb.save(tpl)
        cfg = load_config(tpl)
        check("an untouched template is still a SPLIT configuration",
              cfg.get("kind") != "regionalise" and cfg.get("splits"),
              f"{len(cfg.get('splits', []))} split(s), kind "
              f"{cfg.get('kind', 'split')!r}. The first version read the "
              f"template's own `key | value` header as a setting called 'key', "
              f"so a workbook nobody had edited looked like a regionalisation "
              f"with nothing in it")

        sheet = wb["regionalise"]
        notes = [r[0] for r in sheet.iter_rows(values_only=True)
                 if r and isinstance(r[0], str) and r[0].startswith("#")]
        check("and it carries the sheet, explained, with nothing filled in",
              len(notes) >= 20 and any("0.14 to 0.60" in n for n in notes)
              and any("DOMESTIC" in n for n in notes),
              f"{len(notes)} lines of instruction, which name the delta range "
              f"and the domestic-table requirement — the two things that turn "
              f"a plausible run into a wrong one")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main_())
