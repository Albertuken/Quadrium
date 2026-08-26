"""
No default projection method: the targets sheet says which one you mean.

WHAT WAS WRONG WITH HAVING ONE
--------------------------------
`project_method` defaulted to `sut_euro`, which is the worse of the two methods
on every test that has been run — 61 back-tests against tables the offices later
published, and 54 more with both methods handed identical information
(`OQ-B-16`, `run_projection_backtest.py`). A user who said nothing got the loser
by silence.

Changing the default to `sut_ras` would have moved the problem rather than
removed it, because the two do not answer the same question. They take
DIFFERENT targets:

    sut_euro    gva (per industry, basic prices)
                final_use (per category, purchasers' prices)
    sut_ras     industry_output (per industry)
                use_column_totals (per industry, then per final-use category)
    both        taxes, imports

So the `targets` sheet has already chosen. Nobody reaches a default by accident:
they had to write `gva` rows or `industry_output` rows to get anywhere at all.

WHAT REPLACED IT
------------------
The method is read off the sheet. `project_method` is an OPTIONAL declaration
that has to agree with it, and disagreeing is an error rather than a silent
override in either direction. A sheet that mixes the two, or carries neither
distinctive set, is refused with the question a user can actually answer:

    What do you know about 2022?
        sut_euro  gva, final_use
        sut_ras   industry_output, use_column_totals
        both      taxes, imports

**Picking one for the user would be choosing what they measured.** `sut_ras` is
the better method and it needs industry outputs; if you do not have them, the
answer is not that the engine should guess them.

Run:
    python3 validators/run_project_method_choice.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

DATA = ROOT / "data" / "eurostat"
OUT = ROOT / "outputs" / "_scratch"
FAIL: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def main() -> int:
    import openpyxl

    from quadrium.config import ConfigError, load_config
    from quadrium.eurostat import load_sut

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    need = [DATA / f"naio_10_{d}_ES_{y}.json"
            for y in (2021, 2022) for d in ("cp15", "cp16", "cp1610")]
    if not all(f.exists() for f in need):
        print("  (Spain's 2021 and 2022 pairs are not both here)")
        return 0

    B = load_sut(*[DATA / f"naio_10_{d}_ES_2022.json"
                   for d in ("cp15", "cp16", "cp1610")])
    pi, ai = np.flatnonzero(B.q > 0), np.flatnonzero(B.g > 0)
    live = [B.activity_codes[i] for i in ai]
    u = (np.hstack([B.U_domestic[np.ix_(pi, ai)], B.Y_domestic[pi]]).sum(0)
         + np.hstack([B.U_imported[np.ix_(pi, ai)], B.Y_imported[pi]]).sum(0)
         + np.concatenate([B.taxes_by_activity[ai], B.taxes_by_final_demand]))
    fu = B.Y_domestic.sum(0) + B.Y_imported.sum(0) + B.taxes_by_final_demand
    TAX = float(B.taxes_by_activity.sum() + B.taxes_by_final_demand.sum())
    IMP = float(B.imports[B.q > 0].sum())
    TOTALS = [("taxes", "", TAX), ("imports", "", IMP)]
    EURO = ([("gva", c, float(v)) for c, v in zip(live, B.W.sum(0)[ai])]
            + [("final_use", c, float(v)) for c, v in zip(B.Y_labels, fu)]
            + TOTALS)
    RAS = ([("industry_output", c, float(v)) for c, v in zip(live, B.g[ai])]
           + [("use_column_totals", c, float(v))
              for c, v in zip(live + list(B.Y_labels), u)] + TOTALS)

    OUT.mkdir(parents=True, exist_ok=True)

    def build(tag, rows, declared=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "project"
        kv = [("project_id", "method_choice"), ("table_kind", "eurostat_sut"),
              ("table_path", str(DATA)), ("eurostat_geo", "ES"),
              ("eurostat_year", 2021), ("eurostat_model", "D"),
              ("title", "which method"), ("project_to_year", 2022)]
        if declared:
            kv.append(("project_method", declared))
        for k, v in kv:
            ws.append([k, v])
        ts = wb.create_sheet("targets")
        ts.append(["kind", "code", "value"])
        for r in rows:
            ts.append(list(r))
        for s, h in (("splits", ["parent", "new_code", "new_label", "key_id"]),
                     ("keys", ["key_id", "applies_to", "code", "value",
                               "source", "strength"]),
                     ("scenarios", ["scenario_id", "split_id", "key_id"])):
            wb.create_sheet(s).append(h)
        f = OUT / f"method_{tag}.xlsx"
        wb.save(f)
        return f

    def run(tag, rows, declared=None):
        """Returns the method the engine chose, or the refusal text."""
        import io
        import contextlib
        buf = io.StringIO()
        try:
            with contextlib.redirect_stdout(buf):
                load_config(build(tag, rows, declared), offline=True)
        except ConfigError as exc:
            said = buf.getvalue()
            for m in ("sut_ras", "sut_euro"):
                if f"by {m}:" in said:
                    return m
            return str(exc)
        except Exception:
            pass
        said = buf.getvalue()
        for m in ("sut_ras", "sut_euro"):
            if f"by {m}:" in said:
                return m
        return "(no projection ran)"

    # 1 -- the sheet decides, both ways, with nothing declared.
    print()
    check("a sheet of gva and final_use runs sut_euro, undeclared",
          run("euro", EURO) == "sut_euro",
          "value added by industry and final use by category")
    check("a sheet of industry_output and use_column_totals runs sut_ras",
          run("ras", RAS) == "sut_ras",
          "industry outputs and use column totals — the better method, and "
          "the one that needs more of you")

    # 2 -- a declaration that agrees is accepted; one that disagrees is not.
    print()
    check("declaring the method the sheet already implies is accepted",
          run("ras_declared", RAS, "sut_ras") == "sut_ras",
          "`project_method` is a statement, not a switch")
    msg = run("ras_wrong", RAS, "sut_euro")
    check("declaring the other one is an error, not a silent override",
          isinstance(msg, str) and "project_method says" in msg
          and "make the two agree" in msg,
          "neither the sheet nor the key wins quietly — the disagreement is "
          "the finding")

    # 3 -- a sheet that cannot say is refused with the question, not a guess.
    print()
    mixed = run("mixed", EURO[:3] + RAS[:3] + TOTALS)
    check("a sheet mixing both vocabularies is refused",
          isinstance(mixed, str) and "mixes the two projection methods" in mixed,
          "and it lists what it found")
    bare = run("bare", TOTALS)
    check("and so is one carrying only the rows both methods share",
          isinstance(bare, str) and "no rows that say which projection method"
          in bare,
          "taxes and imports alone say nothing about which you mean")
    for name, m in (("mixed", mixed), ("bare", bare)):
        check(f"the {name} refusal asks what you know, and does not pick",
              isinstance(m, str) and "What do you know about 2022?" in m
              and "picking one for you would be choosing what you measured"
              in m.lower(),
              "the answerable question is what you measured, not which "
              "chapter-18 method you prefer")

    # 4 -- and the thing the removed default was hiding.
    print()
    check("neither method is named the default anywhere in the config",
          "sut_euro" not in
          (ROOT / "src" / "quadrium" / "config.py").read_text(
              encoding="utf-8").split("def _project")[0].split(
                  "raw or None")[0].split("pmethod = ")[-1],
          "`project_method` has no fallback value: it is None when unset and "
          "the sheet is read instead")

    print()
    print("    A default is a choice made for someone who did not make one.")
    print("    Here there was nothing to choose from: the sheet already says")
    print("    what the user has.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
