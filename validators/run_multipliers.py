"""
`OQ-T-05`: the multiplier formula was inside the project's own test table.

The entry says three things: "No multiplier formula, no type I / type II
distinction, no consumption-closure procedure appears anywhere in the loaded
set", and sends the reader to three software packages — CORE_069 (MARIO),
CORE_070 (ioanalysis), CORE_074 (pymrio), none of them extracted.

**The first claim was wrong, and had been since the fixture arrived.**
`UK_IOAT_2023_domestic_ixi.xlsx` carries two sheets the project had never
opened: `Multipliers` (116 × 12) and `Effects` (112 × 18), holding the ONS's own
published Type I multipliers and effects for 103 industries across seven
variables. That is an official reference implementation, sitting inside the table
the engine loads on every run. This project has repeatedly found "next source"
notes pointing outward at something already in hand — CORE_014/015 and the
`OQ-B-08` blocker are two recorded cases — and **this is the first where it was
inside a data file rather than in `library/extracted/`.** The habit of looking
outward before looking at what is loaded has now cost the same kind of finding
more than once; the count is deliberately not given here because nobody has
counted it.

THE TWO FORMULAS, RECOVERED AND VERIFIED
------------------------------------------
Reading the sheets against the engine's own Leontief inverse recovers both
definitions exactly. With `x` industry output, `L = (I − A)^-1`, and `v` any row
of the table (value added, imports, taxes, output itself):

    Type I effect      e = (v ⊘ x)ᵀ L         total impact per unit of final use
    Type I multiplier  m_j = e_j / (v_j / x_j)          total over direct

Checked against every published number — **7 variables × 103 industries, both
sheets, 1,434 published figures — worst error 7.11e-15.** Machine precision.
This is not a plausible reconstruction; it is the ONS's arithmetic, reproduced.

The output multiplier falls out as the special case `v = x`, where the direct
coefficient is 1 and the multiplier equals the effect: **the column sums of the
Leontief inverse**, matching ONS to 2.7e-15.

WHAT IS STILL `NOT SPECIFIED`, AND IT IS THE HALF THAT MATTERS MOST
--------------------------------------------------------------------
Both ONS sheets are headed **"Type I"**. Nothing in the workbook, and nothing in
the loaded methodological set, defines the **induced effect** that CORE_005
¶36.34, p. 1015 names — no type II multiplier, no consumption closure, no rule
for endogenising households. The entry's second claim is now half true and its
third stands entirely. **A closed type I is not a type II**, and an engine that
reported "multipliers" without saying which would be making the more consequential
of the two claims silently.

TWO OBSERVATIONS ABOUT THE SOURCE
-----------------------------------
**The ONS's own subtitle has the ratio backwards.** The `Multipliers` sheet is
headed "Ratio of direct to total impact from final use of an industry", but what
it publishes is total over direct: for A01 the GVA multiplier is 2.136312, and
2.136312 = 0.815078 / 0.381535 — effect over direct coefficient. Direct over
total would be 0.468096. Read as a slip in the caption, like `OQ-B-11`'s two;
nothing depends on it because the numbers are unambiguous.

**One industry is deliberately excluded.** L68A, Owner-Occupiers' Housing, has
output of £260,111 million and no published multiplier. It is imputed rent — no
transaction, no employment, no supply chain — so a multiplier for it would be an
artefact of the imputation. The engine computes one; **the ONS's refusal to
publish it is the better judgement**, and any multiplier report this project
emits should carry the same exclusion rather than quietly filling the cell.

Run:
    python3 validators/run_multipliers.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

FIXTURE = ROOT / "UK_IOAT_2023_domestic_ixi.xlsx"
FAIL: list[str] = []

# Column indices, 0-based, in the two published sheets. Established by reading
# the header rows; re-asserted below against the header text so a re-export
# that moves a column fails loudly instead of comparing the wrong series.
VARIABLES = [
    #  label                                      multipliers  effects
    ("Intermediate consumption at basic prices",       2,   2),
    ("Use of imported products, cif",                  3,   3),
    ("Taxes less subsidies on products",               4,   4),
    ("Compensation of employees",                      6,   6),
    ("Gross operating surplus and mixed income",       7,   7),
    ("Gross value added",                              8,   9),
    ("Output",                                         9,  10),
]
FIRST_DATA_ROW = 8          # 1-based, the row after the SIC/Industry header


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAIL.append(name)


def main() -> int:
    if not FIXTURE.exists():
        print("fixture absent")
        return 0

    import openpyxl
    from quadrium.io_loader import load_uk_analytical_iot

    print(__doc__.strip().split("Run:")[0].rstrip())
    print("\n" + "=" * 78)

    t = load_uk_analytical_iot(FIXTURE)
    X = np.asarray(t.X, float)
    Z = np.asarray(t.Z, float)
    VA = np.asarray(t.VA, float)
    n = len(X)
    safe = np.where(X > 0, X, 1.0)
    L = np.linalg.inv(np.eye(n) - Z / safe)

    wb = openpyxl.load_workbook(FIXTURE, read_only=True, data_only=True)

    def published(sheet: str, c: int) -> np.ndarray:
        rows = wb[sheet].iter_rows(min_row=FIRST_DATA_ROW,
                                   max_row=FIRST_DATA_ROW + n - 1,
                                   values_only=True)
        return np.array([float(r[c]) if isinstance(r[c], (int, float)) else np.nan
                         for r in rows])

    def heading(sheet: str, c: int) -> str:
        return str(list(wb[sheet].iter_rows(min_row=6, max_row=6,
                                            values_only=True))[0][c] or "")

    imports, taxes_products, d1, gos, other_taxes = VA
    series = {
        "Intermediate consumption at basic prices": Z.sum(axis=0),
        "Use of imported products, cif": imports,
        "Taxes less subsidies on products": taxes_products,
        "Compensation of employees": d1,
        "Gross operating surplus and mixed income": gos,
        "Gross value added": d1 + gos + other_taxes,
        "Output": X,
    }

    worst_e = worst_m = 0.0
    compared = 0
    print(f"    {'variable':<44}{'effect':>12}{'multiplier':>14}")
    for label, c_mul, c_eff in VARIABLES:
        assert label in heading("Multipliers", c_mul), \
            f"Multipliers column {c_mul} is no longer {label!r}"
        v = series[label]
        coef = v / safe
        eff = coef @ L
        mul = np.divide(eff, coef, out=np.full(n, np.nan), where=coef != 0)
        pe, pm = published("Effects", c_eff), published("Multipliers", c_mul)
        de = float(np.nanmax(np.abs(eff - pe)))
        dm = float(np.nanmax(np.abs(mul - pm)))
        compared += int(np.isfinite(pe).sum() + np.isfinite(pm).sum())
        worst_e, worst_m = max(worst_e, de), max(worst_m, dm)
        print(f"    {label:<44}{de:>12.1e}{dm:>14.1e}")

    print()
    check("the engine reproduces the ONS's published Type I effects and "
          "multipliers",
          worst_e < 1e-12 and worst_m < 1e-12,
          f"{compared:,} published figures across {len(VARIABLES)} variables, "
          f"worst error {max(worst_e, worst_m):.2e} — machine precision. "
          f"e = (v ⊘ x)ᵀ L and m_j = e_j / (v_j / x_j) are not a plausible "
          f"reconstruction, they are the ONS's own arithmetic")

    out_eff = L.sum(axis=0)
    check("and the output multiplier is exactly the column sums of the "
          "Leontief inverse",
          float(np.nanmax(np.abs(out_eff - published("Effects", 10)))) < 1e-12,
          "the special case v = x, where the direct coefficient is 1 and "
          "multiplier equals effect")

    # ---- what the source gets wrong, and what it wisely refuses -----------
    gva = series["Gross value added"] / safe
    eff_gva = gva @ L
    caption = str(list(wb["Multipliers"].iter_rows(min_row=3, max_row=3,
                                                   values_only=True))[0][0])
    check("the ONS caption states the ratio backwards",
          "direct to total" in caption
          and abs(published("Multipliers", 8)[0] - eff_gva[0] / gva[0]) < 1e-9,
          f'"{caption}" — but A01\'s published GVA multiplier is '
          f'{published("Multipliers", 8)[0]:.6f}, which is total over direct '
          f'({eff_gva[0]:.6f} / {gva[0]:.6f}). Direct over total would be '
          f'{gva[0] / eff_gva[0]:.6f}. A caption slip, like OQ-B-11\'s two')

    codes = [str(r[0]) for r in wb["Effects"].iter_rows(
        min_row=FIRST_DATA_ROW, max_row=FIRST_DATA_ROW + n - 1,
        values_only=True)]
    blank = [codes[i] for i in range(n)
             if not np.isfinite(published("Effects", 10)[i])]
    check("and the one industry the ONS refuses to publish is the right one",
          blank == ["L68A"],
          f"{blank} — Owner-Occupiers' Housing, output "
          f"£{X[codes.index('L68A')]:,.0f} million. Imputed rent has no "
          f"transaction, no employment and no supply chain, so its multiplier "
          f"would be an artefact of the imputation. The engine computes one; "
          f"the ONS's refusal is the better judgement and any report this "
          f"project emits should carry the same exclusion")

    # ---- the induced half: defined, declined, and one false friend -------
    print()
    ext = ROOT / "library" / "extracted"
    import re as _re
    q = (ext / "NSO_UK_01_ONS_IOAT_QMI.txt")
    m = (ext / "NSO_UK_02_ONS_IOAT_methods.txt")
    if q.exists() and m.exists():
        fq = _re.sub(r"\s+", " ", q.read_text())
        fm = _re.sub(r"\s+", " ", m.read_text())
        check("the induced effect IS defined in the loaded set — the entry says "
              "it is not",
              "increased compensation of employees being re-spent on goods and "
              "services" in fm,
              "NSO_UK_02: the induced effect 'measures the effect on the "
              "economy of increased compensation of employees being re-spent "
              "on goods and services'. A definition, from the office that "
              "produced this fixture. What is missing is a procedure, not a "
              "concept")
        check("and the same office states plainly that it does not compute them",
              "which we do not currently produce" in fm
              and "They do not include type 2 effects and multipliers" in fq,
              "NSO_UK_02: 'The induced effect can be estimated by type two "
              "multipliers, which we do not currently produce.' NSO_UK_01 says "
              "the published tables exclude them. So the fixture's silence on "
              "type II is a stated editorial choice, not an omission — and "
              "every multiplier this project derives from it is type I by "
              "construction")
        check("with the user demand acknowledged, which is why it stays open",
              "interested in industry-by- industry tables, and type two "
              "multipliers" in fm,
              "NSO_UK_02 records that many users want exactly this. The gap is "
              "known to the producer and unfilled by them")

    c2 = ext / "CORE_002_SNA2025_CH07_Production_Account.txt"
    if c2.exists():
        f2 = _re.sub(r"\s+", " ", c2.read_text())
        check("and CORE_002's four hits for 'type II' are a false friend",
              bool(_re.search(r'described as .type II. products', f2)),
              "they are type II *products* — goods whose holding generates "
              "output of storage — and have nothing to do with multipliers. "
              "Checked so a future grep does not send anyone to SNA ch. 7")

    print()
    print("    STILL NOT SPECIFIED, and it is the half that matters most: both")
    print("    sheets are headed Type I. Nothing here or in the loaded")
    print("    methodological set defines the INDUCED effect CORE_005 ¶36.34,")
    print("    p. 1015 names — no type II, no consumption closure, no rule for")
    print("    endogenising households. A closed type I is not a type II.")

    print("\n" + "=" * 78)
    if FAIL:
        print(f"{len(FAIL)} check(s) FAILED: {', '.join(FAIL)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
