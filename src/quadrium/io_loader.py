"""
Loading real tables from disk (MVP_0.1 §4).

Four entry points, because there are four situations:

`load_io_table()`   the project's own interchange format — labelled blocks and a
                    `metadata` sheet. Use this for tables you control.

`load_uk_analytical_iot()`  an adapter for the specific published workbook the
                    project ships as its fixture. Real national tables do not
                    come in anyone's interchange format, and pretending
                    otherwise is how loaders acquire silent assumptions. The
                    layout is hard-coded, documented, and asserted on load.

`load_ine_tio()`    the same, for the Spanish table. Product by product, not
                    industry by industry, and split across three sheets rather
                    than one.

`load_ine_tod()`    the Spanish SUPPLY AND USE tables — 110 products by 81
                    activities, the finest the INE publishes, and the only
                    fixture in the project where the valuation identities can
                    be checked at all. It returns a `SupplyUseTables`, not an
                    `IOTable`: a supply-use pair is rectangular and has had no
                    assumption about secondary production applied to it yet.

Each national adapter is written separately on purpose. The two published
tables share every accounting identity and almost no layout convention, and the
one heuristic that could be shared — "a final-demand column whose code is a
strict prefix of another's is a subtotal of it" — is defeated by the INE, which
names its columns in Spanish prose and gives them no codes at all. The concept
generalises; the detection rule does not. So `load_ine_tio` declares its
subtotal groupings and then checks them, which is stronger than the UK
heuristic rather than weaker.

WHAT THE LOADER REFUSES TO DO
-----------------------------
It never repairs. If the table does not balance, it says by how much and stops,
because a table that does not balance is not a table — every number downstream
would inherit the discrepancy silently. MVP_0.1 §5 step 2 makes the original
validation a gate, not a report.

The one crack in that door is `load_ine_tio(unbalanced="residual_column")`, and
it is a crack the caller has to open by name. The INE's interior table genuinely
does not balance for one product (OQ-D-04); the option puts the difference in a
labelled RESIDUAL column that travels into every downstream report, rather than
into a cell that would read as observed.

It never reads the reference period from the filename. The project's own fixture
shipped for twenty-four versions as `UK IO 2022.csv`, was not a CSV, and was for
2023 (`library/specs/D_open_questions.md` OQ-D-01, closed at v1.25 and the file
renamed once the owner confirmed the vintage). The INE's is named
`cne_tio_22.xlsx` and carries only two digits. **A truthful filename today is not
a reason to start trusting filenames** — the rename removed one instance, not the
class.
"""

from __future__ import annotations

import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import numpy as np

from .precision import (assertable_tolerance,
                        assertable_tolerance_mixed,
                        printed_decimals)
from .models import (AllocationKey, IOTable, ProxyStrength,
                     SupplyUseTables)


class LoaderError(ValueError):
    pass


def _open_workbook(path: Path):
    """openpyxl refuses a .csv extension. Copy to a temporary .xlsx rather than
    renaming the user's file."""
    import openpyxl
    path = Path(path)
    if not path.exists():
        raise LoaderError(f"no such file: {path}")
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "wb.xlsx"
        shutil.copy(path, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        sheets = {name: [list(r) for r in wb[name].iter_rows(values_only=True)]
                  for name in wb.sheetnames}
    return sheets


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


# ---------------------------------------------------------------------------
# The UK analytical IOT
# ---------------------------------------------------------------------------

# Layout of the "IOT" sheet, 0-based, established by inspection and re-asserted
# on every load. Mirrors library/validators/run_uk_iot.py, which is where it was
# first worked out.
# The ONS workbook announces its own shape and this loader reads it, because
# the shape is not stable. The 2023 edition merged `CPA_C254` (weapons and
# ammunition) into `CPA_C25`, so the 2022 tables are 105 x 105 and the 2023
# ones 104 x 104; the industry-by-industry edition labels its axis `SIC` /
# `Industry` where the product-by-product one says `CPA` / `Product`. Fixed row
# and column numbers described one file of nine and read the other eight one
# line out of true, which the table's own totals then reported as the DATA
# failing to balance — Construction "off by 406,662" in a file that balances
# exactly.
#
# So: the two axes are found by the `_T` total the sheet prints at the end of
# each, every primary-input row is found by the label beside it, and the row
# codes are required to equal the column codes before anything is read.
_UK_ROWS = {
    "imports": "use of imported products, cif",
    "taxes_products": "taxes less subsidies on products",
    "ic_purchasers": "total intermediate use at purchaser's prices",
    "compensation": "compensation of employees",
    "gos": "gross operating surplus and mixed income",
    "other_taxes": "taxes less subsidies on production",
    "gva": "gross value added",
    "output": "total output at basic prices",
}


def _uk_norm(s) -> str:
    """Lowercase, collapse whitespace — including the sheet's non-breaking
    spaces, which are why several of these labels do not compare equal to
    themselves after an ordinary `.strip()`."""
    import re as _re
    return _re.sub(r"\s+", " ",
                   str(s if s is not None else "").replace("\xa0", " ")
                   ).strip().lower()


def _uk_layout(R, name: str) -> dict:
    """Where the blocks are on the ONS `IOT` sheet, read off the sheet."""
    header = R[3] if len(R) > 3 else []
    end_col = next((j for j in range(2, len(header))
                    if _uk_norm(header[j]) == "_t"), None)
    end_row = next((i for i in range(6, len(R))
                    if _uk_norm(R[i][0] if R[i] else "") == "_t"), None)
    tu_col = next((j for j in range(2, len(header))
                   if _uk_norm(header[j]) == "tu"), None)
    if end_col is None or end_row is None or tu_col is None:
        raise LoaderError(
            f"{name}: the 'IOT' sheet does not print the totals this loader "
            f"navigates by — `_T` at the end of the sector rows and columns "
            f"and `TU` at the end of the final-demand block. Without them "
            f"there is nothing to find the blocks from except fixed offsets, "
            f"and those describe one edition of this workbook.")

    codes = [str(R[i][0] or "").strip() for i in range(6, end_row)]
    head = [str(header[j] or "").strip() for j in range(2, end_col)]
    if codes != head:
        first = next((k for k in range(min(len(codes), len(head)))
                      if codes[k] != head[k]), min(len(codes), len(head)))
        raise LoaderError(
            f"{name}: the 'IOT' sheet's row codes are not its column codes. "
            f"{len(codes)} rows against {len(head)} columns, first "
            f"disagreement at position {first + 1}: row "
            f"{codes[first] if first < len(codes) else '—'!r}, column "
            f"{head[first] if first < len(head) else '—'!r}. A symmetric "
            f"table has one classification on both axes; this one does not.")

    rows: dict[str, int] = {}
    for i in range(end_row, len(R)):
        key = _uk_norm(R[i][1] if len(R[i]) > 1 else "")
        if key and key not in rows:
            rows[key] = i
    out = dict(first_row=6, end_row=end_row, first_col=2, end_col=end_col,
               row_total=end_row, first_fd_col=end_col + 1, tu_col=tu_col,
               axis=str(R[5][0] or "").strip() or "SIC",
               axis_name=str(R[5][1] or "").strip() or "Industry")
    for key, label in _UK_ROWS.items():
        if label not in rows:
            raise LoaderError(
                f"{name}: the 'IOT' sheet has no primary-input row labelled "
                f"{label!r} below its sector block. What it does have: "
                f"{', '.join(sorted(k for k in rows if k)[:8])}…")
        out["row_" + key] = rows[label]
    return out


def load_uk_analytical_iot(path: Path | str) -> IOTable:
    """Load the UK Input-Output Analytical Tables workbook into an `IOTable`.

    Domestic use, at basic prices, GBP million. Both editions the ONS
    publishes are read: **industry by industry**, whose axis the sheet labels
    `SIC` / `Industry`, and **product by product**, labelled `CPA` / `Product`.
    The size is whatever the file says — 104 x 104 for 2023, 105 x 105 for 2016
    to 2022, because the 2023 edition merged `CPA_C254` into `CPA_C25`.

    A note on what goes into `VA`. The model's column identity is
    `Z.sum(0) + VA.sum(0) == X`. For a *domestic* IOT that identity reads

        domestic intermediate + imported intermediate
            + taxes less subsidies on products + GVA  =  output

    so `VA` here holds five rows, of which the first two — imports and taxes
    less subsidies on products — are **not value added** in the SNA sense. They
    are primary inputs to the column of a domestic table. The labels say so;
    do not sum the block and call it GVA.
    """
    path = Path(path)
    sheets = _open_workbook(path)
    if "IOT" not in sheets:
        raise LoaderError(f"{path.name} has no 'IOT' sheet; found: "
                          f"{', '.join(sheets)}")
    R = sheets["IOT"]
    L = _uk_layout(R, path.name)

    rows = range(L["first_row"], L["end_row"])
    cols = range(L["first_col"], L["end_col"])
    codes = [str(R[i][0]) for i in rows]
    names = [str(R[i][1]) for i in rows]
    n = len(codes)

    Z = np.array([[_num(R[i][j]) for j in cols] for i in rows])
    X = np.array([_num(R[L["row_output"]][j]) for j in cols])

    # Final demand. The published sheet mixes components with SUBTOTALS of those
    # components: `P3 S1` (final consumption expenditure) is the sum of
    # `P3 S13`, `P3 S14` and `P3 S15`. Summing the block as printed
    # double-counts household consumption -- for Owner-Occupiers' Housing that
    # is GBP 259,330 million on one row, and the row identity fails by exactly
    # that amount. Aggregates are dropped, not special-cased: a column whose
    # code is a strict prefix of another column's code is a subtotal of it.
    all_fd = list(range(L["first_fd_col"], L["tu_col"]))
    all_codes = [str(R[3][j]) for j in all_fd]
    aggregate = {j for j, c in zip(all_fd, all_codes)
                 if any(o != c and o.startswith(c) for o in all_codes)}
    fd_cols = [j for j in all_fd if j not in aggregate]
    dropped = [c for j, c in zip(all_fd, all_codes) if j in aggregate]
    Y_labels = [str(R[3][j]) for j in fd_cols]
    Y = np.array([[_num(R[i][j]) for j in fd_cols] for i in rows])

    VA_labels = ["Imports of goods and services (cif)",
                 "Taxes less subsidies on products",
                 "Compensation of employees",
                 "Gross operating surplus and mixed income",
                 "Other taxes less subsidies on production"]
    VA = np.array([[_num(R[L[k]][j]) for j in cols] for k in
                   ("row_imports", "row_taxes_products", "row_compensation",
                    "row_gos", "row_other_taxes")])

    # The derivation is checked against the totals the ONS itself prints, so a
    # block found one line out of true fails here and not three identities
    # later as an accusation against the data.
    printed_t = np.array([_num(R[L["row_total"]][j]) for j in cols])
    printed_tu = np.array([_num(R[i][L["tu_col"]]) for i in rows])
    printed_gva = np.array([_num(R[L["row_gva"]][j]) for j in cols])
    for what, a, b in (
            ("the printed `_T` row is the sum of the sector columns",
             printed_t, Z.sum(0)),
            ("the printed `TU` column is intermediate plus final demand",
             printed_tu, Z.sum(1) + Y.sum(1)),
            ("the printed GVA row is its three components", printed_gva,
             VA[2] + VA[3] + VA[4])):
        d = float(np.abs(a - b).max())
        bound = max(1e-6, float(assertable_tolerance(
            np.concatenate([Z.ravel(), Y.ravel(), VA.ravel()]), n + 1)))
        if d > bound:
            raise LoaderError(
                f"{path.name}: {what} — and it is not, by {d:,.3f} against a "
                f"bound of {bound:,.3f}. The blocks are located from the "
                f"labels this sheet prints, so this is the loader having found "
                f"the wrong ones rather than the data failing to add up.")

    menu = []
    if "Menu" in sheets:
        menu = [str(c) for r in sheets["Menu"][:6] for c in r if c]

    year = _infer_year(menu)
    pxp = L["axis"].upper().startswith("CPA")
    table = IOTable(
        table_id=f"UK-IOT-{'PXP' if pxp else 'IXI'}-{year}",
        country="United Kingdom", year=year,
        unit="GBP million, current prices, basic prices",
        classification=(f"{'CPA 2008' if pxp else 'SIC 2007'} "
                        f"({n} {'products' if pxp else 'industries'})"),
        sector_codes=codes, sector_labels=names,
        Z=Z, Y=Y, Y_labels=Y_labels, VA=VA, VA_labels=VA_labels, X=X,
        source=f"ONS, {menu[1] if len(menu) > 1 else path.name}",
        retrieved_at=datetime.now(timezone.utc),
        notes=(f"{'Product-by-product' if pxp else 'Industry-by-industry'}, "
               f"domestic use, basic prices, {n} x {n}. The first two VA rows "
               f"are imports and taxes on products, not value added — see the "
               f"loader docstring. Reference year read from the Menu sheet, "
               f"never from the filename (OQ-D-01); the size and both axes "
               f"read from the sheet's own `_T` totals, never from a fixed "
               f"offset — the ONS changed both between editions."
               + (f" Dropped {len(dropped)} final-demand subtotal column(s) "
                  f"({', '.join(dropped)}) to avoid double counting."
                  if dropped else "")))
    _assert_balances(table, path.name)
    return table


def _infer_year(menu: list[str]) -> int:
    """Read the reference year from the workbook's own banner, not the filename."""
    import re
    for line in menu:
        m = re.search(r"Tables,\s*(\d{4})", line) or re.search(r"\b(20\d\d)\b", line)
        if m:
            return int(m.group(1))
    raise LoaderError("could not read the reference year from the Menu sheet. "
                      "Refusing to guess it from the filename — the project's "
                      "own fixture is named 2022 and is for 2023 (OQ-D-01).")


def _assert_balances(table: IOTable, name: str, inherited: float = 0.0) -> None:
    """Gate, not report (MVP_0.1 §5 step 2).

    The bound is the table's own precision, not a fraction of its largest
    sector. `1e-6 * max|X|` said nothing about how finely the publisher printed
    anything, and it refused Portugal's 2020 symmetric table and Spain's own
    2020 -- both ordinary published files -- over residues an order of
    magnitude inside what their rounding produces. `OQ-B-02`, closed at v1.57:
    an identity summing `n` cells published to `d` decimals cannot be checked
    more tightly than `0.5*10^-d*n`.

    This still catches what it is for. A dropped final-demand column or a block
    pasted at the wrong offset moves a row by thousands; the INE's `interior`
    table, which genuinely does not close, is out by 4,921.6 against a floor of
    about 6 (`OQ-D-04`), and is still refused.
    """
    row = np.abs(table.Z.sum(axis=1) + table.Y.sum(axis=1) - table.X)
    col = np.abs(table.Z.sum(axis=0) + table.VA.sum(axis=0) - table.X)
    # Each block is measured on its own, because a table need not be printed to
    # one precision throughout and the ONS's is not: its intermediate block is
    # full precision and its final demand, output and total use are integers.
    # Pooled, the 105 unrounded cells outvote the 10 rounded ones and the whole
    # identity is judged at float64 accumulation — 5.7e-06 where what the file
    # can distinguish is 5.0. See `precision.assertable_tolerance_mixed`.
    tol = max(1e-6, assertable_tolerance_mixed(
        (table.Z.ravel(), table.n),
        (table.Y.ravel(), table.Y.shape[1]),
        (table.X.ravel(), 1)))
    col_tol = max(1e-6, assertable_tolerance_mixed(
        (table.Z.ravel(), table.n),
        (table.VA.ravel(), table.VA.shape[0]),
        (table.X.ravel(), 1)))
    tol = max(tol, col_tol)
    # A PUBLISHER'S OWN UNCLOSED BOOKS ARE NOT THIS LOADER'S ERROR, and the
    # caller has to have MEASURED the residue against the publisher's own
    # printed totals to claim one -- not guessed a number that makes the gate
    # pass. IDESCAT's 2021 table is the case that added this: its cells are
    # full precision, so the derived floor is zero, and its own `Total usos`
    # column differs from its own `Producció` row by up to 0.84. Two numbers
    # IDESCAT printed, disagreeing with each other.
    tol += max(0.0, float(inherited))
    parts = [f"{'unrounded' if printed_decimals(b) is None else str(printed_decimals(b)) + ' dp'} {lbl}"
             for lbl, b in (("interior", table.Z.ravel()),
                            ("final demand", table.Y.ravel()),
                            ("output", table.X.ravel()))]
    basis = "OQ-B-02, block by block: " + ", ".join(parts)
    if row.max() > tol or col.max() > tol:
        i, j = int(row.argmax()), int(col.argmax())
        raise LoaderError(
            f"{name} does not balance and will not be loaded.\n"
            f"  worst row: {table.sector_codes[i]} off by {row[i]:,.3f} "
            f"({table.sector_labels[i]})\n"
            f"  worst col: {table.sector_codes[j]} off by {col[j]:,.3f}\n"
            f"  tolerance {tol:,.3f} ({basis})\n"
            f"A table that does not balance is not a table: every number "
            f"downstream would inherit the discrepancy without saying so.")


# ---------------------------------------------------------------------------
# The Spanish TIO (INE)
# ---------------------------------------------------------------------------

# Layout of the INE workbook, 0-based. Nothing here is trusted because it was
# read off the screen: `_assert_ine_layout` re-derives every one of these
# indices from an accounting identity on each load, so a moved row fails loudly
# instead of quietly becoming the wrong number.
_INE = dict(
    col_label=1, first_row=9, n=64, first_col=2,
    row_uses_basic=74,        # Total de empleos a precios básicos
    row_taxes_products=76,    # Impuestos netos sobre los productos
    row_uses_purch=77,        # Total de empleos a precios de adquisición
    row_ic_purch=79,          # Consumos intermedios a precios de adquisición
    row_compensation=80,
    row_wages=81, row_social=82,        # components of compensation
    row_other_taxes=83,
    row_gos=84,               # Excedente de explotación bruto / Renta mixta
    row_gva=85,               # subtotal of compensation + other taxes + gos
    row_output=86,            # Producción a precios básicos
    row_imports=87, row_imports_eu=88, row_imports_nonue=89,
    row_supply=90,            # Oferta a precios básicos
)

# Final demand. The same trap as the UK sheet — subtotals sit next to their own
# components — but the UK detection rule does not transfer: the INE labels its
# columns in Spanish prose, and there are no codes to compare as prefixes. The
# concept generalises; the detection rule does not. So the grouping is declared
# here and then *checked* on load: each subtotal must equal the sum of its
# declared components, or nothing is loaded.
_INE_FD_GROUPS = (
    (68, (69, 70, 71), "Total gasto en consumo final"),
    (72, (73, 74), "Formación bruta de capital"),
    (75, (76, 77), "Total exportaciones"),
)
_INE_FD_COLS = (69, 70, 71, 73, 74, 76, 77)
_INE_COL_INTERMEDIATE, _INE_COL_FD, _INE_COL_USES = 67, 78, 79

# THE SAME WORKBOOK IN TWO SHAPES, and the difference is two columns.
#
# From 2021 the INE splits exports into `Exportaciones Unión Europea` (76) and
# `Exportaciones terceros Países` (77). From 2016 to 2020 it publishes only
# `Total exportaciones`, at 75, and everything after it sits two columns
# earlier. Nothing else moves: the intermediate block, the consumption columns
# and the capital columns are identical.
#
# So exports are a SUBTOTAL in the newer shape and a LEAF in the older one,
# which is why the older map has one group fewer rather than a group with one
# member -- a subtotal that equals itself checks nothing.
#
# Detected from the header the workbook prints, not from the file name or the
# sheet count: the year is not what decides this, the publication is.
_INE_FD_GROUPS_NOSPLIT = (
    (68, (69, 70, 71), "Total gasto en consumo final"),
    (72, (73, 74), "Formación bruta de capital"),
)
_INE_FD_COLS_NOSPLIT = (69, 70, 71, 73, 74, 75)
_INE_COLS_NOSPLIT = dict(intermediate=67, fd=76, uses=77)


def _ine_columns(t1):
    """Which column map this workbook uses, read off its own header row.

    Returns `(fd_cols, groups, intermediate, fd_total, uses)`.
    """
    header = t1[7] if len(t1) > 7 else []

    def label(j):
        return str(header[j] or "").strip().lower() if j < len(header) else ""

    if label(76).startswith("exportaciones"):
        return (_INE_FD_COLS, _INE_FD_GROUPS, _INE_COL_INTERMEDIATE,
                _INE_COL_FD, _INE_COL_USES)
    if label(75).startswith("total exportaciones") and \
            label(76).startswith("total demanda final"):
        c = _INE_COLS_NOSPLIT
        return (_INE_FD_COLS_NOSPLIT, _INE_FD_GROUPS_NOSPLIT,
                c["intermediate"], c["fd"], c["uses"])
    raise LoaderError(
        f"the INE workbook's final-demand columns match neither layout this "
        f"loader knows. Column 75 reads {label(75)!r}, 76 {label(76)!r}. The "
        f"two known shapes differ only in whether exports are split into "
        f"European Union and third countries; a third shape needs its own map "
        f"here, checked against the workbook's own subtotals.")

_INE_TOL = 1e-3   # million EUR. Observed residuals are ~1e-11; see M-058.


def _ine_block(R, rows, cols) -> np.ndarray:
    return np.array([[_num(R[i][j]) if j < len(R[i]) else 0.0 for j in cols]
                     for i in rows])


def _ine_row(R, i, cols) -> np.ndarray:
    return np.array([_num(R[i][j]) if j < len(R[i]) else 0.0 for j in cols])


def _ine_vintage(S) -> str:
    """Which of the INE's two workbook layouts this is, from its own index.

    The INE publishes the same statistical revision in two shapes, and the
    difference is what `Tabla 2` MEANS:

        2021 onward   Tabla 2 = the domestic input-output table
                      Tabla 3 = the imports table
        2016 to 2020  Tabla 2 = technical coefficients
                      Tabla 3 = Leontief inverse coefficients

    So for 2016-2020 there is no published domestic/imports split at all —
    only the total table. Reading `Lista_Tablas`, which the loader already
    consults for the reference year, is what tells the two apart; guessing
    from the sheet count would work today and break on the next revision.

    Until 2026-08-25 the loader assumed the newer shape and validated
    `Tabla1 == Tabla2 + Tabla3` on every file, so all five older years failed
    with a layout complaint — off by 29,086.92 for 2020 — when nothing about
    the layout was wrong and the total table was sitting there readable.
    """
    for row in S.get("Lista_Tablas", [])[:16]:
        cells = " ".join(str(c) for c in row if c is not None).lower()
        if "tabla 2" in cells:
            if "interior" in cells:
                return "split"
            if "coeficientes" in cells:
                return "total_only"
    return "split"


def _assert_ine_layout(S, cols, rows, vintage: str = "split",
                       colmap=None) -> None:
    """Re-derive the hard-coded layout from the workbook's own identities.

    Each check pins one hard-coded index. If the INE moves a row, the identity
    it participates in stops holding and the load stops here, rather than
    silently reading (say) gross value added where output was meant.
    """
    L = _INE
    t1 = S["Tabla1"]
    split = vintage == "split"
    t2, t3 = (S["Tabla2"], S["Tabla3"]) if split else (None, None)
    Zt = _ine_block(t1, rows, cols)
    r1 = lambda k: _ine_row(t1, L[k], cols)

    checks = []
    if split:
        Zd, Zm = (_ine_block(x, rows, cols) for x in (t2, t3))
        checks.append(
            ("Tabla1 == Tabla2 + Tabla3 (intermediate block)", Zt, Zd + Zm))
    split_published = bool(
        np.abs(_ine_row(t1, L["row_imports_eu"], cols)).sum()
        + np.abs(_ine_row(t1, L["row_imports_nonue"], cols)).sum())
    if split_published:
        checks.append(("row 'Importaciones'", r1("row_imports"),
                       r1("row_imports_eu") + r1("row_imports_nonue")))
    checks += [
        ("row 'Total de empleos a precios básicos'", r1("row_uses_basic"),
         Zt.sum(0)),
        ("row 'Total de empleos a precios de adquisición'",
         r1("row_uses_purch"), r1("row_uses_basic") + r1("row_taxes_products")),
        ("row 'Consumos intermedios a precios de adquisición'",
         r1("row_ic_purch"), r1("row_uses_purch")),
        ("row 'Remuneración de los asalariados'", r1("row_compensation"),
         r1("row_wages") + r1("row_social")),
        ("row 'Valor añadido bruto a precios básicos'", r1("row_gva"),
         r1("row_compensation") + r1("row_other_taxes") + r1("row_gos")),
        # `Importaciones == UE + terceros países` pins those two rows, and it
        # can only pin rows that carry something. From 2016 to 2020 the INE
        # PRINTS BOTH LABELS AND LEAVES BOTH ROWS EMPTY -- it publishes total
        # imports and not the geographic split -- so the check was comparing a
        # populated row against zero and failing by 33,653.70 on 2020, 58 of 64
        # columns at once. Nothing about the layout was wrong.
        #
        # The rows are used for nothing else, so an absent split costs nothing
        # but this check, and skipping it is recorded on the table rather than
        # passed over.
        ("row 'Oferta a precios básicos'", r1("row_supply"),
         r1("row_output") + r1("row_imports")),
        ("column identity pins 'Producción a precios básicos'", r1("row_output"),
         Zt.sum(0) + r1("row_taxes_products") + r1("row_gva")),
    ]
    fd_cols, groups, col_int, col_fd, _ = colmap or _ine_columns(t1)
    tables = ([("Tabla1", t1), ("Tabla2", t2), ("Tabla3", t3)] if split
              else [("Tabla1", t1)])
    for name, tbl in tables:
        Y = _ine_block(tbl, rows, fd_cols)
        Z = _ine_block(tbl, rows, cols)
        checks.append((f"{name}: final-demand components sum to 'Total demanda "
                       f"final'", Y.sum(1), _ine_row_col(tbl, rows, col_fd)))
        checks.append((f"{name}: 'Total demanda intermedia' column",
                       Z.sum(1), _ine_row_col(tbl, rows, col_int)))
        for sub, comp, lab in groups:
            checks.append((f"{name}: subtotal {lab!r}",
                           _ine_row_col(tbl, rows, sub),
                           sum(_ine_row_col(tbl, rows, c) for c in comp)))

    for what, a, b in checks:
        d = float(np.abs(np.asarray(a) - np.asarray(b)).max())
        if d > _INE_TOL:
            raise LoaderError(
                f"the INE workbook's layout no longer matches the one this "
                f"loader hard-codes.\n  failed check: {what}\n"
                f"  off by {d:,.4f} (tolerance {_INE_TOL:g}, million EUR)\n"
                f"Fix _INE / _INE_FD_GROUPS in io_loader.py against the actual "
                f"sheet. Do not widen the tolerance: these identities hold "
                f"exactly in the published file.")


def _ine_row_col(R, rows, j) -> np.ndarray:
    return np.array([_num(R[i][j]) if j < len(R[i]) else 0.0 for i in rows])


def _ine_codes(R, rows) -> tuple[list[str], list[str]]:
    """Split '44 bis. Alquileres imputados…' into code and label.

    The numeric index row runs 1..64, but the INE's own product numbering does
    not: it carries a '44 bis'. Tabla8 maps the *label* numbering to CPA, so
    that is the one kept as the code.
    """
    # The separator is not reliably ". ". The supply-use workbook writes
    # `5 .Pescado y otros productos de la pesca` -- space BEFORE the dot, none
    # after -- so a `partition(". ")` drops that product and every one after it.
    # Matched on the code instead, which is what actually has a shape.
    import re as _re
    pat = _re.compile(r"^\s*(\d{1,3}(?:\s+bis)?)\s*\.\s*(.+)$")
    codes, labels = [], []
    for i in rows:
        # Whitespace is collapsed FIRST. Some labels carry an embedded newline
        # -- product 26 of the input-output table runs onto a second line --
        # and `.` does not cross one, so a regex written without this silently
        # rejects a label the old `partition(". ")` had accepted.
        raw = _re.sub(r"\s+", " ", str(R[i][_INE["col_label"]] or "")).strip()
        m = pat.match(raw)
        if not m:
            raise LoaderError(f"row {i + 1} is not a numbered product label: "
                              f"{raw!r}")
        codes.append(_re.sub(r"\s+", " ", m.group(1)).strip())
        labels.append(m.group(2).strip())
    if len(set(codes)) != len(codes):
        dup = sorted({c for c in codes if codes.count(c) > 1})
        raise LoaderError(f"duplicate INE product codes: {', '.join(dup)}")
    return codes, labels


def _infer_ine_year(S) -> int:
    """Read the reference year from the workbook's own index sheet.

    Never from the filename: the INE names this file `cne_tio_22.xlsx`, and a
    two-digit year in a filename is exactly the sort of thing OQ-D-01 exists to
    stop the loader from believing.
    """
    import re
    for row in S.get("Lista_Tablas", [])[:12]:
        for cell in row:
            # Two workbooks, two banners: "Tablas Input-Output 2022" and
            # "Tablas de origen y de destino 2022". Deliberately NOT a bare
            # four-digit search — the same sheet says "Revisión Estadística
            # 2024", which is the vintage of the methodology, not of the data.
            m = re.search(r"(?:Input-Output|origen y (?:de )?destino)\s+(\d{4})",
                          str(cell or ""), re.IGNORECASE)
            if m:
                return int(m.group(1))
    raise LoaderError("could not read the reference year from the "
                      "'Lista_Tablas' sheet. Refusing to take it from the "
                      "filename, which carries only two digits (OQ-D-01).")


def load_ine_tio(path: Path | str, variant: str = "interior",
                 unbalanced: str = "refuse") -> IOTable:
    """Load the INE's Spanish input-output table into an `IOTable`.

    The table is **product by product** — unlike the UK fixture, which is
    industry by industry. Both are symmetric, so the engine handles them the
    same way, but a disaggregation of "restaurants" means a different object in
    each, and results are not comparable across the two without saying so.

    `variant` picks which of the INE's three tables is loaded:

    `"interior"`  (default) Tabla 2, domestic output only. This is the
        analytically right one and the one that matches the UK pilot: imports
        enter as a primary input row, so Leontief multipliers count only
        domestic production. As with the UK loader, the first two `VA` rows —
        imported intermediate inputs and taxes less subsidies on products — are
        **not value added**. Do not sum the block and call it GVA.

    `"total"`  Tabla 1, domestic and imported flows together, with imports
        carried as a negative final-demand column so the row identity closes.
        Multipliers from it treat an imported input as if it had been produced
        in Spain, and overstate domestic effects. Offered because it balances
        exactly with nothing derived, but it is not the table you want for
        impact work.

    A DISCREPANCY IN THE PUBLISHED SOURCE
    -------------------------------------
    The interior table does not satisfy the row identity for one product.
    Domestic uses of agricultural products fall 4,921.6 million EUR short of
    domestic agricultural output, and uses of *imported* agricultural products
    exceed recorded imports by exactly the same amount. The two errors cancel,
    which is why Tabla 1 balances to the last decimal while Tabla 2 does not.

    This is a property of the data, not of the parse: every block here is
    verified against the INE's own published totals, and the INE's own
    coefficient matrices (Tabla 4 and Tabla 5) confirm that
    `Producción a precios básicos` is the intended denominator, to zero
    difference in all 64 branches. What the INE does not publish is why —
    OQ-D-04.

    `unbalanced` says what to do about it:

    `"refuse"`  (default) stop, and report the product and the amount. A table
        that does not balance is not a table.

    `"residual_column"`  add one final-demand column holding
        `X - (Z.sum(1) + Y.sum(1))` per product. The residual is *computed
        here, not published by the INE*; the column is labelled as such and the
        amount is recorded in `notes`, so it can be seen in every downstream
        report rather than being absorbed into a cell that looks observed. This
        is the project's RESIDUAL status made explicit, not a repair.
    """
    if variant not in ("interior", "total"):
        raise LoaderError(f"variant must be 'interior' or 'total', "
                          f"not {variant!r}")
    if unbalanced not in ("refuse", "residual_column"):
        raise LoaderError(f"unbalanced must be 'refuse' or 'residual_column', "
                          f"not {unbalanced!r}")

    path = Path(path)
    S = _open_workbook(path)
    missing = [s for s in ("Tabla1", "Tabla2", "Tabla3") if s not in S]
    if missing:
        raise LoaderError(f"{path.name} is missing {', '.join(missing)}; "
                          f"found: {', '.join(S)}. This loader expects the "
                          f"INE's `cne_tio_YY.xlsx` layout.")

    L = _INE
    rows = range(L["first_row"], L["first_row"] + L["n"])
    cols = range(L["first_col"], L["first_col"] + L["n"])
    vintage = _ine_vintage(S)
    if vintage == "total_only" and variant != "total":
        raise LoaderError(
            f"{path.name} carries only the TOTAL table. In this vintage of "
            f"the INE's workbook `Tabla 2` is technical coefficients and "
            f"`Tabla 3` is the Leontief inverse — the domestic and imports "
            f"tables are not published for this year at all, so "
            f"variant={variant!r} has nothing to read.\n"
            f"Load it with variant='total', knowing what that means: an "
            f"imported input is treated as if it had been produced in Spain, "
            f"so multipliers from it overstate domestic effects. The INE "
            f"publishes the split from 2021 onward.")
    colmap = _ine_columns(S["Tabla1"])
    _assert_ine_layout(S, cols, rows, vintage, colmap)
    fd_cols, fd_groups, _col_int, _col_fd, _col_uses = colmap

    t1 = S["Tabla1"]
    t2 = S["Tabla2"] if vintage == "split" else t1
    t3 = S["Tabla3"] if vintage == "split" else None
    codes, labels = _ine_codes(t2, rows)
    r1 = lambda k: _ine_row(t1, L[k], cols)
    X = r1("row_output")

    fd_labels = [str(t2[7][j] or "").strip() for j in fd_cols]
    dropped = [str(t2[7][sub] or "").strip() for sub, _, _ in fd_groups]

    if variant == "interior":
        Z = _ine_block(t2, rows, cols)
        Y = _ine_block(t2, rows, fd_cols)
        Y_labels = list(fd_labels)
        VA_labels = ["Importaciones de bienes y servicios (consumos "
                     "intermedios importados)",
                     "Impuestos netos sobre los productos",
                     "Remuneración de los asalariados",
                     "Otros impuestos netos sobre la producción",
                     "Excedente de explotación bruto / Renta mixta"]
        VA = np.vstack([_ine_block(t3, rows, cols).sum(0),
                        r1("row_taxes_products"), r1("row_compensation"),
                        r1("row_other_taxes"), r1("row_gos")])
        scope = ("Producción interior (Tabla 2). Las importaciones son un input "
                 "primario, no valor añadido: las dos primeras filas del bloque "
                 "VA no son VAB.")
    else:
        Z = _ine_block(t1, rows, cols)
        Y = np.hstack([_ine_block(t1, rows, fd_cols),
                       -r1("row_imports").reshape(-1, 1)])
        Y_labels = fd_labels + ["Importaciones de bienes y servicios "
                                "(columna negativa, DERIVADA)"]
        VA_labels = ["Impuestos netos sobre los productos",
                     "Remuneración de los asalariados",
                     "Otros impuestos netos sobre la producción",
                     "Excedente de explotación bruto / Renta mixta"]
        VA = np.vstack([r1("row_taxes_products"), r1("row_compensation"),
                        r1("row_other_taxes"), r1("row_gos")])
        scope = ("Flujos totales (Tabla 1): interiores e importados juntos. Las "
                 "importaciones van como columna negativa de demanda final para "
                 "cerrar la identidad de fila. Los multiplicadores tratan un "
                 "input importado como si se produjese en España.")

    residual_note = ""
    gap = X - (Z.sum(1) + Y.sum(1))
    if unbalanced == "residual_column" and np.abs(gap).max() > _INE_TOL:
        worst = int(np.abs(gap).argmax())
        affected = int((np.abs(gap) > _INE_TOL).sum())
        Y = np.hstack([Y, gap.reshape(-1, 1)])
        Y_labels = Y_labels + ["Discrepancia estadística (RESIDUAL, calculada "
                               "por el cargador; el INE no la publica)"]
        residual_note = (
            f" Se añadió una columna de demanda final RESIDUAL, calculada aquí "
            f"y no publicada por el INE: {np.abs(gap).sum():,.1f} millones de "
            f"euros en total repartidos en {affected} producto(s), el mayor "
            f"{gap[worst]:,.1f} en «{labels[worst]}». Véase OQ-D-04.")

    year = _infer_ine_year(S)
    table = IOTable(
        table_id=f"ES-TIO-PXP-{year}-{variant}", country="España", year=year,
        unit="millones de euros, precios corrientes, precios básicos",
        classification="CPA 2008 (64 productos, numeración TIO del INE)",
        sector_codes=codes, sector_labels=labels,
        Z=Z, Y=Y, Y_labels=Y_labels, VA=VA, VA_labels=VA_labels, X=X,
        source=(f"INE, Contabilidad Nacional Anual de España, Revisión "
                f"Estadística 2024. Tablas Input-Output {year} ({path.name})"),
        retrieved_at=datetime.now(timezone.utc),
        notes=("Producto x producto, precios básicos. " + scope
               + f" Año de referencia leído de la hoja 'Lista_Tablas', nunca "
                 f"del nombre del fichero, que sólo lleva dos dígitos "
                 f"(OQ-D-01). Se descartaron {len(dropped)} columnas subtotal "
                 f"de demanda final ({', '.join(dropped)}) para no contar dos "
                 f"veces; cada subtotal se comprobó contra sus componentes "
                 f"antes de descartarlo." + residual_note))
    _assert_balances(table, f"{path.name} ({variant})")
    return table


# ---------------------------------------------------------------------------
# The Catalan symmetric table (IDESCAT)
# ---------------------------------------------------------------------------

# `mioc20YYts64.xlsx` carries FOUR tables of the same shape, one per origin of
# the inputs: all origins, Catalan, rest of Spain, rest of world. Only the
# CATALAN one is an input-output table -- see the docstring -- but all four are
# read, because the other three are where its imports come from and are the
# reason this file is worth having at all.
_MIOC_SHEETS = ("ts total", "ts Cat", "ts resta Esp", "ts resta món")

# Rows below the product block, by the number IDESCAT prints in column 1, and
# read by that number rather than by position: the block below the products is
# where a vintage is most likely to gain or lose a line.
_MIOC_TAXES, _MIOC_NONRES = 84, 87
_MIOC_VA = (89, 90, 91)
_MIOC_OUTPUT = 93


def load_idescat_mioc(path: Path | str) -> IOTable:
    """The Catalan symmetric table, at basic prices, domestic origin.

    WHY THIS FILE, AND WHY THE CATALAN SHEET
    ------------------------------------------
    IDESCAT publishes the same symmetric table four times over, split by where
    the inputs came from: `ts total`, `ts Cat`, `ts resta Esp`, `ts resta món`.
    That is a two-region interregional table already compiled, and it is why
    this is worth loading at all: a regionalisation of the Spanish national
    table onto Catalonia can be SCORED against it, at the same year and the
    same 63-branch classification, instead of being an opinion.

    Only the Catalan sheet is an input-output table, and that was measured
    rather than assumed. Against output at basic prices, the row identity comes
    out at

        ts Cat            0.8      <- rounding on 455,354
        ts resta Esp   35,741
        ts resta món   35,653
        ts total       12,303

    The other three cannot close because what they record is supplied by
    imports, not by Catalan production. So `Z` here is Catalan-origin, which is
    the domestic table, and the other two sheets come back inside `VA` as the
    imported inputs each branch buys.

    WHAT GOES IN THE VALUE-ADDED BLOCK, AND WHY IT IS NOT ALL VALUE ADDED
    ----------------------------------------------------------------------
    Column by column, `1·Z_catalan + VA = output` only closes when `VA` carries,
    besides compensation, other production taxes and gross operating surplus:

      * the intermediates bought from the rest of Spain and the rest of the
        world -- the COLUMN sums of the other two sheets, not the import rows
        94 and 95, which are indexed by product and belong to supply. Using
        those instead leaves the identity 6,588 out;
      * net taxes on products, a row of this sheet's intermediate block, as in
        the UK analytical table;
      * resident consumption abroad, zero in 2021 and carried anyway.

    With those it closes to **0.0000**. Labelled honestly, as the UK loader
    labels its own: these rows are not value added and the labels say so.

    ROW 45 IS NOT A BRANCH
    ------------------------
    `dels quals: rendes immobiliàries imputades` is an OF-WHICH memo inside real
    estate, with no CPA code, carrying zero. Sixty-four numbered rows, sixty-
    three branches. Reading the numbering at face value adds a phantom sector;
    that it is zero in 2021 is luck, not a reason. It is dropped for having no
    code and the note says so. The INE publishes the same quantity as a branch
    of its own, `44 bis`, so aligning the two means aggregating the INE's pair.

    Nothing here is an offset. IDESCAT numbers its rows, prints the CPA code
    beside each, and heads every block of columns, so both axes and every
    boundary announce themselves — the rule `OQ-D-01` made after the INE and
    ONS loaders each cost a defect for counting from a corner instead.
    """
    import re as _re

    path = Path(path)
    S = _open_workbook(path)
    missing = [n for n in _MIOC_SHEETS if n not in S]
    if missing:
        raise LoaderError(
            f"{path.name} is missing the sheet(s) {', '.join(missing)}; found: "
            f"{', '.join(S)}. This loader expects IDESCAT's "
            f"`mioc20YYts64.xlsx` layout, whose four sheets are one table per "
            f"origin of the inputs.")
    T, C = S["ts total"], S["ts Cat"]

    def num(R, i, j):
        row = R[i] if i < len(R) else ()
        v = row[j] if j < len(row) else None
        return 0.0 if v is None else float(v)

    def clean(v):
        return _re.sub(r"\s+", " ", str(v if v is not None else "")).strip()

    # --- both axes, off the codes the sheet prints
    hdr = next((i for i, r in enumerate(T)
                if r and any(clean(c) == "Codi" for c in r if c is not None)), None)
    if hdr is None:
        raise LoaderError(
            f"{path.name}: no row carries 'Codi', which is how the product "
            f"columns announce themselves in this layout.")
    col_of = {clean(c): j for j, c in enumerate(T[hdr])
              if c is not None and j > 2}
    prod_rows = [i for i in range(hdr + 1, len(T))
                 if len(T[i]) > 1 and T[i][1] is not None
                 and clean(T[i][1]) in col_of]
    codes = [clean(T[i][1]) for i in prod_rows]
    labels = [clean(T[i][2]) for i in prod_rows]
    n = len(codes)
    if n < 20 or list(col_of) != codes:
        raise LoaderError(
            f"{path.name}: the row codes are not the column codes ({n} rows "
            f"against {len(col_of)} columns, in this order). A symmetric table "
            f"has to be square and in one order.")
    dropped = [f"{clean(T[i][0])} {clean(T[i][2])}" for i in range(hdr + 1, len(T))
               if len(T[i]) > 2 and T[i][0] is not None
               and clean(T[i][0]).isdigit() and i not in prod_rows
               and clean(T[i][2]).lower().startswith("dels quals")]

    # --- final demand: from the CONSUM FINAL head to Total demanda final,
    #     dropping the subtotals printed among them, as the INE loader does.
    heads = {}
    for i in range(hdr):
        for j, c in enumerate(T[i] if i < len(T) else ()):
            if c is not None and clean(c):
                heads.setdefault(clean(c).lower(), j)
    fd_start = heads.get("consum final")
    lab_row = hdr + 1
    label_at = lambda j: clean(T[lab_row][j] if j < len(T[lab_row]) else None)
    fd_end = next((j for j in range(fd_start or 0, len(T[lab_row]))
                   if label_at(j).lower() == "total demanda final"), None)
    if fd_start is None or fd_end is None:
        raise LoaderError(
            f"{path.name}: the final-demand block is not delimited by a "
            f"'CONSUM FINAL' head and a 'Total demanda final' column, so where "
            f"it begins and ends cannot be read off the sheet.")
    fd_cols = [j for j in range(fd_start, fd_end)
               if not label_at(j).lower().startswith("total")
               and label_at(j).lower() != "formació bruta de capital"]
    fd_dropped = [label_at(j) for j in range(fd_start, fd_end) if j not in fd_cols]

    idx = [col_of[c] for c in codes]
    Z = np.array([[num(C, i, j) for j in idx] for i in prod_rows])
    Y = np.array([[num(C, i, j) for j in fd_cols] for i in prod_rows])
    Y_labels = [label_at(j) for j in fd_cols]

    num_of = {int(clean(T[i][0])): i for i in range(hdr, len(T))
              if len(T[i]) and T[i][0] is not None and clean(T[i][0]).isdigit()}
    want = (_MIOC_TAXES, _MIOC_NONRES, *_MIOC_VA, _MIOC_OUTPUT)
    absent = [k for k in want if k not in num_of]
    if absent:
        raise LoaderError(
            f"{path.name}: the rows numbered {', '.join(map(str, absent))} are "
            f"not on the 'ts total' sheet, and they carry the column identity.")

    # The imported inputs each BRANCH buys: the column sums of the other two
    # sheets. Rows 94 and 95 are imports by PRODUCT and belong to supply --
    # using them here leaves the identity 6,588 out.
    va_rows, VA_labels = [], []
    for sheet, lab in (("ts resta Esp", "Consums intermedis importats de la "
                                        "resta d'Espanya (NO és valor afegit)"),
                       ("ts resta món", "Consums intermedis importats de la "
                                        "resta del món (NO és valor afegit)")):
        M = S[sheet]
        va_rows.append(np.array([[num(M, i, j) for j in idx]
                                 for i in prod_rows]).sum(axis=0))
        VA_labels.append(lab)
    for k, extra in ((_MIOC_TAXES, " (NO és valor afegit)"), (_MIOC_NONRES, "")):
        va_rows.append(np.array([num(T, num_of[k], j) for j in idx]))
        VA_labels.append(clean(T[num_of[k]][2]) + extra)
    for k in _MIOC_VA:
        va_rows.append(np.array([num(T, num_of[k], j) for j in idx]))
        VA_labels.append(clean(T[num_of[k]][2]))
    VA = np.vstack(va_rows)
    X = np.array([num(T, num_of[_MIOC_OUTPUT], j) for j in idx])

    # The residue IDESCAT itself admits: its own `Total usos` column against
    # its own `Producció` row, branch by branch. Measured here rather than
    # assumed, and declared so the gate below judges this loader's work and not
    # the publisher's arithmetic.
    published_uses = np.array([num(C, i, fd_end + 1) for i in prod_rows])
    residue = float(np.abs(published_uses - X).max())

    year = _mioc_year(T, path)
    table = IOTable(
        table_id=f"CAT-MIOC-PXP-{year}", country="Catalunya", year=year,
        unit="milions d'euros, preus corrents, preus bàsics",
        classification="CPA 2008 (63 branques, numeració MIOC d'Idescat)",
        sector_codes=codes, sector_labels=labels,
        Z=Z, Y=Y, Y_labels=Y_labels, VA=VA, VA_labels=VA_labels, X=X,
        source=(f"Idescat, Marc Input-Output de Catalunya {year}, taula "
                f"simètrica a 63 branques ({path.name})"),
        retrieved_at=datetime.now(timezone.utc),
        inherited_residue=residue,
        notes=(f"Product by product, basic prices, DOMESTIC origin: `Z` is the "
               f"'ts Cat' sheet, the only one of the four whose row identity "
               f"closes against output. The imported intermediates each branch "
               f"buys are the column sums of 'ts resta Esp' and 'ts resta món' "
               f"and are returned in `VA`, labelled as not being value added, "
               f"together with net taxes on products — without them the column "
               f"identity is 714 out. Dropped {len(dropped)} of-which memo "
               f"row(s) ({', '.join(dropped) or 'none'}): 64 numbered rows, "
               f"{n} branches. Dropped {len(fd_dropped)} final-demand subtotal "
               f"column(s) ({', '.join(fd_dropped)}). Exports are split into "
               f"rest of Spain and rest of world, which is what makes this a "
               f"two-region table rather than a regional one. IDESCAT's own "
               f"books are out by up to {residue:,.4f} between its `Total usos` "
               f"column and its `Producció` row — measured, declared as "
               f"inherited, and not this loader's to close."))
    _assert_balances(table, path.name, inherited=residue)
    return table


def _mioc_year(R, path: Path) -> int:
    """The reference year, from the sheet's own title rows.

    Never from the filename. `mioc2021ts64.xlsx` happens to carry it and the
    2014 and 2016 editions are named differently again, which is the trap
    `OQ-D-01` records.
    """
    import re as _re
    for i in range(min(6, len(R))):
        for c in R[i]:
            m = _re.search(r"\b(19|20)\d{2}\b", str(c or ""))
            if m:
                return int(m.group(0))
    raise LoaderError(
        f"{path.name}: could not read the reference year from the title rows. "
        f"Refusing to take it from the filename (OQ-D-01).")


# ---------------------------------------------------------------------------
# The Spanish supply-use tables (INE)
# ---------------------------------------------------------------------------

# Layout of `cne_tod_YY.xlsx`, 0-based — DERIVED from each workbook, not
# hard-coded. See `_tod_layout`. The constants that used to sit here described
# `cne_tod_22.xlsx` alone and crashed with a bare IndexError on the five older
# years, which is the one failure mode this project does not accept: a loader
# that cannot read a file must say why.

# The three final-demand subtotals, by the label the INE prints. Their
# COMPONENTS are not listed: they are whatever columns sit between one head and
# the next, which is how the same rule reads a vintage that splits exports into
# two columns and one that does not. Each head is then checked against the
# components inferred for it, so a wrong inference fails the load rather than
# silently double-counting.
_TOD_FD_HEADS = ("total gasto en consumo final", "formación bruta de capital",
                 "total exportaciones")


def _tod_norm(s) -> str:
    """Collapse whitespace, lowercase, drop a trailing footnote marker."""
    import re as _re
    t = _re.sub(r"\s+", " ", str(s if s is not None else "")).strip()
    t = _re.sub(r"\s*\(\d+\)\s*$", "", t)
    return t.lower()


def _tod_axes(R, what: str) -> dict:
    """Where the numbers are on one INE supply-use sheet, read off the sheet.

    The INE prints an index row — 1, 2, 3, … — across the activity columns and
    a numbered label down the product column, so both axes announce their own
    extent. Nothing here is a constant except the two header rows, which are
    the same in every vintage.

    The axes are NOT the same in every vintage, and not even the same on both
    sheets of one workbook: in the 2016–2020 files the supply sheet has no
    leading blank column and the use sheet does, so the two are offset by one.
    A loader with a single `first_col` reads the older supply table one column
    to the left of where it is.
    """
    idx = R[8] if len(R) > 8 else []
    first_col = next((j for j, c in enumerate(idx)
                      if _tod_norm(c) == "1"), None)
    if first_col is None or first_col == 0:
        raise LoaderError(
            f"the {what} sheet has no activity index row. Row 9 should number "
            f"the activity columns 1, 2, 3, … and it reads "
            f"{[str(c) for c in idx[:6]]!r}. This loader expects the INE's "
            f"`cne_tod_YY.xlsx` layout.")
    # The index row is NOT 1, 2, 3, …: like the product numbering it carries a
    # continuation, because imputed rents of owner-occupied dwellings are split
    # out of real estate. The label column writes that product '44 bis.' and
    # the index row above it writes '44a' — two conventions for one thing, in
    # one file. A run that accepts only the next integer stops dead at 44 and
    # reads a 65-activity table as a 44-activity one; the block stays
    # rectangular, so nothing but an identity catches it.
    import re as _re
    n_act, last = 0, 0
    while first_col + n_act < len(idx):
        cell = _tod_norm(idx[first_col + n_act])
        if cell == str(last + 1):
            last += 1
        elif last and _re.fullmatch(rf"{last}\s*(bis|a|b)", cell):
            pass
        else:
            break
        n_act += 1

    label_col = first_col - 1
    first_row = next((i for i in range(8, min(len(R), 24))
                      if _tod_norm(R[i][label_col] if label_col < len(R[i])
                                   else "").startswith("1.")), None)
    if first_row is None:
        raise LoaderError(
            f"the {what} sheet has no product labelled '1. …' in column "
            f"{label_col + 1} within its first 24 rows. This loader expects "
            f"the INE's `cne_tod_YY.xlsx` layout.")
    n_prod = 0
    while (first_row + n_prod < len(R)
           and _tod_norm(R[first_row + n_prod][label_col]
                         if label_col < len(R[first_row + n_prod]) else "")):
        n_prod += 1
    return dict(label_col=label_col, first_row=first_row, n_products=n_prod,
                first_col=first_col, n_activities=n_act)


def _tod_header(R, ax, what: str) -> dict:
    """The aggregate columns to the right of the activity block, by label."""
    header = R[7] if len(R) > 7 else []
    out: dict[str, int] = {}
    for j in range(ax["first_col"] + ax["n_activities"], len(header)):
        key = _tod_norm(header[j])
        if key and key not in out:
            out[key] = j
    return out


def _tod_labels(R, ax) -> dict:
    """The rows below the product block, by label, from the label column."""
    out: dict[str, int] = {}
    for i in range(ax["first_row"] + ax["n_products"], len(R)):
        cell = R[i][ax["label_col"]] if ax["label_col"] < len(R[i]) else ""
        key = _tod_norm(cell)
        if key and key not in out:
            out[key] = i
    return out


def _tod_need(d: dict, key: str, what: str, kind: str):
    if key not in d:
        near = ", ".join(sorted(d)[:8])
        raise LoaderError(
            f"the INE supply-use workbook has no {kind} labelled {key!r} on "
            f"its {what} sheet.\n  what it does have: {near}…\n"
            f"Every position this loader uses is found by the label the INE "
            f"prints, so a renamed line stops the load instead of shifting "
            f"the block underneath it.")
    return d[key]


def load_ine_tod(path: Path | str) -> SupplyUseTables:
    """Load the INE's supply and use tables at whatever detail they publish.

    **110 products by 81 activities for 2021 and 2022**, and **65 by 64 for
    2016 to 2020** — the INE moved to the finer table with the 2021 edition,
    and every position in this loader is found by the label the workbook prints
    so that both are read rather than one being hard-coded and the other
    crashing. `OQ-S-05` opened on the assumption that more detail than the
    64-product input-output table would have to be requested; for 2021 onward
    it does not, and this is the file that settled it. **For 2016 to 2020 it
    would**: those years are published at the same 64 activities as the IOT.

    That distinction is not bookkeeping. At 110 products `55 Servicios de
    alojamiento` and `56 Servicios de comidas y bebidas` are separate and their
    outputs sum to the 64-product product 36 to the last decimal, so the split
    this project's pilot estimates is simply published. At 65 products it is
    not, for any year before 2021 — the estimation route is the only one there,
    and a user asking for 2019 must be told that rather than handed a table
    that looks the same and is coarser.

    Two things a supply-use pair gives the project that no IOT can.

    **The margin identities become testable.** A supply table carries trade
    margins, transport margins and taxes less subsidies on products as explicit
    columns, so `ID-19` — that a margin column sums to zero economy-wide and
    that its negatives are mandatory — is arithmetic here rather than
    `NOT APPLICABLE`. That is what `OQ-D-03` said the project needed and could
    not get from an analytical IOT.

    **It balances.** Domestic uses equal domestic output and imported uses equal
    imports, for all 64 products of the aggregated sheets, to 0.0 — where the
    published input-output table fails by 4,921.6 on one of them (`OQ-D-04`).
    Where a clean Spanish domestic table is wanted, this is the source.

    Loaded at purchasers' prices, because that is the valuation at which the
    finest detail exists. The basic-price, domestic and import versions are
    published only at 64 products, and only from 2021 (sheets `Tabla3`,
    `Tabla4`, `Tabla5`); they are not loaded here.
    """
    path = Path(path)
    S = _open_workbook(path)
    for sheet in ("Tabla1", "Tabla2"):
        if sheet not in S:
            raise LoaderError(f"{path.name} has no {sheet!r}; found: "
                              f"{', '.join(S)}. This loader expects the INE's "
                              f"`cne_tod_YY.xlsx` layout.")
    sup, use = S["Tabla1"], S["Tabla2"]
    sa, ua = _tod_axes(sup, "supply"), _tod_axes(use, "use")
    if (sa["n_products"], sa["n_activities"]) != (ua["n_products"],
                                                 ua["n_activities"]):
        raise LoaderError(
            f"{path.name}: the supply and use sheets disagree on the size of "
            f"the table. Supply is {sa['n_products']} products by "
            f"{sa['n_activities']} activities, use is {ua['n_products']} by "
            f"{ua['n_activities']}. A supply-use pair is one table read twice; "
            f"if the two axes differ, nothing downstream means anything.")

    n_p, n_a = sa["n_products"], sa["n_activities"]
    s_rows = range(sa["first_row"], sa["first_row"] + n_p)
    s_cols = range(sa["first_col"], sa["first_col"] + n_a)
    u_rows = range(ua["first_row"], ua["first_row"] + n_p)
    u_cols = range(ua["first_col"], ua["first_col"] + n_a)

    sh = _tod_header(sup, sa, "supply")
    uh = _tod_header(use, ua, "use")
    ur = _tod_labels(use, ua)
    SC = lambda k: _tod_need(sh, k, "supply", "column")
    UC = lambda k: _tod_need(uh, k, "use", "column")
    UR = lambda k: _tod_need(ur, k, "use", "row")

    # Final demand: the heads are named, their components are inferred as the
    # columns between one head and the next, and each inference is checked.
    fd_first, fd_last = UC(_TOD_FD_HEADS[0]), UC("total demanda final")
    heads = [j for j in range(fd_first, fd_last)
             if _tod_norm(use[7][j]) in _TOD_FD_HEADS]
    fd_groups, fd_cols = [], []
    for k, h in enumerate(heads):
        stop = heads[k + 1] if k + 1 < len(heads) else fd_last
        comp = list(range(h + 1, stop))
        if comp:
            fd_groups.append((h, comp))
            fd_cols.extend(comp)
        else:
            fd_cols.append(h)          # a leaf, not a subtotal of itself

    V = _ine_block(sup, s_rows, s_cols)
    U = _ine_block(use, u_rows, u_cols)
    Y = _ine_block(use, u_rows, fd_cols)
    P = lambda g, rows, c: _ine_row_col(g, rows, c)
    q = P(sup, s_rows, SC("total producción"))
    imports = P(sup, s_rows, SC("total importaciones"))
    trade = P(sup, s_rows, SC("márgenes comerciales"))
    transport = P(sup, s_rows, SC("márgenes de transporte"))
    taxes = P(sup, s_rows, SC("impuestos netos sobre los productos"))
    basic = P(sup, s_rows, SC("total oferta a precios básicos"))
    purch = P(sup, s_rows, SC("total oferta a precios de adquisición"))
    uses = P(use, u_rows, UC("total empleos"))
    inter = P(use, u_rows, UC("total demanda intermedia"))

    W = np.vstack([_ine_row(use, UR(k), u_cols) for k in
                   ("remuneración de asalariados",
                    "otros impuestos netos sobre la producción",
                    "excedente bruto de explotación")]
                  + [_ine_row(use, next(i for lbl, i in ur.items()
                                        if lbl.startswith("renta mixta")),
                              u_cols)])
    g = _ine_row(use, UR("producción a precios básicos"), u_cols)
    # Intermediate consumption by activity is the TOTAL line under the product
    # block. The 2021+ files leave its label cell EMPTY and the older ones
    # write 'TOTAL', so it is anchored on the line above it, which both name.
    ic_row = UR("compras de residentes fuera del territorio económico") + 1
    ic = _ine_row(use, ic_row, u_cols)

    checks = [
        ("supply rows sum to product output", V.sum(1), q),
        ("supply at basic prices = output + imports", basic, q + imports),
        ("supply at purchasers' = basic + margins + taxes",
         purch, basic + trade + transport + taxes),
        ("use rows sum to intermediate demand", U.sum(1), inter),
        ("final-demand components sum to total final demand", Y.sum(1),
         P(use, u_rows, fd_last)),
        ("total uses = intermediate + final", uses, inter + Y.sum(1)),
        ("compensation = wages + social contributions",
         _ine_row(use, UR("remuneración de asalariados"), u_cols),
         _ine_row(use, UR("sueldos y salarios brutos"), u_cols)
         + _ine_row(use, UR("cotizaciones sociales"), u_cols)),
        ("gross value added = its four components",
         _ine_row(use, UR("valor añadido bruto a precios básicos"), u_cols),
         W.sum(0)),
        ("use columns sum to intermediate consumption", ic, U.sum(0)),
        ("output = intermediate consumption + value added", g, ic + W.sum(0)),
        # THE central identity of the framework, and the reason a supply-use
        # pair is worth more than an IOT: supply and use meet product by
        # product, at purchasers' prices, before anyone assumes anything.
        ("supply at purchasers' prices == total uses", purch, uses),
    ]
    # Only where the vintage publishes them. An absent split is not a failure;
    # a split that does not add up is.
    if "importaciones países u.e." in sh and "importaciones terceros países" in sh:
        checks.append(("imports EU + non-EU = total imports",
                       P(sup, s_rows, SC("importaciones países u.e."))
                       + P(sup, s_rows, SC("importaciones terceros países")),
                       imports))
    for sub, comp in fd_groups:
        checks.append((f"final-demand subtotal "
                       f"{_tod_norm(use[7][sub])!r} = its {len(comp)} parts",
                       P(use, u_rows, sub),
                       sum(P(use, u_rows, c) for c in comp)))
    for what, a, b in checks:
        d = float(np.abs(np.asarray(a) - np.asarray(b)).max())
        if d > _INE_TOL:
            raise LoaderError(
                f"{path.name} does not satisfy an identity the INE's own "
                f"printed totals should give it.\n  failed check: {what}\n"
                f"  off by {d:,.4f} (tolerance {_INE_TOL:g}, million EUR)\n"
                f"The block positions are read from the labels this workbook "
                f"prints, so this is a disagreement in the figures rather than "
                f"a layout the loader has misread.")

    p_codes, p_labels = _ine_codes(use, u_rows)
    a_labels = [str(sup[7][j] or "").strip() for j in s_cols]
    a_codes = [str(sup[8][j] or "").strip() for j in s_cols]
    year = _infer_ine_year(S)
    finest = n_p > 65
    return SupplyUseTables(
        table_id=f"ES-SUT-{year}", country="España", year=year,
        unit="millones de euros, precios corrientes",
        classification=(f"CPA 2008 ({n_p} productos) x CNAE 2009 "
                        f"({n_a} ramas)"),
        product_codes=p_codes, product_labels=p_labels,
        activity_codes=a_codes, activity_labels=a_labels,
        V=V, U=U, Y=Y,
        Y_labels=[str(use[7][c] or "").strip() for c in fd_cols],
        W=W, W_labels=["Remuneración de los asalariados",
                       "Otros impuestos netos sobre la producción",
                       "Excedente bruto de explotación", "Renta mixta bruta"],
        imports=imports, total_margins=trade + transport,
        trade_margins=trade, transport_margins=transport,
        taxes_on_products=taxes, q=q, g=g,
        source=(f"INE, Contabilidad Nacional Anual de España, Revisión "
                f"Estadística 2024. Tablas de Origen y Destino {year} "
                f"({path.name})"),
        retrieved_at=datetime.now(timezone.utc),
        notes=(f"Origen a precios básicos y destino a precios de adquisición, "
               f"{n_p} productos x {n_a} ramas"
               + (" — el mayor detalle que publica el INE, muy por encima de "
                  "los 64 productos de su propia TIO. "
                  if finest else
                  " — el detalle que el INE publicaba antes de 2021, el mismo "
                  "que su TIO. Alojamiento (CPA 55) y comidas y bebidas "
                  "(CPA 56) NO vienen separados en esta edición: para estos "
                  "años la partición fina no está publicada. ")
               + "Las columnas de márgenes e impuestos vienen explícitas, así "
                 "que las identidades de valoración son comprobables aquí y no "
                 "en una TIO analítica (OQ-D-03). Todas las posiciones se "
                 "localizan por la etiqueta que imprime el INE, no por índices "
                 "fijos, porque las dos ediciones no coinciden ni siquiera en "
                 "la columna donde empieza cada hoja."))


# ---------------------------------------------------------------------------
# The project's own interchange format (MVP_0.1 §4.1)
# ---------------------------------------------------------------------------

_REQUIRED_META = ("country", "year", "unit", "classification", "source")


def load_io_table(path: Path | str, sheet: str = "table") -> IOTable:
    """Load a table in the project's interchange format.

    Blocks are found by their labels, not by fixed positions (§4.1). The sheet
    holds, in order: a header row of sector codes, then one row per sector, then
    a row labelled with each value-added component, then a row `Output`. Final
    demand columns follow the sector columns and are named in the header.

    A separate `metadata` sheet must carry country, year, unit, classification
    and source as key/value rows. Missing metadata is an error: an IO table
    without its price basis and classification cannot be used safely.
    """
    sheets = _open_workbook(path)
    if sheet not in sheets:
        raise LoaderError(f"no sheet named {sheet!r}; found: {', '.join(sheets)}")
    if "metadata" not in sheets:
        raise LoaderError("no 'metadata' sheet. It must carry: "
                          + ", ".join(_REQUIRED_META))

    meta = {str(r[0]).strip().lower(): r[1] for r in sheets["metadata"]
            if r and r[0] is not None}
    missing = [k for k in _REQUIRED_META if k not in meta]
    if missing:
        raise LoaderError(f"metadata sheet is missing: {', '.join(missing)}")

    R = [r for r in sheets[sheet] if r and any(c is not None for c in r)]
    header = [str(c).strip() if c is not None else "" for c in R[0]]
    try:
        out_row = next(i for i, r in enumerate(R)
                       if str(r[0]).strip().lower() in ("output", "total output"))
    except StopIteration:
        raise LoaderError("no row labelled 'Output' or 'Total output'; §4.1 "
                          "requires one so the blocks can be located") from None

    while len(header) > 1 and not header[-1]:
        header.pop()
    labels_col = [str(r[0]).strip() if r[0] is not None else "" for r in R]

    # WHERE THE SECTOR BLOCK ENDS. Between the header and `Output` sit the
    # sector rows and then the value-added rows, and nothing marks the join --
    # so it is found the way §4.1 finds every block, by the labels: the sector
    # rows are the longest run whose labels match the header's leading codes,
    # one for one and in order. What follows that run down the page is value
    # added; what follows it across the header is final demand.
    #
    # Counting non-empty row labels instead, as this did until 2026-08-25, made
    # every value-added row a sector. `n` came out too large, no rows were left
    # for value added, and the loader rejected a well-formed file with a
    # complaint about the rows it had just miscounted. This is the only route
    # into the engine for a table from anywhere other than the UK or Spain, and
    # nothing exercised it -- no test, no validator, no fixture. Writing the
    # user guide is what ran it. `run_interchange_roundtrip.py` now does.
    n = 0
    while (1 + n < out_row and 1 + n < len(header)
           and labels_col[1 + n] and labels_col[1 + n] == header[1 + n]):
        n += 1
    if n == 0:
        raise LoaderError(
            f"no sector rows found. §4.1 locates the sector block by matching "
            f"the leading row labels against the header: row 2 is labelled "
            f"{labels_col[1]!r} and the first sector column is headed "
            f"{header[1]!r}. The sector rows must carry the same codes as the "
            f"sector columns, in the same order.")

    sector_codes = header[1:1 + n]   # == labels_col[1:1 + n], by the match
    va_start = 1 + n
    va_labels = [labels_col[i] for i in range(va_start, out_row)]
    if not va_labels:
        raise LoaderError(
            f"{n} sector row(s) run straight into the Output row, so there are "
            f"no value-added rows; §4.1 requires at least one. Value added is "
            f"not optional: without it a column does not add up to output.")
    stray = [lab for lab in va_labels if lab in header[1:1 + n]]
    if stray:
        raise LoaderError(
            f"the sector code(s) {', '.join(stray)} appear below the sector "
            f"block. Sector rows must be contiguous and in header order — a "
            f"value-added row cannot sit between two of them, because the "
            f"sector block is located by that run.")

    n_cols_total = len(header) - 1
    if n_cols_total <= n:
        raise LoaderError(f"{n_cols_total} data columns for {n} sectors: §4.1 "
                          f"requires at least one final-demand column")
    Y_labels = header[1 + n:1 + n_cols_total]

    Z = np.array([[_num(R[1 + i][1 + j]) for j in range(n)] for i in range(n)])
    Y = np.array([[_num(R[1 + i][1 + n + c]) for c in range(len(Y_labels))]
                  for i in range(n)])
    VA = np.array([[_num(R[va_start + m][1 + j]) for j in range(n)]
                   for m in range(len(va_labels))])
    X = np.array([_num(R[out_row][1 + j]) for j in range(n)])

    table = IOTable(
        table_id=str(meta.get("table_id", Path(path).stem)),
        country=str(meta["country"]), year=int(meta["year"]),
        unit=str(meta["unit"]), classification=str(meta["classification"]),
        sector_codes=sector_codes,
        # `meta` keys were lowercased on the way in, so the lookup must be
        # too. It was not, so every `label_B` row in a metadata sheet was
        # silently ignored and the report named sectors by their codes.
        sector_labels=[str(meta.get(f"label_{c}".lower(), c))
                       for c in sector_codes],
        Z=Z, Y=Y, Y_labels=Y_labels, VA=VA, VA_labels=va_labels, X=X,
        source=str(meta["source"]), retrieved_at=datetime.now(timezone.utc),
        notes=str(meta.get("notes") or "") or None,
        provenance=_read_provenance(sheets, sector_codes),
        lineage=_read_lineage(meta))
    _assert_balances(table, Path(path).name)
    return table


def _read_provenance(sheets: dict, sector_codes: list[str]):
    """Read a `Provenance` sheet back, if the file carries one.

    This engine writes one beside every table it exports. Reading it is what
    keeps a second split from promoting the first split's estimates to
    observations -- a table that has been through a disaggregation balances
    exactly as well as a published one, so nothing in the numbers would give
    it away.

    The sheet stores the §A.1 data status, which is coarser than the internal
    label: `USER_CONSTRAINT` is written as OBSERVED, so a pinned cell comes
    back as an ordinary observation. That loss is deliberate and stated here
    rather than hidden -- on reload the analyst who pinned it is not in the
    room, and the cell is as good as any other observation to whoever is.
    """
    import numpy as np

    from .models import CellLabel

    name = next((s for s in sheets if s.strip().lower() == "provenance"), None)
    if name is None:
        return None
    back = {"OBSERVED": CellLabel.OBSERVED,
            "ESTIMATED": CellLabel.PROXY_ESTIMATED,
            "BALANCED": CellLabel.BALANCED_ADJUSTMENT}
    R = [r for r in sheets[name] if r and any(c is not None for c in r)]
    n = len(sector_codes)
    if len(R) < 1 + n:
        raise LoaderError(
            f"the Provenance sheet has {len(R) - 1} data row(s) for {n} "
            f"sectors. A provenance grid that does not match the table is "
            f"worse than none: it would mislabel cells rather than leave them "
            f"unlabelled. Delete the sheet or correct it.")
    prov = np.empty((n, n), dtype=object)
    for i in range(n):
        for j in range(n):
            cell = R[1 + i][1 + j]
            key = str(cell).strip().upper() if cell is not None else "OBSERVED"
            if key not in back:
                raise LoaderError(
                    f"Provenance cell ({sector_codes[i]}, {sector_codes[j]}) "
                    f"reads {cell!r}; expected one of "
                    f"{', '.join(sorted(back))}.")
            prov[i, j] = back[key]
    return prov


def _read_lineage(meta: dict) -> list[str]:
    """The table's ancestry: `lineage_1`, `lineage_2`, … then `derived_from`.

    Oldest first, so appending this run's own line keeps the order. A table
    from a statistical office has neither key and comes back with an empty
    list, which is the correct answer -- it is the start of a lineage, not a
    step in one.
    """
    older = sorted((k for k in meta if k.startswith("lineage_")),
                   key=lambda k: int(k.rsplit("_", 1)[1]))
    out = [str(meta[k]) for k in older]
    # `derived_from` is a one-line summary for whoever opens the workbook, and
    # on a file this engine wrote it restates the last lineage row with the
    # estimated share appended -- a share the Provenance sheet already carries
    # cell by cell. Taking it as well would double the newest step every time
    # the file went round the loop. It is read only when there is no `lineage_`
    # row at all, which is the hand-written case.
    if not out and meta.get("derived_from"):
        out.append(str(meta["derived_from"]))
    return out


# ---------------------------------------------------------------------------
# Allocation keys (MVP_0.1 §4.2)
# ---------------------------------------------------------------------------

def load_allocation_keys(path: Path | str, applies_to: str = "output",
                         sheet: str | None = None) -> AllocationKey:
    """Load one proxy table: new_sector_code, new_sector_label, value, source,
    source_year, strength.

    One file per block, which is what makes the multi-proxy splitting of §8.C
    possible from the MVP onward.
    """
    sheets = _open_workbook(path)
    name = sheet or next(iter(sheets))
    R = [r for r in sheets[name] if r and any(c is not None for c in r)]
    head = [str(c).strip().lower() if c is not None else "" for c in R[0]]

    def col(*names) -> int:
        for nm in names:
            if nm in head:
                return head.index(nm)
        raise LoaderError(f"{Path(path).name}: no column named "
                          f"{' or '.join(names)}; found {head}")

    c_code, c_val = col("new_sector_code", "code"), col("value", "raw_value")
    c_src, c_yr = col("source"), col("source_year", "year")
    c_str = col("strength")

    codes, vals, srcs, yrs, strs = [], [], [], [], []
    for r in R[1:]:
        if r[c_code] is None:
            continue
        codes.append(str(r[c_code]).strip())
        vals.append(_num(r[c_val]))
        srcs.append(str(r[c_src]))
        yrs.append(int(r[c_yr]))
        strs.append(str(r[c_str]).strip().lower())

    if len(set(srcs)) > 1 or len(set(yrs)) > 1:
        # Not fatal, but the AllocationKey carries one source and one year, so
        # say which one is being recorded rather than picking silently.
        raise LoaderError(
            f"{Path(path).name}: rows disagree on source or source_year "
            f"({sorted(set(srcs))}, {sorted(set(yrs))}). One AllocationKey "
            f"records one source. Split the file, or reconcile it.")

    weakest = min(strs, key=lambda s: ["strong", "medium", "weak"].index(s)
                  if s in ("strong", "medium", "weak") else 0)
    # The key is only as strong as its weakest row: a split resting on one weak
    # proxy is a weak split, whatever the other rows say.
    for s in strs:
        if s not in ("strong", "medium", "weak"):
            raise LoaderError(f"strength {s!r} is not strong/medium/weak")
        if ["strong", "medium", "weak"].index(s) > \
           ["strong", "medium", "weak"].index(weakest):
            weakest = s

    return AllocationKey(
        key_id=Path(path).stem, applies_to=applies_to,
        new_sector_codes=codes, raw_values=vals,
        source=srcs[0], source_year=yrs[0],
        strength=ProxyStrength(weakest),
        notes=f"loaded from {Path(path).name}")


# ---------------------------------------------------------------------------
# The ONS supply-use tables — the second fixture, and NOT a SupplyUseTables
# ---------------------------------------------------------------------------

@dataclass
class OnsSupplyUse:
    """One year of the ONS's published supply and use tables.

    **This is deliberately not a `SupplyUseTables`**, and the reason is the
    whole point of the object. That class promises `V`, the supply matrix of
    products by industries — and the ONS's published workbook does not contain
    one. Its supply table gives each product's *total* domestic output beside
    its imports, margins and taxes; who made what is not published. Filling `V`
    with anything would be an invention the accounting could not detect, which
    is exactly what `load_sut` refuses to do for Eurostat's combined margins.

    So this carries what the source actually has, and the identities that need
    `V` — `ID-07`, and `ID-09`'s margins by industry — stay NOT APPLICABLE on
    this fixture and say why.
    """
    year: int
    product_codes: list[str]
    industry_codes: list[str]
    U: np.ndarray                  # intermediate use, products x industries
    Y: np.ndarray                  # final demand, products x categories
    Y_labels: list[str]
    output_basic: np.ndarray       # by product, total domestic output
    imports: np.ndarray            # by product, goods and services
    margins: np.ndarray            # by product, distributors' trading margins
    taxes_on_products: np.ndarray  # by product, net of subsidies
    total_supply: np.ndarray       # by product, at purchasers' prices
    total_demand: np.ndarray       # by product, as published
    ic_by_industry: np.ndarray     # published total intermediate consumption
    taxes_on_production: np.ndarray
    compensation: np.ndarray
    gos_mixed: np.ndarray
    gva: np.ndarray
    output_by_industry: np.ndarray
    source: str


_ONS_VA_ROWS = {
    "ic": "Total intermediate consumption at purchasers' prices",
    "taxes": "Taxes less subsidies on production",
    "coe": "Compensation of employees",
    "gos": "Gross operating surplus and mixed income",
    "gva": "Gross valued added at basic prices",     # the source's own spelling
    "output": "Total output at basic prices",
}


def load_ons_sut(path: Path | str, year: int) -> OnsSupplyUse:
    """One year out of `NSO_UK_04`, the ONS supply-use workbook (1997-2023).

    The workbook holds five sheets per year. Three are used here: the supply
    table (product totals and valuation columns), the intermediate-consumption
    matrix with the value-added block underneath it, and final demand.

    Products and industries are taken from the sheets themselves rather than
    assumed, and the value-added rows are found **by their printed labels** —
    including the source's own "Gross valued added at basic prices", which is a
    typo in the workbook and is matched as printed rather than corrected.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def sheet(prefix: str):
        name = f"Table {prefix} {year}"
        if name not in wb.sheetnames:
            raise LoaderError(
                f"{path.name} has no sheet {name!r}. Years present: "
                f"{sorted({int(s.rsplit(' ', 1)[1]) for s in wb.sheetnames if s[-4:].isdigit()})}")
        return list(wb[name].iter_rows(values_only=True))

    use, supply, final = sheet("2 - Int Con"), sheet("1 - Supply"), sheet("2 - Final Demand")

    industries = [str(c).strip() for c in use[3][2:106] if c]
    prod_rows = [r for r in use if r and str(r[0] or "").startswith("CPA_")]
    products = [str(r[0]).strip() for r in prod_rows]
    n_ind = len(industries)
    U = np.array([[_num(r[j]) for j in range(2, 2 + n_ind)] for r in prod_rows],
                 float)

    va = {}
    for row in use:
        label = str(row[1] or "").strip()
        for key, want in _ONS_VA_ROWS.items():
            if label == want:
                va[key] = np.array([_num(row[j]) for j in range(2, 2 + n_ind)],
                                   float)
    missing = set(_ONS_VA_ROWS) - set(va)
    if missing:
        raise LoaderError(
            f"{path.name} {year}: the value-added block is missing "
            f"{sorted(_ONS_VA_ROWS[k] for k in missing)}. The labels are matched "
            f"as printed, including the source's own spelling.")

    sup = {str(r[0]).strip(): r for r in supply
           if r and str(r[0] or "").startswith("CPA_")}
    fin = {str(r[0]).strip(): r for r in final
           if r and str(r[0] or "").startswith("CPA_")}
    if set(sup) != set(products) or set(fin) != set(products):
        raise LoaderError(
            f"{path.name} {year}: the three sheets do not carry the same "
            f"products — supply {len(sup)}, use {len(products)}, final demand "
            f"{len(fin)}. They are matched by code, never by position.")

    def col(table, index):
        return np.array([_num(table[p][index]) for p in products], float)

    # Final demand: the four consumption columns, three capital columns and the
    # two export columns, taking the components and never the printed subtotals.
    y_cols = [(2, "households"), (3, "npish"), (4, "central government"),
              (5, "local government"), (8, "gross fixed capital formation"),
              (9, "valuables"), (10, "changes in inventories"),
              (15, "exports of goods"), (16, "exports of services")]
    Y = np.column_stack([col(fin, i) for i, _ in y_cols])

    return OnsSupplyUse(
        year=year, product_codes=products, industry_codes=industries,
        U=U, Y=Y, Y_labels=[lab for _, lab in y_cols],
        output_basic=col(sup, 2), imports=col(sup, 7), margins=col(sup, 8),
        taxes_on_products=col(sup, 9), total_supply=col(sup, 10),
        total_demand=col(fin, 21),
        ic_by_industry=va["ic"], taxes_on_production=va["taxes"],
        compensation=va["coe"], gos_mixed=va["gos"], gva=va["gva"],
        output_by_industry=va["output"],
        source=f"ONS, Input-output supply and use tables, Blue Book 2025 "
               f"({path.name}), {year}")
