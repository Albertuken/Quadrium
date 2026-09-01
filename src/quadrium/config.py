"""
Drive the whole engine from a spreadsheet.

WHY A WORKBOOK AND NOT A USER INTERFACE
---------------------------------------
The inputs to a disaggregation are tables of numbers with metadata: which
sectors to divide, the new codes, the proxies with their source and year, the
relative purchasing intensities by supplier. That is what a spreadsheet is for,
and an economist is faster in one than in any form. Forty intensity values are
a column, not forty input boxes.

It is also the honest choice about verification: a workbook-driven run can be
checked end to end by running it, which a screenful of buttons cannot.

THE WORKBOOK
------------
Five sheets. Only `project` and `splits` are required.

`project`   key/value rows: project_id, table_path, table_kind, title, notes.
            `table_kind` is one of:
              `uk_analytical`  the ONS workbook, industry by industry
              `ine_interior`   the INE workbook, domestic output (product by
                               product) — the Spanish counterpart
              `ine_total`      the INE workbook, total flows
              `interchange`    the project's own format
              `eurostat`       fetched from the Eurostat API by country and
                               year, and cached — see below
            `table_unbalanced` (`refuse` by default, or `residual_column`)
            applies to `ine_interior` alone, whose published table does not
            balance for one product — see OQ-D-04. Setting it on any other kind
            is an error rather than something quietly ignored.

`splits`    one row per NEW subsector:
              sector_code | new_code | new_label | key_id
            Rows sharing a `sector_code` form one split. `key_id` names the
            allocation key for that split; blank falls back to the scenario.

`keys`      one row per subsector per key:
              key_id | new_sector_code | value | source | source_year | strength
            `strength` is strong / medium / weak.

            A KEY YOU DO NOT USE IS NOT WASTED. Any key registered here whose
            `key_id` no split or scenario names becomes an automatic external
            check: the report compares the split it produced against what that
            key measures, and prints the gap. If you have employment AND
            turnover, put both in and drive with one — the other buys you an
            error bar, which nothing else in this system can give you.

`scenarios` one row per scenario:
              scenario_id | label | description | internal_block_alpha
            Blank alpha uses the default 1.0 -- CORE_031 eq. (14), the
            outer product of the weights. Raising it concentrates the block
            on its diagonal and the off-diagonal pays for it, so the parent
            cell is conserved either way; published tables sit around 1.5
            (OQ-S-04). If the sheet is missing, one
            scenario named `S1` is created.

`profiles`  one row per (scenario, subsector, supplier) intensity:
              scenario_id | subsector_code | supplier_code | intensity
            1.0 means the parent sector's average. Scenarios absent from this
            sheet simply have no profiles, which is how a "plain" and a
            "profiled" scenario are written side by side.

FETCHING RATHER THAN NAMING A FILE
-----------------------------------
`table_kind: eurostat` replaces `table_path` with a country and a year:

    eurostat_geo       ES, AT, FR … the two-letter code
    eurostat_year      2022
    eurostat_dataset   product_by_product (default) | industry_by_industry
                       | a raw naio_10_* code
    eurostat_variant   domestic (default) | total
    table_path         optional: where to cache. Defaults to
                       data/eurostat/<dataset>_<GEO>_<year>.json beside the
                       configuration.

**A cached file is never re-fetched.** This is the reproducibility rule, and it
is the reason `eurostat.fetch()` and `eurostat.load_iot()` were built as
separate functions in the first place: statistical offices revise, so a
configuration that downloaded on every run would give one answer in January and
another in June with nothing in the output to say why. The first run downloads
and records the URL, the byte count and the SHA-256; every run after it reads
those same bytes and never touches the network. `--refresh` overrides that
deliberately and says what it is doing; `--offline` refuses to fetch at all and
prints the URL so the file can be brought in by hand.

Everything the workbook cannot express is a project default, and every default
is documented where it lives. Nothing here invents an economic assumption.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np

from .io_loader import LoaderError, _open_workbook, load_ine_tio, \
    load_io_table, load_uk_analytical_iot
from .models import (AllocationKey, Assumption, AssumptionLedger,
                     ProxyStrength, Scenario, SplitSpec)

REQUIRED_SHEETS = ("project",)
# A workbook describes ONE of two jobs: dividing a sector or estimating a
# region. It needs `splits` for the first and `regionalise` for the second, and
# `splits` was unconditionally required until v1.86 -- which is why
# regionalisation only existed as command-line flags, in a tool whose guide
# opens by promising you will not need Python.
# THREE jobs, not two: `targets` with `table_kind: eurostat_sut` describes a
# projection, which has no splits either. The first version of this check
# listed only the first two and refused every projection workbook.
ONE_OF_SHEETS = ("splits", "regionalise", "targets")
TABLE_KINDS = ("uk_analytical", "interchange",
               "ine_interior", "ine_total", "eurostat", "eurostat_sut")


class ConfigError(ValueError):
    """Something in the workbook is wrong, said in the analyst's terms."""


def _eurostat_cache_path(meta: dict, base_dir: Path) -> tuple[Path, dict]:
    """Where the download lives, and what to ask for if it is not there yet.

    Returns the cache path and the request it stands for. Nothing here touches
    the network: this only resolves what the workbook asked for into a filename
    and a set of API parameters, so that a configuration can be READ, and its
    mistakes reported, without a connection.
    """
    from .eurostat import DATASETS

    geo = str(_need(meta, "eurostat_geo", "project", 0)).strip().upper()
    if not (len(geo) == 2 and geo.isalpha()):
        raise ConfigError(
            f"eurostat_geo is {geo!r}. It must be a two-letter country code — "
            f"ES, AT, FR, PT. Eurostat answers an unknown code with an empty "
            f"result rather than an error, so this is checked here instead.")
    try:
        year = int(str(_need(meta, "eurostat_year", "project", 0)).strip())
    except ValueError:
        raise ConfigError(
            f"eurostat_year is {meta.get('eurostat_year')!r}, which is not a "
            f"year.") from None

    name = str(meta.get("eurostat_dataset")
               or "product_by_product").strip().lower()
    dataset = DATASETS.get(name, name)
    if not dataset.startswith("naio_10_"):
        raise ConfigError(
            f"eurostat_dataset {name!r} is neither one of "
            f"{', '.join(sorted(DATASETS))} nor a naio_10_* code.")

    variant = str(meta.get("eurostat_variant") or "domestic").strip().lower()
    if variant not in ("domestic", "total"):
        # `imports` is refused for a reason worth repeating here rather than
        # letting the loader raise it after the download: Eurostat publishes no
        # output vector for the imported block, so it is an input table and not
        # a symmetric IOT.
        raise ConfigError(
            f"eurostat_variant {variant!r} must be 'domestic' (the default, "
            f"and what Leontief analysis wants) or 'total'. 'imports' is an "
            f"input table, not a symmetric IOT — Eurostat publishes no output "
            f"vector for it, so the column identity has nothing to close "
            f"against.")

    raw = meta.get("table_path")
    if raw and str(raw).strip():
        path = Path(str(raw).strip())
    else:
        path = Path("data") / "eurostat" / f"{dataset}_{geo}_{year}.json"
    if not path.is_absolute():
        path = (Path(base_dir) / path).resolve()

    return path, {"dataset": dataset, "geo": geo, "year": year,
                  "variant": variant}


def _eurostat_sut_paths(meta: dict, base_dir: Path) -> tuple[Path, dict]:
    """Where the three files of a supply-use system live, and what to ask for.

    A symmetric table is one download. A supply-use SYSTEM is three, and they
    are not interchangeable:

        naio_10_cp15    supply at basic prices, with its valuation columns
        naio_10_cp16    use at purchasers' prices
        naio_10_cp1610  use at BASIC prices, split DOM / IMP

    The third is the one that makes a transformation possible. Without it the
    domestic and imported halves would have to be derived, which means assuming
    every user of a product imports the same share of it -- an economic
    hypothesis, not bookkeeping, and not one this engine makes for anybody.
    """
    from .eurostat import DATASETS

    geo = str(_need(meta, "eurostat_geo", "project", 0)).strip().upper()
    if not (len(geo) == 2 and geo.isalpha()):
        raise ConfigError(
            f"eurostat_geo is {geo!r}. It must be a two-letter country code.")
    try:
        year = int(str(_need(meta, "eurostat_year", "project", 0)).strip())
    except ValueError:
        raise ConfigError(
            f"eurostat_year is {meta.get('eurostat_year')!r}, not a year."
        ) from None

    model = str(meta.get("eurostat_model") or "D").strip().upper()
    if model not in ("A", "B", "C", "D"):
        raise ConfigError(
            f"eurostat_model {model!r} must be one of the four in CORE_013 "
            f"Figure 12.2, p. 378:\n"
            f"  A  product technology            product x product\n"
            f"  B  industry technology           product x product\n"
            f"  C  fixed industry sales          industry x industry\n"
            f"  D  fixed product sales           industry x industry\n"
            f"A and C need a square supply table and may produce negative "
            f"cells; B and D cannot produce them. Which is right is not a "
            f"question the data answers.")

    raw = meta.get("table_path")
    root = Path(str(raw).strip()) if raw and str(raw).strip() \
        else Path("data") / "eurostat"
    if not root.is_absolute():
        root = (Path(base_dir) / root).resolve()
    if root.suffix:                       # a file was named; use its folder
        root = root.parent
    # A supply-use system caches THREE files, so `table_path` names a folder.
    # If something is already there and is not one, say so: the alternative is
    # a FileExistsError from `mkdir` several frames down, which is what a
    # `table_path` shared with a single-file `eurostat` run produced.
    if root.exists() and not root.is_dir():
        raise ConfigError(
            f"table_path is {root}, which exists and is a file. A supply-use "
            f"system caches three downloads, so this names the FOLDER they go "
            f"in — not one of them. Leave it blank for data/eurostat, or point "
            f"it at a directory.")

    files = {name: root / f"{DATASETS[name]}_{geo}_{year}.json"
             for name in ("supply", "use_purchasers", "use_basic")}
    unbalanced = str(meta.get("sut_unbalanced") or "refuse").strip().lower()
    if unbalanced not in ("refuse", "cancelling"):
        raise ConfigError(
            f"sut_unbalanced {unbalanced!r} must be 'refuse' (the default) or "
            f"'cancelling'.\n\n"
            f"`cancelling` admits ONE case: a closing identity out beyond what "
            f"the source's own precision allows, whose residues SUM TO ZERO "
            f"and sit in lines the message names — a boundary between two "
            f"industries rather than a table that fails to add up.\n\n"
            f"No source this project holds currently needs it. It was written "
            f"for Belgium's 2022 pair, +0.8 on L68A and -0.8 on L68B and 0.000 "
            f"on the other 87, which turned out to be INSIDE what Belgium's "
            f"one-decimal printing allows once the precision was read off the "
            f"figures the file actually uses rather than off two anomalous "
            f"cells in 2,829.\n\n"
            f"It does not admit residues that accumulate. Whatever is missing "
            f"from a table as a whole stays missing, and the load stops.")

    to_year = meta.get("project_to_year")
    # NO DEFAULT, AND THAT IS THE POINT.
    #
    # It defaulted to `sut_euro`, which is the worse of the two methods on
    # every test that has been run (`OQ-B-16`), so a user who said nothing got
    # the loser by silence. Changing the default to `sut_ras` would have moved
    # the problem rather than removed it.
    #
    # The two methods take DIFFERENT targets, so the `targets` sheet already
    # says which one the user means: nobody can reach a default by accident,
    # because they had to write `gva` rows or `industry_output` rows to get
    # anywhere. `project_method` is now an OPTIONAL declaration that has to
    # agree with the sheet, and `_project` reads the method off the sheet.
    raw = str(meta.get("project_method") or "").strip().lower()
    if raw and raw not in ("sut_euro", "sut_ras"):
        raise ConfigError(
            f"project_method {raw!r} must be 'sut_euro' or 'sut_ras', or "
            f"left out — the `targets` sheet says which one you mean.\n\n"
            f"They project onto different quantities:\n"
            f"  sut_euro   gva (one per industry, BASIC prices)\n"
            f"             final_use (one per category, PURCHASERS' prices)\n"
            f"             taxes, imports (one row each, totals)\n"
            f"  sut_ras    industry_output (one per industry)\n"
            f"             use_column_totals (one per industry, then one per\n"
            f"                 final-use category — each industry's output\n"
            f"                 LESS its value added)\n"
            f"             taxes, imports (one row each; their sum is the "
            f"total this method balances against)")
    pmethod = raw or None
    if to_year not in (None, ""):
        try:
            to_year = int(str(to_year).strip())
        except ValueError:
            raise ConfigError(
                f"project_to_year is {meta.get('project_to_year')!r}, not a "
                f"year.") from None
    else:
        to_year = None

    return root, {"geo": geo, "year": year, "model": model, "files": files,
                  "took_default_model": not meta.get("eurostat_model"),
                  "project_to_year": to_year, "project_method": pmethod,
                  "unbalanced": unbalanced}


def _project(sut, req: dict, targets: list, defaults: list):
    """Move a supply-use pair to a later year, onto totals from the `targets`
    sheet.

    THE SHEET IS FOUR KINDS OF ROW, and each is a different shape of fact:

        gva         one row per industry, its code, value added at BASIC prices
        final_use   one row per final-use category, at PURCHASERS' prices
        taxes       one row, the total
        imports     one row, the total

    The price bases are not decoration. The method carries taxes as a row of
    the use table, so a final-use target has to include them; getting that
    wrong does not fail loudly, it runs to the iteration ceiling with every
    value-added deviation reading 1.00003, which looks like success. What tells
    the two apart is projecting a pair onto its OWN totals and requiring the
    pair back -- exactly, in one iteration, which is what happens when the
    bases are right.
    """
    import numpy as np

    VOCAB = {"sut_euro": ("gva", "final_use", "taxes", "imports"),
             "sut_ras": ("industry_output", "use_column_totals",
                         "taxes", "imports")}
    SHARED = ("taxes", "imports")

    kinds = []
    for n, r in enumerate(targets, start=2):
        kind = str(_need(r, "kind", "targets", n)).strip().lower()
        known = {k for v in VOCAB.values() for k in v}
        if kind not in known:
            raise ConfigError(
                f"targets row {n}: kind {kind!r} is not one of "
                f"{', '.join(sorted(known))}.")
        kinds.append(kind)

    # THE SHEET SAYS WHICH METHOD, BECAUSE ONLY THE SHEET KNOWS WHAT YOU HAVE.
    #
    # `gva` and `final_use` belong to SUT-EURO, `industry_output` and
    # `use_column_totals` to SUT-RAS, and `taxes` and `imports` to both. A
    # sheet carrying rows of one distinctive set has already chosen. The
    # question a user can answer is "what do I know about the later year?",
    # not "which of two methods from chapter 18 do I want", and this is the
    # place that difference shows up.
    seen = set(kinds)
    votes = {m: seen & (set(v) - set(SHARED)) for m, v in VOCAB.items()}
    chosen = [m for m, hit in votes.items() if hit]
    if len(chosen) != 1:
        lines = "\n".join(
            f"      {m:<9} {', '.join(k for k in VOCAB[m] if k not in SHARED)}"
            for m in VOCAB)
        raise ConfigError(
            ("the `targets` sheet mixes the two projection methods' rows: "
             + ", ".join(sorted(seen)) if len(chosen) > 1 else
             "the `targets` sheet carries no rows that say which projection "
             "method you mean")
            + f".\n\n  What do you know about {req['project_to_year']}?\n"
            + lines
            + "\n      both      taxes, imports\n\n"
            + "  Write the rows for what you have. There is no default: "
              "`sut_ras` is the better\n  of the two on every test run "
              "(OQ-B-16) but it needs industry outputs, and\n  picking one "
              "for you would be choosing what you measured.")
    method = chosen[0]
    if req.get("project_method") and req["project_method"] != method:
        raise ConfigError(
            f"project_method says {req['project_method']!r} and the `targets` "
            f"sheet carries {', '.join(sorted(votes[method]))}, which is what "
            f"{method!r} takes.\n\nLeave `project_method` out and the sheet "
            f"decides, or make the two agree.")
    wanted = VOCAB[method]
    by_kind: dict = {}
    for n, (r, kind) in enumerate(zip(targets, kinds), start=2):
        try:
            value = float(r.get("value"))
        except (TypeError, ValueError):
            raise ConfigError(
                f"targets row {n}: value {r.get('value')!r} is not a number."
            ) from None
        by_kind.setdefault(kind, []).append(
            (str(r.get("code") or "").strip(), value))

    for kind in wanted:
        if kind not in by_kind:
            raise ConfigError(
                f"the `targets` sheet has no {kind!r} row(s). A "
                f"`{method}` projection needs all four: "
                f"{', '.join(wanted)}.")

    def vector(kind, codes, what):
        given = dict(by_kind[kind])
        missing = [c for c in codes if c not in given]
        extra = [c for c in given if c not in codes]
        if missing or extra:
            raise ConfigError(
                f"the {kind!r} targets do not match this pair's {what}.\n"
                + (f"  missing: {', '.join(missing[:8])}"
                   + ("…" if len(missing) > 8 else "") + "\n" if missing else "")
                + (f"  not in the table: {', '.join(extra[:8])}"
                   + ("…" if len(extra) > 8 else "") if extra else ""))
        return np.array([given[c] for c in codes], float)

    live_a = [c for c, g in zip(sut.activity_codes, sut.g) if g > 0]

    if method == "sut_ras":
        # SUT-RAS is given industry output and use column totals and imposes
        # them, rather than approaching value added iteratively. Wired here
        # because leaving the better method reachable only from Python is the
        # same "built, verified and unreachable" fault the engine keeps finding
        # in itself. See OQ-B-16 and `run_projection_backtest.py`.
        cols = list(live_a) + list(sut.Y_labels)
        projected = sut.project(
            method="sut_ras", year=req["project_to_year"],
            taxes=by_kind["taxes"][0][1], imports=by_kind["imports"][0][1],
            industry_output=vector("industry_output", live_a,
                                   "industries with output"),
            use_column_totals=vector("use_column_totals", cols,
                                     "use columns (industries then "
                                     "final-use categories)"))
        print(f"    Projected {sut.year} -> {req['project_to_year']} by "
              f"sut_ras: output {sut.q.sum():,.0f} -> "
              f"{projected.q.sum():,.0f} "
              f"({100 * (projected.q.sum() / sut.q.sum() - 1):+.2f} %)")
        return projected

    gva = vector("gva", sut.activity_codes, "industries") \
        if len(by_kind["gva"]) == len(sut.activity_codes) \
        else vector("gva", live_a, "industries with output")
    if len(gva) == len(live_a) and len(live_a) != len(sut.activity_codes):
        full = np.zeros(len(sut.activity_codes))
        full[[i for i, g in enumerate(sut.g) if g > 0]] = gva
        gva = full
        defaults.append(
            f"the gva targets cover the {len(live_a)} industries with output "
            f"and not the {len(sut.activity_codes)} in the table; the rest "
            f"were taken as zero, which is what they already are")

    projected = sut.project(
        gva=gva, final_use=vector("final_use", list(sut.Y_labels),
                                  "final-use categories"),
        taxes=by_kind["taxes"][0][1], imports=by_kind["imports"][0][1],
        method="sut_euro", year=req["project_to_year"])
    print(f"    Projected {sut.year} -> {req['project_to_year']} by "
          f"{method}: output "
          f"{sut.q.sum():,.0f} -> {projected.q.sum():,.0f} "
          f"({100 * (projected.q.sum() / sut.q.sum() - 1):+.2f} %)")
    return projected


def _load_eurostat_sut(req: dict, offline: bool, refresh: bool,
                       defaults: list, targets: list):
    """Fetch or read the three files, then transform by the named model."""
    import json

    from .eurostat import EurostatError, fetch, load_sut

    if req["took_default_model"]:
        defaults.append(
            "no `eurostat_model` was named, so model D (fixed product sales "
            "structure) was used. CORE_013 par. 12.76, p. 393 recommends it "
            "for rectangular tables and it cannot produce negative cells, but "
            "IT IS STILL A CHOICE ABOUT SECONDARY PRODUCTION and the four "
            "models give four different tables from the same data")

    for name, path in req["files"].items():
        side = path.with_suffix(path.suffix + ".provenance")
        if path.exists() and not refresh:
            continue
        if offline:
            raise ConfigError(
                f"--offline was given and {path.name} is not cached yet. A "
                f"supply-use system needs all three of "
                f"{', '.join(p.name for p in req['files'].values())}.")
        try:
            rec = fetch(name, req["geo"], req["year"], path)
        except EurostatError as exc:
            raise ConfigError(
                f"the Eurostat download failed on {name}:\n{exc}\n\n"
                f"A supply-use system needs all three files; nothing was "
                f"transformed.") from None
        side.write_text(json.dumps(rec, indent=2))
        print(f"    Downloaded {rec['dataset']} {req['geo']} {req['year']} — "
              f"{rec['bytes']:,} bytes, SHA-256 {rec['sha256'][:16]}…")

    try:
        sut = load_sut(req["files"]["supply"], req["files"]["use_purchasers"],
                       req["files"]["use_basic"],
                       unbalanced=req.get("unbalanced", "refuse"))
    except EurostatError as exc:
        raise ConfigError(f"the supply-use pair could not be built:\n{exc}"
                          ) from None

    print(f"    Supply-use: {sut.V.shape[0]} products x {sut.V.shape[1]} "
          f"activities, {sut.q.sum():,.0f} {sut.unit}")
    for chunk in (sut.notes or "").split("CLOSURE:")[1:]:
        for part in chunk.strip().rstrip(".").split("; "):
            if part.startswith("ADMITTED"):
                # Split on the marker, not on the first colon: the note
                # itself contains one, inside `sut_unbalanced: cancelling`.
                marker = "`sut_unbalanced: cancelling`: "
                body = part.split(marker, 1)[-1] if marker in part else part
                defaults.append(
                    "`sut_unbalanced: cancelling` was set, and it admitted a "
                    "real discrepancy — " + body)
            else:
                defaults.append("the source's closing identities lean rather "
                                "than cancel — " + part)

    if req.get("project_to_year"):
        sut = _project(sut, req, targets, defaults)
    try:
        table = sut.to_iot(req["model"])
    except (ValueError, ArithmeticError) as exc:
        raise ConfigError(
            f"model {req['model']} could not transform this system:\n{exc}"
        ) from None
    print(f"    Transformed by model {req['model']} -> {table.n} sectors, "
          f"{int((table.Z < 0).sum())} negative cell(s) in Z")

    # THE THREE FILES, ON THE TABLE. A supply-use system has no single source
    # file to checksum, so the manifest records none; the provenance has to
    # travel with the table instead, or a reader cannot tell which vintage of
    # which download produced the figures they are holding.
    stamps = []
    for name, path in req["files"].items():
        side = path.with_suffix(path.suffix + ".provenance")
        try:
            rec = json.loads(side.read_text())
            # `.get`, not `[...]`, and for a reason found on 2026-08-26: a
            # sidecar written by hand rather than by `fetch()` used the key
            # `retrieved` where this reads `retrieved_at`, and the whole run
            # died with `KeyError: 'retrieved_at'` after the table had loaded,
            # transformed and projected. A provenance stamp is a note about
            # the data; a missing field in it must degrade the note, not kill
            # the run. `run_provenance_sidecars.py` now checks the fields too.
            stamps.append(f"{rec.get('dataset', path.name)} "
                          f"{str(rec.get('retrieved_at', 'date unrecorded'))[:10]} "
                          f"{str(rec.get('sha256', ''))[:16]}…")
        except (ValueError, OSError):
            stamps.append(f"{path.name} (no provenance sidecar)")
    table.notes = ((table.notes or "") + " Built from " + "; ".join(stamps)
                   + ".").strip()
    return table


def _load_eurostat(path: Path, req: dict, offline: bool, refresh: bool):
    """Load the table, downloading it first if it is not already here.

    THE RULE IS THAT A CACHED FILE IS NEVER RE-FETCHED. A statistical office
    revises; a configuration that downloaded on every run would answer one way
    in January and another in June, with nothing in the output to say why, and
    the whole point of this engine is that a number can be traced back to
    something someone else could repeat. So the first run downloads and records
    the URL, the byte count and the SHA-256 in a sidecar; every run after it
    reads exactly those bytes and never opens a socket.

    `refresh` overrides that on purpose and says so. `offline` refuses to
    download at all and prints the URL, which is what a machine behind a
    firewall — or an air-gapped review of someone else's result — needs.
    """
    import json

    from .eurostat import EurostatError, fetch, load_iot

    # NOT `.json`. The sidecar sits beside its table, and `data/eurostat/` is
    # globbed for `*.json` by more than one validator, each of which would then
    # try to read a provenance record as a JSON-stat cube. `run_uk_classification`
    # did exactly that on 2026-08-25, five minutes after the first sidecar was
    # written, and died on `KeyError: 'id'`.
    side = path.with_suffix(path.suffix + ".provenance")

    if path.exists() and not refresh:
        note = None
        if side.exists():
            try:
                rec = json.loads(side.read_text())
                note = (f"Eurostat {rec.get('dataset')} {req['geo']} "
                        f"{req['year']}, {req['variant']} variant. Downloaded "
                        f"{str(rec.get('retrieved_at', ''))[:10]}, "
                        f"{rec.get('bytes')} bytes, SHA-256 "
                        f"{str(rec.get('sha256', ''))[:16]}…. Read from the "
                        f"local cache; the network was not used.")
            except (ValueError, OSError):
                note = None
        if note is None:
            # Still say it came from the cache — that is the fact the reader
            # needs either way — and then say what is missing rather than
            # letting the absence pass as if nothing were.
            note = (f"Eurostat {req['dataset']} {req['geo']} {req['year']}, "
                    f"{req['variant']} variant. Read from the local cache "
                    f"({path.name}); the network was not used. NO PROVENANCE "
                    f"SIDECAR was found beside it, so when it was downloaded, "
                    f"from what URL, and with what checksum are not recorded.")
        return _tag(load_iot(path, req["variant"]), note)

    if offline:
        raise ConfigError(
            f"--offline was given and {path} is not cached yet.\n\n"
            f"Either drop --offline, or fetch it once by hand:\n"
            f"    {_eurostat_url(req)}\n"
            f"and save the response as that file.")

    # WHAT THE OLD FILE HELD, read before it is overwritten.
    #
    # A re-download that changes the SHA-256 has NOT necessarily changed a
    # number. Measured on this project's own cache on 2026-08-25: the Spanish
    # symmetric table fetched on 2026-08-10 and again fifteen days later hashed
    # differently, and all 17,957 values were identical -- what moved was
    # Eurostat's own `updated` stamp and an `extension` block. Hashing the raw
    # bytes is the right integrity check because it is conservative, but it
    # answers "are these the same bytes", not "are these the same figures", and
    # only the second question is the analyst's. So `--refresh` compares.
    previous = None
    if refresh and path.exists():
        try:
            previous = json.loads(path.read_text()).get("value")
        except (ValueError, OSError):
            previous = None

    try:
        rec = fetch(req["dataset"], req["geo"], req["year"], path)
    except EurostatError as exc:
        raise ConfigError(
            f"the Eurostat download failed:\n{exc}\n\n"
            f"Nothing was written. The cache path was {path}.") from None
    side.write_text(json.dumps(rec, indent=2))

    verb = "Re-downloaded" if refresh else "Downloaded"
    print(f"    {verb} {rec['dataset']} {req['geo']} {req['year']} — "
          f"{rec['bytes']:,} bytes, {rec['n_values']:,} values")
    print(f"    cached at {path}")
    print(f"    SHA-256 {rec['sha256']}")
    if refresh:
        print("    --refresh replaced a file that was already here. The "
              "figures below are the new ones.")
        if previous is None:
            print("    The previous file could not be read for comparison, so "
                  "whether any figure moved is unknown.")
        else:
            now = json.loads(path.read_text()).get("value") or {}
            moved = [k for k in set(previous) | set(now)
                     if previous.get(k) != now.get(k)]
            if not moved:
                print(f"    {len(now):,} values, NONE of them changed. The "
                      f"checksum moved and the data did not — Eurostat "
                      f"restamps a release without revising it.")
            else:
                worst = max(moved, key=lambda k: abs((now.get(k) or 0)
                                                     - (previous.get(k) or 0)))
                print(f"    {len(moved):,} of {len(now):,} values CHANGED. "
                      f"Largest move: {previous.get(worst)} -> "
                      f"{now.get(worst)}. Any earlier result from this "
                      f"configuration was computed on different figures.")
    return _tag(load_iot(path, req["variant"]),
                f"Eurostat {rec['dataset']} {req['geo']} {req['year']}, "
                f"{req['variant']} variant. Downloaded "
                f"{str(rec.get('retrieved_at', 'date unrecorded'))[:10]} from "
                f"{rec.get('url', 'url unrecorded')}, "
                f"{rec.get('bytes', '?')} bytes, SHA-256 "
                f"{str(rec.get('sha256', ''))[:16]}….")


def _eurostat_url(req: dict) -> str:
    from .eurostat import API
    return (API.format(dataset=req["dataset"])
            + f"&geo={req['geo']}&time={req['year']}&unit=MIO_EUR")


def _tag(table, note: str):
    """Put the download's provenance where the report will print it.

    The reader of a result has to be able to see which file it came from and
    when it arrived, and `notes` is the field the report already surfaces under
    'What the loader decided when reading this file'.
    """
    table.notes = f"{table.notes} {note}".strip() if table.notes else note
    return table


def _rows(sheets: dict, name: str) -> list[dict]:
    """Sheet -> list of dicts keyed by the header row, blank rows dropped."""
    if name not in sheets:
        return []
    raw = [r for r in sheets[name] if r and any(c is not None for c in r)]
    if not raw:
        return []
    head = [str(c).strip().lower() if c is not None else "" for c in raw[0]]
    out = []
    for r in raw[1:]:
        # A row whose first cell starts with '#' is a comment. The template
        # writes its instructions underneath the data in exactly that form, so
        # the sheet explains itself without the explanation becoming data.
        if r and r[0] is not None and str(r[0]).lstrip().startswith("#"):
            continue
        row = {h: (r[i] if i < len(r) else None)
               for i, h in enumerate(head) if h}
        if any(v is not None and str(v).strip() != "" for v in row.values()):
            out.append(row)
    return out


def _need(row: dict, field: str, sheet: str, n: int):
    v = row.get(field)
    if v is None or str(v).strip() == "":
        raise ConfigError(f"sheet '{sheet}', row {n}: '{field}' is empty and "
                          f"is required")
    return v


def _strength(v, sheet: str, n: int) -> ProxyStrength:
    s = str(v).strip().lower()
    if s not in ("strong", "medium", "weak"):
        raise ConfigError(f"sheet '{sheet}', row {n}: strength {v!r} must be "
                          f"strong, medium or weak. It is not decoration — a "
                          f"weak proxy makes the whole split weak, and the "
                          f"report says so.")
    return ProxyStrength(s)


def load_config(path: Path | str, *, offline: bool = False,
                refresh: bool = False) -> dict:
    """Read a configuration WORKBOOK and return everything `IOProject` needs.

    `offline` refuses any network access; `refresh` forces a re-fetch of a
    cached download. Both are no-ops unless `table_kind` is `eurostat`.
    """
    path = Path(path)
    sheets = _open_workbook(path)
    missing = [s for s in REQUIRED_SHEETS if s not in sheets]
    if missing:
        raise ConfigError(
            f"{path.name} is missing the sheet(s) {', '.join(missing)}. "
            f"Run `python3 run_quadrium.py --template my_config.xlsx` to get a "
            f"workbook with the right shape and comments in it.")
    if not any(s in sheets and _rows(sheets, s) for s in ONE_OF_SHEETS):
        raise ConfigError(
            f"{path.name} says which table to use and then nothing to do with "
            f"it. Fill in `splits` to divide a sector, `regionalise` to "
            f"estimate a region from it, or `targets` to project a supply-use "
            f"pair. A workbook with none of the three describes no job.")
    meta = {}
    for r in sheets["project"]:
        if r and r[0] is not None and str(r[0]).strip():
            k = str(r[0]).strip()
            if k.startswith("#"):
                continue
            meta[k.lower()] = r[1] if len(r) > 1 else None
    tables = {name: _rows(sheets, name)
              for name in ("splits", "keys", "scenarios", "profiles",
                           "targets")}

    # `regionalise` is key/value like `project`, not a table of rows: the job
    # has one set of parameters, not one per sector.
    reg = {}
    for r in sheets.get("regionalise", []):
        if r and r[0] is not None and str(r[0]).strip():
            k = str(r[0]).strip()
            # `#` is a comment; `key` is the template's own header row, which
            # the first version read as a setting called "key" and so treated
            # an untouched template as a regionalisation with nothing in it.
            if k.startswith("#") or k.lower() in ("key", "sector_code"):
                continue
            reg[k.lower()] = r[1] if len(r) > 1 else None
    if reg:
        return build_regionalisation(meta, reg, base_dir=path.parent,
                                     label=path.name, offline=offline,
                                     refresh=refresh)
    return build_config(meta, tables, base_dir=path.parent, label=path.name,
                        offline=offline, refresh=refresh)


def _load_declared_table(meta: dict, base_dir, tables: dict, offline: bool,
                         refresh: bool, defaults_taken: list):
    """The table the `project` sheet names, however it names it.

    Extracted at v1.86 so a regionalisation reads its national table exactly
    the way a split reads the table it divides -- including `table_kind:
    eurostat`, which the command-line route cannot do at all. Two ways of
    naming the same table would have been two places to fix a loader.
    """
    kind = str(meta.get("table_kind") or "uk_analytical").strip().lower()
    if kind not in TABLE_KINDS:
        raise ConfigError(f"table_kind {kind!r} must be one of {TABLE_KINDS}")

    # `eurostat` names a country and a year instead of a file, and `table_path`
    # becomes where the download is KEPT rather than where it already is. So
    # the existence check below cannot apply to it: on a first run the file is
    # supposed to be missing.
    if kind == "eurostat_sut":
        table_path, fetch_note = _eurostat_sut_paths(meta, base_dir)
    elif kind == "eurostat":
        table_path, fetch_note = _eurostat_cache_path(meta, base_dir)
    else:
        table_path = Path(str(_need(meta, "table_path", "project", 0)).strip())
        if not table_path.is_absolute():
            table_path = (Path(base_dir) / table_path).resolve()
        fetch_note = None
        if not table_path.exists():
            raise ConfigError(f"table_path points at {table_path}, which does "
                              f"not exist. Paths may be absolute or relative "
                              f"to the config file.")

    # `table_unbalanced` means something for exactly one kind. A workbook that
    # sets it anywhere else is refused rather than quietly ignored: an analyst
    # who typed `residual_column` and got a silent `refuse` would have no way of
    # telling, and an analyst who typed it on a table that balances would think
    # they had authorised something they had not.
    unbalanced = str(meta.get("table_unbalanced") or "refuse").strip().lower()
    if unbalanced not in ("refuse", "residual_column"):
        raise ConfigError(f"table_unbalanced {unbalanced!r} must be 'refuse' "
                          f"(the default) or 'residual_column'")
    if unbalanced != "refuse" and kind != "ine_interior":
        raise ConfigError(
            f"table_unbalanced={unbalanced!r} applies only to "
            f"table_kind 'ine_interior', not {kind!r}. Refusing rather than "
            f"ignoring a setting you would never see was ignored.")

    loaders = {
        "uk_analytical": lambda p: load_uk_analytical_iot(p),
        "interchange": lambda p: load_io_table(p),
        "ine_interior": lambda p: load_ine_tio(p, "interior", unbalanced),
        "ine_total": lambda p: load_ine_tio(p, "total"),
        "eurostat": lambda p: _load_eurostat(p, fetch_note, offline, refresh),
        "eurostat_sut": lambda p: _load_eurostat_sut(
            fetch_note, offline, refresh, defaults_taken,
            tables.get("targets", [])),
    }
    try:
        table = loaders[kind](table_path)
    except LoaderError as exc:
        raise ConfigError(f"the table could not be loaded:\n{exc}") from None
    return table, table_path, kind


def build_config(meta: dict, tables: dict, base_dir: Path = Path("."),
                 label: str = "<configuration>", *, offline: bool = False,
                 refresh: bool = False) -> dict:
    """Build a run from plain Python data, with no spreadsheet involved.

    `meta` is the `project` sheet as a dict; `tables` maps 'splits', 'keys',
    'scenarios' and 'profiles' to lists of row-dicts with the same column names
    the workbook uses.

    WHY THIS EXISTS SEPARATELY FROM THE WORKBOOK.
    A configuration is a declarative description of a run: which table, which
    sectors, which proxies with which sources, which scenarios. A human writes
    that in a spreadsheet. A program — a connector to a statistical institute,
    a script, or a model asked to set up a run — writes the same thing as data.
    Keeping the two entry points on one code path means both get identical
    validation and identical error messages, so a machine-written configuration
    cannot take a shortcut a human-written one could not.

    Every check below is deliberately the same for both. In particular a
    generated configuration still has to name a real source and a real strength
    for every proxy: there is no path into this engine that produces a number
    without saying where it came from.
    """
    meta = {str(k).strip().lower(): v for k, v in (meta or {}).items()}
    # Declared here rather than beside the scenarios below, because the loaders
    # take defaults too and the analyst must hear about those as well.
    defaults_taken: list[str] = []
    project_id = str(_need(meta, "project_id", "project", 0)).strip()
    table, table_path, kind = _load_declared_table(
        meta, base_dir, tables, offline, refresh, defaults_taken)

    # ---- keys ---------------------------------------------------------
    grouped: dict[str, list[dict]] = {}
    for n, r in enumerate(tables.get("keys", []), start=2):
        grouped.setdefault(str(_need(r, "key_id", "keys", n)).strip(),
                           []).append((n, r))
    keys = {}
    for key_id, entries in grouped.items():
        codes, values, srcs, yrs, strengths = [], [], [], [], []
        for n, r in entries:
            codes.append(str(_need(r, "new_sector_code", "keys", n)).strip())
            try:
                values.append(float(_need(r, "value", "keys", n)))
            except (TypeError, ValueError):
                raise ConfigError(f"sheet 'keys', row {n}: value "
                                  f"{r.get('value')!r} is not a number") from None
            srcs.append(str(_need(r, "source", "keys", n)).strip())
            yrs.append(int(_need(r, "source_year", "keys", n)))
            strengths.append(_strength(_need(r, "strength", "keys", n),
                                       "keys", n))
        order = ["strong", "medium", "weak"]
        weakest = max(strengths, key=lambda s: order.index(s.value))
        keys[key_id] = AllocationKey(
            key_id=key_id, applies_to="output", new_sector_codes=codes,
            raw_values=values, source=srcs[0], source_year=yrs[0],
            # A key is only as strong as its weakest row: a split resting on
            # one weak proxy is a weak split, whatever the other rows say.
            strength=weakest,
            notes=(f"from {label}"
                   + ("; rows disagree on source, the first is recorded"
                      if len(set(srcs)) > 1 else "")))

    # ---- splits -------------------------------------------------------
    by_sector: dict[str, dict] = {}
    for n, r in enumerate(tables.get("splits", []), start=2):
        sector = str(_need(r, "sector_code", "splits", n)).strip()
        entry = by_sector.setdefault(sector, {"codes": [], "labels": [],
                                              "keys": set(), "rows": []})
        entry["codes"].append(str(_need(r, "new_code", "splits", n)).strip())
        entry["labels"].append(str(r.get("new_label")
                                   or entry["codes"][-1]).strip())
        k = r.get("key_id")
        if k is not None and str(k).strip():
            entry["keys"].add(str(k).strip())
        entry["rows"].append(n)
    if not by_sector:
        raise ConfigError("sheet 'splits' has no rows: nothing to divide")

    splits = []
    for sector, e in by_sector.items():
        try:
            table.index_of(sector)
        except KeyError:
            raise ConfigError(
                f"sheet 'splits' rows {e['rows']}: sector {sector!r} is not in "
                f"the loaded table. Check the code against the table's own "
                f"classification ({table.classification}).") from None
        if len(e["keys"]) > 1:
            # Silently taking one of them would make the result depend on row
            # order, which is exactly the kind of thing that never gets noticed.
            raise ConfigError(
                f"split '{sector}' names more than one allocation key "
                f"({', '.join(sorted(e['keys']))}) across its rows. One split "
                f"takes one key. Use separate splits, or leave key_id blank on "
                f"all but one row.")
        e["key"] = next(iter(e["keys"]), None)
        if e["key"] and e["key"] not in keys:
            raise ConfigError(
                f"split '{sector}' names key '{e['key']}', which is not in the "
                f"'keys' sheet. Available: {', '.join(sorted(keys)) or 'none'}")
        if e["key"]:
            kk = keys[e["key"]]
            if sorted(kk.new_sector_codes) != sorted(e["codes"]):
                raise ConfigError(
                    f"split '{sector}' lists subsectors {e['codes']} but key "
                    f"'{e['key']}' covers {kk.new_sector_codes}. Every "
                    f"subsector needs a weight and every weight needs a "
                    f"subsector.")
        splits.append(SplitSpec(
            sector_code=sector, new_codes=e["codes"], new_labels=e["labels"],
            keys_by_block={"output": e["key"]} if e["key"] else {}))

    # ---- profiles, grouped by scenario ---------------------------------
    profiles: dict[str, dict] = {}
    all_new = {c for s in splits for c in s.new_codes}
    for n, r in enumerate(tables.get("profiles", []), start=2):
        sid = str(_need(r, "scenario_id", "profiles", n)).strip()
        sub = str(_need(r, "subsector_code", "profiles", n)).strip()
        sup = str(_need(r, "supplier_code", "profiles", n)).strip()
        if sub not in all_new:
            raise ConfigError(
                f"sheet 'profiles', row {n}: subsector {sub!r} is not created "
                f"by any split. Created: {', '.join(sorted(all_new))}")
        try:
            table.index_of(sup)
        except KeyError:
            raise ConfigError(f"sheet 'profiles', row {n}: supplier {sup!r} is "
                              f"not a sector of the loaded table") from None
        try:
            val = float(_need(r, "intensity", "profiles", n))
        except (TypeError, ValueError):
            raise ConfigError(f"sheet 'profiles', row {n}: intensity "
                              f"{r.get('intensity')!r} is not a number") from None
        profiles.setdefault(sid, {}).setdefault(sub, {})[sup] = val

    # ---- scenarios -----------------------------------------------------
    # DEFAULTS TAKEN ARE COLLECTED AND REPORTED, not applied quietly. A sheet
    # that is absent and a sheet whose NAME was mistyped are the same thing to
    # this loader, and the second is a mistake the analyst would never see: the
    # run succeeds, on a configuration they did not write (2026-08-10).
    scen_rows = list(tables.get("scenarios", []))
    if not scen_rows:
        scen_rows = [{"scenario_id": "S1", "label": "Default"}]
        defaults_taken.append(
            "no 'scenarios' sheet (or it was empty), so ONE scenario named 'S1' "
            "was created with alpha 0.5. If you wrote a scenarios sheet, check "
            "its name is exactly 'scenarios' — a mistyped sheet name looks "
            "identical to an absent one from here")
    if not tables.get("profiles"):
        defaults_taken.append(
            "no 'profiles' sheet, so every subsector buys the same mix as its "
            "parent and all subsectors of a split share its multiplier")
    if not tables.get("keys"):
        defaults_taken.append(
            "no 'keys' sheet, so the split has no allocation key of its own")
    scenarios = []
    for n, r in enumerate(scen_rows, start=2):
        sid = str(_need(r, "scenario_id", "scenarios", n)).strip()
        alpha = r.get("internal_block_alpha")
        scenarios.append(Scenario(
            scenario_id=sid, label=str(r.get("label") or sid).strip(),
            description=(str(r["description"]).strip()
                         if r.get("description") else None),
            internal_block_alpha=(float(alpha) if alpha not in (None, "")
                                  else 0.5),
            input_profiles=profiles.get(sid, {})))
    named = {s.scenario_id for s in scenarios}
    orphan = set(profiles) - named
    if orphan:
        raise ConfigError(
            f"sheet 'profiles' names scenario(s) {sorted(orphan)} that the "
            f"'scenarios' sheet does not define. A profile with no scenario "
            f"would be silently ignored, which is worse than an error.")

    # ---- ledger --------------------------------------------------------
    ledger = AssumptionLedger(project_id=project_id)
    for key in keys.values():
        ledger.add(Assumption(
            assumption_id=f"KEY-{key.key_id}",
            description=f"Subsectors split by '{key.key_id}' with weights "
                        + ", ".join(f"{c} {w:.1%}" for c, w
                                    in zip(key.new_sector_codes, key.weights)),
            applies_to="allocation", source=key.source,
            validated_by="declared in the configuration workbook",
            confidence=key.strength,
            impact_on_results=("high — a weak proxy makes the whole split weak"
                               if key.strength is ProxyStrength.WEAK
                               else "medium")))
    for sid, prof in profiles.items():
        ledger.add(Assumption(
            assumption_id=f"PROFILE-{sid}",
            description=f"Scenario '{sid}' gives {len(prof)} subsector(s) "
                        f"purchasing intensities that differ from their "
                        f"parent's average.",
            applies_to="input structure",
            source="analyst judgement, declared in the configuration workbook",
            validated_by="NOT VALIDATED — no source states these intensities",
            confidence=ProxyStrength.WEAK,
            impact_on_results="this is what makes the subsectors' multipliers "
                              "differ; without it they are identical by "
                              "construction"))
    ledger.add(Assumption(
        assumption_id="ALPHA",
        description="Trade among new subsectors is estimated by double "
                    "proportionality, damped on the diagonal by alpha.",
        applies_to="internal block",
        source="MVP_0.1 §6.3 — project convention, no methodological source",
        validated_by="pending analyst review", confidence=ProxyStrength.WEAK,
        impact_on_results="high — the only part of the table with no "
                          "observation behind it"))

    return {"project_id": project_id,
            "title": str(meta.get("title") or f"{project_id} — sector split"),
            "table": table, "splits": splits, "scenarios": scenarios,
            "keys": keys, "ledger": ledger, "source_file": table_path,
            "notes": str(meta.get("notes") or ""),
            "defaults_taken": defaults_taken}


def write_template(path: Path | str) -> Path:
    """Write a workbook the analyst can fill in, with the comments inside it."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    path = Path(path)
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    head = PatternFill("solid", fgColor="DDEBF7")
    note = Font(italic=True, color="808080")

    def sheet(name, header, rows, notes):
        ws = wb.create_sheet(name)
        for j, h in enumerate(header, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font, c.fill = bold, head
            ws.column_dimensions[c.column_letter].width = max(16, len(h) + 4)
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=v)
        r = len(rows) + 3
        for line in notes:
            ws.cell(row=r, column=1,
                    value=(f"# {line}" if line else "#")).font = note
            r += 1
        return ws

    ws = wb.active
    ws.title = "project"
    for i, (k, v) in enumerate([
            ("project_id", "my_first_split"),
            # Relative to THIS workbook, and it resolves only if the
            # workbook was written inside a checkout. Someone who installed the
            # package does not have that file; the note beside the field says
            # so and sends them to --sources, which lists what they do have.
            ("table_path", "../UK_IOAT_2023_domestic_ixi.xlsx"),
            ("table_kind", "uk_analytical"),
            ("title", "My sector split"),
            ("notes", "")], start=1):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 46
    for i, line in enumerate([
            "#",
            "# table_path may be absolute, or relative to THIS file.",
            "# The seeded value points at a table that ships with the SOURCE",
            "# CHECKOUT. If you installed the package you do not have it: run",
            "#     quadrium --sources",
            "# to list every table this engine can load on your machine, and",
            "#     quadrium --find <CODE> --geo <XX>",
            "# if you know the sector but not the table.",
            "# table_kind: uk_analytical  the ONS workbook (industry x industry)",
            "#             ine_interior  the INE workbook, domestic output",
            "#             ine_total     the INE workbook, total flows",
            "#             interchange   the project's own format, and what",
            "#                           this engine writes -- point at a",
            "#                           result to split a second sector.",
            "#             eurostat      symmetric IOT, downloaded by",
            "#                           country and year.",
            "#             eurostat_sut  supply-use pair, downloaded and",
            "#                           TRANSFORMED into a symmetric table.",
            "#",
            "# For table_kind: eurostat, delete table_path (or use it to say",
            "# where to cache) and add instead:",
            "#     eurostat_geo       ES        two-letter country code",
            "#     eurostat_year      2022",
            "#     eurostat_dataset   product_by_product | industry_by_industry",
            "#     eurostat_variant   domestic | total",
            "#",
            "# For table_kind: eurostat_sut, the same geo and year, and:",
            "#     eurostat_model     A | B | C | D   (default D)",
            "#   A product technology       product x product, may go negative",
            "#   B industry technology      product x product, cannot",
            "#   C fixed industry sales     industry x industry, may go negative",
            "#   D fixed product sales      industry x industry, cannot",
            "# The four are four assumptions about secondary production and",
            "# give four different tables from the same data. CORE_013",
            "# Figure 12.2, p. 378.",
            "# The first run downloads and records the URL and SHA-256; every",
            "# run after it reads those same bytes offline. --refresh forces a",
            "# new download and warns that your results may move.",
            "# table_unbalanced: refuse (default) or residual_column.",
            "#             Only for ine_interior, which does not balance for",
            "#             one product -- see OQ-D-04. Anywhere else it is an error.",
            "#",
            "# Fill in the other sheets, then run:",
            "#     quadrium <this file>",
            "# (from a checkout, without installing: python3 run_quadrium.py)",
    ], start=7):
        ws.cell(row=i, column=1, value=line).font = note

    sheet("splits", ["sector_code", "new_code", "new_label", "key_id"],
          [["I56", "I561", "Restaurants and mobile food service", "k56"],
           ["I56", "I562", "Event catering", "k56"],
           ["I56", "I563", "Beverage serving activities", "k56"]],
          ["One row per NEW subsector. Rows sharing sector_code form one split.",
           "You may divide several sectors: just add rows with another",
           "sector_code. key_id names the allocation key in the 'keys' sheet."])

    sheet("keys", ["key_id", "new_sector_code", "value", "source",
                   "source_year", "strength"],
          [["k56", "I561", 720000, "ONS BRES (REPLACE — illustrative)", 2023,
            "weak"],
           ["k56", "I562", 120000, "ONS BRES (REPLACE — illustrative)", 2023,
            "weak"],
           ["k56", "I563", 380000, "ONS BRES (REPLACE — illustrative)", 2023,
            "weak"]],
          ["The proxy that decides how big each subsector is. Values are",
           "relative; they are normalised for you.",
           "strength: strong / medium / weak. A key is recorded at its WEAKEST",
           "row, because a split resting on one weak proxy is a weak split.",
           "Write the real source. It goes into the report and the ledger."])

    sheet("scenarios", ["scenario_id", "label", "description",
                        "internal_block_alpha"],
          [["S1_plain", "Size only",
            "No input profiles: every subsector inherits its parent's "
            "purchasing pattern.", None],
           ["S2_profiled", "Differentiated input structures",
            "Subsectors buy different mixes.", None]],
          ["One row per scenario. Leave internal_block_alpha blank for the",
           "default 0.5 (MVP_0.1 §6.3 — a project convention, not a source).",
           "Two or more scenarios let you see how much the answer depends on",
           "your choices, which is the honest way to present it."])

    sheet("profiles", ["scenario_id", "subsector_code", "supplier_code",
                       "intensity"],
          [["S2_profiled", "I563", "C1101T1106 & C12", 2.1],
           ["S2_profiled", "I563", "C101", 0.45],
           ["S2_profiled", "I561", "C101", 1.35]],
          ["How intensively each subsector buys from each supplier, RELATIVE",
           "to the parent sector's average. 1.0 = the average, 2.1 = buys",
           "2.1 times as intensively, 0.45 = less than half.",
           "",
           "WITHOUT this sheet every subsector gets a scaled copy of its",
           "parent's input structure, and they all end up with the SAME",
           "multiplier — arithmetic, not economics. This sheet is what makes",
           "them genuinely different as buyers.",
           "",
           "There is a limit: how far you can push it is bounded by how much",
           "the parent sector trades with itself. The report prints the",
           "headroom, and an impossible set is rejected with an explanation."])

    path.parent.mkdir(parents=True, exist_ok=True)
    # The other job this workbook can describe. Left EMPTY of values on
    # purpose: a workbook with both `splits` and `regionalise` filled in
    # describes two jobs, and the loader takes the regionalisation. The
    # comments say what to put here when that is what you want.
    sheet("regionalise", ["key", "value"], [], [
        '',
        "LEAVE THIS SHEET EMPTY unless you want to estimate a REGION's",
        'table from a national one, instead of dividing a sector.',
        '',
        'Fill in three keys in column A, their values in column B:',
        'method          SLQ | CILQ | RLQ | FLQ   (default FLQ)',
        "delta           the FLQ's convexity parameter, 0 <= d < 1.",
        'REQUIRED for FLQ and there is no default:',
        'measured across ten regions in two countries',
        'it runs from 0.14 to 0.60, median 0.26, so a',
        'default would be a guess wearing a number.',
        'activity_path   a CSV, `sector_code,regional` and optionally',
        '`,national`, one row per sector of the table',
        'named on the `project` sheet. Output or',
        'employment; without the third column the',
        "table's own output is used, which is right",
        'only when the second column is output too.',
        '',
        'The table comes from the `project` sheet, as it does for a split,',
        'so `table_kind: eurostat` works here too. It must be the DOMESTIC',
        "table: a total-flow one regionalises the country's imports as",
        'though they were local supply, and nothing downstream catches it.',
        '',
        'Every run prints what the method is known to get wrong -- the',
        'family overstates local multipliers, and cross-hauling is not',
        'reproduced in any amount anyone chose. There is no way to',
        'suppress that, and the absence of one is deliberate.',
    ])

    wb.save(path)
    return path


def read_activity(path: Path | str, table) -> tuple:
    """The region's activity by sector, against the table it will be scaled from.

    Shared by the workbook route and the command-line one, so the two cannot
    drift into different refusals for the same mistake. Returns
    `(Q_region, Q_national, national_from)`.
    """
    import csv as _csv

    path = Path(path)
    if not path.exists():
        raise ConfigError(f"activity_path points at {path}, which does not "
                          f"exist. Paths may be absolute or relative to the "
                          f"config file.")
    rows = list(_csv.reader(path.open()))
    head = [str(c).strip().lower() for c in (rows[0] if rows else [])]
    if head[:2] != ["sector_code", "regional"]:
        raise ConfigError(
            f"{path.name} must start with a header row reading "
            f"`sector_code,regional` and optionally `,national`. Regional "
            f"activity is output or employment; without the third column the "
            f"table's own output is used, which is only right when the second "
            f"column is output too.")
    has_national = len(head) > 2 and head[2] == "national"

    seen = {str(r[0]).strip(): r for r in rows[1:] if r and str(r[0]).strip()}
    missing = [c for c in table.sector_codes if c not in seen]
    extra = [c for c in seen if c not in table.sector_codes]
    if missing or extra:
        raise ConfigError(
            f"{path.name} and the table do not describe the same sectors.\n"
            f"  in the table but not the file: "
            f"{', '.join(missing[:8]) or 'none'}{' …' if len(missing) > 8 else ''}\n"
            f"  in the file but not the table: "
            f"{', '.join(extra[:8]) or 'none'}{' …' if len(extra) > 8 else ''}\n"
            f"Align the classifications; do not pad.")
    try:
        q_reg = np.array([float(seen[c][1]) for c in table.sector_codes])
        q_nat = (np.array([float(seen[c][2]) for c in table.sector_codes])
                 if has_national else table.X.copy())
    except (ValueError, IndexError) as exc:
        raise ConfigError(f"{path.name} has a value that is not a number: "
                          f"{exc}") from None
    return q_reg, q_nat, ("file" if has_national else "table output")


def build_regionalisation(meta: dict, reg: dict, base_dir: Path = Path("."),
                          label: str = "<configuration>", *,
                          offline: bool = False,
                          refresh: bool = False) -> dict:
    """A regionalisation described by a workbook rather than by flags.

    The `project` sheet already says which table to use and how to load it, so
    this sheet only says what to do with it -- which is why `table_kind:
    eurostat` works here and does not on the command line.
    """
    meta = {str(k).strip().lower(): v for k, v in (meta or {}).items()}
    defaults_taken: list[str] = []
    project_id = str(_need(meta, "project_id", "project", 0)).strip()
    table, table_path, kind = _load_declared_table(
        meta, base_dir, {}, offline, refresh, defaults_taken)

    method = str(reg.get("method") or "FLQ").strip().upper()
    if method not in ("SLQ", "CILQ", "RLQ", "FLQ"):
        raise ConfigError(f"method {method!r} on the `regionalise` sheet must "
                          f"be SLQ, CILQ, RLQ or FLQ")
    delta = reg.get("delta")
    if delta is not None and str(delta).strip() != "":
        try:
            delta = float(delta)
        except (TypeError, ValueError):
            raise ConfigError(f"delta {delta!r} is not a number") from None
    else:
        delta = None

    ap = reg.get("activity_path")
    if not ap or not str(ap).strip():
        raise ConfigError(
            "the `regionalise` sheet needs `activity_path`: a CSV of the "
            "region's activity by sector. There is nothing to scale the "
            "national table by without it.")
    ap = Path(str(ap).strip())
    if not ap.is_absolute():
        ap = (Path(base_dir) / ap).resolve()
    q_reg, q_nat, national_from = read_activity(ap, table)

    return {"kind": "regionalise", "project_id": project_id,
            "title": str(meta.get("title") or f"{project_id} — regionalisation"),
            "table": table, "source_file": table_path, "table_kind": kind,
            "method": method, "delta": delta,
            "Q_region": q_reg, "Q_national": q_nat,
            "national_activity_from": national_from,
            "activity_file": ap,
            "notes": str(meta.get("notes") or ""),
            "defaults_taken": defaults_taken}
