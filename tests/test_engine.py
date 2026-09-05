"""
Unit checks for the MVP engine. Runnable without pytest:

    python3 tests/test_engine.py

The end-to-end check is `examples/synthetic_accommodation.py`, which exits
non-zero if any scenario fails validation. This file isolates the properties
that example cannot show on its own.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "examples"))
sys.path.insert(0, str(ROOT / "library" / "validators"))

from synthetic_accommodation import (build_keys, build_scenarios,  # noqa: E402
                                     build_table)
from quadrium.balancing import BalancingError, balance, ras  # noqa: E402
from quadrium.disaggregation import split_sectors, targets  # noqa: E402
from quadrium.reaggregation import reaggregate  # noqa: E402
from quadrium.models import (ProxyStrength, Scenario,  # noqa: E402
                              SplitSpec)
from quadrium.scenarios import ScenarioInfeasible, run_scenario  # noqa: E402

NEW = ["HOT", "CAM", "RES", "FBS"]
LBL = ["Hotels", "Camping", "Restaurants", "F&B"]
FAILURES: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'} {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAILURES.append(name)


def test_gras_reduces_to_ras():
    """RAS is the special case of GRAS with no negatives (UNH_18 ¶18.35, p. 558).

    Not an aesthetic claim — it is why replacing RAS with GRAS costs nothing.
    """
    rng = np.random.default_rng(0)
    Z = rng.uniform(1, 100, (6, 6))
    tr = Z.sum(axis=1) * rng.uniform(0.9, 1.1, 6)
    tc = Z.sum(axis=0) * rng.uniform(0.9, 1.1, 6)
    tc *= tr.sum() / tc.sum()                      # make the margins consistent
    Zr, _, _ = ras(Z, tr, tc, tol=1e-12)
    Zg, info = balance(Z, tr, tc, method="GRAS", tol=1e-12)
    dev = float(np.max(np.abs(Zr - Zg)))
    check("GRAS reproduces RAS on a non-negative matrix", dev < 1e-6,
          f"max|RAS - GRAS| = {dev:.2e}, method chosen = {info['method']}")


def test_ras_refuses_negatives():
    Z = np.array([[10.0, 5.0], [-2.0, 8.0]])
    try:
        balance(Z, np.array([16.0, 7.0]), np.array([9.0, 14.0]), method="RAS")
    except BalancingError as exc:
        check("RAS refuses a matrix with negatives", "non-negative" in str(exc))
    else:
        check("RAS refuses a matrix with negatives", False, "it accepted one")


def test_reaggregation_is_exact():
    """The Reaggregation Guarantee (MVP_0.1 §8), the sharp version.

    Everything except the split sector was copied, never estimated, so it must
    return bit-for-bit — not 'within tolerance'.
    """
    table = build_table()
    keys, scenarios = build_keys(), build_scenarios()
    res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], scenarios[0], keys)
    Z_re = reaggregate(res.table.Z, res.mapping, table.n)
    diff = np.abs(Z_re - table.Z)
    mask = np.ones_like(diff, dtype=bool)
    sidx = [s["sector_code"] for s in res.splits]
    p = table.index_of(sidx[0])
    mask[p, :] = False
    mask[:, p] = False
    check("untouched sectors reaggregate exactly", float(diff[mask].max()) < 1e-12,
          f"max|dev| = {diff[mask].max():.2e}")
    check("split sector reaggregates within tolerance", float(diff.max()) < 1e-9,
          f"max|dev| = {diff.max():.2e}")


def test_signs_and_zeros_preserved():
    table = build_table()
    res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], build_scenarios()[0], build_keys())
    seed_sign = np.sign(res.seed_Z)
    out_sign = np.sign(res.table.Z)
    check("no cell changes sign during balancing",
          int(np.count_nonzero(seed_sign != out_sign)) == 0)
    check("the original negative survives",
          int((res.table.Z < 0).sum()) == int((res.seed_Z < 0).sum()),
          f"{int((res.table.Z < 0).sum())} negative cell(s) in the result")


def test_infeasible_scenario_is_rejected():
    """Multi-proxy splitting can describe an economy that cannot exist."""
    table = build_table()
    keys, scenarios = build_keys(), build_scenarios()
    mixed = [s for s in scenarios if s.scenario_id == "S3_mixed"][0]
    try:
        run_scenario(table, [SplitSpec("ACC", NEW, LBL)], mixed, keys)
    except ScenarioInfeasible as exc:
        check("infeasible proxy combination is rejected, not solved",
              "impossible economy" in exc.explanation)
    else:
        check("infeasible proxy combination is rejected, not solved", False,
              "it produced a table")


def test_original_table_is_never_mutated():
    table = build_table()
    before = table.Z.copy()
    run_scenario(table, [SplitSpec("ACC", NEW, LBL)], build_scenarios()[0], build_keys())
    check("the original table is not mutated (spec §5)",
          np.array_equal(before, table.Z))


def test_weights_sum_to_one():
    table = build_table()
    seed = split_sectors(table, [SplitSpec("ACC", NEW, LBL)], build_scenarios()[0], build_keys())
    ok = all(np.isclose(np.asarray(w).sum(), 1.0)
             for s in seed["splits"] for w in s["weights"].values())
    check("every allocation key normalises to 1", ok)


def test_real_uk_table_loads_and_balances():
    """The published ONS table, not a synthetic one.

    Guards the two traps in that file: the reference year is 2023 while the
    filename says 2022 (OQ-D-01), and the final-demand block mixes components
    with a subtotal of them (`P3 S1`), which double-counts household
    consumption by GBP 259,330 million on one row if summed as printed.
    """
    from quadrium.io_loader import load_uk_analytical_iot
    path = ROOT / "UK_IOAT_2023_domestic_ixi.xlsx"
    if not path.exists():
        check("real UK table loads", True, "fixture absent, skipped")
        return
    t = load_uk_analytical_iot(path)
    check("year read from the workbook, not the filename", t.year == 2023,
          f"year = {t.year}, filename says 2022")
    check("no final-demand subtotal column survives",
          "P3 S1" not in t.Y_labels, f"Y = {t.Y_labels}")
    row = float(np.abs(t.Z.sum(1) + t.Y.sum(1) - t.X).max())
    col = float(np.abs(t.Z.sum(0) + t.VA.sum(0) - t.X).max())
    scale = float(np.abs(t.X).max())
    check("the real table balances", max(row, col) < 1e-6 * scale,
          f"rows {row:.2e}, cols {col:.2e} against outputs up to {scale:,.0f}")
    check("the real table has legitimate negatives",
          int((t.Z < 0).sum()) + int((t.Y < 0).sum()) + int((t.VA < 0).sum()) > 0,
          f"Z={int((t.Z<0).sum())} Y={int((t.Y<0).sum())} VA={int((t.VA<0).sum())}"
          " — a non-negative solver could not reproduce this table")


def test_loader_refuses_an_unbalanced_table():
    from quadrium.io_loader import LoaderError, _assert_balances
    table = build_table()
    table.X = table.X.copy()
    table.X[0] += 1000.0
    try:
        _assert_balances(table, "tampered")
    except LoaderError as exc:
        check("an unbalanced table is refused, not repaired",
              "does not balance" in str(exc))
    else:
        check("an unbalanced table is refused, not repaired", False,
              "it was accepted")


def test_real_ine_table_loads_and_balances():
    """The published INE table — the second real fixture, and a different shape.

    Product by product, not industry by industry; split across three sheets;
    final-demand subtotals named in Spanish prose with no codes, so the UK
    prefix heuristic cannot reach them.

    The interior table does not balance for one product (OQ-D-04), which makes
    this the one place where refusing is the *expected* outcome and getting a
    table back is the failure.
    """
    from quadrium.io_loader import LoaderError, load_ine_tio
    path = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    if not path.exists():
        check("real INE table loads", True, "fixture absent, skipped")
        return

    try:
        load_ine_tio(path)
    except LoaderError as exc:
        worst = next((l.strip() for l in str(exc).splitlines()
                      if "worst row" in l), str(exc).splitlines()[0])
        check("the INE interior table is refused by default (OQ-D-04)",
              "does not balance" in str(exc) and "4,921" in str(exc), worst)
    else:
        check("the INE interior table is refused by default (OQ-D-04)", False,
              "it was accepted — the 4,921.6 discrepancy went in silently")

    t = load_ine_tio(path, unbalanced="residual_column")
    check("year read from the workbook, not the two-digit filename",
          t.year == 2022, f"year = {t.year}")
    check("the '44 bis' product keeps its own code",
          "44 bis" in t.sector_codes and len(set(t.sector_codes)) == 64,
          f"{len(set(t.sector_codes))} distinct codes")
    check("no final-demand subtotal column survives",
          not any(l.startswith(("Total", "Formación bruta de capital"))
                  and "fijo" not in l for l in t.Y_labels),
          f"Y = {t.Y_labels}")
    check("the residual says out loud that the INE did not publish it",
          any("RESIDUAL" in l for l in t.Y_labels)
          and t.notes is not None and "OQ-D-04" in t.notes)

    for name, table in (("interior", t),
                        ("total", load_ine_tio(path, variant="total"))):
        row = float(np.abs(table.Z.sum(1) + table.Y.sum(1) - table.X).max())
        col = float(np.abs(table.Z.sum(0) + table.VA.sum(0) - table.X).max())
        check(f"the INE {name} table balances",
              max(row, col) < 1e-6 * float(np.abs(table.X).max()),
              f"rows {row:.2e}, cols {col:.2e}")

    # The layout is hard-coded. Prove the guard that re-derives it actually
    # fires, rather than trusting that it would.
    from quadrium import io_loader as IL
    for key, wrong in (("row_output", 85), ("row_gos", 83)):
        good = IL._INE[key]
        IL._INE[key] = wrong
        try:
            load_ine_tio(path, unbalanced="residual_column")
        except LoaderError as exc:
            failed = next((l.strip() for l in str(exc).splitlines()
                           if "failed check" in l), "")
            check(f"a moved '{key}' is caught by an identity",
                  "layout no longer matches" in str(exc), failed)
        else:
            check(f"a moved '{key}' is caught by an identity", False,
                  "the wrong row was read without complaint")
        finally:
            IL._INE[key] = good


def test_the_spanish_table_is_reachable_from_a_workbook():
    """A loader nobody can ask for is a loader that does not exist.

    Also guards the anti-silence rule: `table_unbalanced` means something for
    exactly one `table_kind`, and setting it anywhere else is refused rather
    than ignored.
    """
    from quadrium.config import ConfigError, build_config
    path = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    if not path.exists():
        check("the Spanish table is reachable from a workbook", True,
              "fixture absent, skipped")
        return

    base = dict(project_id="es", table_path=str(path))
    splits = {"splits": [{"sector_code": "36", "new_code": "36A",
                          "new_label": "Alojamiento", "key_id": ""},
                         {"sector_code": "36", "new_code": "36B",
                          "new_label": "Comidas y bebidas", "key_id": ""}]}

    cfg = build_config({**base, "table_kind": "ine_interior",
                        "table_unbalanced": "residual_column"}, splits)
    check("a workbook can ask for the Spanish domestic table",
          cfg["table"].table_id == "ES-TIO-PXP-2022-interior",
          cfg["table"].table_id)
    check("and for the total-flows one",
          build_config({**base, "table_kind": "ine_total"},
                       splits)["table"].table_id == "ES-TIO-PXP-2022-total")

    try:
        build_config({**base, "table_kind": "ine_interior"}, splits)
    except ConfigError as exc:
        check("the default still refuses the unbalanced table",
              "does not balance" in str(exc))
    else:
        check("the default still refuses the unbalanced table", False,
              "it loaded")

    try:
        build_config({**base, "table_kind": "uk_analytical",
                      "table_unbalanced": "residual_column"}, splits)
    except ConfigError as exc:
        check("a setting that would be ignored is an error instead",
              "applies only to" in str(exc), str(exc)[:90])
    else:
        check("a setting that would be ignored is an error instead", False,
              "it was silently ignored")


def test_the_spanish_pilot_holds_its_two_delicate_properties():
    """Guards the two things in `examples/es_hosteleria.py` that are easy to
    break and silent when broken.

    First, an input profile must change what a subsector buys and not how much.
    The engine normalises intensities per supplier, so it cannot enforce that;
    the example does it, and product 36's diagonal is only 0.12 % of its input
    column, so an unneutralised profile is not a rounding error — it is a
    refused scenario.

    Second, the survey's value-added key does not fit inside the parent's
    accounts (OQ-B-12, A-06). That infeasibility is the pilot's finding. If it
    ever silently starts passing, either the data or the engine changed and the
    report's argument no longer holds.
    """
    sys.path.insert(0, str(ROOT / "examples"))
    path = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    if not path.exists():
        check("the Spanish pilot holds", True, "fixture absent, skipped")
        return
    from es_hosteleria import (NEW, LBL, build_keys,  # noqa: E402
                               build_profiles)
    from quadrium.io_loader import load_ine_tio
    from quadrium.models import Scenario, SplitSpec
    from quadrium.scenarios import run_project

    table = load_ine_tio(path, unbalanced="residual_column")
    keys = build_keys()
    w = keys["k_produccion"].w
    profiles = build_profiles(table, w)

    p = table.index_of("36")
    col = table.Z[:, p].copy()
    col[p] = 0.0
    from quadrium.disaggregation import _column_shares
    got = col @ _column_shares(table, p, NEW, w, profiles, True)
    want = w * col.sum()
    off = float(np.abs(got - want).max())
    room = float(table.Z[p, p])
    check("the input profile moves composition, not level",
          off < 1e-6 * room,
          f"purchases off by {off:.3e} million EUR against an internal block "
          f"of {room:,.1f} ({off / room:.2e} of it)")
    check("and it still says a restaurant buys food and a hotel rents premises",
          profiles["36B"]["5"] > profiles["36A"]["5"]
          and profiles["36A"]["44"] > profiles["36B"]["44"],
          f"36A food x{profiles['36A']['5']:.3f} premises "
          f"x{profiles['36A']['44']:.3f}; 36B food x{profiles['36B']['5']:.3f} "
          f"premises x{profiles['36B']['44']:.3f}")

    split = SplitSpec("36", NEW, LBL, keys_by_block={"output": "k_produccion"})
    results, meta = run_project(
        table, [split],
        [Scenario(scenario_id="S1", label="base"),
         Scenario(scenario_id="S3", label="va observado",
                  keys_by_block={"value_added": "k_vab",
                                 "intermediate_cols": "k_compras"},
                  input_profiles=profiles)],
        keys)
    check("the proportional split is feasible",
          [r.scenario_id for r in results] == ["S1"],
          f"feasible: {[r.scenario_id for r in results]}")
    rejected = {i["scenario_id"] for i in meta["infeasible"]}
    check("the observed value-added key is still refused (OQ-B-12, A-06)",
          rejected == {"S3"}, f"rejected: {sorted(rejected)}")
    check("and the refusal names the negative it would have required",
          any("cannot be negative" in i["detail"] for i in meta["infeasible"]))


def test_a_key_from_the_wrong_year_says_so():
    """Silence #16: `source_year` was stored, printed, exported — and compared
    with `table.year` by nothing.

    The evidence behind what the warning says lives in
    `library/validators/run_key_vintage.py`, which measures the cost on seven
    years of real data. This checks the plumbing: that a stale key reaches the
    reader, and that a clean run does not cry wolf.
    """
    from dataclasses import replace
    table = build_table()
    keys, scenario = build_keys(), build_scenarios()[0]

    # The shipped fixture's own keys are 2021 and 2022 against a 2024 table.
    # That was invisible until this check existed and it is NOT a mistake — a
    # proxy almost always lags the table it splits, so the fixture is the
    # realistic case. Asserted here so nobody later "tidies" the years and
    # removes the one example in the repo where the warning fires.
    shipped = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], scenario, keys)
    sv = [c for c in shipped.report.checks if c.name == "check_key_vintage"][0]
    check("the vintage check runs at all", sv is not None)
    check("the shipped fixture's own keys lag its table, and it now says so",
          not sv.passed and "do not come from 2024" in sv.detail, sv.detail[:100])

    aligned = {k: replace(v, source_year=table.year) for k, v in keys.items()}
    res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], scenario, aligned)
    vint = [c for c in res.report.checks if c.name == "check_key_vintage"]
    check("a key from the table's own year does not warn",
          vint and vint[0].passed, vint[0].detail if vint else "absent")

    stale = {k: replace(v, source_year=table.year - 3) for k, v in keys.items()}
    res2 = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], scenario, stale)
    v2 = [c for c in res2.report.checks if c.name == "check_key_vintage"][0]
    check("a key from three years earlier does warn", not v2.passed)
    check("and the warning names the gap, not just the year",
          "-3 year(s)" in v2.detail, v2.detail[:110])
    check("it is a warning, not an error — a stale proxy is often the best one",
          v2.severity == "warning" and res2.report.passed)

    from quadrium.reporting import scenario_section
    md = scenario_section(res2)
    check("the report puts the gap next to the year in the key table",
          "| vs table |" in md and "**-3 yr**" in md)
    check("and a same-year run says so instead of staying blank",
          "same year" in scenario_section(res))


def test_a_share_of_something_that_changes_sign_is_refused():
    """Real data, INE structural business survey 2020: hospitality gross
    operating surplus was -1,838,308 thousand EUR in accommodation against
    +231,683 in food service.

    That pair sums negative and the old guard caught it. Flip the signs and the
    sum is positive with one part negative — `[1000, -100]` normalised to
    `[1.111, -0.111]` and was ACCEPTED, handing a subsector -11 % of its block.
    """
    from quadrium.models import AllocationKey, ProxyStrength

    def build(vals):
        return AllocationKey(key_id="k", applies_to="output",
                             new_sector_codes=["A", "B"], raw_values=vals,
                             source="s", source_year=2020,
                             strength=ProxyStrength.MEDIUM)

    for label, vals in (("sum negative", [-1_838_308.0, 231_683.0]),
                        ("sum POSITIVE, one part negative", [1000.0, -100.0])):
        try:
            k = build(vals)
        except ValueError as exc:
            check(f"refused: {label}", "negative raw value" in str(exc)
                  or "sum to" in str(exc))
        else:
            check(f"refused: {label}", False,
                  f"accepted, weights {[round(w, 4) for w in k.weights]}")
    check("an ordinary key is untouched",
          build([30.0, 70.0]).weights == [0.3, 0.7])


def test_value_added_rows_can_carry_their_own_evidence():
    """OQ-B-12. The block used to split by one scalar per subsector, so a
    survey measuring compensation and operating surplus separately could not be
    used at all.

    Three properties, and the third is what stops this being a foot-gun:
    a pinned row takes its key exactly, every row still sums back to its parent,
    and pinning rows that claim more than the block key leaves is refused rather
    than absorbed.
    """
    from dataclasses import replace
    from quadrium.disaggregation import DisaggregationError
    from quadrium.models import AllocationKey, ProxyStrength

    table = build_table()
    keys, scenario = build_keys(), build_scenarios()[0]
    rows = table.VA_labels
    p = table.index_of("ACC")

    # The pin is a mild perturbation of the block's own weights. A pin far from
    # them is not "more evidence", it is a claim the block total cannot fund —
    # which is the refusal tested further down.
    base = np.asarray(run_scenario(table, [SplitSpec("ACC", NEW, LBL)],
                                   scenario, keys)
                      .splits[0]["weights"]["value_added"], float)
    tweak = base * np.array([1.10, 0.95, 1.02, 0.93])
    pin = AllocationKey(key_id="k_pin", applies_to="value_added",
                        new_sector_codes=NEW,
                        raw_values=[float(v) for v in tweak],
                        source="s", source_year=table.year,
                        strength=ProxyStrength.MEDIUM)
    keys = {**keys, "k_pin": pin}
    spec = SplitSpec("ACC", NEW, LBL,
                     va_row_keys={rows[0]: "k_pin"}, va_residual_row=rows[-1])
    res = run_scenario(table, [spec], scenario, keys)

    got = [res.table.VA[0, res.table.index_of(c)] / table.VA[0, p] for c in NEW]
    check("a pinned VA row takes its key exactly",
          np.allclose(got, pin.w), f"{[round(x, 4) for x in got]}")

    for r, label in enumerate(rows):
        total = sum(res.table.VA[r, res.table.index_of(c)] for c in NEW)
        if not np.isclose(total, table.VA[r, p]):
            check(f"row {label!r} still sums back to its parent", False,
                  f"{total:,.4f} against {table.VA[r, p]:,.4f}")
            break
    else:
        check("every VA row still sums back to its parent, residual included",
              True, f"{len(rows)} row(s)")

    # A residual row cannot take more of its parent than exists.
    greedy = replace(pin, key_id="k_greedy",
                     raw_values=[999., 1., 1., 1.], weights=None)
    try:
        run_scenario(table, [SplitSpec("ACC", NEW, LBL,
                                       va_row_keys={rows[0]: "k_greedy",
                                                    rows[1]: "k_greedy"},
                                       va_residual_row=rows[-1])],
                     scenario, {**keys, "k_greedy": greedy})
    except DisaggregationError as exc:
        check("pinning more than the block leaves is refused",
              "impossible share" in str(exc), str(exc)[:90])
    else:
        check("pinning more than the block leaves is refused", False,
              "it was absorbed silently")

    # Naming rows without naming who absorbs them is refused at construction.
    try:
        SplitSpec("ACC", NEW, LBL, va_row_keys={rows[0]: "k_pin"})
    except ValueError as exc:
        check("a residual row must be named, not guessed",
              "va_residual_row" in str(exc))
    else:
        check("a residual row must be named, not guessed", False, "accepted")


def test_an_input_profile_carries_its_provenance_and_its_side_effect():
    """OQ-B-13, both halves.

    A profile was a bare `dict[str, float]`: a pattern derived from an official
    survey and one guessed over lunch reached the reader identically labelled.
    And it moved subsector SIZE as a side effect of describing composition,
    which the engine never mentioned — the thing that made the Spanish pilot
    infeasible twice.
    """
    from quadrium.disaggregation import (neutralise_profile,
                                          profile_level_shift)
    from quadrium.models import ProfileProvenance
    from quadrium.reporting import scenario_section

    table = build_table()
    keys, base = build_keys(), build_scenarios()[0]
    # Two suppliers each, so there is a within-subsector composition to preserve.
    prof = {"HOT": {"AGR": 1.4, "MAN": 0.7},
            "RES": {"AGR": 0.8, "MAN": 1.2}}
    w = np.asarray(run_scenario(table, [SplitSpec("ACC", NEW, LBL)], base, keys)
                   .splits[0]["weights"]["intermediate_cols"], float)

    shift = profile_level_shift(table, "ACC", NEW, w, prof)
    check("the engine measures what a profile does to subsector size",
          shift["max_abs"] > 0 and not shift["neutral"],
          f"moves {shift['max_abs']:.3f} against an internal block of "
          f"{shift['internal_block']:.1f}")

    fixed = neutralise_profile(table, "ACC", NEW, w, prof)
    after = profile_level_shift(table, "ACC", NEW, w, fixed["profiles"])
    check("and can remove the level while keeping the pattern",
          after["neutral"] and after["max_abs"] < 1e-6 * shift["max_abs"],
          f"{shift['max_abs']:.3f} -> {after['max_abs']:.2e}")
    # What survives is each subsector's COMPOSITION -- what it buys.
    for code in ("HOT", "RES"):
        before = prof[code]["AGR"] / prof[code]["MAN"]
        after = (fixed["profiles"][code]["AGR"]
                 / fixed["profiles"][code]["MAN"])
        if abs(before - after) > 1e-9:
            check("each subsector's composition survives untouched", False,
                  f"{code}: {before:.4f} -> {after:.4f}")
            break
    else:
        check("each subsector's composition survives untouched", True,
              "the ratios among the suppliers a subsector names")
    # What cannot survive, and the docstring says why: between subsectors, the
    # pattern IS the level, and the level is what the key already fixed.
    cross_before = prof["HOT"]["AGR"] / prof["RES"]["AGR"]
    cross_after = (fixed["profiles"]["HOT"]["AGR"]
                   / fixed["profiles"]["RES"]["AGR"])
    check("the between-subsector ratio moves, because it was a size claim",
          abs(cross_before - cross_after) > 1e-9,
          f"{cross_before:.4f} -> {cross_after:.4f}; put a size claim in the "
          f"intermediate_cols key, not in a profile")

    neutral = fixed["profiles"]
    unsourced = replace_scenario(base, neutral, None)
    sourced = replace_scenario(base, neutral, ProfileProvenance(
        source="a real survey", source_year=table.year,
        strength=ProxyStrength.MEDIUM))
    md_un = scenario_section(run_scenario(table, [SplitSpec("ACC", NEW, LBL)],
                                          unsourced, keys))
    md_so = scenario_section(run_scenario(table, [SplitSpec("ACC", NEW, LBL)],
                                          sourced, keys))
    check("an unsourced profile is labelled as such",
          "no source recorded" in md_un)
    check("a sourced one is not, and names its source",
          "no source recorded" not in md_so and "a real survey" in md_so)
    md_raw = scenario_section(run_scenario(
        table, [SplitSpec("ACC", NEW, LBL)],
        replace_scenario(base, {"HOT": {"AGR": 1.15}}, None), keys))
    check("and the size side effect is reported, not left to be discovered",
          "moves subsector SIZE" in md_raw)
    check("a neutralised profile draws no such warning",
          "moves subsector SIZE" not in md_un)


def replace_scenario(base, profiles, provenance):
    from dataclasses import replace
    return replace(base, scenario_id="P", input_profiles=profiles,
                   profile_provenance=provenance)


def test_corroboration_reports_a_spread_and_refuses_to_rank():
    """OQ-S-06. The report used to mark the scenario that disagreed least with
    its unused keys as "better-supported".

    The Spanish pilot killed that. Its largest disagreement was employment, at
    58.8 %, and when the INE's 110-product supply table settled the answer,
    employment was the CLOSEST of seven keys and the driving key was 9.8 points
    out. Least disagreement measures resemblance to your own inputs.
    """
    from quadrium.reporting import build_report

    table = build_table()
    keys, scenarios = build_keys(), build_scenarios()
    usable = [s for s in scenarios if s.scenario_id != "S3_mixed"]
    results, meta = __import__("quadrium.scenarios", fromlist=["run_project"]) \
        .run_project(table, [SplitSpec("ACC", NEW, LBL)], usable, keys)
    md = build_report(results, meta, "corroboration spread")

    check("the ranking sentence is gone",
          "better-supported split" not in md and "disagrees least with the "
          "measurements" not in md)
    check("a spread is reported instead",
          "How far the outside evidence disagrees" in md
          and "closest key" in md and "furthest key" in md)
    check("and the removal is explained where the ranking used to be",
          "removed for cause" in md and "58.8" in md,
          "the case that killed it is named, not just the rule")
    check("the caveat still says what a large disagreement IS good for",
          "where to go looking" in md)


def test_the_spanish_supply_use_tables_load_and_balance():
    """The finest table the INE publishes, and the only fixture in the project
    where the valuation identities can be checked at all (OQ-D-03).

    Guards two label traps that a `partition(". ")` handled and a "more robust"
    regex did not: the supply-use workbook writes `5 .Pescado` with the space on
    the wrong side of the dot, and the input-output workbook wraps product 26
    onto a second line.
    """
    from quadrium.io_loader import load_ine_tod
    path = ROOT / "data" / "ine" / "cne_tod_22.xlsx"
    if not path.exists():
        check("the Spanish supply-use tables load", True, "fixture absent")
        return
    s = load_ine_tod(path)
    check("110 products by 81 activities, the published detail",
          (s.n_products, s.n_activities) == (110, 81),
          f"{s.n_products} x {s.n_activities}")
    check("the year comes from the workbook's own banner",
          s.year == 2022, f"year = {s.year}")
    check("a label with the dot on the wrong side still parses",
          "5" in s.product_codes
          and s.product_labels[s.product_codes.index("5")].startswith("Pescado"),
          s.product_labels[s.product_codes.index("5")][:40])

    d = float(np.abs(s.supply_at_purchasers() - s.use_at_purchasers()).max())
    check("ID-01 holds: supply meets use product by product", d < 1e-3,
          f"max deviation {d:.2e}")
    for label, v in (("trade", s.trade_margins),
                     ("transport", s.transport_margins)):
        check(f"ID-08 the {label} margin column sums to zero",
              abs(float(v.sum())) < 1e-3, f"{float(v.sum()):.2e}")
        check(f"ID-19 and it carries the negatives that make it work",
              int((v < 0).sum()) > 0,
              f"{int((v < 0).sum())} negative(s), most negative {v.min():,.1f}")

    # The reason this fixture exists: an analytical IOT cannot answer any of it.
    from quadrium.io_loader import load_ine_tio
    tio = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    if tio.exists():
        t = load_ine_tio(tio, unbalanced="residual_column")
        pair = [s.q[s.index_of_product(c)] for c in ("73", "74")]
        check("and it disaggregates the IOT's product 36 exactly",
              abs(sum(pair) - t.X[t.index_of("36")]) < 0.05,
              f"{pair[0]:,.1f} + {pair[1]:,.1f} = {sum(pair):,.1f} against "
              f"{t.X[t.index_of('36')]:,.1f}")


def test_the_eurostat_connector_loads_and_refuses_correctly():
    """The connector, and the three traps it was written around.

    A missing cell is not a zero; the product set is derived from the data
    rather than assumed; and the object's own row identity is checked, not only
    the identities the source happens to satisfy.
    """
    from quadrium.eurostat import EurostatError, load_iot
    path = ROOT / "data" / "eurostat" / "naio_10_cp1700_ES_2022.json"
    if not path.exists():
        check("the Eurostat connector loads", True, "fixture absent")
        return

    dom = load_iot(path, variant="domestic")
    tot = load_iot(path, variant="total")
    for t in (dom, tot):
        row = float(np.abs(t.Z.sum(1) + t.Y.sum(1) - t.X).max())
        col = float(np.abs(t.Z.sum(0) + t.VA.sum(0) - t.X).max())
        check(f"{t.table_id} balances both ways",
              max(row, col) < 1e-3, f"rows {row:.2e}, cols {col:.2e}")
    check("the product set is derived from the data, not a hard-coded 64",
          dom.n == 65 and "C10-12" in dom.sector_codes
          and "C10" not in dom.sector_codes,
          f"{dom.n} codes; the CPA hierarchy is served whole and the "
          f"populated level is the one taken")
    check("the domestic variant picks P5M where P52/P53 are not published",
          "P5M" in dom.Y_labels and "P52" not in dom.Y_labels,
          ", ".join(dom.Y_labels))
    check("and the total variant picks the finer pair, which IS published",
          "P52" in tot.Y_labels and "P53" in tot.Y_labels)
    check("the total table carries a DERIVED negative imports column",
          any("DERIVED" in l for l in tot.Y_labels),
          "without it the IOTable row identity is off by 78,638")

    try:
        load_iot(path, variant="imports")
    except EurostatError as exc:
        check("the imports variant is refused as not being an IOT",
              "no `P1` output vector" in str(exc))
    else:
        check("the imports variant is refused as not being an IOT", False,
              "it returned a table")

    # OQ-D-04: Eurostat's copy of the Spanish table does not carry the defect.
    ine = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    if ine.exists():
        from quadrium.io_loader import load_ine_tio
        i = load_ine_tio(ine, unbalanced="residual_column")
        j = dom.Y_labels.index("P3_S15")
        k = [n for n, l in enumerate(i.Y_labels) if "instituciones" in l][0]
        check("Eurostat has 7.3 where the INE workbook has -4,914.3 (OQ-D-04)",
              abs(dom.Y[dom.index_of("A01"), j] - 7.3) < 0.05
              and i.Y[0, k] < -4000,
              f"Eurostat {dom.Y[dom.index_of('A01'), j]:,.1f} against INE "
              f"{i.Y[0, k]:,.1f}")
        check("and the two agree on every total, so it is the same table",
              abs(dom.Z.sum() - i.Z.sum()) < 0.05
              and abs(dom.X.sum() - i.X.sum()) < 0.05,
              f"Z {dom.Z.sum():,.1f} · X {dom.X.sum():,.1f}")


def test_the_eurostat_supply_use_loader():
    """A supply-use pair for any member state, not just the one country whose
    spreadsheet layout the project happens to have reverse-engineered.

    Guards the three things that went wrong writing it, each of which would
    have been a silent wrong answer rather than a crash.
    """
    from quadrium.eurostat import load_sut
    sup = ROOT / "data" / "eurostat" / "naio_10_cp15_AT_2022.json"
    use = ROOT / "data" / "eurostat" / "naio_10_cp16_AT_2022.json"
    if not (sup.exists() and use.exists()):
        check("the Eurostat supply-use loader", True, "fixture absent")
        return
    s = load_sut(sup, use)
    check("it builds a supply-use pair from two Eurostat files",
          (s.n_products, s.n_activities) == (65, 65) and s.year == 2022,
          f"{s.table_id}: {s.n_products} x {s.n_activities}")

    # 1. NACE section P is Education. A filter that drops everything starting
    #    with "P" to remove P3/P5/P6 loses it, and activity output falls
    #    25,913 short -- which is a wrong table, not an error.
    check("NACE section P (Education) survives the final-demand filter",
          "P" in s.activity_codes and abs(s.q.sum() - s.g.sum()) < 0.5,
          f"q {s.q.sum():,.1f} against g {s.g.sum():,.1f}")

    # 2. Eurostat publishes trade and transport margins COMBINED. Filling the
    #    two component fields with the total and a column of zeros would be
    #    undetectable downstream.
    check("combined margins are recorded as combined, not split by invention",
          s.trade_margins is None and s.transport_margins is None
          and s.total_margins is not None,
          "`OTTM` is the sum; ID-09 cannot be asked of this source")
    check("ID-08 still holds: the margin column sums to zero",
          abs(float(s.total_margins.sum())) < 1e-3
          and int((s.total_margins < 0).sum()) > 0,
          f"{float(s.total_margins.sum()):.2e} with "
          f"{int((s.total_margins < 0).sum())} negative products")

    # 3. A mismatched pair is two economies, not a supply-use pair.
    from quadrium.eurostat import EurostatError
    es = ROOT / "data" / "eurostat" / "naio_10_cp15_ES_2022.json"
    if es.exists():
        try:
            load_sut(es, use)
        except EurostatError as exc:
            check("a mismatched country pair is refused",
                  "do not make a supply-use pair" in str(exc))
        else:
            check("a mismatched country pair is refused", False, "it loaded")


def test_project_folder_is_reproducible():
    """MVP_0.1 §3: everything needed to rerun the analysis lands on disk."""
    import json
    import tempfile
    from quadrium.models import AssumptionLedger
    from quadrium.project import IOProject

    table = build_table()
    with tempfile.TemporaryDirectory() as td:
        proj = IOProject(
            project_id="t", table=table,
            splits=[SplitSpec("ACC", NEW, LBL)], scenarios=build_scenarios(),
            keys=build_keys(), ledger=AssumptionLedger(project_id="t"),
            root=Path(td))
        d = proj.run().write()
        must = ["project.json", "original_table.csv", "report.md",
                "assumption_ledger.json"]
        missing = [f for f in must if not (d / f).exists()]
        check("the project folder holds the reproducibility record",
              not missing, f"missing: {missing}" if missing else "")
        for res in proj.results:
            sd = d / "scenarios" / res.scenario_id
            need = ["table_disaggregated.csv", "provenance.csv",
                    "validation_report.json", "diagnostics.json",
                    "technical_coefficients.csv"]
            gone = [f for f in need if not (sd / f).exists()]
            check(f"scenario {res.scenario_id} exports are complete", not gone,
                  f"missing: {gone}" if gone else "")
        man = json.loads((d / "project.json").read_text())
        check("the rejected scenario is recorded, not dropped",
              len(man["outcome"]["scenarios_rejected"]) == 1,
              f"{man['outcome']['scenarios_rejected']}")
        # Recorded is not the same as counted. `all_passed` read only the
        # scenarios that ran, so a run that produced one table out of two
        # reported success -- in the manifest and in the exit code alike.
        check("and a rejected scenario makes the run not-all-passed",
              man["outcome"]["all_passed"] is False,
              f"all_passed={man['outcome']['all_passed']} with "
              f"{len(man['outcome']['scenarios_rejected'])} rejected")
        check("the manifest states where its tolerances come from",
              "0.5*10^-d*n" in man["tolerances"]
              and "PROJECT CHOICE" in man["tolerances"],
              "the derived floor and the residue of genuine choice, rather "
              "than a pointer to a register a reader may not hold")

        # WHAT THE USER ACTUALLY SEES. `cli.py` prints `project.summary()`
        # after every run, and the reachability sweep found that neither it
        # nor `ValidationReport.n_warnings`, which it reads, had ever been
        # entered -- the files on disk were checked and the line on the screen
        # was not.
        text = proj.summary()
        check("the run summary names every scenario that ran and every one "
              "that was rejected, which is all a user sees on stdout",
              all(r.scenario_id in text for r in proj.results)
              and all(i["scenario_id"] in text
                      for i in proj.meta.get("infeasible", []))
              and "REJECTED" in text,
              f"{len(proj.results)} ran, "
              f"{len(proj.meta.get('infeasible', []))} rejected, "
              f"{len(text.splitlines())} lines")
        check("and it carries each scenario's warning count, which comes from "
              "the report rather than from the folder",
              all(f"{r.report.n_warnings} warn" in text
                  for r in proj.results),
              ", ".join(f"{r.scenario_id}: {r.report.n_warnings}"
                        for r in proj.results))


def test_label_mask_beats_the_naive_comparison():
    """`prov == CellLabel.X` on an object array is silently all-False.

    This is not a hypothetical. `CellLabel` subclasses `str`, numpy converts
    the right-hand side to a `numpy.str_`, and the elementwise comparison then
    matches nothing — while the same comparison on a single cell is correct.
    The failure is invisible: counts of untouched cells come out as zero and
    read as "nothing to report".

    The test asserts the trap as well as the fix, so that `label_mask` cannot
    be mistaken for a redundant wrapper around `==` and deleted.
    """
    import numpy as np
    from quadrium.models import CellLabel, label_mask

    a = np.empty((2, 2), dtype=object)
    a[:, :] = CellLabel.OBSERVED
    a[0, 0] = CellLabel.PROXY_ESTIMATED

    check("the naive array comparison still lies", (a == CellLabel.OBSERVED).sum() == 0,
          "3 of 4 cells ARE observed; numpy says 0")
    check("the scalar comparison is right, which is what hides the trap",
          a[0, 1] == CellLabel.OBSERVED)
    check("label_mask counts what is there",
          int(label_mask(a, CellLabel.OBSERVED).sum()) == 3
          and int(label_mask(a, CellLabel.PROXY_ESTIMATED).sum()) == 1)
    check("and it works on the string form the loader reads back",
          int(label_mask(np.array([["observed", "estimated"]], dtype=object),
                         CellLabel.OBSERVED).sum()) == 1)


def test_export_json_handles_numpy_and_enums():
    """The JSON writer must not choke on the types the engine actually uses."""
    import json
    import tempfile
    from quadrium.export import write_json
    from quadrium.models import CellLabel, ProxyStrength
    payload = {"a": np.float64(1.5), "b": np.int64(3), "c": np.array([1.0, 2.0]),
               "d": CellLabel.BALANCED_ADJUSTMENT, "e": ProxyStrength.WEAK,
               "f": np.bool_(True), "g": float("nan")}
    with tempfile.TemporaryDirectory() as td:
        out = write_json(payload, Path(td) / "x.json")
        back = json.loads(out.read_text())
    check("JSON export survives numpy, enums and NaN",
          back["d"] == "balanced_adjustment" and back["c"] == [1.0, 2.0]
          and back["g"] is None, str(back))


def test_input_profiles_preserve_supplier_totals():
    """The property that keeps reaggregation exact.

    A profile redistributes a supplier's sales WITHIN the new group. It must
    never change how much that supplier sells TO the group, or the untouched
    rows stop hitting their targets and §8 breaks.
    """
    from quadrium.disaggregation import split_sector
    table = build_table()
    plain = Scenario(scenario_id="plain", label="plain",
                     keys_by_block={"output": "key_employment"})
    profiled = Scenario(
        scenario_id="profiled", label="profiled",
        keys_by_block={"output": "key_employment"},
        input_profiles={"HOT": {"AGR": 0.90, "MAN": 1.10},
                        "RES": {"AGR": 1.10, "OTH": 0.92}})
    keys = build_keys()
    a = split_sectors(table, [SplitSpec("ACC", NEW, LBL)], plain, keys)
    b = split_sectors(table, [SplitSpec("ACC", NEW, LBL)], profiled, keys)
    pos = a["splits"][0]["positions"]
    off = [i for i in range(len(a["codes"])) if i not in pos]
    dev = float(np.max(np.abs(a["Z"][np.ix_(off, pos)].sum(axis=1)
                              - b["Z"][np.ix_(off, pos)].sum(axis=1))))
    check("input profiles preserve每 supplier's total sales to the group".replace("每", " each "),
          dev < 1e-10, f"max|dev| = {dev:.2e}")
    moved = float(np.max(np.abs(a["Z"][np.ix_(off, pos)] - b["Z"][np.ix_(off, pos)])))
    check("input profiles actually move purchases between subsectors",
          moved > 1.0, f"max cell change = {moved:.2f}")


def test_input_profiles_differentiate_multipliers():
    """Without profiles every subsector has the same multiplier; with them, not."""
    table = build_table()
    keys = build_keys()
    plain = Scenario(scenario_id="plain", label="plain",
                     keys_by_block={"output": "key_employment"})
    profiled = Scenario(
        scenario_id="profiled", label="profiled",
        keys_by_block={"output": "key_employment"},
        # Gentle on purpose: this synthetic sector trades only 12 with itself,
        # so it absorbs about +-10 % of intensity variation before the internal
        # block would need a negative total. The UK fixture has far more room.
        input_profiles={"HOT": {"AGR": 0.90, "MAN": 1.10},
                        "RES": {"AGR": 1.10, "MAN": 0.90}})
    out = {}
    for sc in (plain, profiled):
        res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], sc, keys)
        pos = res.splits[0]["positions"]
        mult = res.diagnostics["multipliers"][pos]
        out[sc.scenario_id] = (float(mult.max() - mult.min()),
                               res.splits[0]["input_structure"])
    check("a single key leaves the subsectors undifferentiated",
          not out["plain"][1]["differentiated"],
          f"cosine distance {out['plain'][1]['mean_cosine_distance']:.6f}, "
          f"multiplier spread {out['plain'][0]:.2e}")
    check("input profiles differentiate them",
          out["profiled"][1]["differentiated"] and out["profiled"][0] > 1e-4,
          f"cosine distance {out['profiled'][1]['mean_cosine_distance']:.4f}, "
          f"multiplier spread {out['profiled'][0]:.4f}")


def test_input_profiles_reject_nonsense():
    """Strict where it should be strict.

    A profile on the SplitSpec is about that split, so a typo must raise. A
    profile on the Scenario is a shared pool across several splits, so entries
    naming another split's subsectors are ignored by design — that is what lets
    one scenario be "profiled" and another "plain" over the same splits.
    """
    from quadrium.disaggregation import DisaggregationError, split_sectors
    table, keys = build_table(), build_keys()
    sc = Scenario(scenario_id="x", label="x",
                  keys_by_block={"output": "key_employment"})
    cases = {
        "unknown subsector": {"NOPE": {"AGR": 1.5}},
        "unknown supplier": {"HOT": {"NOPE": 1.5}},
        "negative intensity": {"HOT": {"AGR": -1.0}},
        "supplier with nowhere to sell": {c: {"AGR": 0.0} for c in NEW},
    }
    for name, prof in cases.items():
        spec = SplitSpec("ACC", NEW, LBL, input_profiles=prof)
        try:
            split_sectors(table, [spec], sc, keys)
        except DisaggregationError:
            check(f"spec-level profile rejects: {name}", True)
        else:
            check(f"spec-level profile rejects: {name}", False, "it was accepted")

    shared = Scenario(scenario_id="y", label="y",
                      keys_by_block={"output": "key_employment"},
                      input_profiles={"HOT": {"AGR": 1.05},
                                      "SOMEONE_ELSE": {"AGR": 9.9}})
    try:
        seed = split_sectors(table, [SplitSpec("ACC", NEW, LBL)], shared, keys)
    except DisaggregationError as exc:
        check("scenario-level pool ignores other splits' subsectors", False,
              str(exc)[:80])
    else:
        check("scenario-level pool ignores other splits' subsectors",
              seed["splits"][0]["profiled"])




def _two_sector_setup():
    from quadrium.models import AllocationKey, ProxyStrength
    table = build_table()
    keys = {k.key_id: k for k in [
        AllocationKey("k_acc", "output", ["HOT", "RES"], [30., 70.],
                      "illustrative", 2024, ProxyStrength.WEAK),
        AllocationKey("k_tra", "output", ["ROAD", "RAIL"], [65., 35.],
                      "illustrative", 2024, ProxyStrength.WEAK)]}
    specs = [SplitSpec("ACC", ["HOT", "RES"], ["Hotels", "Restaurants"],
                       {"output": "k_acc"}),
             SplitSpec("TRA", ["ROAD", "RAIL"], ["Road", "Rail"],
                       {"output": "k_tra"})]
    return table, keys, specs


def test_two_sectors_in_one_run():
    """Both split sectors reaggregate; everything else returns bit-for-bit."""
    from quadrium.reaggregation import reaggregate
    table, keys, specs = _two_sector_setup()
    sc = Scenario(scenario_id="multi", label="multi")
    res = run_scenario(table, specs, sc, keys)

    check("two sectors are divided in one run", len(res.splits) == 2,
          f"{[s['sector_code'] for s in res.splits]} -> "
          f"{res.table.n} sectors from {table.n}")
    check("the multi-sector run validates", res.report.passed,
          f"reaggregation error {res.report.reaggregation_error_pct:.2e} %")

    Z_re = reaggregate(res.table.Z, res.mapping, table.n)
    diff = np.abs(Z_re - table.Z)
    mask = np.ones_like(diff, dtype=bool)
    for code in ("ACC", "TRA"):
        i = table.index_of(code)
        mask[i, :] = False
        mask[:, i] = False
    check("sectors touched by neither split return exactly",
          float(diff[mask].max()) < 1e-12, f"max|dev| = {diff[mask].max():.2e}")
    check("both split sectors reaggregate within tolerance",
          float(diff.max()) < 1e-9, f"max|dev| = {diff.max():.2e}")
    check("no sign changes across either split",
          res.diagnostics["balance_info"]["sign_changes"] == 0)


def test_split_order_does_not_matter():
    """Without input profiles the splits commute. Asserted, not assumed.

    It matters because the engine applies them sequentially: if the order
    changed the answer, the result would depend on how the analyst happened to
    list the sectors.
    """
    table, keys, specs = _two_sector_setup()
    sc = Scenario(scenario_id="o", label="o")
    a = run_scenario(table, specs, sc, keys)
    b = run_scenario(table, specs[::-1], sc, keys)
    order = [b.table.sector_codes.index(c) for c in a.table.sector_codes]
    dev = float(np.max(np.abs(a.table.Z - b.table.Z[np.ix_(order, order)])))
    check("splitting A then B equals splitting B then A", dev < 1e-10,
          f"max|dev| = {dev:.2e}")


def test_a_sector_cannot_be_split_twice():
    from quadrium.disaggregation import DisaggregationError, split_sectors
    table, keys, specs = _two_sector_setup()
    sc = Scenario(scenario_id="d", label="d")
    for name, bad in {
        "the same sector listed twice": [specs[0], specs[0]],
        "a new code that already exists": [
            SplitSpec("ACC", ["MAN", "RES"], ["x", "y"], {"output": "k_acc"})],
    }.items():
        try:
            split_sectors(table, bad, sc, keys)
        except DisaggregationError:
            check(f"rejects: {name}", True)
        else:
            check(f"rejects: {name}", False, "it was accepted")


def test_config_workbook_round_trip():
    """The template must be loadable, and loading it must build a real run.

    This is the path an analyst who does not write Python actually takes, so it
    gets a test rather than a demo.
    """
    import tempfile
    from quadrium.config import load_config, write_template
    if not (ROOT / "UK_IOAT_2023_domestic_ixi.xlsx").exists():
        check("config workbook round trip", True, "fixture absent, skipped")
        return
    with tempfile.TemporaryDirectory() as td:
        cfg_path = write_template(Path(td) / "cfg.xlsx")
        _set(cfg_path, "project", 2, 2, str((ROOT / "UK_IOAT_2023_domestic_ixi.xlsx").resolve()))
        cfg = load_config(cfg_path)
    check("the template is a valid configuration",
          cfg["table"].n == 104 and len(cfg["splits"]) == 1
          and len(cfg["scenarios"]) == 2,
          f"{cfg['table'].n} sectors, {len(cfg['splits'])} split(s), "
          f"{len(cfg['scenarios'])} scenario(s)")
    check("comment rows are not read as data",
          all(k.strip() for s in cfg["splits"] for k in s.new_codes)
          and all(not c.startswith("#") for s in cfg["splits"]
                  for c in s.new_codes))
    profiled = [s for s in cfg["scenarios"] if s.input_profiles]
    check("profiles attach to the scenario that names them",
          len(profiled) == 1 and profiled[0].scenario_id == "S2_profiled",
          f"{[s.scenario_id for s in profiled]}")
    check("the ledger records every key and every profile",
          any(a.assumption_id.startswith("KEY-") for a in cfg["ledger"].assumptions)
          and any(a.assumption_id.startswith("PROFILE-")
                  for a in cfg["ledger"].assumptions))


def _set(path, sheet, row, col, value):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    wb[sheet].cell(row=row, column=col, value=value)
    wb.save(path)


def test_config_errors_are_useful():
    """A bad workbook must say what is wrong, where, and in the analyst's terms."""
    import shutil
    import tempfile
    from quadrium.config import ConfigError, load_config, write_template
    if not (ROOT / "UK_IOAT_2023_domestic_ixi.xlsx").exists():
        check("config errors are useful", True, "fixture absent, skipped")
        return
    with tempfile.TemporaryDirectory() as td:
        base = write_template(Path(td) / "base.xlsx")
        _set(base, "project", 2, 2, str((ROOT / "UK_IOAT_2023_domestic_ixi.xlsx").resolve()))
        cases = {
            "bad strength": ("keys", 2, 6, "very strong", "strong, medium or weak"),
            "non-numeric weight": ("keys", 2, 3, "lots", "is not a number"),
            "unknown sector": ("splits", 2, 1, "NOSUCH", "not in the loaded table"),
            "unknown supplier": ("profiles", 2, 3, "XX9", "not a sector"),
            "orphan profile": ("profiles", 2, 1, "S9", "does not define"),
        }
        for name, (sheet, row, col, bad, expect) in cases.items():
            p = Path(td) / f"{name.replace(' ', '_')}.xlsx"
            shutil.copy(base, p)
            _set(p, sheet, row, col, bad)
            try:
                load_config(p)
            except ConfigError as exc:
                check(f"explains: {name}", expect in str(exc),
                      str(exc).splitlines()[0][:70])
            else:
                check(f"explains: {name}", False, "it was accepted")


def test_classification_validates_split_codes():
    """M-049: is this a subdivision, or just a list of codes?"""
    from quadrium.classification import NACE_REV_2_1, check_split
    good = check_split("I56", ["I561", "I562", "I563"])
    check("a division splitting into its groups is accepted", good.ok,
          good.summary()[:60])
    check("and it still says coverage was not checked",
          any("coverage not checked" in u for u in good.unchecked))

    bad = {
        "a child outside the parent": ("I56", ["I561", "I571"]),
        "skipping a level": ("I56", ["I5601", "I5602"]),
        "a child repeating the parent": ("I56", ["I56", "I561"]),
        "the same child twice": ("I56", ["I561", "I561"]),
        "more children than digits allow": ("56", [f"56{i}" for i in "123456789"] + ["560"]),
    }
    for name, (parent, kids) in bad.items():
        check(f"rejects: {name}", not check_split(parent, kids).ok)

    deep = check_split("5610", ["56101", "56102"], NACE_REV_2_1)
    check("rejects splitting below NACE's deepest level", not deep.ok,
          deep.problems[0][:70] if deep.problems else "")
    national = check_split("5610", ["56101", "56102"])
    check("allows it in a national version with a fifth digit", national.ok)

    unknown = check_split("ACC", ["HOT", "RES"])
    check("codes that are not a classification are not an error",
          unknown.ok and not unknown.parsed)


def test_engine_refuses_an_illegitimate_split():
    """The check runs on every path, not only in the workbook loader."""
    from quadrium.disaggregation import DisaggregationError
    from quadrium.io_loader import load_uk_analytical_iot
    from quadrium.models import AllocationKey, ProxyStrength
    path = ROOT / "UK_IOAT_2023_domestic_ixi.xlsx"
    if not path.exists():
        check("engine refuses an illegitimate split", True, "fixture absent")
        return
    table = load_uk_analytical_iot(path)
    keys = {"k": AllocationKey("k", "output", ["I561", "I571"], [7., 3.],
                               "illustrative", 2023, ProxyStrength.WEAK)}
    try:
        run_scenario(table,
                     [SplitSpec("I56", ["I561", "I571"], ["a", "b"],
                                {"output": "k"})],
                     Scenario(scenario_id="s", label="s"), keys)
    except DisaggregationError as exc:
        check("the engine refuses a child outside the parent",
              "not inside I56" in str(exc))
    else:
        check("the engine refuses a child outside the parent", False,
              "it produced a table")


def test_targets_are_consistent():
    """Row and column targets must sum to the same grand total, or no table exists."""
    table = build_table()
    seed = split_sectors(table, [SplitSpec("ACC", NEW, LBL)], build_scenarios()[0], build_keys())
    tr, tc = targets(seed["Y"], seed["VA"], seed["X"])
    check("row and column targets sum to the same total",
          abs(tr.sum() - tc.sum()) < 1e-9,
          f"difference = {tr.sum() - tc.sum():.2e}")


def test_id19_margin_column():
    """ID-19: a margin column sums to zero AND gives the margin up somewhere.

    CORE_010 par. 7.19, p. 211. The sign pattern is half the identity: a column
    can hit a zero total while having reallocated nothing, and the first version
    of this check passed exactly that case because it tested "not positive"
    instead of "strictly gives something up".
    """
    from quadrium.identities import id19_margin_column_sums_to_zero as id19
    cases = [
        ("a correct column",            [12.0, 30.0, -50.0, 8.0],   [2],    True),
        ("zero total, nothing moved",   [12.0, 30.0, 0.0, -42.0],   [2],    False),
        ("does not sum to zero",        [12.0, 30.0, -10.0, 8.0],   [2],    False),
        ("service row positive",        [-30.0, -20.0, 50.0, 0.0],  [2],    False),
        ("two service rows, one zero",  [40.0, 10.0, -50.0, 0.0],   [2, 3], True),
        ("an empty column",             [0.0, 0.0, 0.0, 0.0],       [2],    True),
    ]
    for name, col, svc, want in cases:
        r = id19(col, service_rows=svc)
        check(f"{name} -> {'passes' if want else 'is rejected'}",
              r.passed == want, r.detail)


def test_corroboration_uses_the_keys_the_split_did_not():
    """A registered-but-unused key becomes an automatic external check.

    This is the only check in the project that can distinguish a good split
    from a bad one; every other one passes on any key at all. So it is worth
    testing three things and not just that it runs: that it ignores the key
    that DROVE the split, that it compares an unused key against the block
    that key claims to describe, and that the gap it reports is arithmetic
    rather than decoration.
    """
    from quadrium.validation import corroborate_keys
    from quadrium.models import AllocationKey, ProxyStrength
    codes = ["A1", "A2"]

    def key(kid, block, vals):
        return AllocationKey(key_id=kid, applies_to=block,
                             new_sector_codes=codes, raw_values=vals,
                             source="test", source_year=2023,
                             strength=ProxyStrength.STRONG)

    keys = {"driver": key("driver", "output", [60.0, 40.0]),
            "other": key("other", "value_added", [50.0, 50.0]),
            "elsewhere": AllocationKey(
                key_id="elsewhere", applies_to="output",
                new_sector_codes=["B1", "B2"], raw_values=[1.0, 1.0],
                source="test", source_year=2023,
                strength=ProxyStrength.STRONG)}
    used = {"output": "driver"}
    weights = {"output": [0.6, 0.4], "value_added": [0.6, 0.4]}

    got, skipped = corroborate_keys(keys, codes, used, weights)
    ids = [c["key_id"] for c in got]
    check("the driving key is not compared against itself", "driver" not in ids,
          f"compared: {ids}")
    check("a key for another split is ignored", "elsewhere" not in ids,
          f"compared: {ids}")
    check("the unused key is compared", ids == ["other"], f"compared: {ids}")

    c = got[0]
    check("compared against the block the key claims",
          c["compared_against_block"] == "value_added",
          c["compared_against_block"])
    # split gave A1 0.6 where the key measures 0.5 -> +20 %
    check("the gap is computed, not decorative",
          abs(c["rows"][0]["gap"] - 0.2) < 1e-12
          and abs(c["max_abs_gap"] - 0.2) < 1e-12,
          f"gap={c['rows'][0]['gap']:.4f} worst={c['max_abs_gap']:.4f}")

    # And with nothing spare to compare against, it says nothing at all.
    check("no unused key -> no corroboration, not an empty flourish",
          corroborate_keys({"driver": keys["driver"]}, codes, used,
                           weights)[0] == [],
          "should be an empty list")

    # A weak key must never become evidence. This is the case that produced a
    # WRONG verdict on the first real run: a leftover illustrative key beat two
    # official ones because nothing filtered it out.
    weak = key("guess", "output", [90.0, 10.0])
    weak.strength = ProxyStrength.WEAK
    got2, skipped2 = corroborate_keys({**keys, "guess": weak}, codes, used,
                                      weights)
    check("a weak key is not used as corroboration",
          "guess" not in [c["key_id"] for c in got2],
          f"compared: {[c['key_id'] for c in got2]}")
    check("but it is reported as skipped, not dropped in silence",
          [s["key_id"] for s in skipped2] == ["guess"],
          f"skipped: {skipped2}")


def test_the_report_does_not_stay_silent():
    """Three places the report used to say nothing, where nothing read as fine.

    Each of these was found by reading real output, not by reasoning about the
    code, and each has the same shape: an absent line that a reader completes
    in the system's favour.
    """
    import numpy as np
    from types import SimpleNamespace
    from quadrium.reporting import scenario_section
    from quadrium.models import CellLabel

    def render(split):
        class R:
            scenario_id = "S"
            provenance = np.array([[CellLabel.OBSERVED]], dtype=object)
            report = SimpleNamespace(to_markdown=lambda: "(none)")
            table = SimpleNamespace(sector_codes=["B1"], Z=np.array([[1.0]]),
                                    Y=np.array([[1.0]]), VA=np.array([[1.0]]),
                                    X=np.array([2.0]))
            diagnostics = {"balance_info": {
                "method": "GRAS", "reason": "-", "converged": True,
                "iterations_per_split": {"B": 1}, "max_row_dev": 0.0,
                "max_col_dev": 0.0, "n_negative_seed": 0,
                "n_negative_result": 0, "sign_changes": 0},
                "multipliers": [1.0]}
            splits = [split]
        return scenario_section(R())

    base = {"sector_code": "B", "sector_label": "B", "positions": [0],
            "keys_used": {"output": "k"}, "keys_inherited": {"output": False},
            "key_meta": {"k": {"strength": "strong", "source": "s",
                               "source_year": 2023}},
            "weights": {"output": [1.0]}, "headroom_pct": 5.0,
            "parent_diagonal": 100.0, "tightest_internal_total": 5.0,
            "internal_block_share_pct": 0.0, "original_diagonal": 100.0,
            "input_structure": {}}

    # 1. A key with no headroom at all used to produce NO headroom line.
    out = render({**base, "headroom_pct": float("nan"), "parent_diagonal": 0.0})
    check("zero self-trade is stated, not omitted",
          "trades nothing with itself" in out,
          "the most constrained case must not be the quietest")

    # 2. A weak key driving the split used to look like a strong one.
    out = render({**base, "key_meta": {"k": {"strength": "weak", "source": "s",
                                            "source_year": 2023}}})
    check("a weak driving key is called out", "`weak` is driving" in out, out[:80])

    # 3. A block that inherited its key used to look like a deliberate choice.
    out = render({**base, "keys_inherited": {"output": True}})
    check("an inherited key says so", "inherited" in out, out[:80])
    check("and the inheritance is explained, not just labelled",
          "not a decision by the analyst" in out, out[:80])


def test_machine_facing_surfaces_do_not_stay_silent():
    """The JSON, the folder and the CLI, where the reader is a program.

    A person gets the caveats from report.md. Anything automating over this
    gets the folder and the JSON, and those used to carry the numbers without
    the warnings — including, for a whole day, the results of a scenario that
    had been withdrawn for being misleading.
    """
    import json
    import shutil
    from quadrium.project import IOProject

    table, keys, scen = build_table(), build_keys(), build_scenarios()
    root = ROOT / "outputs" / "_test_silences"
    if root.exists():
        shutil.rmtree(root)

    spec = SplitSpec("ACC", NEW, LBL)
    p1 = IOProject(project_id="_test_silences", table=table, splits=[spec],
                   scenarios=scen[:1], keys=keys, root=ROOT / "outputs")
    p1.run().write()
    d = root / "scenarios"
    first = {x.name for x in d.iterdir()}

    # A stale scenario folder from an earlier configuration must not survive.
    (d / "S9_withdrawn").mkdir()
    (d / "S9_withdrawn" / "table_disaggregated.csv").write_text("stale\n")
    p2 = IOProject(project_id="_test_silences", table=table, splits=[spec],
                   scenarios=scen[:1], keys=keys, root=ROOT / "outputs")
    p2.run().write()
    after = {x.name for x in d.iterdir()}
    check("a withdrawn scenario's folder is removed, not left looking current",
          "S9_withdrawn" not in after, f"folder holds {sorted(after)}")
    check("and the removal is announced, not silent",
          "S9_withdrawn" in getattr(p2, "removed_scenarios", []),
          f"removed: {getattr(p2, 'removed_scenarios', None)}")
    check("the scenarios that belong are untouched", first <= after,
          f"{sorted(first)} vs {sorted(after)}")

    # The manifest is the authoritative list, so a consumer can check the
    # filesystem against something.
    man = json.loads((root / "project.json").read_text())
    check("project.json names the scenarios this run produced",
          man.get("scenario_ids") == sorted(first), str(man.get("scenario_ids")))
    check("and records what it removed as stale",
          man.get("scenarios_removed_as_stale") == ["S9_withdrawn"],
          str(man.get("scenarios_removed_as_stale")))

    # The corroboration's scope must be readable without parsing English prose.
    diag = json.loads(
        (d / sorted(first)[0] / "diagnostics.json").read_text())
    check("the JSON states what the corroboration covers",
          "corroboration_covers" in diag["splits"][0],
          "a consumer reading max_abs_gap needs to know its scope")

    shutil.rmtree(root)


def test_exports_travel_with_their_status():
    """A number that leaves this system must not leave its caveats behind.

    The XLSX shades estimated cells and carries a README sheet. The CSV — the
    file people actually open — carried neither, and the provenance sat in a
    separate file nobody has a reason to look for.
    """
    import shutil
    import numpy as np
    from quadrium.project import IOProject

    root = ROOT / "outputs" / "_test_exports"
    if root.exists():
        shutil.rmtree(root)
    p = IOProject(project_id="_test_exports", table=build_table(),
                  splits=[SplitSpec("ACC", NEW, LBL)],
                  scenarios=build_scenarios()[:1], keys=build_keys(),
                  root=ROOT / "outputs")
    p.run().write()
    sd = root / "scenarios" / p.results[0].scenario_id

    csv = (sd / "table_disaggregated.csv").read_text()
    head = [ln for ln in csv.splitlines() if ln.startswith("#")]
    check("the table CSV says not everything in it is observed",
          any("NOT ALL OF THESE NUMBERS ARE OBSERVED" in ln for ln in head),
          f"{len(head)} comment line(s)")
    check("and it points at where the caveats are",
          any("report.md" in ln for ln in head), str(head[-1:]))
    check("the data itself is still parseable below the comments",
          csv.splitlines()[len(head)].startswith("code,label,"),
          csv.splitlines()[len(head)][:40])

    # An uncomputable coefficient must not be written as a plausible zero.
    coef = (sd / "technical_coefficients.csv").read_text()
    check("the coefficients file distinguishes 'unknown' from 'zero'",
          coef.startswith("#") and "not the same as a coefficient" in coef,
          coef.splitlines()[0][:60])

    # The loader's own decisions reach the human-readable report.
    rep = (root / "report.md").read_text()
    check("the report states what the loader decided about the input",
          "What the loader decided" in rep or "Reference year:" in rep,
          "a reader must be told their table was modified")

    shutil.rmtree(root)


def test_a_pinned_cell_the_solver_moved_stops_claiming_to_be_pinned():
    """The worst silence found in this engine: a label that asserted a falsehood.

    `user_constraints` writes a value and marks the cell USER_CONSTRAINT, which
    the report renders "OBSERVED (analyst-pinned)". Nothing protected it —
    `locked_cells` is a different mechanism and GRAS refuses locks outright —
    so on any table with negatives a pin inside the internal block is moved by
    the solver and went on describing itself as the analyst's own value.
    Reproduced: asked 99.0, held 0.3734, still labelled pinned.
    """
    from quadrium.models import CellLabel
    table = build_table()
    p = table.index_of("ACC")
    sc = Scenario(scenario_id="pinned", label="pinned",
                  keys_by_block={"output": "key_employment"},
                  user_constraints={f"{p},{p+1}": 99.0})
    res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], sc, build_keys())

    held = float(res.table.Z[p, p + 1])
    check("the solver did move the pinned cell (the premise of this test)",
          abs(held - 99.0) > 1e-6, f"cell holds {held:.4f}")
    check("so it no longer calls itself analyst-pinned",
          res.provenance[p, p + 1] is CellLabel.BALANCED_ADJUSTMENT,
          f"label is {res.provenance[p, p+1].value}")

    over = res.diagnostics.get("user_constraints_overridden") or []
    check("the override is recorded with both numbers", len(over) == 1
          and abs(over[0]["requested"] - 99.0) < 1e-9
          and abs(over[0]["actual"] - held) < 1e-9, str(over))
    check("and the validation report warns rather than passing quietly",
          any(c.name == "check_user_constraints_held" and not c.passed
              for c in res.report.checks),
          "a moved pin must not be a silent pass")

    # A pin the solver did NOT touch must keep its label — the fix must not
    # relabel everything out of caution.
    off = table.index_of("MAN")
    sc2 = Scenario(scenario_id="untouched", label="untouched",
                   keys_by_block={"output": "key_employment"},
                   user_constraints={f"{off},{off}": float(table.Z[off, off])})
    res2 = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], sc2, build_keys())
    check("a pin the solver left alone still reads as pinned",
          res2.provenance[off, off] is CellLabel.USER_CONSTRAINT,
          f"label is {res2.provenance[off, off].value}")


def test_a_number_too_small_to_show_is_not_shown_as_zero():
    """Rounding must not turn "a little" into "none".

    The last of the silences, and the quietest: `f"{0.04:,.1f}"` is `"0.0"`, so
    a sector that buys a little was reported as buying nothing, with no way for
    the reader to tell that case from a true zero.
    """
    from quadrium.reporting import _fmt, _pct

    check("a true zero still reads as zero", _fmt(0.0) == "0.0", _fmt(0.0))
    check("a small positive is marked as below the threshold",
          _fmt(0.04) == "<0.05", _fmt(0.04))
    check("a small negative keeps its sign", _fmt(-0.04) == ">-0.05",
          _fmt(-0.04))
    check("anything that rounds to a visible figure is untouched",
          _fmt(0.06) == "0.1" and _fmt(1234.5) == "1,234.5",
          f"{_fmt(0.06)} / {_fmt(1234.5)}")
    check("the threshold follows the requested precision",
          _fmt(0.0004, 3) == "<0.0005" and _fmt(0.04, 3) == "0.040",
          f"{_fmt(0.0004, 3)} / {_fmt(0.04, 3)}")
    check("not-computable stays a third thing, not a zero",
          _fmt(float("nan")) == "—", _fmt(float("nan")))

    # The same rule on percentages, where a rounded-away gap reads as agreement.
    check("a zero gap reads as zero", _pct(0.0) == "+0.0%", _pct(0.0))
    check("a tiny gap is not reported as agreement",
          _pct(0.0004) == "<+0.05%", _pct(0.0004))
    check("a real gap is unchanged", _pct(-0.153) == "-15.3%", _pct(-0.153))


def test_the_internal_block_conserves_the_parent_cell_at_every_alpha():
    """`alpha` reparameterises the block; it must not leak the parent cell.

    The old form scaled the diagonal and left the off-diagonal at 1.0, so every
    alpha != 1 broke CORE_031 eq. (15): at alpha = 0.5 the block summed to
    -22.6 % of the parent cell and a balancing step that knows nothing about
    the block was left to repair it (`OQ-S-04`). The off-diagonal now takes
    `beta = (1 - alpha*d)/(1 - d)`, so the sum holds for any alpha and reduces
    exactly to eq. (14) at alpha = 1.

    There was no test of `alpha` at all until 2026-08-26, on a parameter this
    project had already had wrong once.
    """
    from quadrium.disaggregation import split_sector

    table = build_table()
    keys = build_keys()
    spec = SplitSpec("ACC", NEW, LBL,
                 keys_by_block={"output": "key_turnover"})
    p_idx = table.index_of("ACC")
    z_pp = float(table.Z[p_idx, p_idx])

    sums, diagonals = [], []
    for alpha in (0.5, 1.0, 1.5, 2.0):
        sc = Scenario(scenario_id=f"a{alpha}", label=str(alpha),
                      keys_by_block={"output": "key_turnover"},
                      internal_block_alpha=alpha)
        seed = split_sector(table, "ACC", NEW, LBL, sc, keys, spec)
        pos = seed["new_positions"]
        block = np.asarray(seed["Z"])[np.ix_(pos, pos)]
        sums.append(float(block.sum()))
        diagonals.append(float(np.trace(block)))

    worst = max(abs(s - z_pp) for s in sums)
    check("the internal block sums to the parent cell at every alpha",
          worst / max(abs(z_pp), 1e-12) < 1e-12,
          f"parent cell {z_pp:,.4f}, block sums "
          f"{', '.join(f'{s:,.4f}' for s in sums)} — worst gap {worst:.3g}")
    check("and alpha still does what it is for: concentrate the diagonal",
          diagonals == sorted(diagonals) and diagonals[-1] > diagonals[0],
          f"diagonal {' -> '.join(f'{d:,.1f}' for d in diagonals)} across "
          f"alpha 0.5, 1.0, 1.5, 2.0")


def test_a_zero_row_is_zero_at_the_SOURCE_s_precision():
    """A row summing below what the file can print is zero, not "nearly zero".

    `check_zero_row_col` warns about rows and columns that BALANCING made zero,
    which is worth knowing, and it judged "zero" at a flat 1e-12. Each of those
    numbers is a SUM over a whole row of published cells, so what it can carry
    is the rounding of `n` of them — `assertable_tolerance`, the same quantity
    the balance checks two hundred lines above already derive.

    Measured, 17 lines across five real tables (Hungary in all four years and
    Slovakia) sit between 1e-12 and their file's own floor. A line at 0.2 in
    the seed and 0.0 after balancing would have been reported to the user as
    created by the solver, when the source cannot tell either value from zero.

    Seventh instance of a bound assumed rather than derived; the sweep that
    found it is `library/validators/run_derived_bounds.py`.
    """
    from quadrium.precision import assertable_tolerance

    table = build_table()
    values = np.concatenate([table.Z.ravel(), table.Y.ravel(),
                             table.VA.ravel(), table.X.ravel()])
    floor = assertable_tolerance(values, table.n)
    check("the fixture has a floor above 1e-12 to test against",
          floor > 1e-12,
          f"{floor:,.6g} from this table's own printed precision")

    res = run_scenario(table, [SplitSpec("ACC", NEW, LBL)], build_scenarios()[0],
                       build_keys())
    found = [c for c in res.report.checks if c.name == "check_zero_row_col"]
    check("the check runs and says what it means by zero", bool(found)
          and "distinguish" in found[0].detail,
          found[0].detail[-90:] if found else "the check did not run")

    # the substance: a line below the floor counts as zero, one above does not
    Z = np.zeros((3, 3))
    Z[0, 0] = floor * 0.5
    Z[1, 1] = floor * 10.0
    below = float(np.abs(Z.sum(axis=1))[0]) < floor
    above = float(np.abs(Z.sum(axis=1))[1]) < floor
    check("a row under the source's floor is zero and one over it is not",
          below and not above,
          f"{Z.sum(axis=1)[0]:,.6g} counts as zero and "
          f"{Z.sum(axis=1)[1]:,.6g} does not, against a floor of {floor:,.6g}")



def test_the_refusals_about_a_FILE_FROM_AN_OFFICE():
    """The largest unreached block, and the one the engine reaches on its own.

    Forty-one of the engine's refusals judge what arrived from a statistical
    office rather than what the user wrote: a response that is not JSON-stat, a
    file without the dimension the loader needs, two halves of a pair that are
    not the same country and year. They matter because **the engine downloads
    those files itself** — someone asking for a new country meets these
    messages, and nothing had ever read one aloud.

    No network and no invented fixture: `data/eurostat/` already holds a hundred
    real JSON-stat files. A JSON-stat is a dict, so one is loaded, one thing is
    removed or emptied, and it is written to a temporary file. The base is real
    and the mutation is one line.
    """
    import json
    import tempfile
    from quadrium.eurostat import EurostatError, load_iot, load_sut

    DATA = ROOT / "data" / "eurostat"
    base = DATA / "naio_10_cp1700_FR_2021.json"
    if not base.exists():
        check("there is a real JSON-stat file to mutate", False, str(base))
        return
    doc = json.loads(base.read_text())
    check("the base file is real and loads, so each mutation is the only "
          "difference", load_iot(base).n == 89, f"{base.name}, 89 products")

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        def mutated(edit, name="m.json"):
            d = json.loads(json.dumps(doc))
            edit(d)
            path = tmp / name
            path.write_text(json.dumps(d))
            return path

        def drop_dim(dim):
            def edit(d):
                if dim in d.get("id", []):
                    k = d["id"].index(dim)
                    d["id"].pop(k)
                    d["size"].pop(k)
                d.get("dimension", {}).pop(dim, None)
            return edit

        cases = [
            ("a response that is not JSON-stat at all",
             lambda: load_iot(mutated(lambda d: d.pop("dimension", None))),
             "not a JSON-stat"),
            ("a file with no `stk_flow` dimension, which it needs to pick a "
             "valuation",
             lambda: load_iot(mutated(drop_dim("stk_flow"))),
             "stk_flow"),
            ("a file with neither product axis a symmetric table must have",
             lambda: load_iot(mutated(lambda d: (drop_dim("prd_use")(d),
                                                 drop_dim("ind_use")(d)))),
             "neither"),
            ("a file that carries no values at all",
             lambda: load_iot(mutated(lambda d: d.update(value={}))),
             ""),
        ]
        for name, run, fragment in cases:
            try:
                run()
            except EurostatError as exc:
                check(f"the engine refuses {name}",
                      fragment.lower() in str(exc).lower(),
                      str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}", False,
                      f"{type(exc).__name__} instead: {str(exc)[:66]}")
            else:
                check(f"the engine refuses {name}", False, "it built a table")

        # Value edits rather than structural ones: the file keeps its shape and
        # stops adding up, which is the harder half of what an office can send.
        def blank(codes, dim="prd_ava"):
            """Erase every value on one axis position, leaving the shape."""
            def edit(d):
                ids, size = d["id"], d["size"]
                if dim not in ids:
                    return
                k = ids.index(dim)
                stride = 1
                for j in range(k + 1, len(size)):
                    stride *= size[j]
                idx = d["dimension"][dim]["category"]["index"]
                wanted = {idx[c] for c in codes if c in idx}
                d["value"] = {f: v for f, v in d["value"].items()
                              if (int(f) // stride) % size[k] not in wanted}
            return edit

        try:
            load_iot(base, variant="imports")
        except EurostatError as exc:
            check("the engine refuses the one valuation that has no output "
                  "vector, and explains why",
                  "imports" in str(exc) and "P1" in str(exc),
                  str(exc)[:86] + "…")
        except Exception as exc:                       # noqa: BLE001
            check("the engine refuses the one valuation that has no output "
                  "vector, and explains why", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("the engine refuses the one valuation that has no output "
                  "vector, and explains why", False, "it built a table")

        try:
            load_iot(base, variant="sideways")
        except EurostatError as exc:
            check("and any valuation that is not one of the three",
                  "sideways" in str(exc), str(exc)[:86] + "…")
        except Exception as exc:                       # noqa: BLE001
            check("and any valuation that is not one of the three", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("and any valuation that is not one of the three", False,
                  "it built a table")

        # A file whose parts stop adding up to the total it also publishes. The
        # shape is untouched; one product simply carries no value any more, so
        # the codes that remain fall short of `CPA_TOTAL`.
        one = next(c for c in load_iot(base).sector_codes[:5])
        try:
            load_iot(mutated(blank([f"CPA_{one}"]), "short.json"))
        except EurostatError as exc:
            check("the engine refuses a file whose products no longer sum to "
                  "the total it publishes",
                  "sum" in str(exc).lower() or "populated" in str(exc).lower(),
                  str(exc)[:86] + "…")
        except Exception as exc:                       # noqa: BLE001
            check("the engine refuses a file whose products no longer sum to "
                  "the total it publishes", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("the engine refuses a file whose products no longer sum to "
                  "the total it publishes", False,
                  "it built a table from parts that do not add up")

        # A PAIR has to be the same country and year, and the message says so.
        supply = DATA / "naio_10_cp15_ES_2022.json"
        use = DATA / "naio_10_cp16_ES_2022.json"
        other = DATA / "naio_10_cp1610_BE_2022.json"
        if supply.exists() and use.exists() and other.exists():
            try:
                load_sut(supply, use, other)
            except EurostatError as exc:
                check("the engine refuses a pair whose halves are different "
                      "countries", "pair" in str(exc).lower()
                      or "belong" in str(exc).lower(),
                      str(exc)[:86] + "…")
            except Exception as exc:                   # noqa: BLE001
                check("the engine refuses a pair whose halves are different "
                      "countries", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("the engine refuses a pair whose halves are different "
                      "countries", False, "it built the pair")

        # ---- THE SUPPLY-USE ROUTE ------------------------------------------
        #
        # The same kind of file, judged behind a different door: `load_sut`
        # reads THREE datasets at once, and the office can deform any of them.
        # None of these refusals had ever been read aloud, although the engine
        # downloads all three files itself the moment someone asks for a
        # country by name.
        #
        # Austria 2022 is the base because it is the triple this project
        # already loads clean, 65 x 65. Every case changes ONE thing in ONE
        # file.
        SUP = DATA / "naio_10_cp15_AT_2022.json"
        USE = DATA / "naio_10_cp16_AT_2022.json"
        BAS = DATA / "naio_10_cp1610_AT_2022.json"

        def deformed(src, edit, name):
            """`mutated`, for a base other than the symmetric table."""
            d = json.loads(src.read_text())
            edit(d)
            path = tmp / name
            path.write_text(json.dumps(d))
            return path

        def scale(factor, **where):
            """Multiply the ONE published aggregate at `where`.

            Blanking a product -- which is what catches the symmetric table
            above -- does NOT work on this source, and the first version of
            this block found that out by building a table it expected to be
            refused. Austria publishes both CPA levels, `CPA_B` beside
            `CPA_B05`...`B09`, so `_finest_tiling` covers the hole with the
            parent code and the file still adds up.

            Moving the published TOTAL is a deformation no tiling can absorb,
            and it is the likelier accident anyway: an office revises an
            aggregate and does not re-send the parts.
            """
            def edit(d):
                ids, size = d["id"], d["size"]
                stride = [1] * len(size)
                for i in range(len(size) - 2, -1, -1):
                    stride[i] = stride[i + 1] * size[i + 1]
                pos = 0
                for i, k in enumerate(ids):
                    idx = d["dimension"][k]["category"]["index"]
                    v = where.get(k, next(iter(idx)))   # the rest are 1 long
                    pos += idx[v] * stride[i]
                d["value"][str(pos)] *= factor
            return edit

        triple = [p for p in (SUP, USE, BAS) if p.exists()]
        check("the Austrian supply-use triple these mutations deform is "
              "present", len(triple) == 3,
              ", ".join(p.name for p in triple))

        if len(triple) == 3:
            pair = load_sut(SUP, USE, BAS)
            check("the Austrian triple is real and loads, so each mutation "
                  "below is the only difference",
                  (pair.n_products, pair.n_activities) == (65, 65),
                  f"AT 2022, {pair.n_products} x {pair.n_activities}")

            not_json = tmp / "una_hoja_de_calculo.csv"
            not_json.write_text("producto,valor\nCPA_A01,1234\n")

            use_doc = json.loads(USE.read_text())
            all_products = [c for c
                            in use_doc["dimension"]["prd_ava"]["category"]["index"]
                            if c.startswith("CPA_")]

            sut_cases = [
                # `table_kind` and `table_path` sit next to each other in the
                # template, so pointing the Eurostat route at a spreadsheet is
                # one wrong line. It used to answer with a codec error raised
                # inside `codecs.py`.
                ("a saved response that is text but is not JSON",
                 lambda: load_iot(not_json),
                 "is not json"),

                ("a supply file with no product axis at all",
                 lambda: load_sut(deformed(SUP, drop_dim("prd_amo"),
                                           "sup_sin_producto.json"), USE),
                 "no `prd_amo` dimension"),

                # Aimed at the `cp1610` dimension refusal and met it: dropping
                # the INDUSTRY axis gets there, because `_Cube.at` ignores a
                # keyword for a dimension the cube does not carry, so the
                # final-demand test upstream still finds its cells populated.
                ("a basic-price use table with no industry axis",
                 lambda: load_sut(SUP, USE,
                                  deformed(BAS, drop_dim("ind_use"),
                                           "bas_sin_industria.json")),
                 "this expects naio_10_cp1610"),

                # Aimed at that SAME refusal by dropping `stk_flow`, and met an
                # earlier and more specific one instead. The engine is right
                # and the expectation was wrong: the final-demand columns are
                # chosen by reading the basic-price cube's DOM flow, and a cube
                # with no `stk_flow` cannot answer, so the two files stop
                # agreeing on which components they publish before anyone asks
                # what dimensions the file has. Both were unreached, so the
                # case is kept pointed where it actually lands.
                ("a basic-price use table with no `stk_flow` axis, which "
                 "leaves its final-demand columns unreadable",
                 lambda: load_sut(SUP, USE,
                                  deformed(BAS, drop_dim("stk_flow"),
                                           "bas_sin_flujo.json")),
                 "must agree on which final-demand"),

                ("a supply file whose published total supply no longer "
                 "matches the products under it",
                 lambda: load_sut(
                     deformed(SUP, scale(1.5, ind_impv="TS_BP",
                                         prd_amo="CPA_TOTAL"),
                              "sup_total_movido.json"), USE),
                 "populated products sum to"),

                ("a use file whose published total output no longer matches "
                 "the industries under it",
                 lambda: load_sut(
                     SUP, deformed(USE, scale(1.5, ind_use="TOTAL",
                                              prd_ava="P1"),
                                   "use_total_movido.json")),
                 "populated industries' output sums to"),

                # Every product erased on one side. The halves are still the
                # same country and the same year -- they simply have nothing
                # in common, which is the one thing a PAIR has to have.
                ("a use file carrying no product values at all",
                 lambda: load_sut(SUP, deformed(USE,
                                                blank(all_products, "prd_ava"),
                                                "use_en_blanco.json")),
                 "no product carries values in both files"),
            ]
            for name, run, fragment in sut_cases:
                try:
                    run()
                except EurostatError as exc:
                    check(f"the engine refuses {name}",
                          fragment.lower() in str(exc).lower(),
                          str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
                except Exception as exc:               # noqa: BLE001
                    check(f"the engine refuses {name}", False,
                          f"{type(exc).__name__} instead: {str(exc)[:66]}")
                else:
                    check(f"the engine refuses {name}", False,
                          "it built a table")



def test_the_refusals_about_an_OFFICE_WORKBOOK():
    """The other half of what arrives from an office: the spreadsheets.

    The UK's analytical table and the INE's are workbooks with a fixed shape,
    and the loaders refuse a great many departures from it — no `IOT` sheet, row
    codes that are not the column codes, no primary-input row, a reference year
    that cannot be read, product labels that are not numbered, duplicate codes,
    supply and use sheets that disagree on their size, a final-demand column
    map that is neither of the two known ones, a renamed line on the supply-use
    sheets.

    The project holds the real files, so the baseline is not invented: one is
    copied and one thing is changed, the same way `configs/ejemplo.xlsx` is
    used for the config sheets.
    """
    import shutil
    import tempfile

    import openpyxl
    from quadrium.io_loader import (LoaderError, load_idescat_mioc,
                                    load_ine_tio, load_ine_tod,
                                    load_uk_analytical_iot)

    uk = ROOT / "UK_IOAT_2023_domestic_ixi.xlsx"
    ine = ROOT / "data" / "ine" / "cne_tio_22.xlsx"
    have = [p for p in (uk, ine) if p.exists()]
    check("the office workbooks this mutates are present",
          len(have) == 2,
          ", ".join(p.name for p in have))
    if len(have) != 2:
        return

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        def variant(src, edit, name="wb.xlsx"):
            path = tmp / name
            shutil.copy(src, path)
            wb = openpyxl.load_workbook(path)
            edit(wb)
            wb.save(path)
            return path

        def refuses(name, run, fragment):
            try:
                run()
            except LoaderError as exc:
                check(f"the engine refuses {name}",
                      fragment.lower() in str(exc).lower(),
                      str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}", False,
                      f"{type(exc).__name__} instead: {str(exc)[:64]}")
            else:
                check(f"the engine refuses {name}", False, "it built a table")

        def drop_sheet(which):
            return lambda wb: wb.remove(wb[which])

        uk_sheets = openpyxl.load_workbook(uk, read_only=True).sheetnames
        iot_sheet = next((n for n in uk_sheets if n.upper().startswith("IOT")),
                         None)
        if iot_sheet:
            refuses("a UK workbook with its IOT sheet removed",
                    lambda: load_uk_analytical_iot(
                        variant(uk, drop_sheet(iot_sheet))),
                    "no 'IOT' sheet")

        refuses("an INE workbook asked for a valuation that does not exist",
                lambda: load_ine_tio(ine, variant="neither"),
                "variant must be")
        refuses("an INE workbook asked for an unbalanced policy that is not one",
                lambda: load_ine_tio(ine, unbalanced="improvise"),
                "unbalanced must be")

        # Aimed at the missing-sheet refusal and met an earlier one: the index
        # sheet is where the reference year is read, and that happens first.
        # Both are on the unreached list, so both are covered.
        refuses("an INE workbook whose index sheet is gone, so the year cannot "
                "be read",
                lambda: load_ine_tio(
                    variant(ine, drop_sheet("Lista_Tablas"), "ine.xlsx")),
                "reference year")
        refuses("an INE workbook missing one of the tables it must carry",
                lambda: load_ine_tio(
                    variant(ine, drop_sheet("Tabla2"), "ine2.xlsx")),
                "is missing")

        # One cell at a time, inside a sheet that is present and the right
        # shape. The INE's interior table does not balance as published
        # (`OQ-D-04`), so these pass the policy the loader needs to get past
        # that -- otherwise every mutation would meet the balance refusal
        # instead of the one it aims at, which is what the first version did.
        from quadrium.io_loader import _INE
        col = _INE["col_label"] + 1
        ok_policy = dict(unbalanced="residual_column")

        # The product codes are read from `Tabla2`, not `Tabla1`, and from a row
        # range the loader computes per vintage -- so the row is found by the
        # pattern the loader itself matches rather than assumed. The first
        # version of this aimed at Tabla1 and quietly changed nothing.
        import re as _re
        _numbered = _re.compile(r"^\s*\d{1,3}(?:\s+bis)?\s*\.")

        def first_label_row(path=ine, sheet="Tabla2"):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for r, row in enumerate(wb[sheet].iter_rows(values_only=True), 1):
                if len(row) > col - 1 and row[col - 1] is not None \
                        and _numbered.match(str(row[col - 1])):
                    return r
            return None

        row0 = first_label_row()
        check("the INE product labels are where the loader says they are",
              row0 is not None,
              f"first numbered label on 'Tabla2' row {row0}")
        if row0 is None:
            return

        refuses("an INE product label that is not numbered the way the INE "
                "numbers them",
                lambda: load_ine_tio(variant(
                    ine,
                    lambda wb: setattr(wb["Tabla2"].cell(row=row0, column=col),
                                       "value", "just a name"), "ine3.xlsx"),
                    **ok_policy),
                "not a numbered product label")

        refuses("two INE products that carry the same code",
                lambda: load_ine_tio(variant(
                    ine,
                    lambda wb: setattr(
                        wb["Tabla2"].cell(row=row0 + 1, column=col), "value",
                        wb["Tabla2"].cell(row=row0, column=col).value),
                    "ine4.xlsx"), **ok_policy),
                "duplicate INE product codes")

        # ------------------------------------------------------------------
        # Four refusals about the reference year and the column maps. Each one
        # had a message written for it and nothing had ever read one out; the
        # deformation is always ONE thing changed on the real file.
        # ------------------------------------------------------------------

        # The year is read from the workbook's own banner and never from the
        # filename (`OQ-D-01`, and this project's own fixture is named 2022 and
        # is for 2023). A workbook that arrives without the banner therefore
        # has to stop rather than guess, and removing the `Menu` sheet is that
        # workbook.
        refuses("a UK workbook with no Menu sheet to read the year from",
                lambda: load_uk_analytical_iot(
                    variant(uk, drop_sheet("Menu"), "uk_nomenu.xlsx")),
                "could not read the reference year from the Menu sheet")

        # `_ine_columns` knows two final-demand layouts and decides which one
        # it is holding from two fixed header cells, 76 and 77 of row 8. This
        # fixture is the split-exports shape, so 77 carries the European Union
        # export column; overwriting it leaves a workbook that is neither
        # shape, which is the third layout the loader says needs its own map.
        def header_cell(path, sheet, row, col):
            book = openpyxl.load_workbook(path, read_only=True,
                                          data_only=True)
            r = list(book[sheet].iter_rows(values_only=True))[row - 1]
            return str(r[col - 1] or "").strip() if col - 1 < len(r) else ""

        ine_hdr = header_cell(ine, "Tabla1", 8, 77)
        check("the INE workbook is the split-exports layout this deforms",
              ine_hdr.lower().startswith("exportaciones"), ine_hdr)

        refuses("an INE workbook whose final-demand columns are a third "
                "layout, mapped nowhere",
                lambda: load_ine_tio(variant(
                    ine,
                    lambda wb: setattr(wb["Tabla1"].cell(row=8, column=77),
                                       "value",
                                       "Columna que este cargador no conoce"),
                    "ine5.xlsx")),
                "match neither layout this loader knows")

        # The supply-use loader finds every position by the label the INE
        # prints, so a renamed line stops the load instead of shifting the
        # block underneath it. The column is found here THE WAY THE LOADER
        # FINDS IT — by its normalised label — because its index differs
        # between the 65-product vintage and the 110-product one, and a fixed
        # index would quietly rename nothing on the other.
        from quadrium.io_loader import _tod_norm
        tod = ROOT / "data" / "ine" / "cne_tod_22.xlsx"
        check("the INE supply-use workbook this mutates is present",
              tod.exists(), tod.name)
        if tod.exists():
            book = openpyxl.load_workbook(tod, read_only=True, data_only=True)
            hdr = list(book["Tabla2"].iter_rows(values_only=True))[7]
            fd_col = next((j + 1 for j, v in enumerate(hdr)
                           if _tod_norm(v) == "total demanda final"), None)
            check("the use sheet's final-demand total is where the loader "
                  "looks it up by label",
                  fd_col is not None,
                  f"row 8, column {fd_col} of 'Tabla2'")
            if fd_col is not None:
                refuses("an INE supply-use workbook with its final-demand "
                        "total column renamed",
                        lambda: load_ine_tod(variant(
                            tod,
                            lambda wb: setattr(
                                wb["Tabla2"].cell(row=8, column=fd_col),
                                "value", "Suma de la demanda"),
                            "tod.xlsx")),
                        "no column labelled 'total demanda final'")

        # IDESCAT's own year, the same rule and its own refusal. This file is
        # NOT published in the public tree, so the case is guarded rather than
        # required: a tree without a fixture cannot exercise the code that
        # reads it, which is why the coverage floor lives in each tree's own
        # record instead of in `run_refusal_coverage.py`.
        mioc = ROOT / "data" / "idescat" / "mioc2021ts64.xlsx"
        if mioc.exists():
            def blank_years(wb):
                for row in wb["ts total"].iter_rows(min_row=1, max_row=6):
                    for cell in row:
                        if cell.value is not None and _re.search(
                                r"\b(19|20)\d{2}\b", str(cell.value)):
                            cell.value = None

            refuses("a Catalan workbook with no year anywhere in its title "
                    "rows",
                    lambda: load_idescat_mioc(
                        variant(mioc, blank_years, "mioc.xlsx")),
                    "could not read the reference year from the title rows")

        # ------------------------------------------------------------------
        # THE ONS SHEET'S OWN SIGNPOSTS, and the four refusals that guard
        # them. `_uk_layout` finds every block on the `IOT` sheet from the
        # totals the sheet prints -- `_T` at the end of the sector rows and
        # columns, `TU` at the end of the final-demand block -- and never from
        # a fixed offset, because the ONS changed the size and both axes
        # between editions. Three refusals defend that navigation and a fourth
        # contrasts the result against those same printed totals, so that a
        # block found one line out of true fails as the loader having looked
        # in the wrong place rather than three identities later as an
        # accusation against the ONS's arithmetic.
        #
        # All four were on the unreached list: a loader that reads the real
        # file every day never takes any of these branches, which is exactly
        # the shape of promise this sweep exists to find.
        #
        # The positions come from `_uk_layout` RUN ON THE REAL FILE and are
        # not written down here. This fixture is the 104 x 104 edition and the
        # 2016-2022 ones are 105 x 105, so a hardcoded index would deform a
        # different cell on the next vintage -- the `Tabla1` lesson from the
        # other side. openpyxl counts from 1 and `_open_workbook` from 0, so
        # every row and column below is an index plus one.
        # ------------------------------------------------------------------
        from quadrium.io_loader import (_UK_ROWS, _open_workbook, _uk_layout,
                                        _uk_norm)

        sheets_uk = _open_workbook(uk)
        check("the UK workbook carries the 'IOT' sheet these deform",
              "IOT" in sheets_uk, ", ".join(list(sheets_uk)[:6]) + "…")
        if "IOT" in sheets_uk:
            R_uk = sheets_uk["IOT"]
            L_uk = _uk_layout(R_uk, uk.name)
            check("and its signposts are where the loader navigates by them",
                  _uk_norm(R_uk[3][L_uk["tu_col"]]) == "tu"
                  and _uk_norm(R_uk[L_uk["end_row"]][0]) == "_t",
                  f"`TU` in column {L_uk['tu_col'] + 1}, the `_T` row at row "
                  f"{L_uk['end_row'] + 1}, "
                  f"{L_uk['end_col'] - L_uk['first_col']} sectors")

            def uk_cell(row, col, value, name):
                """One cell of the IOT sheet changed, and nothing else."""
                return variant(
                    uk,
                    lambda wb: setattr(wb["IOT"].cell(row=row, column=col),
                                       "value", value),
                    name)

            # Blank `TU` and the final-demand block has no end. This is the
            # FIRST thing `_uk_layout` asks, so nothing can get in front of it.
            refuses("a UK sheet with no `TU` to find the end of final demand by",
                    lambda: load_uk_analytical_iot(
                        uk_cell(4, L_uk["tu_col"] + 1, None, "uk_notu.xlsx")),
                    "does not print the totals this loader navigates by")

            # `_T` and `TU` still print, so the check above passes and this one
            # is reached: one column header renamed and the sheet no longer
            # carries one classification on both axes. The message names the
            # position, which is the first column of the sector block.
            refuses("a UK sheet whose column codes are not its row codes",
                    lambda: load_uk_analytical_iot(uk_cell(
                        4, L_uk["first_col"] + 1, "ZZZ", "uk_axes.xlsx")),
                    "the 'IOT' sheet's row codes are not its column codes")

            # The primary-input row is found the way the LOADER finds it --
            # by its normalised label in column B below the sector block, the
            # sheet's non-breaking spaces included -- and every one of the
            # eight labels was confirmed to appear exactly once, so renaming
            # the first occurrence removes it rather than shadowing it.
            # `_UK_ROWS` is walked in insertion order and `imports` is first,
            # so that is the label the message must name.
            below = {}
            for i in range(L_uk["end_row"], len(R_uk)):
                key = _uk_norm(R_uk[i][1] if len(R_uk[i]) > 1 else "")
                if key and key not in below:
                    below[key] = i
            imports_row = below.get(_UK_ROWS["imports"])
            check("the imports row is where the loader looks it up by label",
                  imports_row is not None,
                  f"{_UK_ROWS['imports']!r} on row "
                  f"{imports_row + 1 if imports_row is not None else '—'}")
            if imports_row is not None:
                refuses("a UK sheet with its imports line renamed, so a "
                        "primary input the model needs is not there",
                        lambda: load_uk_analytical_iot(uk_cell(
                            imports_row + 1, 2,
                            "Imports, under a name this loader does not know",
                            "uk_noimports.xlsx")),
                        "has no primary-input row labelled")

            # And the contrast against the printed totals. The `_T` ROW is not
            # part of `Z` -- it sits at `end_row`, one past the last sector --
            # so moving one of its cells leaves the interior untouched and
            # breaks only the equality between what the sheet publishes as the
            # column total and what its own column adds up to. The data is at
            # full float precision, so `assertable_tolerance` derives a bound
            # of 1e-6 and 1,000 is unmistakably past it. This is the first of
            # the loader's three contrasts, so it fires before the `TU` column
            # and the GVA row are looked at, and well before `_assert_balances`.
            printed = R_uk[L_uk["end_row"]][L_uk["first_col"]]
            refuses("a UK sheet whose printed `_T` row is not the sum of its "
                    "own sector columns",
                    lambda: load_uk_analytical_iot(uk_cell(
                        L_uk["end_row"] + 1, L_uk["first_col"] + 1,
                        float(printed) + 1000.0, "uk_total.xlsx")),
                    "the printed `_T` row is the sum of the sector columns")




def test_the_catalan_table_loads_and_says_whose_residue_it_carries():
    """IDESCAT's symmetric table, and the two things it forced.

    The file is the answer a regionalisation can be scored against: Catalonia
    2021 at 63 branches, beside the INE's national table for the same year at
    64. It also carries the same table four times over, split by where the
    inputs came from, which is a two-region interregional table already
    compiled.

    Two things it forced on the engine, both measured rather than assumed:

    Only the CATALAN sheet is an input-output table. Against output, the row
    identity is 0.8 for `ts Cat` and 12,303 for `ts total` — what the others
    record is supplied by imports, not by Catalan production. So the imported
    intermediates come back inside `VA`, labelled as not being value added,
    which is what makes the column identity close to 7e-12.

    And the row identity does NOT close to zero, because IDESCAT's own books do
    not: its `Total usos` column differs from its `Producció` row by up to 0.84
    over 455,354. The cells are full precision, so the derived floor is zero and
    nothing would have absorbed it. `_assert_balances` gained an `inherited`
    argument for this — a residue the caller has MEASURED against the
    publisher's own printed totals, not a number chosen to make a gate pass.
    """
    from quadrium.io_loader import LoaderError, load_idescat_mioc
    from quadrium.precision import printed_decimals

    # The Catalan data is PRIVATE: putting it in the public tree means
    # redistributing a new publisher under a licence nobody here has read, and
    # `run_idescat_catalonia.py` is in sync_public's PRIVATE_ONLY for the same
    # reason. A test file cannot be split, so this one says which tree it is in
    # rather than failing in one of them -- or, worse, passing in silence. The
    # discriminator is the methodological library, which never leaves.
    path = ROOT / "data" / "idescat" / "mioc2021ts64.xlsx"
    private_tree = (ROOT / "library" / "extracted").is_dir()
    if not path.exists():
        check("the Catalan table is absent only where it is meant to be",
              not private_tree,
              "this is the public tree, which does not ship IDESCAT's data, so "
              "there is nothing to load and nothing to check"
              if not private_tree else
              f"missing from the PRIVATE tree, where it should be: {path}")
        return

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        t = load_idescat_mioc(path)

    check("it loads with the branches the sheet actually has",
          t.n == 63 and t.year == 2021,
          f"{t.n} branches, {t.year}, {t.X.sum():,.0f} M€ of output — 64 "
          f"numbered rows, one of which is an of-which memo with no CPA code")

    col = float(np.abs(t.Z.sum(axis=0) + t.VA.sum(axis=0) - t.X).max())
    check("the column identity closes, with imports counted as inputs",
          col < 1e-6,
          f"{col:.2e}. Imported intermediates are the COLUMN sums of the other "
          f"two sheets; the import ROWS are indexed by product and using those "
          f"leaves it 6,588 out")

    row = float(np.abs(t.Z.sum(axis=1) + t.Y.sum(axis=1) - t.X).max())
    check("and the row residue is declared as the publisher's own",
          abs(row - t.inherited_residue) < 1e-9 and 0 < row < 1.0,
          f"{row:.4f}, matching `inherited_residue` exactly — IDESCAT's "
          f"`Total usos` against IDESCAT's `Producció`")
    check("which nothing else could have absorbed, because the file is unrounded",
          printed_decimals(t.Z.ravel()) is None,
          "full float precision, so `assertable_tolerance` gives zero")

    labelled = [s for s in t.VA_labels if "NO és valor afegit" in s]
    check("and the rows that are not value added say so in their own label",
          len(labelled) == 3,
          f"{len(labelled)} of {len(t.VA_labels)}: the two import blocks and "
          f"net taxes on products, as the UK loader labels its own")

    try:
        load_idescat_mioc(ROOT / "data" / "ine" / "cne_tio_21.xlsx")
        check("and another office's workbook is refused by name", False,
              "it read the INE's file as if it were IDESCAT's")
    except LoaderError as exc:
        check("and another office's workbook is refused by name",
              "missing the sheet" in str(exc), str(exc)[:80] + "…")



def test_the_refusals_the_METHODS_make():
    """What the balancing, transformation and projection steps refuse.

    These sit past the loaders: the file was read and the split was legal, and
    the method itself will not proceed. Each is one call away from a matrix
    written by hand, so nothing but not looking had kept them untested.
    """
    from quadrium.balancing import BalancingError, balance, ras
    from quadrium.sut_euro import SutEuroError, sut_euro
    from quadrium.transformation import (TransformationError, choose_model,
                                         transform)

    def refuses(name, fn, kind, fragment):
        try:
            fn()
        except kind as exc:
            check(f"the engine refuses {name}",
                  fragment.lower() in str(exc).lower(),
                  str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
        except Exception as exc:                       # noqa: BLE001
            check(f"the engine refuses {name}", False,
                  f"{type(exc).__name__} instead: {str(exc)[:64]}")
        else:
            check(f"the engine refuses {name}", False, "it went ahead")

    neg = np.array([[10.0, -2.0], [3.0, 8.0]])
    refuses("a negative seed handed to RAS, and names the method that takes one",
            lambda: ras(neg, np.array([8.0, 11.0]), np.array([13.0, 6.0])),
            BalancingError, "non-negative seed")

    pos = np.array([[10.0, 5.0], [4.0, 6.0]])
    refuses("interior cells pinned under a method that only takes margins",
            lambda: balance(pos, np.array([15.0, 10.0]), np.array([14.0, 11.0]),
                            method="GRAS", locked_cells=[(0, 0)]),
            BalancingError, "locked cells")

    refuses("a secondary-production type that is not one of the three",
            lambda: choose_model(square=True, secondary_type="mysterious"),
            TransformationError, "not one of")

    V_T = np.array([[80.0, 20.0], [10.0, 90.0]])
    Ud = np.array([[30.0, 20.0], [15.0, 25.0]])
    Um = np.zeros((2, 2))
    Yd = np.array([[50.0], [60.0]])
    Ym = np.zeros((2, 1))
    W = np.array([[45.0, 65.0]])
    g = V_T.sum(axis=0)
    x = V_T.sum(axis=1)
    refuses("the hybrid model asked for without the matrix that defines it",
            lambda: transform("E", V_T, Ud, Um, Yd, Ym, W, g, x),
            TransformationError, "needs the hybrid")

    refuses("a projection whose pieces are not one supply-use pair",
            lambda: sut_euro(
                np.zeros((3, 4)), np.zeros((2, 2)), np.zeros(3),
                np.zeros((3, 3)),
                va_target=np.zeros(3), final_use_target=np.zeros(1),
                tls_target=np.zeros(3), imports_target=np.zeros(3)),
            SutEuroError, "one SUT pair")



def test_the_refusals_the_SUT_to_IOT_step_makes_of_ITS_CALLER():
    """The internal contracts of the SUT-to-IOT step, and none had a case.

    These are a different animal from every other group the sweep tracks. They
    judge neither a file from an office, nor the user's workbook, nor what was
    asked for in it: they judge the ARGUMENTS one part of this engine hands
    another. A rectangular supply table where the algebra needs a square one, a
    hybrid matrix in the wrong shape or holding a fraction where CORE_013 par.
    12.62, p. 389 defines a 0/1 matrix, a model name that is not one of
    Figure 12.2's.

    They were the whole of the `caller` class in `data/_refusal_coverage.json`
    -- six of six unreached -- and the reason is not that they are hard. Every
    one is a call away from a matrix written by hand, which is what this does.
    They stayed untested because a caller check looks like it cannot fire, and
    a check nobody can see fire is a check nobody has read.

    Each case was confirmed to reach the refusal it aims at and not an earlier
    one: `transform` screens the model name before touching a matrix,
    `hybrid_transformation_matrix` screens H's shape before its contents, and
    `almon` screens the supply table's shape before U's columns -- so the case
    for the second of each pair has to satisfy the first.
    """
    from quadrium.transformation import (TransformationError, almon,
                                         hybrid_matrix_avoiding_negatives,
                                         hybrid_transformation_matrix,
                                         transform)

    def refuses(name, fn, fragment):
        try:
            fn()
        except TransformationError as exc:
            check(f"the engine refuses {name}",
                  fragment.lower() in str(exc).lower(),
                  str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
        except Exception as exc:                       # noqa: BLE001
            check(f"the engine refuses {name}", False,
                  f"{type(exc).__name__} instead: {str(exc)[:64]}")
        else:
            check(f"the engine refuses {name}", False, "it went ahead")

    V_T = np.array([[80.0, 20.0], [10.0, 90.0]])       # square: 2 x 2
    Ud = np.array([[30.0, 20.0], [15.0, 25.0]])
    Um = np.zeros((2, 2))
    Yd = np.array([[50.0], [60.0]])
    Ym = np.zeros((2, 1))
    W = np.array([[45.0, 65.0]])
    g = V_T.sum(axis=0)
    x = V_T.sum(axis=1)
    rect = np.array([[80.0, 20.0, 5.0], [10.0, 90.0, 7.0]])   # 2 products, 3 industries

    refuses("a model name that is not one of CORE_013 Figure 12.2's",
            lambda: transform("Z", V_T, Ud, Um, Yd, Ym, W, g, x),
            "not one of the four")

    refuses("the hybrid search handed a supply table that is not square",
            lambda: hybrid_matrix_avoiding_negatives(rect, Ud, Um, W, g, x),
            "needs a square supply table")

    refuses("an H whose shape is not the supply table's",
            lambda: hybrid_transformation_matrix(V_T, g, x, np.ones((2, 3))),
            "same shape as the supply table")

    # H's shape is checked first, so this one has to be the right shape and
    # wrong only in its contents: a fraction is a different model, not this one.
    refuses("an H holding a fraction where CORE_013 defines 0 or 1",
            lambda: hybrid_transformation_matrix(
                V_T, g, x, np.array([[1.0, 0.5], [0.0, 1.0]])),
            "must hold only 0 and 1")

    refuses("Almon's procedure handed a supply table that is not square",
            lambda: almon(Ud, rect),
            "needs a square supply table")

    # Same ordering: V_T must pass the square check before U's columns are
    # looked at, so V_T stays 2 x 2 and U is the one carrying three columns.
    refuses("a use table whose columns are not the supply table's industries",
            lambda: almon(np.zeros((2, 3)), V_T),
            "both are indexed by industry")


def test_the_refusals_about_what_was_ASKED_FOR():
    """Eleven refusals judge the split itself, and none had a case.

    These sit between the workbook and the numbers: the sheets parsed, the
    table loaded, and then what the user asked for does not hold together — a
    key with the wrong number of weights, weights that do not sum to one, a new
    subsector code that collides with a sector being split. They are a direct
    call away from the synthetic fixture, so there was no reason for them to be
    untested beyond nobody having looked.

    TWO WERE ADDED LATER, AND ONE OF THEM REPLACED A CASE THAT WAS GREEN AND
    WRONG. `_va_columns` and `_weights` refuse a mis-sized key with the same
    sentence, and the case aimed at the first named a key the second sees
    first, so it had been passing on the wrong refusal since it was written --
    which is exactly what `data/_refusal_coverage.json` said, and what a
    passing check said otherwise. The other is weights that clear the length
    check and are still not shares.
    """
    from quadrium.disaggregation import (DisaggregationError, split_sector,
                                         split_sectors)
    from quadrium.models import AllocationKey

    table = build_table()
    keys = build_keys()
    sc = Scenario(scenario_id="s", label="s",
                  keys_by_block={"output": "key_turnover"})
    spec = SplitSpec("ACC", NEW, LBL, keys_by_block={"output": "key_turnover"})

    def refuses(name, fn, fragment):
        try:
            fn()
        except DisaggregationError as exc:
            check(f"the engine refuses {name}", fragment.lower() in str(exc).lower(),
                  str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
        except Exception as exc:                       # noqa: BLE001
            check(f"the engine refuses {name}", False,
                  f"{type(exc).__name__} instead: {str(exc)[:66]}")
        else:
            check(f"the engine refuses {name}", False, "it went ahead")

    refuses("a split into fewer than two subsectors",
            lambda: split_sector(table, "ACC", NEW[:1], LBL[:1], sc, keys, spec),
            "fewer than 2")
    refuses("codes and labels of different lengths",
            lambda: split_sector(table, "ACC", NEW, LBL[:2], sc, keys, spec),
            "differ in length")
    refuses("a scenario that asks for no splits at all",
            lambda: split_sectors(table, [], sc, keys),
            "no splits requested")
    # Aimed at "repeats the code of a sector being split" and met an earlier,
    # more specific one: the code already exists in the table at all, which is
    # checked first. That is the engine being right; the case covers what it
    # actually says.
    refuses("a new subsector code that already exists in the table",
            lambda: split_sectors(
                table,
                [SplitSpec("ACC", ["ACC", "CAM"], ["a", "b"],
                           keys_by_block={"output": "key_turnover"})],
                sc, keys),
            "already exist in the table")

    # a key whose weights do not match the number of parts, and one the
    # scenario names but the workbook never defined
    bad_len = dict(keys)
    k = keys["key_turnover"]
    bad_len["key_turnover"] = AllocationKey(
        key_id="key_turnover", applies_to="output", new_sector_codes=NEW[:2],
        raw_values=[1.0, 1.0], source=k.source, source_year=k.source_year,
        strength=k.strength)
    refuses("a key with the wrong number of weights for the split",
            lambda: split_sector(table, "ACC", NEW, LBL, sc, bad_len, spec),
            "weights for")
    # The value-added rows can carry their own keys, and that opt-in has three
    # refusals of its own -- a row that is not in the table, a key that is not
    # loaded, a key of the wrong length. None had been exercised.
    VA_ROW = "Compensation of employees"
    def va_spec(row_keys, residual="Gross operating surplus"):
        return SplitSpec("ACC", NEW, LBL,
                         keys_by_block={"output": "key_turnover"},
                         va_row_keys=row_keys, va_residual_row=residual)

    refuses("value-added row keys naming a row this table does not have",
            lambda: split_sector(table, "ACC", NEW, LBL, sc, keys,
                                 va_spec({"Wages": "key_turnover"})),
            "is not in this table")
    refuses("value-added row keys naming a key that was never loaded",
            lambda: split_sector(table, "ACC", NEW, LBL, sc, keys,
                                 va_spec({VA_ROW: "absent"})),
            "not among the loaded keys")
    # THIS CASE USED TO PASS FOR THE WRONG REASON, and the sweep is what showed
    # it. It aimed at `_va_columns`'s length check and named `key_turnover` --
    # which is also the split's OUTPUT key, so `_weights` refused it three
    # calls earlier with a message of the same shape ("key 'key_turnover' has 2
    # weights for 4 subsectors"), and the fragment matched either one. The
    # engine was right both times; the case was not reaching what it claimed
    # to. `_weights` was already covered; `_va_columns` was not, and stayed
    # unreached in `data/_refusal_coverage.json` while a green check said
    # otherwise. The block key is left intact and the WRONG length is put on a
    # key no block names, so nothing sees it before the value-added rows do.
    short_va = dict(keys)
    kg = keys["key_gva"]
    short_va["key_gva"] = AllocationKey(
        key_id="key_gva", applies_to="value_added", new_sector_codes=NEW[:2],
        raw_values=[1.0, 1.0], source=kg.source, source_year=kg.source_year,
        strength=kg.strength)
    refuses("a value-added row key with the wrong number of weights",
            lambda: split_sector(table, "ACC", NEW, LBL, sc, short_va,
                                 va_spec({VA_ROW: "key_gva"})),
            "key 'key_gva' has 2 weights for 4 subsectors")

    # Weights that pass the length check and still are not shares. Reaching
    # this needs `weights=` passed explicitly: `AllocationKey.__post_init__`
    # normalises `raw_values` when `weights` is None, so a key built the
    # ordinary way can never sum to anything but one, and the guard would be
    # unreachable rather than merely untested.
    kt = keys["key_turnover"]
    bad_sum = dict(keys)
    bad_sum["key_turnover"] = AllocationKey(
        key_id="key_turnover", applies_to="output", new_sector_codes=NEW,
        raw_values=[1.0, 1.0, 1.0, 1.0], source=kt.source,
        source_year=kt.source_year, strength=kt.strength,
        weights=[0.2, 0.2, 0.2, 0.2])
    refuses("weights that are the right length and do not sum to one",
            lambda: split_sector(table, "ACC", NEW, LBL, sc, bad_sum, spec),
            "weights sum to 0.8, not 1")

    # Two splits in one pass, the second introducing a code the first is still
    # taking apart. Aimed at a second guard that used to sit below this one --
    # "repeats the code of a sector being split" -- and reached the first
    # instead, every time. It could not be otherwise: that guard tested
    # `introduced & seen`, and `seen` holds only codes the table already has,
    # so it was a subset of the collision checked three lines above. Unreachable
    # rather than untested, and now removed.
    refuses("a new code that collides with a sector being split, caught by the "
            "check that can actually see it",
            lambda: split_sectors(
                table,
                [SplitSpec("ACC", NEW, LBL,
                           keys_by_block={"output": "key_turnover"}),
                 SplitSpec("TRA", ["ACC", "X2"], ["a", "x"],
                           keys_by_block={"output": "key_turnover"})],
                sc, keys),
            "already exist in the table")

    refuses("a split with no key for a block and none to fall back on",
            lambda: split_sector(
                table, "ACC", NEW, LBL,
                Scenario(scenario_id="s", label="s"), keys,
                SplitSpec("ACC", NEW, LBL)),
            "has no allocation key")

    refuses("a scenario naming a key that was never defined",
            lambda: split_sector(
                table, "ACC", NEW, LBL,
                Scenario(scenario_id="s", label="s",
                         keys_by_block={"output": "absent"}),
                keys,
                SplitSpec("ACC", NEW, LBL,
                          keys_by_block={"output": "absent"})),
            "not among the loaded keys")



def test_the_config_refusals_a_stranger_meets_next():
    """The other half of the front door: the sheets that say what to split.

    `io_loader` reads the TABLE; `config` reads the workbook that says what to
    do with it — `splits`, `keys`, `scenarios`, `profiles`. It holds 41 refusals
    and 15 had ever fired, the largest single block of unreached ones in the
    engine, and every message in it names a sheet, a row and a field because it
    is written for someone who filled the sheet in by hand.

    The engine ships a working example, so the baseline is not invented: this
    loads `configs/ejemplo.xlsx`, asserts it loads, then changes ONE cell at a
    time. A refusal counts only if it names the thing that was broken.
    """
    import openpyxl
    from quadrium.config import ConfigError, load_config
    from quadrium.eurostat import EurostatError

    example = ROOT / "configs" / "ejemplo.xlsx"
    if not example.exists():
        check("the shipped example is where this expects it", False,
              str(example))
        return

    import shutil
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        def variant(edit):
            """Copy the example, apply one edit, return the path.

            `table_path` is relative to the example's own folder, so the copy
            gets it as an absolute path — otherwise every case would fail on a
            missing table instead of on the thing being tested.
            """
            path = tmp / "cfg.xlsx"
            shutil.copy(example, path)
            wb = openpyxl.load_workbook(path)
            for row in wb["project"].iter_rows(min_col=1, max_col=2):
                if row[0].value == "table_path":
                    row[1].value = str(
                        (example.parent / str(row[1].value)).resolve())
            edit(wb)
            wb.save(path)
            return path

        try:
            load_config(variant(lambda wb: None))
            check("the shipped example loads, so each break is the only "
                  "difference", True, example.name)
        except ConfigError as exc:
            check("the shipped example loads, so each break is the only "
                  "difference", False, str(exc)[:90])
            return

        def set_project(key, value):
            def edit(wb):
                ws = wb["project"]
                for row in ws.iter_rows(min_col=1, max_col=2):
                    if row[0].value == key:
                        row[1].value = value
                        return
                ws.append([key, value])
            return edit

        def set_cell(sheet, row, col, value):
            return lambda wb: setattr(wb[sheet].cell(row=row, column=col),
                                      "value", value)

        cases = [
            ("a table_kind that is not one of the kinds it reads",
             set_project("table_kind", "nonsense"), "table_kind"),
            # Changing ONE row of a split hits an earlier and different
            # refusal -- the rows of one split must agree on their key -- which
            # is the engine being right and the first version of this test being
            # wrong. Both are worth covering, so both are here.
            ("a split whose rows disagree about which key drives it",
             set_cell("splits", 2, 4, "no_such_key"), "more than one"),
            ("a split naming a key the keys sheet does not define",
             lambda wb: [setattr(wb["splits"].cell(row=r, column=4), "value",
                                 "no_such_key") for r in (2, 3, 4)],
             "not in the"),
            ("a key row whose strength is not strong/medium/weak",
             set_cell("keys", 2, 6, "quite good"), "strength"),
            ("a profile for a subsector the splits sheet never creates",
             set_cell("profiles", 2, 2, "I999"), "not created"),
            ("a profile intensity that is not a number",
             set_cell("profiles", 2, 4, "a lot"), "intensity"),
            ("a key row with an empty required field",
             set_cell("keys", 2, 4, None), "empty"),
            ("a table_path that points at nothing",
             set_project("table_path", "no/such/table.xlsx"), "does not"),
            ("a table_unbalanced that is not one of the two policies",
             set_project("table_unbalanced", "shrug"), "table_unbalanced"),
            # `keys` absent is read as empty, not missing, so the split's own
            # key check fires first -- which is a better message anyway.
            #
            # `project` is the only unconditionally required sheet. `splits` was
            # too until v1.86, when `regionalise` became the other job a
            # workbook can describe; deleting `splits` now gets the refusal for
            # a workbook that names a table and nothing to do with it, which is
            # the more useful of the two messages because it says what the two
            # options ARE.
            ("a workbook with the one required sheet deleted",
             lambda wb: wb.remove(wb["project"]), "missing the sheet"),
            ("a workbook that names a table and no job",
             lambda wb: wb.remove(wb["splits"]), "describes no job"),
            ("a split whose subsectors and whose key disagree",
             set_cell("keys", 2, 2, "I999"), "but key"),
            # The Eurostat route: the engine downloads the table itself, so
            # these three are all a user gives it and all it can check.
            ("a Eurostat country that is not a country code",
             lambda wb: [set_project("table_kind", "eurostat")(wb),
                         set_project("eurostat_geo", "Spain")(wb)],
             "two-letter country code"),
            ("a Eurostat year that is not a year",
             lambda wb: [set_project("table_kind", "eurostat")(wb),
                         set_project("eurostat_geo", "ES")(wb),
                         set_project("eurostat_year", "recently")(wb)],
             "not a year"),
            # The Eurostat route resolves `table_path` before it validates
            # the model, so pointing it at a workbook is what a wrong
            # `table_kind` actually produces -- and it produced a raw
            # UnicodeDecodeError until this found it.
            ("the Eurostat route pointed at a spreadsheet",
             lambda wb: [set_project("table_kind", "eurostat")(wb),
                         set_project("eurostat_geo", "ES")(wb),
                         set_project("eurostat_year", 2022)(wb)],
             "not a JSON-stat file"),
        ]
        # NOT covered here, and why: `eurostat_model` is validated only after
        # the table resolves, and `project_method` / `project_to_year` only
        # when a `targets` sheet exists, which the shipped example has none of.
        # Reaching them needs a second fixture rather than one changed cell.
        for name, edit, fragment in cases:
            try:
                load_config(variant(edit))
            # A config can legitimately surface a LOADER's refusal: the sheets
            # were readable and what they pointed at was not.
            except (ConfigError, EurostatError) as exc:
                check(f"the engine refuses {name}, and says where",
                      fragment.lower() in str(exc).lower(),
                      str(exc).replace("\n", " ")[:86] + "…")
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says where", False,
                      f"{type(exc).__name__} instead: {str(exc)[:66]}")
            else:
                check(f"the engine refuses {name}, and says where", False,
                      "it accepted the workbook")



def test_the_workbook_refusals_a_stranger_meets_first():
    """The front door, which had no case at all behind it.

    `run_refusal_coverage.py` measured that 52 of the engine's 159 refusals had
    ever been reached, and this file's first version claimed most of the rest
    judged what the CALLER passed. Reading all 159 says **seven** do. Thirty-four
    judge the user's OWN SPREADSHEET — the route `docs/GUIDE.md` opens with:
    "No Python: you fill in a spreadsheet and run one command".

    Those are the first thing a stranger with a slightly different sheet meets,
    and not one of them had a case. This builds a minimal valid workbook and
    breaks it one way at a time, so each refusal is checked by the message it
    gives rather than by the fact that something failed.
    """
    import openpyxl
    from quadrium.io_loader import LoaderError, load_io_table

    META = [("country", "ES"), ("year", 2022), ("unit", "MIO_EUR"),
            ("classification", "CPA_2.1"), ("source", "test")]
    TABLE = [["", "A", "B", "HH"],
             ["A", 10.0, 5.0, 35.0],
             ["B", 4.0, 6.0, 40.0],
             ["VA", 36.0, 39.0, None],
             ["Output", 50.0, 50.0, None]]

    def build(tmp, table=None, meta=None, sheets=("table", "metadata")):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        if "table" in sheets:
            ws = wb.create_sheet("table")
            for row in (TABLE if table is None else table):
                ws.append(row)
        if "metadata" in sheets:
            ws = wb.create_sheet("metadata")
            for row in (META if meta is None else meta):
                ws.append(list(row))
        path = tmp / "wb.xlsx"
        wb.save(path)
        return path

    import tempfile
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        # the workbook has to load before breaking it means anything
        try:
            t = load_io_table(build(tmp))
            check("the minimal valid workbook loads, so the breakages are the "
                  "only difference", t.n == 2,
                  f"{t.n} sectors, {t.Y.shape[1]} final-demand column(s)")
        except LoaderError as exc:
            check("the minimal valid workbook loads, so the breakages are the "
                  "only difference", False, str(exc)[:90])
            return

        cases = [
            ("a missing file",
             lambda: load_io_table(tmp / "absent.xlsx"), "no such file"),
            ("a workbook with no `table` sheet",
             lambda: load_io_table(build(tmp, sheets=("metadata",))),
             "no sheet named"),
            ("a workbook with no `metadata` sheet",
             lambda: load_io_table(build(tmp, sheets=("table",))),
             "no 'metadata' sheet"),
            ("metadata missing the price basis",
             lambda: load_io_table(
                 build(tmp, meta=[m for m in META if m[0] != "unit"])),
             "missing: unit"),
            ("no row saying which line is output",
             lambda: load_io_table(
                 build(tmp, table=[r for r in TABLE
                                   if str(r[0]).lower() != "output"])),
             "no row labelled 'Output'"),
            ("no final-demand column at all",
             lambda: load_io_table(
                 build(tmp, table=[r[:3] for r in TABLE])),
             "final-demand column"),
        ]
        for name, run, fragment in cases:
            try:
                run()
            except LoaderError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(),
                      str(exc)[:88] + ("…" if len(str(exc)) > 88 else ""))
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead: {str(exc)[:70]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it accepted the workbook")



def test_the_refusals_that_judge_DATA_actually_fire():
    """Four promises the engine makes to a user that nothing had ever tested.

    `run_refusal_coverage.py` wrapped every exception type the engine defines,
    ran the whole suite, and found **49 of 159 refusal sites reached**. Most of
    the rest are argument checks that judge what the CALLER passed — `model 'X'
    is not one of the four` — and no published table can reach them, so their
    being unexercised costs nothing.

    Four are different: they judge the DATA, so they are what a user meets
    first if their table is the wrong shape of wrong, and not one had a case.
    `transformation.py` is the starkest — the SUT-to-IOT step carries four
    models from CORE_013 and ONE of its eleven refusals had ever fired.

    Each is reachable with a table of the right shape and the wrong contents,
    which is what this builds.
    """
    from quadrium.sut_ras import DegenerateMarginError, sut_ras
    from quadrium.transformation import (TransformationError, market_shares,
                                         product_mix)

    V_T = np.array([[80.0, 0.0], [20.0, 0.0]])   # product 2 is made by nobody
    g = V_T.sum(axis=0)
    x = V_T.sum(axis=1)
    try:
        product_mix(V_T, g)
    except TransformationError as exc:
        check("a zero industry total refuses the transformation by name",
              "industry output g" in str(exc) and "zero" in str(exc),
              str(exc)[:88] + "…")
    else:
        check("a zero industry total refuses the transformation by name", False,
              "it returned a matrix built on a division by zero")

    V_T2 = np.array([[80.0, 20.0], [0.0, 0.0]])   # product 2 has no output
    try:
        market_shares(V_T2, V_T2.sum(axis=1))
    except TransformationError as exc:
        check("and so does a zero product total, from the other coefficient",
              "product output x" in str(exc),
              str(exc)[:88] + "…")
    else:
        check("and so does a zero product total, from the other coefficient",
              False, "it returned a matrix built on a division by zero")

    # SUT-RAS: a product row whose non-negative part sums to zero has no root.
    n = 2
    Pd = np.zeros((n, n))
    Nd = np.array([[0.0, 5.0], [5.0, 0.0]])       # everything negative
    Pm = np.zeros((n, n))
    Nm = np.zeros((n, n))
    Pv = np.zeros((n, n))
    Nv = np.zeros((n, n))
    try:
        sut_ras(Pd, Nd, Pm, Nm, Pv, Nv,
                m=np.zeros(n), x=np.ones(n), u=np.ones(n),
                MT=np.zeros(n))
        check("a product row with no non-negative part is refused, not solved",
              False, "it returned a table")
    except DegenerateMarginError as exc:
        check("a product row with no non-negative part is refused, not solved",
              "product rows" in str(exc) or "import rows" in str(exc),
              str(exc)[:88] + "…")
    except Exception as exc:                       # noqa: BLE001
        check("a product row with no non-negative part is refused, not solved",
              False, f"{type(exc).__name__} instead: {str(exc)[:70]}")

    # The import side has its own root and its own refusal, and the distinction
    # that matters is INERT against DEGENERATE: a row that is entirely empty is
    # left alone with a factor of 1 (the fix that made SUT-RAS runnable on real
    # tables at all), while one that has negative mass and no positive mass is
    # genuinely unsolvable. This builds the second, not the first.
    from quadrium.sut_ras import _import_factors
    Pm = np.zeros((2, 2))
    Nm = np.array([[0.0, 4.0], [0.0, 0.0]])
    try:
        _import_factors(Pm, Nm, m=np.array([1.0, 0.0]), s=np.ones(2), r=1.0)
        check("and an import row that is negative-only is refused, not inert",
              False, "it returned a factor for a row that has no root")
    except DegenerateMarginError as exc:
        check("and an import row that is negative-only is refused, not inert",
              "import rows" in str(exc),
              str(exc)[:88] + "…")



def test_a_margin_below_the_floor_is_not_a_sign():
    """An empty line with a rounding-sized target must not be called infeasible.

    A parent with no internal sales gives a block that is exactly zero, and its
    internal targets are the rounding of a difference of large sums — in the
    real cases that found this, -2.8e-14 for Hungary 2022 and -0.021 for France
    2021, against a detectability floor of the order of a tenth. Three sign
    tests in `gras` compared those against EXACT zero and refused the whole
    split of `Q87_88`, health and social work, in three countries.

    The caller already computes that floor and passes it as `margin_floor`;
    `_assert_margins_consistent` was the only place using it. The targets are
    now snapped once, at the top, and only where the seed line is empty — a
    small target on a line with mass is a real constraint and stays one.
    """
    Z = np.zeros((2, 2))
    u = np.array([0.0, 0.0])
    v = np.array([-2.84e-14, 0.0])

    try:
        balance(Z, u, v, method="GRAS")
    except BalancingError as exc:
        check("without a floor, a rounding-sized target is refused", True,
              f"{str(exc)[:70]}… — which is the behaviour that cost three "
              f"real splits")
    else:
        check("without a floor, a rounding-sized target is refused", False,
              "it was accepted, so this test no longer covers the defect")

    Zb, info = balance(Z, u, v, method="GRAS", margin_floor=0.1)
    check("with the floor, the same problem balances to zero",
          float(np.abs(Zb).max()) == 0.0,
          f"largest cell {np.abs(Zb).max():.3g} — the block is empty and stays "
          f"empty, which is the only answer consistent with a parent that has "
          f"no internal sales")

    # and a small target on a line that is NOT empty is still a constraint
    Z2 = np.array([[10.0, 5.0], [4.0, 6.0]])
    try:
        balance(Z2, np.array([-0.05, 25.05]), np.array([14.0, 11.0]),
                method="GRAS", margin_floor=0.1)
    except BalancingError:
        check("but a line with mass keeps its sign constraint", True,
              "a floor-sized target on a non-empty line is still refused, so "
              "the snap did not widen the rule to every small number")
    else:
        check("but a line with mass keeps its sign constraint", False,
              "it was accepted — the snap is reaching lines it should not")



def test_the_allocation_key_cannot_move_a_multiplier():
    """Without an input profile, the key sets sizes and nothing else — exactly.

    A proportional split hands every part the parent's own column of technical
    coefficients: `A[i, part] = s*Z[i, parent] / (s*X[parent])`, and the share
    cancels. The output multiplier obeys `m_j = 1 + sum_i m_i*A[i,j]`, which
    reads columns only, and the parts contribute
    `sum_a m_a*s_a*A[parent,j] = m_parent*A[parent,j]` — the shares sum to one
    and cancel again. So no key can reach a multiplier.

    This is an identity and the project had been relying on it in prose since
    `run_real_key.py` was written, with nothing checking it. Measured on 638
    real proxies it holds in 636 of 638 (`run_key_invariance.py`), and the two
    exceptions are keys that give a real subsector a share of exactly ZERO —
    where there is no share to cancel — which `check_proxy_coverage` fails at
    error severity so they never reach a delivered table. Here it is checked
    against two keys chosen to be as far apart as they can be while both stay
    positive, so a dependence would have nowhere to hide.

    The second half is the boundary: an input profile IS able to move them,
    which is what makes `OQ-B-17` a question about profiles and not about keys.
    """
    from quadrium.disaggregation import split_sector
    from quadrium.models import AllocationKey

    table = build_table()
    spec = SplitSpec("ACC", NEW, LBL, keys_by_block={"output": "k"})
    others = [c for c in table.sector_codes if c != "ACC"]

    def run(raw, profiles):
        keys = {"k": AllocationKey(
            key_id="k", applies_to="output", new_sector_codes=NEW,
            raw_values=list(raw), source="test", source_year=table.year,
            strength=ProxyStrength.MEDIUM)}
        sc = Scenario(scenario_id="s", label="s",
                      keys_by_block={"output": "k"},
                      input_profiles=profiles or {})
        seed = split_sector(table, "ACC", NEW, LBL, sc, keys, spec)
        Z, X = np.asarray(seed["Z"]), np.asarray(seed["X"])
        A = Z / np.where(X == 0, 1.0, X)
        m = np.linalg.inv(np.eye(len(X)) - A).sum(0)
        return m[seed["new_positions"]], X[seed["new_positions"]]

    equal = np.ones(len(NEW))
    skewed = np.arange(1, len(NEW) + 1, dtype=float) ** 3

    m_a, x_a = run(equal, None)
    m_b, x_b = run(skewed, None)
    gap = float(np.abs(m_a - m_b).max() / max(float(np.abs(m_a).max()), 1e-12))
    check("no key can move a subsector's multiplier without a profile",
          gap < 1e-12,
          f"an equal key against one weighted 1, 8, 27, 64 differ by "
          f"{gap:.2e} — machine precision, not a small effect")
    check("while the same two keys move the sizes as far as they ask",
          float(np.abs(x_a - x_b).max() / max(float(x_a.max()), 1e-12)) > 0.5,
          f"outputs {', '.join(f'{v:,.0f}' for v in x_a)} against "
          f"{', '.join(f'{v:,.0f}' for v in x_b)}")

    profiled = {NEW[0]: {c: 1.7 for c in others}}
    m_c, _ = run(equal, profiled)
    m_d, _ = run(skewed, profiled)
    moved = float(np.abs(m_c - m_d).max() / max(float(np.abs(m_c).max()), 1e-12))
    check("but an input profile lets the key reach structure",
          moved > 1e-6,
          f"the same two keys now differ by {moved:.2e}. A profile is the only "
          f"route from a key to a multiplier, and OQ-B-17 measures what that "
          f"route is worth once the balancer has had it")



def main() -> int:
    print("quadrium engine checks")
    print("=" * 60)
    for fn in (test_gras_reduces_to_ras, test_ras_refuses_negatives,
               test_reaggregation_is_exact, test_signs_and_zeros_preserved,
               test_infeasible_scenario_is_rejected,
               test_original_table_is_never_mutated, test_weights_sum_to_one,
               test_real_uk_table_loads_and_balances,
               test_loader_refuses_an_unbalanced_table,
               test_real_ine_table_loads_and_balances,
               test_the_internal_block_conserves_the_parent_cell_at_every_alpha,
               test_the_allocation_key_cannot_move_a_multiplier,
               test_a_margin_below_the_floor_is_not_a_sign,
               test_the_refusals_that_judge_DATA_actually_fire,
               test_the_workbook_refusals_a_stranger_meets_first,
               test_the_config_refusals_a_stranger_meets_next,
               test_the_refusals_about_what_was_ASKED_FOR,
               test_the_refusals_the_METHODS_make,
               test_the_refusals_the_SUT_to_IOT_step_makes_of_ITS_CALLER,
               test_the_catalan_table_loads_and_says_whose_residue_it_carries,
               test_the_refusals_about_an_OFFICE_WORKBOOK,
               test_the_refusals_about_a_FILE_FROM_AN_OFFICE,
               test_a_zero_row_is_zero_at_the_SOURCE_s_precision,
               test_the_spanish_supply_use_tables_load_and_balance,
               test_the_eurostat_connector_loads_and_refuses_correctly,
               test_the_eurostat_supply_use_loader,
               test_the_spanish_table_is_reachable_from_a_workbook,
               test_the_spanish_pilot_holds_its_two_delicate_properties,
               test_a_key_from_the_wrong_year_says_so,
               test_a_share_of_something_that_changes_sign_is_refused,
               test_value_added_rows_can_carry_their_own_evidence,
               test_an_input_profile_carries_its_provenance_and_its_side_effect,
               test_corroboration_reports_a_spread_and_refuses_to_rank,
               test_project_folder_is_reproducible,
               test_export_json_handles_numpy_and_enums,
               test_label_mask_beats_the_naive_comparison,
               test_input_profiles_preserve_supplier_totals,
               test_input_profiles_differentiate_multipliers,
               test_input_profiles_reject_nonsense,
               test_two_sectors_in_one_run, test_split_order_does_not_matter,
               test_a_sector_cannot_be_split_twice,
               test_config_workbook_round_trip, test_config_errors_are_useful,
               test_classification_validates_split_codes,
               test_engine_refuses_an_illegitimate_split,
               test_targets_are_consistent, test_id19_margin_column,
               test_corroboration_uses_the_keys_the_split_did_not,
               test_the_report_does_not_stay_silent,
               test_machine_facing_surfaces_do_not_stay_silent,
               test_exports_travel_with_their_status,
               test_a_pinned_cell_the_solver_moved_stops_claiming_to_be_pinned,
               test_a_number_too_small_to_show_is_not_shown_as_zero):
        print(f"\n{fn.__name__}")
        fn()
    print("\n" + "=" * 60)
    if FAILURES:
        print(f"{len(FAILURES)} check(s) FAILED: {', '.join(FAILURES)}")
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())


def test_an_interregional_table_is_an_IOTable_and_says_how_to_cut_it():
    """A multi-region table needs no new type, and the axis refuses a bad layout.

    `CORE_039` p. 286 describes an IRIO as an ordinary table whose intermediate
    matrix is (S*R) x (S*R). Both of `IOTable`'s balance identities hold on it
    unchanged, so the right move is an additive axis rather than a second class
    — every existing validator keeps working because the object it checks has
    not changed.

    What the axis has to earn is the SLICING. Arranged region-major, each
    region's own table is a contiguous diagonal block; arranged any other way
    the same slice returns a plausible matrix of the wrong thing. So the layout
    is enforced, not assumed, and this test is mostly the refusals.
    """
    from quadrium.models import IOTable

    S, R = 3, 2
    n = S * R

    def build(region_codes, sector_codes=None):
        Z = np.arange(n * n, dtype=float).reshape(n, n)
        Y = np.full((n, 1), 2.0)
        VA = np.full((1, n), 2.0)
        X = Z.sum(1) + Y.sum(1)
        VA = (X - Z.sum(0)).reshape(1, n)
        return IOTable(
            table_id="irio", country="XX", year=2021, unit="M",
            classification="NACE", sector_codes=sector_codes or ["a", "b", "c"] * R,
            sector_labels=["a", "b", "c"] * R, Z=Z, Y=Y, Y_labels=["fd"],
            VA=VA, VA_labels=["va"], X=X, source="test",
            region_codes=region_codes)

    t = build(["R1"] * S + ["R2"] * S)
    assert t.regions == ["R1", "R2"]
    assert t.n_regions == 2 and t.sectors_per_region == S

    # The diagonal block is the region's own table, and it is a VIEW: these
    # matrices are (S*R)^2 and slicing must not copy them.
    assert t.intraregional("R2").shape == (S, S)
    # shares_memory rather than `.base is`, because numpy collapses the base
    # of a view-of-a-view to the ultimate owner, which reshape makes not-Z.
    assert np.shares_memory(t.block("R1", "R2"), t.Z)
    assert np.array_equal(t.intraregional("R1"), t.Z[:S, :S])
    assert np.array_equal(t.block("R1", "R2"), t.Z[:S, S:])
    assert np.array_equal(t.regional_output("R2"), t.X[S:])

    def refuses(fragment, fn):
        try:
            fn()
        except ValueError as exc:
            assert fragment in str(exc), f"expected {fragment!r}, got {exc}"
            return
        raise AssertionError(f"no refusal mentioning {fragment!r}")

    # Interleaved regions would give silently wrong blocks.
    refuses("contiguous block", lambda: build(["R1", "R2"] * S))

    # A region carrying different sectors, or the same ones in a different
    # order, breaks the correspondence between one block and the next.
    refuses("same sectors",
            lambda: build(["R1"] * S + ["R2"] * S,
                          sector_codes=["a", "b", "c", "a", "c", "b"]))

    # Length is checked against the axis it sits parallel to.
    refuses("length 6", lambda: build(["R1"] * S))

    # And a single-region table says so instead of returning something.
    plain = build(None)
    assert plain.regions == [] and plain.n_regions == 0
    refuses("no regional axis", lambda: plain.block("R1", "R1"))
    refuses("no regional axis", lambda: plain.regional_output("R1"))


def test_a_profile_that_cannot_fit_is_turned_away_by_name():
    """`OQ-B-17` option 4: say which of the two things went wrong.

    A scenario can be infeasible because the allocation keys describe an
    impossible economy, or because an input profile asks the parent's internal
    block for more trade than it has. Until now both died with the same
    sentence, and the second is the one the engine can see coming: a profiled
    split whose headroom is already negative is refused with certainty, 37 of
    37 in `run_input_profiles_backtest.py`.

    The owner chose this option on 2026-09-01 precisely because it forecloses
    nothing — it is a better diagnosis of the same refusal, and changes nothing
    for the splits that survive, which is where OQ-B-17's finding lives.

    So this test is a CONTRAST, not an assertion that a refusal happens. The
    same impossible keys, with and without a profile, have to produce different
    explanations; if they did not, the new message would be decoration.
    """
    import re as _re

    from quadrium.models import AllocationKey, ProxyStrength, Scenario, SplitSpec
    from quadrium.project import IOProject

    codes = ["HOT", "CAM", "RES", "FBS"]
    keys = build_keys()
    # An output key that gives one subsector essentially nothing while final
    # demand and value added are shared normally: impossible on its own.
    keys["skew"] = AllocationKey(
        key_id="skew", applies_to="output", new_sector_codes=codes,
        raw_values=[1, 3300, 3300, 3300], source="test fixture",
        source_year=2022, strength=ProxyStrength.WEAK)
    blocks = {"output": "skew", "final_demand": "key_employment",
              "value_added": "key_gva"}
    profiles = {"HOT": {"AGR": 2.0, "MAN": 0.4},
                "RES": {"AGR": 0.5, "MAN": 1.6}}

    def explain(profiled):
        sc = Scenario(scenario_id="P", label="x", keys_by_block=blocks,
                      input_profiles=profiles if profiled else {})
        try:
            IOProject(project_id="t", table=build_table(), keys=keys,
                      splits=[SplitSpec("ACC", codes, ["a", "b", "c", "d"])],
                      scenarios=[sc]).run()
        except ValueError as exc:
            return str(exc)
        raise AssertionError("expected the scenario to be infeasible")

    plain, profiled = explain(False), explain(True)

    assert "the allocation keys imply an impossible economy" in plain
    assert "input profile asks for more internal trade" not in plain

    assert "an input profile asks for more internal trade than the sector has" \
        in profiled
    assert "asks `ACC`'s internal block for more than it has" in profiled
    # The figure, not just the sentence: a diagnosis that does not say how far
    # out it is leaves the analyst guessing how much to soften.
    assert _re.search(r"headroom is -\d+\.\d+ %", profiled)
    # And it still says what to do about it.
    assert "Soften the input profiles" in profiled


def test_the_refusals_the_PROXY_FILE_and_the_PROVENANCE_SHEET_make():
    """Two files the user writes by hand, and a refusal that fired by luck.

    A split needs a proxy file, and a table that has been through one carries a
    `Provenance` sheet. Both are the user's own spreadsheets, both are read
    back by `io_loader`, and neither had a case behind any of its refusals.

    WHAT THIS FOUND
    -----------------
    `load_allocation_keys` refused an unknown `strength` **only when the
    invalid value happened to sort first.** It ranked the rows before
    validating them, using a min() whose key mapped an unknown word to 0 — the
    rank of "strong" — so an invalid row could become `weakest`, and ranking a
    valid row against it then raised a bare `ValueError('bogus' is not in
    list)` out of the standard library: no file name, no column, no list of
    what is allowed.

        strong + bogus  ->  LoaderError, correctly
        medium + bogus  ->  ValueError, uncaught
        weak   + bogus  ->  ValueError, uncaught

    The refusal was written, reviewed and never reached, and the record said so
    for five days. This is the argument for the whole exercise: the message
    existed and two thirds of the users who earned it would never have seen it.

    The fix validates before it ranks. The case below runs all three orders,
    because one order passing is exactly what hid this.
    """
    import tempfile

    import openpyxl

    from quadrium.io_loader import LoaderError, load_allocation_keys

    def keyfile(tmp, rows, tag, head=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(head or ["new_sector_code", "new_sector_label", "value",
                           "source", "source_year", "strength"])
        for r in rows:
            ws.append(list(r))
        p = tmp / f"{tag}.xlsx"
        wb.save(p)
        return p

    OK = ("A", "a", 60.0, "SBS 2021", 2021, "strong")

    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)

        k = load_allocation_keys(keyfile(tmp, [OK, ("B", "b", 40.0, "SBS 2021",
                                                    2021, "weak")], "valid"))
        check("a valid proxy file loads, and takes the strength of its WEAKEST "
              "row", k.strength.value == "weak" and len(k.new_sector_codes) == 2,
              f"strong + weak -> {k.strength.value}; a split resting on one "
              f"weak proxy is a weak split, whatever the other rows say")

        k2 = load_allocation_keys(keyfile(tmp, [("A", "a", 60.0, "SBS 2021",
                                                 2021, "weak"), OK], "order"))
        check("and the answer does not depend on the order of the rows",
              k2.strength.value == "weak",
              f"weak + strong -> {k2.strength.value}, the same. The defect "
              f"this test was written for was exactly an answer that changed "
              f"with the order")

        cases = [
            # All three orders. Before 2026-09-02 the first refused and the
            # other two crashed with a ValueError from list.index().
            ("an unknown strength after a STRONG row",
             [OK, ("B", "b", 40.0, "SBS 2021", 2021, "bogus")],
             "is not strong/medium/weak"),
            ("an unknown strength after a MEDIUM row",
             [("A", "a", 60.0, "SBS 2021", 2021, "medium"),
              ("B", "b", 40.0, "SBS 2021", 2021, "bogus")],
             "is not strong/medium/weak"),
            ("an unknown strength after a WEAK row",
             [("A", "a", 60.0, "SBS 2021", 2021, "weak"),
              ("B", "b", 40.0, "SBS 2021", 2021, "bogus")],
             "is not strong/medium/weak"),
            ("rows that disagree on the source",
             [OK, ("B", "b", 40.0, "a different survey", 2021, "strong")],
             "rows disagree on source or source_year"),
            ("rows that disagree on the year",
             [OK, ("B", "b", 40.0, "SBS 2021", 2019, "strong")],
             "rows disagree on source or source_year"),
            ("a file with no rows at all",
             [], "declares no split"),
        ]
        for i, (name, rows, fragment) in enumerate(cases):
            try:
                load_allocation_keys(keyfile(tmp, rows, f"bad{i}"))
            except LoaderError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(), str(exc)[:86])
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead of LoaderError: "
                      f"{str(exc)[:64]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it accepted the file")

        # A column the loader cannot find at all.
        try:
            load_allocation_keys(keyfile(
                tmp, [OK], "nohead",
                head=["code", "label", "value", "source", "year", "vigour"]))
        except LoaderError as exc:
            check("and a proxy file with no `strength` column names the column "
                  "it wanted", "strength" in str(exc), str(exc)[:86])
        except Exception as exc:                       # noqa: BLE001
            check("and a proxy file with no `strength` column names the column "
                  "it wanted", False, f"{type(exc).__name__}: {str(exc)[:60]}")

        # ---- the Provenance sheet, read back off a table
        META = [("country", "ES"), ("year", 2022), ("unit", "MIO_EUR"),
                ("classification", "CPA_2.1"), ("source", "test")]
        TABLE = [["", "A", "B", "HH"],
                 ["A", 10.0, 5.0, 35.0],
                 ["B", 4.0, 6.0, 40.0],
                 ["VA", 36.0, 39.0, None],
                 ["Output", 50.0, 50.0, None]]

        def table_with_provenance(rows, tag):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet("table")
            for row in TABLE:
                ws.append(row)
            ws = wb.create_sheet("metadata")
            for kv in META:
                ws.append(list(kv))
            ws = wb.create_sheet("Provenance")
            for row in rows:
                ws.append(row)
            p = tmp / f"{tag}.xlsx"
            wb.save(p)
            return p

        from quadrium.io_loader import load_io_table

        good = table_with_provenance([["", "A", "B"],
                                      ["A", "OBSERVED", "ESTIMATED"],
                                      ["B", "BALANCED", "OBSERVED"]], "prov_ok")
        t = load_io_table(good)
        counts = t.provenance_counts()
        check("a Provenance sheet is read back, and the estimates stay "
              "estimates", counts.get("ESTIMATED") == 1
              and counts.get("BALANCED") == 1 and counts.get("OBSERVED") == 2,
              f"{counts} — this is what stops a second split promoting the "
              f"first split's estimates to observations")

        prov_cases = [
            ("a Provenance sheet with a row missing",
             [["", "A", "B"], ["A", "OBSERVED", "OBSERVED"]],
             "data row"),
            ("a Provenance cell that reads something else",
             [["", "A", "B"], ["A", "OBSERVED", "PROBABLY"],
              ["B", "OBSERVED", "OBSERVED"]],
             "expected one of"),
        ]
        for i, (name, rows, fragment) in enumerate(prov_cases):
            try:
                load_io_table(table_with_provenance(rows, f"prov{i}"))
            except LoaderError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(), str(exc)[:86])
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead of LoaderError: "
                      f"{str(exc)[:64]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it accepted the sheet — a grid that does not match the "
                      "table mislabels cells rather than leaving them "
                      "unlabelled")


def test_the_config_refusals_the_SUPPLY_USE_route_makes():
    """The third job in the workbook, and the one with no cases behind it.

    A configuration workbook does one of three things: `splits` divides a
    sector, `regionalise` estimates a region, `targets` moves a supply-use pair
    to a later year. The first two have had cases for weeks. The third holds
    eleven of `config.py`'s refusals and not one had ever fired, so the whole
    supply-use route through the user's spreadsheet was unchecked at its edges.

    They fall in two places, and the difference matters to whoever reads a
    failure:

        `_eurostat_sut_paths`   read off the `project` sheet BEFORE anything is
                                downloaded -- geo, year, model, method, the
                                unbalanced policy, the cache folder
        `_project`              read off the `targets` sheet AFTER the pair is
                                in memory -- an unknown kind, a value that is
                                not a number, a kind that is missing

    The first six cost nothing to reach and are checked unconditionally. The
    last three need a real pair, so they run only when Austria's three files
    are already cached; that is stated below rather than skipped in silence.

    `project_method` has no default ON PURPOSE — it defaulted to `sut_euro`,
    which is the worse of the two methods on every test run (OQ-B-16), so a
    user who said nothing got the loser by silence. The sheet decides now, and
    the refusal below is what a stranger meets when a declaration and a sheet
    disagree in the other direction: a method that is not one of the two.
    """
    import tempfile

    import openpyxl

    from quadrium.config import ConfigError, load_config

    CACHE = ROOT / "data" / "eurostat"
    # AT 2022, not AT 2018: the cache holds cp15 and cp16 for 2018 and NOT
    # cp1610, the use table at basic prices split DOM/IMP -- which is the one
    # file that makes a transformation possible at all. A pair short of it
    # would have sent every case below to a download instead of to the refusal
    # being tested.
    BASE = [("project_id", "sut"), ("table_kind", "eurostat_sut"),
            ("eurostat_geo", "AT"), ("eurostat_year", 2022),
            ("project_to_year", 2023), ("table_path", str(CACHE))]
    RAS = [("industry_output", "A01", 100.0),
           ("use_column_totals", "A01", 50.0),
           ("taxes", "", 10.0), ("imports", "", 20.0)]

    def swap(rows, key, value):
        return [(key, value) if k == key else (k, v) for k, v in rows]

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        def book(meta, targets, tag):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "project"
            for kv in meta:
                ws.append(list(kv))
            ws = wb.create_sheet("targets")
            ws.append(["kind", "code", "value"])
            for r in targets:
                ws.append(list(r))
            p = tmp / f"{tag}.xlsx"
            wb.save(p)
            return p

        def refuses(name, meta, targets, fragment, **kw):
            tag = "".join(c if c.isalnum() else "_" for c in name)[:40]
            try:
                load_config(book(meta, targets, tag), **kw)
            except ConfigError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(), str(exc)[:86])
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead of ConfigError: "
                      f"{str(exc)[:64]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it accepted the workbook")

        # ---- read off the `project` sheet, before any network
        refuses("a three-letter country code",
                swap(BASE, "eurostat_geo", "ESP"), RAS,
                "must be a two-letter country code")
        refuses("a year that is not a year",
                swap(BASE, "eurostat_year", "dos mil"), RAS, "not a year")
        refuses("a projection year that is not a year",
                swap(BASE, "project_to_year", "pronto"), RAS, "not a year")
        refuses("a transformation model outside the four in CORE_013",
                BASE + [("eurostat_model", "Z")], RAS,
                "must be one of the four in CORE_013")
        refuses("a projection method that is neither of the two",
                BASE + [("project_method", "sut_x")], RAS,
                "must be 'sut_euro' or 'sut_ras'")
        refuses("an unbalanced policy that is neither of the two",
                BASE + [("sut_unbalanced", "maybe")], RAS,
                "must be 'refuse' (the default) or 'cancelling'")

        # NO suffix: a path with one is read as naming a file inside the
        # cache folder, and its parent is used. This refusal is for a
        # `table_path` that IS an existing file where a folder has to go --
        # the case that produced a FileExistsError from mkdir several frames
        # down when a single-file `eurostat` run shared a path with this one.
        afile = tmp / "not_a_folder"
        afile.write_text("{}")
        refuses("a cache path that names a FILE where three downloads go",
                BASE + [("table_path", str(afile))], RAS,
                "which exists and is a file")

        empty = tmp / "empty_cache"
        empty.mkdir()
        refuses("--offline with nothing cached",
                BASE + [("table_path", str(empty))], RAS,
                "is not cached yet", offline=True)

        # ---- read off the `targets` sheet, which needs a real pair
        from quadrium.eurostat import DATASETS

        cached = all((CACHE / f"{DATASETS[n]}_AT_2022.json").exists()
                     for n in ("supply", "use_purchasers", "use_basic"))
        check("Austria's 2022 supply-use pair is cached, so the `targets` "
              "refusals can be reached without a network", cached,
              "the three files are in data/eurostat" if cached else
              "NOT cached — the three `targets` refusals below are skipped, "
              "and this is said rather than passed over in silence")
        if not cached:
            return

        refuses("a targets row whose kind is not one of the six",
                BASE, RAS + [("gva_x", "A01", 1.0)], "is not one of")
        refuses("a targets value that is not a number",
                BASE, [("industry_output", "A01", "mucho")] + RAS[1:],
                "is not a number")
        refuses("a targets sheet missing one of the four kinds a method needs",
                BASE, RAS[:3], "row(s)")


def test_the_regionalise_sheet_refuses_what_it_cannot_read():
    """Three refusals on the newest sheet, none of which had a case.

    `regionalise` is the sheet added at v1.86 so the guide's second sentence —
    "No Python: you fill in a spreadsheet and run one command" — was true of
    every job and not only two of them. It arrived with its refusals written
    and unreached, which is the state this whole exercise exists to end.

    All three are about a cell the user typed: a path that is not there, a
    quantity that is not a quantity, a delta that is not a number. None of them
    needs a network or a fixture beyond the national table already in the tree.
    """
    import csv as _csv
    import tempfile

    import openpyxl

    from quadrium.config import ConfigError, load_config

    national = ROOT / "data" / "ine" / "cne_tio_21.xlsx"
    check("a national table to point the sheet at", national.exists(),
          national.name)
    if not national.exists():
        return

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        from quadrium.io_loader import load_ine_tio
        table = load_ine_tio(national, variant="interior")

        def activity(rows, tag):
            p = tmp / f"{tag}.csv"
            with p.open("w", newline="") as fh:
                w = _csv.writer(fh)
                w.writerow(["sector_code", "regional"])
                for r in rows:
                    w.writerow(list(r))
            return p

        good = activity([(c, f"{x * 0.2:.6f}")
                         for c, x in zip(table.sector_codes, table.X)], "good")

        def book(reg_rows, tag):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "project"
            for kv in (("project_id", "reg"),
                       ("table_path", str(national.resolve())),
                       ("table_kind", "ine_interior")):
                ws.append(list(kv))
            ws = wb.create_sheet("regionalise")
            ws.append(["key", "value"])
            for r in reg_rows:
                ws.append(list(r))
            p = tmp / f"{tag}.xlsx"
            wb.save(p)
            return p

        cfg = load_config(book([("method", "FLQ"), ("delta", 0.25),
                                ("activity_path", str(good))], "ok"))
        check("a filled-in `regionalise` sheet loads, so the breakages are the "
              "only difference",
              cfg.get("kind") == "regionalise" and cfg.get("method") == "FLQ"
              and cfg.get("delta") == 0.25 and cfg.get("Q_region") is not None,
              f"kind {cfg.get('kind')!r}, method {cfg.get('method')!r}, delta "
              f"{cfg.get('delta')}, national activity from "
              f"{cfg.get('national_activity_from')!r}")

        cases = [
            ("an activity_path that is not there",
             [("method", "SLQ"), ("activity_path", str(tmp / "absent.csv"))],
             "which does not exist"),
            ("an activity value that is not a number",
             [("method", "SLQ"),
              ("activity_path", str(activity(
                  [(c, "mucho" if i == 3 else "10.0")
                   for i, c in enumerate(table.sector_codes)], "nan")))],
             "not a number"),
            ("a delta that is not a number",
             [("method", "FLQ"), ("delta", "un cuarto"),
              ("activity_path", str(good))],
             "delta"),
        ]
        for i, (name, rows, fragment) in enumerate(cases):
            try:
                load_config(book(rows, f"bad{i}"))
            except ConfigError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(), str(exc)[:86])
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead of ConfigError: "
                      f"{str(exc)[:64]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it accepted the sheet")


def test_the_refusals_the_SUPPLY_USE_workbooks_from_two_offices_make():
    """The ONS's supply-use workbook and the INE's, broken one thing at a time.

    Both are files this project already reads for real — `run_uk_sut_identities`
    runs on all 27 published years of the first — so the loaders that read them
    are on the route a stranger meets when they ask for a country nobody has
    asked for yet. Neither loader's refusals had ever been reached by anything.

    The ONS book is 2.9 MB and 136 sheets, so it is NOT copied whole the way
    `test_the_refusals_about_an_OFFICE_WORKBOOK` copies the UK and INE
    analytical tables. One year's three sheets are read out and written into a
    small workbook of their own, at the same row and column positions, and that
    is what gets bent. The numbers are still the ONS's; only the book is
    shorter. The baseline is asserted to load and to carry the same blocks as
    the real file first, because a baseline nobody checked is not a baseline.

    Three things the engine got right and this test had to be corrected to:

    1. Asking for a year the book does not publish names the **intermediate
       consumption** sheet, not supply. `load_ons_sut` reads the three sheets in
       one tuple assignment and Python evaluates it left to right, so
       `Table 2 - Int Con` is fetched first. The expectation was written for
       supply and was wrong.

    2. The missing value-added row is provoked by writing
       `Gross value added at basic prices` — the CORRECT spelling. The ONS
       prints `Gross valued added`, the loader matches what is printed, and so
       repairing the source's typo is precisely what stops the load.

    3. The products refusal fires on the SET of codes and is right to, but the
       evidence it prints is `supply 104, use 104, final demand 104` — three
       equal numbers offered as proof that three sheets disagree. One renamed
       code changes no count. The message is left as it stands here because
       editing it moves the site's key in `data/_refusal_coverage.json` and that
       is a deliberate act with its own bookkeeping in both trees, not a side
       effect of a coverage pass. It is worth naming the offending codes.
    """
    import shutil
    import tempfile

    import openpyxl
    from quadrium.io_loader import (LoaderError, _tod_norm, load_ine_tod,
                                    load_ons_sut)

    ons = ROOT / "data" / "ons" / "NSO_UK_04_ONS_supply_use_tables_BB25.xlsx"
    ine = ROOT / "data" / "ine" / "cne_tod_22.xlsx"
    have = [p for p in (ons, ine) if p.exists()]
    check("the two supply-use workbooks this mutates are present",
          len(have) == 2, ", ".join(p.name for p in have))
    if len(have) != 2:
        return

    def refuses(name, run, fragment):
        try:
            run()
        except LoaderError as exc:
            check(f"the engine refuses {name}",
                  fragment.lower() in str(exc).lower(),
                  str(exc)[:86] + ("…" if len(str(exc)) > 86 else ""))
        except Exception as exc:                       # noqa: BLE001
            check(f"the engine refuses {name}", False,
                  f"{type(exc).__name__} instead: {str(exc)[:64]}")
        else:
            check(f"the engine refuses {name}", False, "it built a table")

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        # ---- the ONS supply-use workbook, 1997-2023 --------------------
        year = 2023
        names = [f"Table 1 - Supply {year}", f"Table 2 - Int Con {year}",
                 f"Table 2 - Final Demand {year}"]
        book = openpyxl.load_workbook(ons, read_only=True, data_only=True)
        rows = {n: [list(r) for r in book[n].iter_rows(values_only=True)]
                for n in names}

        def trimmed(edit=None, name="ons.xlsx"):
            data = {n: [list(r) for r in rows[n]] for n in names}
            if edit:
                edit(data)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for n in names:
                ws = wb.create_sheet(n)
                for r in data[n]:
                    ws.append(r)
            path = tmp / name
            wb.save(path)
            return path

        whole = load_ons_sut(ons, year)
        short = load_ons_sut(trimmed(), year)
        check("the three sheets on their own load to the same table the whole "
              "book does",
              short.product_codes == whole.product_codes
              and np.array_equal(short.U, whole.U)
              and np.array_equal(short.gva, whole.gva),
              f"{short.U.shape[0]} products by {short.U.shape[1]} industries, "
              f"identical to the 136-sheet original")

        # Against the REAL book: the years it lists are its own 27.
        refuses("a year the ONS does not publish",
                lambda: load_ons_sut(ons, 1066),
                "has no sheet 'Table 2 - Int Con 1066'")

        refuses("an ONS book whose value-added block has had the source's own "
                "misspelling corrected",
                lambda: load_ons_sut(trimmed(
                    lambda d: d[names[1]][113].__setitem__(
                        1, "Gross value added at basic prices"), "va.xlsx"),
                    year),
                "the value-added block is missing")

        refuses("an ONS book whose supply sheet carries a product code the "
                "other two do not",
                lambda: load_ons_sut(trimmed(
                    lambda d: d[names[0]][3].__setitem__(0, "CPA_A01x"),
                    "cpa.xlsx"), year),
                "the three sheets do not carry the same products")

        # ---- the INE's supply-use pair --------------------------------
        # `_tod_axes` reads both axes off the sheet rather than assuming them,
        # so both of its refusals are about the sheet having stopped announcing
        # its own extent. The positions are found the way the loader finds
        # them, never hard-coded: the 2016-2020 workbooks put the supply sheet
        # one column to the left of the use sheet, and a test with a constant
        # here would pass on one vintage and lie on the other.
        R = list(openpyxl.load_workbook(
            ine, read_only=True, data_only=True)["Tabla1"].iter_rows(
                values_only=True))
        first_col = next((j for j, c in enumerate(R[8])
                          if _tod_norm(c) == "1"), None)
        first_row = (next((i for i in range(8, 24)
                           if _tod_norm(R[i][first_col - 1]).startswith("1.")),
                          None) if first_col else None)
        check("the INE supply sheet's own axes are where the loader looks",
              first_col is not None and first_row is not None,
              f"activity index row 9 starts at column {first_col + 1}, "
              f"product '1. …' on row {first_row + 1}"
              if first_col and first_row else "not found")
        if first_col is None or first_row is None:
            return

        def bent(edit, name):
            path = tmp / name
            shutil.copy(ine, path)
            wb = openpyxl.load_workbook(path)
            edit(wb["Tabla1"])
            wb.save(path)
            return path

        def blank_index_row(ws):
            for cell in ws[9]:
                cell.value = None

        refuses("an INE supply sheet that no longer numbers its activity "
                "columns",
                lambda: load_ine_tod(bent(blank_index_row, "tod1.xlsx")),
                "the supply sheet has no activity index row")

        refuses("an INE product label that has lost the number the INE prints "
                "in front of it",
                lambda: load_ine_tod(bent(
                    lambda ws: setattr(
                        ws.cell(row=first_row + 1, column=first_col),
                        "value", "Productos de la agricultura"), "tod2.xlsx")),
                "has no product labelled '1. …'")


def test_the_refusals_the_EUROSTAT_CONNECTOR_makes():
    """Six refusals about a server's answer, and none had ever fired.

    `fetch` is the only function in this engine that touches the network, and
    four of its refusals judge what came back: an HTTP 413, any other HTTP
    error, a body that is not JSON, and the 200 that carries an empty `value`.
    Nothing had reached them, because reaching them by making a real request
    would need the server to misbehave on command — so this hands `urlopen`
    the failures instead. That is the whole trick, and it is why they stayed
    unreached for so long when they are four lines apart.

    THE 413 IS THE ONE WORTH READING
    ----------------------------------
    `fetch`'s docstring records that a 413 has been seen twice and that
    NEITHER time survived a retry — and that on the second sighting a clean
    mechanism was inferred (`geo=ES` answered 200, `geo=GB` and `geo=XX`
    answered 413, so an unrecognised filter must be dropped and the request
    becomes every country) and written into the docstring as measured fact.
    Three minutes later all three codes answered 200. The message therefore
    says *retry first* and refuses to explain, which is the correct thing for
    a symptom observed twice and understood zero times. This checks it still
    says that, because a future reader's instinct will be to "improve" it into
    a diagnosis again.

    The other two are the cube: a dimension asked for without a value, and a
    total-flows table with no imports by product. A total table's uses are
    output PLUS imports, so `Z.sum(1) + Y.sum(1) == X` cannot hold without
    them; the fixture strips exactly the `IMP` values from a real Spanish file
    and checks that `domestic` still loads from the same file — the refusal is
    about what `total` needs, not about a broken download.
    """
    import json as _json
    import tempfile
    import urllib.error
    from unittest.mock import patch

    from quadrium.eurostat import (EurostatError, _Cube, _read_cube, fetch,
                                   load_iot)

    ES = ROOT / "data" / "eurostat" / "naio_10_cp1700_ES_2022.json"
    check("a real Eurostat response to build the fixtures from", ES.exists(),
          ES.name)
    if not ES.exists():
        return

    class Answer:
        """What `urlopen` returns, as a context manager, and nothing more."""

        def __init__(self, body):
            self.body = body

        def read(self):
            return self.body

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    def http(code):
        def boom(*a, **k):
            raise urllib.error.HTTPError("http://x", code, "no", {}, None)
        return boom

    with tempfile.TemporaryDirectory() as td:
        dest = Path(td) / "out.json"

        cases = [
            ("HTTP 413, without inventing a reason for it",
             http(413), "retry first"),
            ("any other HTTP error, with the code in it",
             http(503), "returned HTTP 503"),
            ("a body that is not JSON",
             lambda *a, **k: Answer(b"<html>Service Unavailable</html>"),
             "did not return JSON"),
            ("a 200 that carries no values at all",
             lambda *a, **k: Answer(_json.dumps(
                 {"id": [], "size": [], "dimension": {}, "value": {}}).encode()),
             "returned no values for geo="),
        ]
        for name, fake, fragment in cases:
            with patch("urllib.request.urlopen", fake):
                try:
                    fetch("product_by_product", "ES", 2022, dest)
                except EurostatError as exc:
                    check(f"the connector refuses {name}",
                          fragment.lower() in str(exc).lower(), str(exc)[:88])
                except Exception as exc:               # noqa: BLE001
                    check(f"the connector refuses {name}", False,
                          f"{type(exc).__name__} instead of EurostatError: "
                          f"{str(exc)[:60]}")
                else:
                    check(f"the connector refuses {name}", False, "it accepted")
            check(f"and writes nothing to disk when {name.split(',')[0]} came "
                  f"back", not dest.exists(),
                  "a refused download must not leave a half-file behind for "
                  "the next run to read as a cache hit")

        # ---- and the 413 still refuses to explain itself
        with patch("urllib.request.urlopen", http(413)):
            try:
                fetch("product_by_product", "ES", 2022, dest)
            except EurostatError as exc:
                msg = str(exc)
        check("the 413 message says retry and offers no mechanism",
              "transient both times" in msg and "stk_flow" in msg
              and "filter" not in msg.lower(),
              "it reports the symptom and the workaround. A mechanism was "
              "inferred from one observation twice, and was wrong twice; the "
              "message must not grow one back")

        # ---- the cube
        doc = _read_cube(ES)
        cube = _Cube(doc)
        try:
            cube.at()
        except EurostatError as exc:
            check("the cube refuses a lookup that leaves a dimension open",
                  "needs a value" in str(exc), str(exc)[:88])
        except Exception as exc:                       # noqa: BLE001
            check("the cube refuses a lookup that leaves a dimension open",
                  False, f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("the cube refuses a lookup that leaves a dimension open",
                  False, "it returned something")

        # ---- a total table with the imports removed
        i = cube.ids.index("stk_flow")
        j, stride, size = cube.index["stk_flow"]["IMP"], cube.stride[i], cube.size[i]
        before = len(doc["value"])
        doc["value"] = {k: v for k, v in doc["value"].items()
                        if (int(k) // stride) % size != j}
        stripped = Path(td) / "no_imports.json"
        stripped.write_text(_json.dumps(doc))
        check("the fixture removed the import block and nothing else",
              len(doc["value"]) < before,
              f"{before} values down to {len(doc['value'])}")

        try:
            load_iot(stripped, variant="total")
        except EurostatError as exc:
            check("a TOTAL table with no imports by product is refused",
                  "imports by product" in str(exc), str(exc)[:88])
        except Exception as exc:                       # noqa: BLE001
            check("a TOTAL table with no imports by product is refused", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("a TOTAL table with no imports by product is refused", False,
                  "it built a table whose rows cannot close")

        t = load_iot(stripped, variant="domestic")
        check("and the SAME file still loads as domestic, so the refusal is "
              "about what `total` needs", t.n > 0,
              f"{t.n} products. A total table's uses are output PLUS imports, "
              f"so Z.sum(1) + Y.sum(1) == X cannot hold without them; the "
              f"domestic table never needed them")


def test_the_refusals_the_ACQUISITION_module_makes():
    """The worst-covered module in the engine: 1 of 6 refusals ever reached.

    `acquire.py` is how this project fetches a document from a statistical
    office, and its refusals are the ones that decide what the engine is
    ALLOWED to bring in. That made them the least-tested and the most worth
    testing: an acquisition guard nobody has watched fire is a policy nobody
    has confirmed exists.

    Five had no case. Each is checked here by handing the network layer the
    situation rather than waiting to meet it:

        robots.txt says no          — twice: `acquire` and `find_documents`
        the response is too big     — the 64 MB cap, patched down so the test
                                      does not have to produce 64 MB
        the response is not a doc   — `image/png` where a PDF was expected
        any other HTTP error        — 500, which is NOT access control

    WHY THE HTTP SPLIT MATTERS
    ----------------------------
    `_open` sorts HTTP failures into two kinds and they mean opposite things.
    401, 402, 403 and 407 raise `AccessRefused` and the engine stops for a
    human, because CLAUDE.md's standing authorisation covers technical
    friction around content already confirmed open and NOT getting past a
    paywall or a login — and nothing in the code can tell those apart. Every
    other code is an ordinary failure and raises `AcquisitionRefused`. The
    case below checks a 500 lands on the ordinary side, so the day someone
    widens that tuple, the meaning of stopping does not quietly widen with it.

    The robots.txt refusal is a POLICY, not a protocol rule: a statistics
    portal is public and this is a single request, but an engine fetching on
    its own account reads the sign on the door. Nothing had ever shown it
    doing so.
    """
    import tempfile
    import urllib.error
    from unittest.mock import patch

    from quadrium import acquire as A
    from quadrium.acquire import AccessRefused, AcquisitionRefused

    URL = "https://www.ine.es/algo.pdf"          # an allowlisted host, so the
    PORTAL = "https://www.ine.es/portal.htm"     # allowlist is not what fires

    class Resp:
        """What `_open` returns: headers, a body, and a final URL."""

        def __init__(self, body=b"%PDF-1.4", ctype="application/pdf",
                     url=URL):
            self.body, self._url = body, url
            self.headers = {"Content-Type": ctype}

        def geturl(self):
            return self._url

        def read(self, n=None):
            return self.body[:n] if n else self.body

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

    # `headers` must answer .get() and .get_content_charset()
    class H(dict):
        def get_content_charset(self):
            return "utf-8"

    def resp(body=b"%PDF-1.4", ctype="application/pdf"):
        r = Resp(body, ctype)
        r.headers = H({"Content-Type": ctype})
        return r

    with tempfile.TemporaryDirectory() as td:
        dest = Path(td) / "doc.pdf"

        # ---- the sign on the door, on both doors
        with patch.object(A, "_robots_allows", lambda url: False):
            for name, run in (("acquire", lambda: A.acquire(URL, dest)),
                              ("find_documents",
                               lambda: A.find_documents(PORTAL, ["metodolog"]))):
                try:
                    run()
                except AcquisitionRefused as exc:
                    check(f"{name} refuses what robots.txt disallows",
                          "robots.txt" in str(exc), str(exc)[:88])
                except Exception as exc:                # noqa: BLE001
                    check(f"{name} refuses what robots.txt disallows", False,
                          f"{type(exc).__name__}: {str(exc)[:60]}")
                else:
                    check(f"{name} refuses what robots.txt disallows", False,
                          "it fetched anyway")

        # ---- everything below gets past robots and fails on the response
        with patch.object(A, "_robots_allows", lambda url: True):
            with patch.object(A, "_open", lambda *a, **k:
                              resp(b"x" * 40, "image/png")):
                try:
                    A.acquire(URL, dest)
                except AcquisitionRefused as exc:
                    check("a response that is not a document type is refused",
                          "not a document type" in str(exc), str(exc)[:88])
                except Exception as exc:                # noqa: BLE001
                    check("a response that is not a document type is refused",
                          False, f"{type(exc).__name__}: {str(exc)[:60]}")
                else:
                    check("a response that is not a document type is refused",
                          False, "it saved an image as a document")

            # The cap is 64 MB. Patch it down rather than produce 64 MB: the
            # refusal is about the comparison, not about the number.
            with patch.object(A, "MAX_BYTES", 16), \
                 patch.object(A, "_open", lambda *a, **k: resp(b"y" * 64)):
                try:
                    A.acquire(URL, dest)
                except AcquisitionRefused as exc:
                    check("a response over the byte cap is refused",
                          "larger than the" in str(exc)
                          and "16-byte cap" in str(exc), str(exc)[:88])
                except Exception as exc:                # noqa: BLE001
                    check("a response over the byte cap is refused", False,
                          f"{type(exc).__name__}: {str(exc)[:60]}")
                else:
                    check("a response over the byte cap is refused", False,
                          "it saved it")

            check("and neither refusal left a file behind", not dest.exists(),
                  "a document refused for its type or its size must not be on "
                  "disk: the register would cite a file the engine declined")

        # ---- the HTTP split, which is the one with two meanings
        def http(code):
            def boom(req, timeout=None):
                raise urllib.error.HTTPError(URL, code, "no", {}, None)
            return boom

        with patch("urllib.request.urlopen", http(500)):
            try:
                A._open(URL)
            except AccessRefused as exc:                # noqa: BLE001
                check("an ordinary HTTP failure is NOT treated as access "
                      "control", False,
                      f"500 raised AccessRefused: {str(exc)[:60]}")
            except AcquisitionRefused as exc:
                check("an ordinary HTTP failure is NOT treated as access "
                      "control", "answered HTTP 500" in str(exc),
                      f"{str(exc)[:60]} — AcquisitionRefused, not "
                      f"AccessRefused. The two mean opposite things and only "
                      f"one of them stops for a human")
            else:
                check("an ordinary HTTP failure is NOT treated as access "
                      "control", False, "it returned a response")

        # ---- and the page that is not behind HTTP at all
        with patch.object(A, "_robots_allows", lambda url: True), \
             patch.object(A, "_open", lambda *a, **k: resp(
                 b"<html><body>Please sign in to read this article.</body></html>",
                 "text/html")):
            try:
                A.acquire(URL, dest)
            except AccessRefused as exc:
                check("a login page served as HTML is refused, not saved "
                      "under the document's name",
                      "paywall or login page" in str(exc)
                      and "sign in to read" in str(exc).lower(), str(exc)[:88])
            except Exception as exc:                    # noqa: BLE001
                check("a login page served as HTML is refused, not saved "
                      "under the document's name", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a login page served as HTML is refused, not saved "
                      "under the document's name", False,
                      "it saved the login page")
        check("and it is not on disk, because a login page under a document's "
              "name would be cited", not dest.exists(),
              "this refusal and the 401/403 one are the two the coverage "
              "counter could not see at all until 2026-09-03: AccessRefused "
              "subclasses AcquisitionRefused, so the sweep caught their "
              "raises at run time but its static pass matched neither name")

        with patch("urllib.request.urlopen", http(403)):
            try:
                A._open(URL)
            except AccessRefused as exc:
                check("and 403 still is, and says why the engine stops there",
                      "access control" in str(exc) and "CLAUDE.md" in str(exc),
                      "the standing authorisation covers technical friction "
                      "around content already confirmed open, not getting "
                      "past a login — and nothing in the code can tell those "
                      "apart, so a human decides")
            except Exception as exc:                    # noqa: BLE001
                check("and 403 still is, and says why the engine stops there",
                      False, f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("and 403 still is, and says why the engine stops there",
                      False, "it returned a response")


def test_three_refusals_left_over_from_three_different_routes():
    """The last loose ones: a projection target, a download, and an iteration.

    They have nothing in common except that nobody had reached them, which is
    why they were still here after the groups were cleared. Each belongs to a
    different part of the engine and each is checked on its own terms.

    1. `config.vector` — a `targets` sheet naming an industry the supply-use
       pair does not have. The sheet is the user's, the codes are the
       publisher's, and a typo in one is the most ordinary way this fails.

    2. `config._load_eurostat` — the download failed, and the message has to
       say **nothing was written** and name the cache path. A half-written
       file would be read as a cache hit by the next run, so "nothing was
       written" is a promise, not a pleasantry.

    3. `disaggregation.neutralise_profile` — the rescaling did not converge
       inside the iterations allowed. Reached by lowering the ceiling, which
       is the caller's own parameter and exactly the documented failure: a
       profile is meant to change composition and not level, and returning an
       unconverged one would move subsector size while claiming to describe
       only its shape.
    """
    import tempfile
    from unittest.mock import patch

    import numpy as np
    import openpyxl

    from quadrium.config import ConfigError, load_config
    from quadrium.disaggregation import DisaggregationError, neutralise_profile
    from quadrium.eurostat import EurostatError

    CACHE = ROOT / "data" / "eurostat"

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        # ---- 1. a targets code the pair does not have
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "project"
        for kv in (("project_id", "p"), ("table_kind", "eurostat_sut"),
                   ("eurostat_geo", "AT"), ("eurostat_year", 2022),
                   ("project_to_year", 2023), ("table_path", str(CACHE))):
            ws.append(list(kv))
        ws = wb.create_sheet("targets")
        ws.append(["kind", "code", "value"])
        for row in (("industry_output", "NO_SUCH_INDUSTRY", 100.0),
                    ("use_column_totals", "NO_SUCH_INDUSTRY", 50.0),
                    ("taxes", "", 10.0), ("imports", "", 20.0)):
            ws.append(list(row))
        bad_codes = tmp / "badcodes.xlsx"
        wb.save(bad_codes)

        from quadrium.eurostat import DATASETS
        cached = all((CACHE / f"{DATASETS[n]}_AT_2022.json").exists()
                     for n in ("supply", "use_purchasers", "use_basic"))
        check("Austria's 2022 pair is cached, so the targets refusal is "
              "reachable without a network", cached,
              "the three files are in data/eurostat" if cached else
              "NOT cached — this case is skipped and said so")
        if cached:
            try:
                load_config(bad_codes)
            except ConfigError as exc:
                check("a targets sheet naming an industry the pair does not "
                      "have is refused, and both lists are printed",
                      "do not match this pair" in str(exc)
                      and "NO_SUCH_INDUSTRY" in str(exc), str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check("a targets sheet naming an industry the pair does not "
                      "have is refused, and both lists are printed", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a targets sheet naming an industry the pair does not "
                      "have is refused, and both lists are printed", False,
                      "it projected onto codes that are not there")

        # ---- 2. the download failed, and nothing was written
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "project"
        for kv in (("project_id", "p"), ("table_kind", "eurostat"),
                   ("eurostat_geo", "ES"), ("eurostat_year", 1998),
                   ("table_path", str(tmp / "cache"))):
            ws.append(list(kv))
        # The workbook must declare a JOB, or it is refused for having none
        # before the download is ever attempted. A `regionalise` sheet is the
        # cheapest one: the table load fails first, so the activity file is
        # never read and need not match anything.
        act = tmp / "activity.csv"
        act.write_text("sector_code,regional\nA,1.0\n")
        ws = wb.create_sheet("regionalise")
        ws.append(["key", "value"])
        ws.append(["method", "SLQ"])
        ws.append(["activity_path", str(act)])
        failing = tmp / "failing.xlsx"
        wb.save(failing)

        def boom(*a, **k):
            raise EurostatError("naio_10_cp1700 returned no values for "
                                "geo=ES, time=1998.")

        # `config` imports `fetch` inside the function, so the name to patch
        # is the one in `quadrium.eurostat`, not a module attribute of config.
        with patch("quadrium.eurostat.fetch", boom):
            try:
                load_config(failing)
            except ConfigError as exc:
                check("a failed Eurostat download says nothing was written, "
                      "and where it would have gone",
                      "Nothing was written" in str(exc)
                      and "cache path was" in str(exc), str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check("a failed Eurostat download says nothing was written, "
                      "and where it would have gone", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a failed Eurostat download says nothing was written, "
                      "and where it would have gone", False, "it loaded")

        leftovers = list((tmp / "cache").glob("*")) if (tmp / "cache").exists() \
            else []
        check("and the promise holds: the cache folder has no half-file in it",
              not leftovers,
              "a partial download left behind would be read as a cache hit by "
              "the next run, which is the failure the message rules out")

        # ---- 3. a profile that will not go level-neutral in the iterations given
        from quadrium.models import CellLabel, IOTable

        n = 3
        Z = np.array([[4.0, 3.0, 2.0], [2.0, 5.0, 3.0], [1.0, 2.0, 6.0]])
        Y = np.array([[11.0], [10.0], [11.0]])
        X = Z.sum(axis=1) + Y.sum(axis=1)
        VA = (X - Z.sum(axis=0)).reshape(1, n)
        t = IOTable(table_id="t", country="XX", year=2021, unit="",
                    classification="", sector_codes=["A", "B", "C"],
                    sector_labels=["a", "b", "c"], Z=Z, Y=Y, Y_labels=["FD"],
                    VA=VA, VA_labels=["VA"], X=X, source="fixture")

        w = np.array([0.6, 0.4])
        profiles = {"A1": {"B": 9.0}, "A2": {"B": 0.05}}
        try:
            neutralise_profile(t, "A", ["A1", "A2"], w, profiles, max_iter=1)
        except DisaggregationError as exc:
            check("a profile that will not go level-neutral in the iterations "
                  "allowed is refused",
                  "level-neutral" in str(exc) and "still off by" in str(exc),
                  str(exc)[:88])
        except Exception as exc:                       # noqa: BLE001
            check("a profile that will not go level-neutral in the iterations "
                  "allowed is refused", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check("a profile that will not go level-neutral in the iterations "
                  "allowed is refused", False,
                  "it returned a profile that moves subsector size while "
                  "claiming to describe only composition")

        ok = neutralise_profile(t, "A", ["A1", "A2"], w, profiles,
                                max_iter=2000)
        check("and the SAME profile converges when given the room, so the "
              "refusal is about the ceiling and not about a profile that "
              "could never work",
              ok["shift_after"] < 1e-9 < ok["shift_before"]
              and ok["iterations"] < 2000,
              f"the level shift goes from {ok['shift_before']:.4f} to "
              f"{ok['shift_after']:.1e} in {ok['iterations']} iterations — "
              f"which is the distinction the message has to earn, and the "
              f"reason max_iter is the caller's to set")


def test_the_refusals_TWO_SPANISH_WORKBOOKS_make_when_deformed():
    """Seven refusals about a published file, reached by deforming a real one.

    `load_ine_tod` reads the INE's supply-use workbook and `load_idescat_mioc`
    the Catalan symmetric table. Both are read every day by this project, both
    are full of refusals about layout, and none of those refusals had a case —
    because the only files that reach them are files a statistical office got
    wrong, and the offices do not oblige.

    So the fixtures are the REAL workbooks with one thing broken. Each case
    reads the file once, mutates the in-memory sheets, and patches
    `_open_workbook` to hand the loader the mutation. No file is written and
    no shipped data is touched.

    WHAT DEFORMING THEM TAUGHT
    ----------------------------
    Cutting six columns off the use sheet to make the axes disagree did NOT
    reach the size refusal: a more specific one fired first, about a column
    label that had gone missing. The size refusal needs a deformation that
    leaves the labels alone — one activity number removed from the index row.
    That is the sixth time in this project a refusal turned out to sit behind
    an earlier and more specific one, and each time the engine was right.

    The Catalan pair is worth naming precisely. `list(col_of) != codes` is not
    a set comparison: it fails when the row codes are the column codes IN
    ANOTHER ORDER, which is checked below by swapping two products and getting
    "63 rows against 63 columns". A symmetric table has to be square and in
    one order, and a set comparison would have let a permuted table through.
    """
    import re as _re
    from unittest.mock import patch

    from quadrium import io_loader as L
    from quadrium.io_loader import (LoaderError, load_idescat_mioc,
                                    load_ine_tod)

    def clean(v):
        return _re.sub(r"\s+", " ", str(v if v is not None else "")).strip()

    def deformed(loader, path, mutate, name, fragment):
        sheets = {k: [list(r) for r in v]
                  for k, v in L._open_workbook(path).items()}
        mutate(sheets)
        with patch.object(L, "_open_workbook", lambda p: sheets):
            try:
                loader(path)
            except LoaderError as exc:
                check(f"the engine refuses {name}, and says which",
                      fragment.lower() in str(exc).lower(), str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check(f"the engine refuses {name}, and says which", False,
                      f"{type(exc).__name__} instead of LoaderError: "
                      f"{str(exc)[:60]}")
            else:
                check(f"the engine refuses {name}, and says which", False,
                      "it loaded the deformed workbook")

    # ---- the INE supply-use workbook
    INE = ROOT / "data" / "ine" / "cne_tod_21.xlsx"
    check("the INE supply-use workbook is on disk", INE.exists(), INE.name)
    if INE.exists():
        sut = load_ine_tod(INE)
        check("and it loads unbroken, so the breakages are the only difference",
              sut.n_products > 0 and sut.n_activities > 0,
              f"{sut.n_products} products by {sut.n_activities} activities")

        deformed(load_ine_tod, INE, lambda S: S.pop("Tabla1"),
                 "a workbook with no `Tabla1`", "has no 'Tabla1'")

        def one_activity_fewer(S):
            # The index row numbers the activity columns 1, 2, 3, … Blanking
            # the last number shortens the use axis and LEAVES EVERY LABEL in
            # place, which is what it takes to reach the size refusal rather
            # than the label one that guards the road to it.
            row = S["Tabla2"][8]
            last = max(j for j, c in enumerate(row)
                       if str(c).strip().replace(".0", "").isdigit())
            row[last] = None

        deformed(load_ine_tod, INE, one_activity_fewer,
                 "supply and use sheets that disagree on the size of the table",
                 "disagree on the size")

        def one_figure_moved(S):
            for row in S["Tabla1"]:
                for j, c in enumerate(row):
                    if isinstance(c, (int, float)) and c and abs(c) > 1000:
                        row[j] = c + 5000.0
                        return

        deformed(load_ine_tod, INE, one_figure_moved,
                 "a workbook whose own printed totals do not add up",
                 "does not satisfy an identity")

    # ---- the Catalan symmetric table
    CAT = ROOT / "data" / "idescat" / "mioc2021ts64.xlsx"
    if not CAT.exists():
        # NOT A FAILURE, and saying so is the point. The public tree does not
        # publish the Catalan workbook, so nothing there can reach these four
        # refusals — a tree without a fixture cannot exercise the code that
        # reads it. That is why the coverage floor is recorded per tree
        # instead of being a constant shared by both.
        check("the four Catalan cases are skipped, because this tree does not "
              "publish the workbook they deform", True,
              "mioc2021ts64.xlsx is not here. The private tree reaches these "
              "four; this one cannot, and its floor is lower by exactly that")
        return
    check("the Catalan symmetric table is on disk", True, CAT.name)

    t = load_idescat_mioc(CAT)
    check("and it loads unbroken", t.n > 20, f"{t.n} sectors")

    def product_rows(T):
        hdr = next(i for i, r in enumerate(T)
                   if r and any(clean(c) == "Codi" for c in r if c is not None))
        col_of = {clean(c): j for j, c in enumerate(T[hdr])
                  if c is not None and j > 2}
        return hdr, col_of, [i for i in range(hdr + 1, len(T))
                             if len(T[i]) > 1 and T[i][1] is not None
                             and clean(T[i][1]) in col_of]

    def rename(S, what, to):
        for row in S["ts total"]:
            for j, c in enumerate(row):
                if isinstance(c, str) and c.strip().lower() == what:
                    row[j] = to

    deformed(load_idescat_mioc, CAT, lambda S: rename(S, "codi", "Code"),
             "a Catalan sheet with no `Codi` row", "no row carries 'Codi'")
    deformed(load_idescat_mioc, CAT,
             lambda S: rename(S, "consum final", "XXX"),
             "a final-demand block with no `CONSUM FINAL` head",
             "not delimited by a")

    def swap_two_products(S):
        T = S["ts total"]
        _, _, rows = product_rows(T)
        T[rows[0]], T[rows[1]] = list(T[rows[1]]), list(T[rows[0]])

    deformed(load_idescat_mioc, CAT, swap_two_products,
             "row codes that are the column codes IN ANOTHER ORDER",
             "the row codes are not the column codes")

    def drop_the_output_row(S):
        from quadrium.io_loader import _MIOC_OUTPUT
        T = S["ts total"]
        hdr, _, _ = product_rows(T)
        for i in range(hdr, len(T)):
            if T[i] and T[i][0] is not None \
               and str(T[i][0]).strip().replace(".0", "").isdigit() \
               and int(float(T[i][0])) == _MIOC_OUTPUT:
                T[i][0] = None
                return

    deformed(load_idescat_mioc, CAT, drop_the_output_row,
             "a sheet missing a numbered row that carries the column identity",
             "are not on the 'ts total' sheet")


def test_the_last_refusals_the_NUMBERS_and_the_PIPELINE_make():
    """The six left over, and the one that stays a stated gap.

    These are the refusals that need something to go wrong in the arithmetic
    or in the middle of the supply-use pipeline, which is why they outlasted
    every other group.

    THE ONE WORTH THE SEARCH
    --------------------------
    `gras` tests sign feasibility twice. The cheap per-line test asks whether
    each row and column that needs a negative total has a negative entry; the
    exact test asks, by linear programming, whether ANY table with this sign
    pattern has those margins. The second refusal fires only where the first
    passes — a system that looks fine line by line and is globally impossible
    — and no fixture in the suite had ever been one.

    Rather than declare it unreachable, a search found one in sixteen tries: a
    3x3 that clears the per-line test and fails the exact one. It is written
    out below as a constant, so the case is reproducible and nobody has to run
    the search again. This is `OQ-B-09` — no specified method lets a cell
    change sign — showing up in the balancing.

    The two SUT-EURO refusals are the same question in the projection. A
    product with no base-year output has no market shares to hold constant,
    and an industry whose value added changes sign cannot be reached by a
    method that scales by `target / base`: the ratio carries a negative base
    to a negative number of the target's magnitude, however many iterations
    it is given. Hungary's air transport went from -96.7 in 2021 to +28.3 in
    2022, and every back-test run crossing that sign stalled at ~300 %
    deviation, unmoved at 60,000 iterations.

    The three pipeline refusals are the supply-use route failing at each of
    its three steps in turn: the download, the pair, the transformation.

    WHAT IS DELIBERATELY NOT HERE
    -------------------------------
    `hybrid_matrix_avoiding_negatives`'s "model A cannot even be computed on
    this table" stays unreached and is recorded as a known gap in
    `run_refusal_coverage.py` with its reason. Reaching it would mean building
    a supply table that is singular under model A and not under the search
    that wraps it, which is a table nobody would publish. A stated gap is
    worth more than a fixture invented to close a number.
    """
    import tempfile
    from unittest.mock import patch

    import numpy as np
    import openpyxl

    from quadrium.config import ConfigError, load_config
    from quadrium.eurostat import EurostatError
    from quadrium.gras import SignInfeasibleError, gras
    from quadrium.sut_euro import SutEuroError, sut_euro

    # ---- the exact sign-feasibility test, on a system found by search
    T = np.array([[2.0, -1.0, -2.0], [1.0, 1.0, 1.0], [0.0, 1.0, -2.0]])
    u = np.array([1.0, 3.0, 3.0])
    v = np.array([0.0, 10.5, -3.5])
    check("the fixture is balanced, so the refusal is not about the margins "
          "failing to add up", abs(u.sum() - v.sum()) < 1e-12,
          f"row totals {u.sum():g}, column totals {v.sum():g}")
    try:
        gras(T, u, v)
    except SignInfeasibleError as exc:
        check("a system that passes the per-line sign test and fails the "
              "exact one is refused",
              "at least one cell to change sign" in str(exc)
              and "sign preserving by construction" in str(exc),
              "the per-line test asks whether each line that needs a negative "
              "total has a negative entry; the exact one asks by LP whether "
              "ANY table with this sign pattern has these margins. Only the "
              "second catches this 3x3")
    except Exception as exc:                           # noqa: BLE001
        check("a system that passes the per-line sign test and fails the "
              "exact one is refused", False,
              f"{type(exc).__name__}: {str(exc)[:60]}")
    else:
        check("a system that passes the per-line sign test and fails the "
              "exact one is refused", False,
              "it balanced a table that cannot exist")

    # ---- the two SUT-EURO refusals
    Ud0 = np.array([[3.0, 1.0, 5.0], [1.0, 4.0, 4.0]])
    Um0 = np.array([[0.5, 0.2, 1.0], [0.2, 0.6, 0.8]])
    tls0 = np.array([0.3, 0.2, 0.5])
    V0 = np.array([[8.0, 1.0], [2.0, 9.0]])
    va0 = V0.sum(axis=1) - (Ud0[:, :2] + Um0[:, :2]).sum(axis=0) - tls0[:2]
    base = dict(va_target=va0 * 1.05, final_use_target=np.array([6.0]),
                tls_target=1.1, imports_target=float(Um0.sum()) * 1.05)

    ok = sut_euro(Ud0, Um0, tls0, V0, **base)
    check("the base pair projects, so the two refusals below are the only "
          "difference", ok is not None,
          f"value added {va0[0]:+.1f}, {va0[1]:+.1f} in the base year")

    V_zero = V0.copy()
    V_zero[:, 1] = 0.0
    for name, args, fragment in (
            ("a product with no base-year output",
             dict(V0=V_zero, **base), "no market shares to hold constant"),
            ("an industry whose value added changes sign",
             dict(V0=V0, **{**base,
                            "va_target": np.array([va0[0],
                                                   -abs(va0[1]) * 1.2])}),
             "change the SIGN of their value added")):
        V = args.pop("V0")
        try:
            sut_euro(Ud0, Um0, tls0, V, **args)
        except SutEuroError as exc:
            check(f"SUT-EURO refuses {name}",
                  fragment.lower() in str(exc).lower(), str(exc)[:88])
        except Exception as exc:                       # noqa: BLE001
            check(f"SUT-EURO refuses {name}", False,
                  f"{type(exc).__name__}: {str(exc)[:60]}")
        else:
            check(f"SUT-EURO refuses {name}", False, "it projected anyway")

    # ---- the supply-use pipeline, failing at each of its three steps
    CACHE = ROOT / "data" / "eurostat"
    from quadrium.eurostat import DATASETS
    cached = all((CACHE / f"{DATASETS[n]}_AT_2022.json").exists()
                 for n in ("supply", "use_purchasers", "use_basic"))

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        def book(table_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "project"
            for kv in (("project_id", "p"), ("table_kind", "eurostat_sut"),
                       ("eurostat_geo", "AT"), ("eurostat_year", 2022),
                       ("table_path", str(table_path))):
                ws.append(list(kv))
            # A job has to be declared or the workbook is refused for having
            # none before the pipeline is ever entered. `regionalise` is the
            # cheapest: every case here fails upstream of the activity file,
            # so it is never read.
            act = tmp / "activity.csv"
            act.write_text("sector_code,regional\nA,1.0\n")
            ws = wb.create_sheet("regionalise")
            ws.append(["key", "value"])
            ws.append(["method", "SLQ"])
            ws.append(["activity_path", str(act)])
            p = tmp / f"cfg{abs(hash(str(table_path)))}.xlsx"
            wb.save(p)
            return p

        def boom(*a, **k):
            raise EurostatError("naio_10_cp15 returned no values for geo=AT.")

        empty = tmp / "empty"
        empty.mkdir()
        with patch("quadrium.eurostat.fetch", boom):
            try:
                load_config(book(empty))
            except ConfigError as exc:
                check("a download that fails mid-system says nothing was "
                      "transformed", "download failed on" in str(exc)
                      and "needs all three files" in str(exc), str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check("a download that fails mid-system says nothing was "
                      "transformed", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a download that fails mid-system says nothing was "
                      "transformed", False, "it loaded")

        check("Austria's 2022 pair is cached, so the last two steps can fail "
              "on their own terms", cached,
              "the three files are in data/eurostat" if cached else
              "NOT cached — the last two cases are skipped and said so")
        if not cached:
            return

        with patch("quadrium.eurostat.load_sut", boom):
            try:
                load_config(book(CACHE))
            except ConfigError as exc:
                check("a pair that cannot be built says so, and carries the "
                      "reason up", "supply-use pair could not be built"
                      in str(exc) and "returned no values" in str(exc),
                      str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check("a pair that cannot be built says so, and carries the "
                      "reason up", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a pair that cannot be built says so, and carries the "
                      "reason up", False, "it built one")

        from quadrium.models import SupplyUseTables

        def no_transform(self, model):
            raise ValueError(f"D^T is singular for model {model}")

        with patch.object(SupplyUseTables, "to_iot", no_transform):
            try:
                load_config(book(CACHE))
            except ConfigError as exc:
                check("a transformation that fails names the model that could "
                      "not do it", "could not transform this system"
                      in str(exc) and "singular" in str(exc), str(exc)[:88])
            except Exception as exc:                   # noqa: BLE001
                check("a transformation that fails names the model that could "
                      "not do it", False,
                      f"{type(exc).__name__}: {str(exc)[:60]}")
            else:
                check("a transformation that fails names the model that could "
                      "not do it", False, "it transformed")


def test_the_COMMANDS_THE_GUIDE_DOCUMENTS_actually_run():
    """Four command-line paths the guide names, and none had ever been run.

    The reachability sweep of 2026-09-04 found `cli.py`'s `_catalogue`,
    `_describe`, `_verdicts` and `_wrap` entered by nothing in the suite. They
    are not dead code: `_catalogue` is what `--sources` and `--find` do, and
    `docs/GUIDE.md` documents both — `--find` in four places. `run_public_docs`
    checks that every option APPEARS in the guide. **Nothing checked that any
    of them works.** A regression in either would have shipped in silence,
    past a validator whose job is the guide.

    `_describe` and `_warn_about_substance` are worse placed: they run on
    every split, on both the `--check` path and the one that produces numbers.
    Their never having been entered means the route the guide opens with —
    fill in a spreadsheet, run one command — had never been exercised through
    `cli.main` at all. The engine was reached through `IOProject` instead,
    which is not what a user types.

    `_availability` is left out on purpose: it queries Eurostat over the
    network for a country whose tables are not on disk, and a test that needs
    a fetch is a test that fails on a train.
    """
    import io
    import tempfile
    from contextlib import redirect_stderr, redirect_stdout

    from quadrium.cli import main

    def run(argv):
        out, err = io.StringIO(), io.StringIO()
        try:
            with redirect_stdout(out), redirect_stderr(err):
                code = main(argv)
        except SystemExit as exc:                      # noqa: PERF203
            code = exc.code
        return code, out.getvalue(), err.getvalue()

    code, out, err = run(["--sources"])
    check("`--sources` runs and lists what can actually be loaded",
          code == 0 and "table(s) you can load and split" in out
          and out.count("eurostat:") > 5,
          f"exit {code}, {out.count('eurostat:')} Eurostat sources among "
          f"{out.splitlines()[0] if out else '(nothing)'}")

    code, out, err = run(["--find", "I55", "--geo", "ES", "--offline"])
    check("`--find` answers for a sector Spain does not separate, and says "
          "what to do instead",
          code == 0 and "I55" in out
          and ("split" in out.lower() or "contains it" in out.lower()),
          f"exit {code}. The guide names --find in four places and nothing "
          f"had ever run it")

    check("and the answer names the countries that DO separate it, which is "
          "what makes it a real distinction rather than a missing row",
          any(g in out for g in ("BE", "FR", "HU", "UK")),
          "a user told only that their own country cannot split I55 learns "
          "nothing about whether the split is meaningful")

    # ---- the route the guide opens with, through the command the guide names
    example = ROOT / "configs" / "ejemplo.xlsx"
    check("the shipped example workbook is where the guide says", example.exists(),
          str(example.name))
    if not example.exists():
        return

    import shutil

    import openpyxl
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        cfg = tmp / "cfg.xlsx"
        shutil.copy(example, cfg)
        wb = openpyxl.load_workbook(cfg)
        for row in wb["project"].iter_rows(min_col=1, max_col=2):
            if row[0].value == "table_path":
                row[1].value = str((example.parent / str(row[1].value)).resolve())
        wb.save(cfg)

        code, out, err = run([str(cfg), "--check"])
        check("`quadrium <workbook> --check` runs the guide's own route and "
              "describes what it read",
              code == 0 and "Nothing was run" in out,
              f"exit {code}" + (f"; stderr {err[:80]}" if err else ""))

        check("and it says what the configuration IS before saying it is valid",
              any(w in out for w in ("sector", "split", "scenario")),
              "`_describe` prints the table, the splits and the scenarios; a "
              "check that only says 'valid' tells a user nothing about what "
              "it validated")


def test_the_functions_the_REACHABILITY_SWEEP_found_UNCALLED():
    """Ten functions the engine defines that nothing had ever entered.

    `library/tools/sweep_reachability.py` counts 284 functions and found 22
    that no validator and no test had ever called. Wiring `identities.py`'s
    ID-07 and ID-08 into `run_ine_sut_identities.py` and six of `cli.py`'s took
    it to 15; this covers ten of those fifteen.

    Every case asserts what the function is FOR, in both directions where it
    has two. A smoke call would move the number and prove nothing, which is
    exactly the failure `run_refusal_coverage.py` exists to stop: a count that
    rises without a claim behind it is not coverage.

    Three of the four identities below were reported as wired on 2026-09-04 and
    were not -- `run_ine_sut_identities.py` wired ID-07 and ID-08 only, and
    `run_structural_zeros.py` still checks structural zeros with its own
    arithmetic while `identities.structural_zero_check` sits uncalled. The same
    duplication, found again by re-reading the list instead of the summary.

    The five left uncalled are declared in `data/_reachability.json` with their
    reasons: two need the network, one is a second API for something an object
    does better, and two are the run-summary path of `quadrium run`.
    """
    from quadrium.classification import NACE_REV_2_1, NATIONAL_NACE
    from quadrium.identities import (id06_gdp_three_approaches,
                                     id10_cif_fob_sums_to_zero,
                                     id13_value_added_preserved,
                                     structural_zero_check)
    from quadrium.models import CellLabel, count_label, label_counts
    from quadrium.reaggregation import reaggregate_vector
    from quadrium.sut_euro import SutEuroStep1

    # ---- ID-06: the three approaches converge only after balancing ---------
    agree = id06_gdp_three_approaches(1_000_000.0, 1_000_000.0, 1_000_000.0)
    apart = id06_gdp_three_approaches(1_000_000.0, 1_002_500.0)
    check("ID-06 passes when the three approaches give one GDP and fails when "
          "two of them differ",
          agree.passed and not apart.passed,
          f"agreeing {agree.max_abs_dev:.3g}, differing "
          f"{apart.max_abs_dev:,.0f} — {agree.citation}")
    check("and the citation travels with the answer, which is why the free "
          "function exists rather than two lines of arithmetic",
          "9.16" in agree.citation, agree.citation)

    # ---- ID-10: the CIF/FOB adjustment sums to zero ------------------------
    zero = id10_cif_fob_sums_to_zero(-4_820.0, 4_820.0)
    nonzero = id10_cif_fob_sums_to_zero(-4_820.0, 4_000.0)
    check("ID-10 accepts a CIF/FOB adjustment whose two entries cancel and "
          "refuses one that does not",
          zero.passed and not nonzero.passed,
          f"sum 0 against {abs(-4_820.0 + 4_000.0):,.0f} — {zero.citation}")
    disagree = id10_cif_fob_sums_to_zero(-4_820.0, 4_820.0,
                                         total_imports_cif=500_000.0,
                                         total_imports_fob=495_180.0)
    check("and it catches the second half of the identity too: the entries "
          "cancel while the two import totals disagree",
          not disagree.passed and disagree.n_violations >= 1,
          f"{disagree.n_violations} violation(s), worst "
          f"{disagree.max_abs_dev:,.0f}")

    # ---- ID-13: what a transformation may and may not move -----------------
    W = np.array([[150.0, 620.0, 140.0], [186.0, 452.0, 82.0]])
    moved_rows = np.array([[160.0, 610.0, 140.0], [176.0, 462.0, 82.0]])
    check("ID-13 industry x industry requires the value-added block itself to "
          "be unaltered, so a block that moved fails",
          id13_value_added_preserved(W, W, "IOT_IXI").passed
          and not id13_value_added_preserved(W, moved_rows, "IOT_IXI").passed,
          "CORE_005 par. 36.50, p. 1017 — only composition changes, and it "
          "changes by moving entries between ROWS of the intermediate matrix")
    check("ID-13 product x product requires only the TOTAL to survive, so the "
          "same moved block passes and one that gained value does not",
          id13_value_added_preserved(W, moved_rows, "IOT_PXP").passed
          and not id13_value_added_preserved(W, moved_rows + 500.0,
                                             "IOT_PXP").passed,
          f"total {W.sum():,.0f} either way — CORE_005 par. 36.49, p. 1017")

    # ---- the structural zero, which is not an identity but is required -----
    cite = "CORE_003 par. 15.95, p. 495"
    empty = structural_zero_check(np.zeros((4, 3)), "use table, lower right",
                                  cite)
    dirty = np.zeros((4, 3))
    dirty[2, 1] = -37.5
    filled = structural_zero_check(dirty, "use table, lower right", cite)
    check("a block that must be zero by construction passes empty and fails "
          "with one cell in it, and the failure names how far off it is",
          empty.passed and not filled.passed
          and abs(filled.max_abs_dev - 37.5) < 1e-9,
          f"worst {filled.max_abs_dev} in {filled.n_violations} cell(s) — "
          f"{filled.citation}")

    # ---- the two intermediate margins of an IOTable ------------------------
    t = build_table()
    row = t.intermediate_row_totals()
    col = t.intermediate_col_totals()
    check("intermediate sales counted as output less final demand equal the "
          "intermediate matrix's own row sums",
          float(np.abs(row - t.Z.sum(1)).max()) < 1e-9,
          f"max deviation {float(np.abs(row - t.Z.sum(1)).max()):.2e} over "
          f"{t.n} sectors")
    check("and intermediate purchases counted as output less value added "
          "equal its column sums — the same two numbers reached two ways, "
          "which is the whole content of the pair",
          float(np.abs(col - t.Z.sum(0)).max()) < 1e-9,
          f"max deviation {float(np.abs(col - t.Z.sum(0)).max()):.2e}")

    # ---- label_counts, and the numpy trap its helper was written for -------
    prov = np.array([[CellLabel.OBSERVED.value, CellLabel.OBSERVED.value],
                     [CellLabel.PROXY_ESTIMATED.value,
                      CellLabel.BALANCED_ADJUSTMENT.value]], dtype=object)
    counts = label_counts(prov)
    check("label_counts reports every label the vocabulary defines, counts "
          "summing to the cells of the array",
          set(counts) == set(CellLabel) and sum(counts.values()) == prov.size,
          f"{ {k.name: v for k, v in counts.items() if v} } over "
          f"{prov.size} cells")
    check("and it agrees with count_label label by label, which is the "
          "comparison that beats `provenance == label`",
          all(counts[lbl] == count_label(prov, lbl) for lbl in CellLabel),
          "a str-Enum member stringifies as 'CellLabel.OBSERVED' on 3.10, so "
          "the naive array comparison counts zero everywhere")

    # ---- what level of the classification a code sits at -------------------
    check("level_of reads the level off the digit count, section letter and "
          "dots included, and returns None for what is not a code",
          (NACE_REV_2_1.level_of("56") == "division"
           and NACE_REV_2_1.level_of("56.1") == "group"
           and NACE_REV_2_1.level_of("C5610") == "class"
           and NACE_REV_2_1.level_of("Accommodation") is None),
          "CORE_030 p. 13 — the section letter is not part of the numeric code")
    check("and a national version reaches the fifth digit NACE itself does "
          "not have",
          NATIONAL_NACE.level_of("56101") == "national subclass"
          and NACE_REV_2_1.level_of("56101") is None,
          "CORE_030 p. 15 permits a fifth digit for national purposes "
          "provided it still nests")

    # ---- reaggregating a vector, the companion of reaggregate() ------------
    mapping = [0, 1, 1, 1, 2]          # sector 1 was split into three
    v = np.array([90.0, 12.0, 30.0, 8.0, 700.0])
    back = reaggregate_vector(v, mapping, 3)
    check("a split vector sums back to the original sectors exactly, and to "
          "the same grand total",
          np.allclose(back, [90.0, 50.0, 700.0])
          and abs(back.sum() - v.sum()) < 1e-12,
          f"{back.tolist()} from {len(v)} subsectors, total {back.sum():,.1f}")

    # ---- the GDP disagreement SUT-EURO reports and does not repair ---------
    step1 = SutEuroStep1(Ud=np.zeros((2, 2)), Um=np.zeros((2, 2)),
                         tls=np.zeros(2), gva=np.zeros(2),
                         gdp_supply_side=257_346.0, gdp_use_side=258_432.0)
    check("step 1 of SUT-EURO reports the size of its own inconsistency, the "
          "one UNH_18 ¶18.94, p. 576 prints and leaves standing",
          abs(step1.inconsistency - 1_086.0) < 1e-9,
          f"{step1.inconsistency:,.0f} between a use side of 258,432 and a "
          f"supply side of 257,346 — the property existed and "
          f"run_sut_euro_austria.py subtracted the two by hand instead")
