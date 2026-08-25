"""
Export: Excel, CSV and JSON metadata (MVP_0.1 §1.13).

One rule governs every writer here: **a number never leaves without its data
status.** An exported spreadsheet is the artefact that outlives the run and gets
mailed around, so it is the last place where a `BALANCED` cell can quietly
become an observation. Every sheet that carries values has a sibling sheet or
column carrying provenance.
"""

from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from datetime import date, datetime
from enum import Enum
from pathlib import Path

import numpy as np

from .models import DATA_STATUS, CellLabel, DisaggregationResult, IOTable


class _Encoder(json.JSONEncoder):
    """numpy, datetimes, enums and dataclasses, none of which json knows."""

    def default(self, o):
        if isinstance(o, Enum):
            return o.value
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, (np.floating,)):
            return None if np.isnan(o) else float(o)
        if isinstance(o, (np.bool_,)):
            return bool(o)
        if is_dataclass(o) and not isinstance(o, type):
            return asdict(o)
        if isinstance(o, Path):
            return str(o)
        return super().default(o)


def _sanitize(o):
    """Replace non-finite floats with None, recursively.

    `json.dumps` writes bare `NaN` and `Infinity` by default. Neither is valid
    JSON: RFC 8259 has no such literals, and strict parsers — including
    `pandas.read_json`, most JavaScript, and `json.loads(..., parse_constant=…)`
    in other people's code — reject or mangle them. The engine really does
    produce them: `solver_step` is NaN on a RAS run, and technical coefficients
    are NaN wherever output is zero.

    `allow_nan=False` would raise instead, which turns a reporting detail into a
    failed run. Encoding them as null is the honest option: JSON has a way to
    say "no value" and this is it.
    """
    if isinstance(o, float):
        return None if (o != o or o in (float("inf"), float("-inf"))) else o
    if isinstance(o, np.floating):
        f = float(o)
        return None if (f != f or f in (float("inf"), float("-inf"))) else f
    if isinstance(o, dict):
        return {k: _sanitize(v) for k, v in o.items()}
    if isinstance(o, (list, tuple)):
        return [_sanitize(v) for v in o]
    if isinstance(o, np.ndarray):
        return _sanitize(o.tolist())
    if is_dataclass(o) and not isinstance(o, type):
        return _sanitize(asdict(o))
    return o


def write_json(obj, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(_sanitize(obj), cls=_Encoder, indent=2,
                      ensure_ascii=False, allow_nan=False)
    path.write_text(text, encoding="utf-8")
    return path


def write_table_csv(table: IOTable, path: Path,
                    provenance_note: str | None = None) -> Path:
    """The full system as one CSV: Z, final demand, value added, output.

    A COMMENT HEADER GOES FIRST, AND IT IS NOT DECORATION.
    The XLSX shades estimated cells and carries a README sheet; this file used
    to carry neither, just a grid of numbers in which an observed cell and an
    estimated one are indistinguishable. The provenance existed -- in a
    SEPARATE file, provenance.csv, which nobody opening "table_disaggregated"
    has any reason to look for (2026-08-10).

    Comment lines are prefixed `#`. Excel shows them as text rows at the top,
    which is exactly the visibility wanted, and `pandas.read_csv(..., comment="#")`
    skips them. Nothing in this project reads these files back.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    header = ["code", "label"] + table.sector_codes + list(table.Y_labels) + ["Output"]
    lines = []
    if provenance_note:
        lines += [f"# {ln}" for ln in provenance_note.splitlines()]
    lines.append(",".join(_q(h) for h in header))
    for i, code in enumerate(table.sector_codes):
        row = ([_q(code), _q(table.sector_labels[i])]
               + [f"{v:.6f}" for v in table.Z[i]]
               + [f"{v:.6f}" for v in table.Y[i]]
               + [f"{table.X[i]:.6f}"])
        lines.append(",".join(row))
    for m, lbl in enumerate(table.VA_labels):
        lines.append(",".join([_q(lbl), _q("")]
                              + [f"{v:.6f}" for v in table.VA[m]]
                              + [""] * len(table.Y_labels) + [""]))
    lines.append(",".join([_q("Output"), _q("")]
                          + [f"{v:.6f}" for v in table.X]
                          + [""] * len(table.Y_labels) + [""]))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_provenance_csv(res: DisaggregationResult, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    codes = res.table.sector_codes
    lines = [",".join(["code"] + [_q(c) for c in codes])]
    for i, code in enumerate(codes):
        lines.append(",".join([_q(code)]
                              + [str(res.provenance[i, j]).split(".")[-1].lower()
                                 if not isinstance(res.provenance[i, j], str)
                                 else res.provenance[i, j]
                                 for j in range(len(codes))]))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def _q(s) -> str:
    s = "" if s is None else str(s)
    return f'"{s}"' if ("," in s or '"' in s) else s


def write_xlsx(res: DisaggregationResult, path: Path) -> Path | None:
    """Workbook with one sheet per block, plus provenance and a README sheet.

    Returns None if openpyxl is unavailable rather than failing the run: the
    CSV and JSON outputs are the reproducible ones, the workbook is a
    convenience.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    t = res.table
    wb = openpyxl.Workbook()

    bold = Font(bold=True)
    warn = PatternFill("solid", fgColor="FFF2CC")     # estimated
    bal = PatternFill("solid", fgColor="DDEBF7")      # balanced

    ws = wb.active
    ws.title = "README"
    notes = [
        ("IO Model Foundry — disaggregated table", True),
        ("", False),
        (f"Table: {t.table_id}", False),
        (f"Country / year: {t.country} {t.year}", False),
        (f"Unit: {t.unit}", False),
        (f"Classification: {t.classification}", False),
        (f"Source: {t.source}", False),
        (f"Scenario: {res.scenario_id}", False),
        ("Sectors divided:", False),
    ] + [
        (f"  {s['sector_code']} -> {', '.join(s['new_codes'])}", False)
        for s in res.splits
    ] + [
        ("", False),
        ("HOW TO READ THE NUMBERS", True),
        ("Shaded cells are NOT observations.", False),
        ("  yellow = ESTIMATED (from a proxy; no direct measurement)", False),
        ("  blue   = BALANCED (moved by the solver to satisfy the identities)",
         False),
        ("Unshaded cells were copied from the original table unchanged.", False),
        ("", False),
        ("Solver convergence is necessary but NOT sufficient for statistical",
         False),
        ("validity (CORE_006 par. 9.51, p. 288). Read validation_report.json.",
         False),
        ("", False),
        ("Every tolerance in this pipeline is a project choice. No loaded", False),
        ("methodological source states a numerical accounting tolerance.", False),
    ]
    for r, (text, is_bold) in enumerate(notes, start=1):
        c = ws.cell(row=r, column=1, value=text)
        if is_bold:
            c.font = bold
    ws.column_dimensions["A"].width = 78

    prov = res.provenance
    ws = wb.create_sheet("Z")
    ws.cell(row=1, column=1, value="code").font = bold
    for j, code in enumerate(t.sector_codes):
        ws.cell(row=1, column=2 + j, value=code).font = bold
    for i, code in enumerate(t.sector_codes):
        ws.cell(row=2 + i, column=1, value=code).font = bold
        for j in range(len(t.sector_codes)):
            c = ws.cell(row=2 + i, column=2 + j, value=float(t.Z[i, j]))
            c.number_format = "#,##0.00"
            label = prov[i, j]
            if label == CellLabel.PROXY_ESTIMATED.value or \
               label is CellLabel.PROXY_ESTIMATED:
                c.fill = warn
            elif label == CellLabel.BALANCED_ADJUSTMENT.value or \
                    label is CellLabel.BALANCED_ADJUSTMENT:
                c.fill = bal
    ws.freeze_panes = "B2"

    for name, arr, rowlabels, collabels in (
            ("FinalDemand", t.Y, t.sector_codes, t.Y_labels),
            ("ValueAdded", t.VA, t.VA_labels, t.sector_codes)):
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value="").font = bold
        for j, cl in enumerate(collabels):
            ws.cell(row=1, column=2 + j, value=str(cl)).font = bold
        for i, rl in enumerate(rowlabels):
            ws.cell(row=2 + i, column=1, value=str(rl)).font = bold
            for j in range(len(collabels)):
                c = ws.cell(row=2 + i, column=2 + j, value=float(arr[i, j]))
                c.number_format = "#,##0.00"
        ws.freeze_panes = "B2"

    ws = wb.create_sheet("Output")
    ws.cell(row=1, column=1, value="code").font = bold
    ws.cell(row=1, column=2, value="label").font = bold
    ws.cell(row=1, column=3, value="output").font = bold
    for i, code in enumerate(t.sector_codes):
        ws.cell(row=2 + i, column=1, value=code)
        ws.cell(row=2 + i, column=2, value=t.sector_labels[i])
        ws.cell(row=2 + i, column=3, value=float(t.X[i])).number_format = "#,##0.00"
    ws.column_dimensions["B"].width = 45

    ws = wb.create_sheet("Provenance")
    ws.cell(row=1, column=1, value="code").font = bold
    for j, code in enumerate(t.sector_codes):
        ws.cell(row=1, column=2 + j, value=code).font = bold
    for i, code in enumerate(t.sector_codes):
        ws.cell(row=2 + i, column=1, value=code).font = bold
        for j in range(len(t.sector_codes)):
            lab = prov[i, j]
            lab = lab if isinstance(lab, str) else lab.value
            c = ws.cell(row=2 + i, column=2 + j, value=DATA_STATUS[CellLabel(lab)])
            c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"

    wb.save(path)
    return path
